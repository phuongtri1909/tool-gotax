import requests,random
import os,time,shutil,io,base64,json
import zipfile
import logging
from datetime import datetime, timedelta
from .base_service import BaseService
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
# Lazy import playwright - chỉ import khi cần dùng
# from playwright.sync_api import sync_playwright

# ✅ Thiết lập logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Tạo log file
log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'logs')
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(log_dir, f'invoice_process_{datetime.now().strftime("%Y%m%d")}.log')
file_handler = logging.FileHandler(log_file, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)

# Format log
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

class BackendService(BaseService):
    def __init__(self, proxy_url=None, job_id=None):
        super().__init__(proxy_url=proxy_url)
        self.proxy_url = proxy_url  # ✅ Lưu proxy URL để recreate session
        self.job_id = job_id  # ✅ Lưu job_id để check cancelled flag
    
    def _check_cancelled(self):
        """Check if job is cancelled from Redis"""
        if not self.job_id:
            return False
        
        try:
            import sys
            import os
            import time
            sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            from shared.redis_client import get_redis_client
            
            redis_client = get_redis_client()
            
            # ✅ Check cancelled flag
            cancelled = redis_client.get(f"job:{self.job_id}:cancelled")
            if cancelled:
                cancelled = cancelled.decode('utf-8') if isinstance(cancelled, bytes) else str(cancelled).strip()
                if cancelled == '1':
                    return True
            
            # ✅ Check status
            status = redis_client.get(f"job:{self.job_id}:status")
            if status:
                status = status.decode('utf-8') if isinstance(status, bytes) else str(status).strip()
                if status == 'cancelled':
                    return True
            
            # ✅ Check nếu client đã disconnect (không poll trong 10 giây)
            # Chỉ check nếu job đang processing
            if status == 'processing':
                last_poll_time = redis_client.get(f"job:{self.job_id}:last_poll_time")
                if last_poll_time is not None:
                    last_poll_time = int(last_poll_time) if isinstance(last_poll_time, bytes) else int(last_poll_time)
                    current_time = int(time.time())
                    time_since_last_poll = current_time - last_poll_time
                    
                    # Nếu không có poll trong 10 giây → client đã reload/đóng tab → auto cancel
                    if time_since_last_poll > 10:
                        try:
                            redis_client.set(f"job:{self.job_id}:cancelled", "1")
                            redis_client.set(f"job:{self.job_id}:status", "cancelled")
                            
                            # Publish progress message
                            progress_data = {
                                'percent': 0,
                                'message': 'Client đã đóng kết nối - Job đã bị hủy',
                                'data': []
                            }
                            import json
                            redis_client.rpush(f"job:{self.job_id}:progress:list", json.dumps(progress_data, ensure_ascii=False).encode('utf-8'))
                            
                            return True
                        except:
                            pass
        except Exception as e:
            # Silent error - nếu không check được thì tiếp tục
            pass
        
        return False
    
    def _safe_get(self, url, **kwargs):
        """
        Wrapper cho session.get() với check cancelled flag
        Check cancelled flag trước và sau request
        """
        # Check cancelled flag trước khi gọi request
        if self._check_cancelled():
            raise Exception("Job đã bị hủy (Ctrl+C)")
        
        # Gọi request
        response = self.session.get(url, **kwargs)
        
        # Check cancelled flag sau khi request hoàn thành
        if self._check_cancelled():
            raise Exception("Job đã bị hủy (Ctrl+C)")
        
        return response
    
    def _recreate_session_with_new_proxy(self):
        """
        ✅ Tạo session mới + add proxy lại (IP tự đổi)
        Thay vì delay/backoff khi 429
        """
        print(f"🔄 Recreating session + rotating proxy IP...")
        # ✅ Tạo session mới bằng requests.Session()
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "PostmanRuntime/7.43.4",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "vi-VN,vi;q=0.9",
            "Connection": "close"
        })
        # Add proxy lại (Luna Proxy tự đổi IP)
        if self.proxy_url:
            self.session.proxies = {
                'http': self.proxy_url
            }
        
        return self.session
    
    def day_split(self,start_date, end_date):
        date_format = "%d/%m/%Y" 
        date1 = datetime.strptime(start_date, date_format)  
        date2 = datetime.strptime(end_date, date_format) 
        one_month = timedelta(days=27)  
        date_ranges = []  
        while date1 <= date2:
            sub_array = []  
            sub_array.append(date1.strftime(date_format))  
            date1 += one_month 
            if date1 > date2:
                date1 = date2 
            sub_array.append(date1.strftime(date_format)) 
            date_ranges.append(sub_array) 
            date1 += timedelta(days=1)
        return date_ranges
    def remove_duplicate_elements(self,data):
        seen_elements = set()
        unique_elements = []
        for element in data:
            element_json = json.dumps(element, sort_keys=True)
            if element_json not in seen_elements:
                seen_elements.add(element_json)
                unique_elements.append(element)
        unique_json_array = unique_elements
        return unique_json_array
    def increase_date(self,date_string):
        try:
            date = datetime.strptime(date_string, "%d/%m/%Y")
            increased_date = date + timedelta(days=1)
            increased_date_string = increased_date.strftime("%d/%m/%Y")
            return increased_date_string
        except ValueError:
            return "Định dạng ngày không hợp lệ!"
    def tongquat_(self,type_invoice:int = 0,headers: dict = {},start_date:str = "",end_date:str = "",progress_callback=None):
        tout = 15
        self.arr_ed = self.day_split(start_date,end_date)
        self.progress_callback = progress_callback  # Lưu callback để sử dụng sau
        # Use os.path.join for cross-platform path
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_thongke_a = os.path.join(base_dir, '__pycache__', 'template', 'Thống kê tổng quát.xlsx')
        # Load workbook mẫu
        wb = load_workbook(excel_thongke_a)
        sheet = wb.active
        br = 1
        if type_invoice == 1:
            br = 2
            type_hoadon = 'sold'
        d = ""
        e = ""
        type_list = {""}
        if type_invoice == 2:
            type_list = {'5','6','8'}
            d = f';ttxly=='
            type_hoadon = 'purchase'
        if 1:
            datas_first = ""
            count = len(self.arr_ed)
            start_index = 0  # ✅ Bắt đầu từ 0 để hiển thị 0/0 0% ban đầu
            
            # ✅ Gọi progress_callback ban đầu với 0/0 0%
            if self.progress_callback:
                self.progress_callback(
                    current_step="Đang khởi tạo...",
                    processed=0,
                    total=count
                )
            
            for self.i in range(len(self.arr_ed)):
                # ✅ Check cancelled flag trước khi xử lý mỗi month range
                if self._check_cancelled():
                    raise Exception("Job đã bị hủy (Ctrl+C)")
                
                # ✅ Tăng start_index trước khi gọi callback
                start_index += 1
                
                if self.progress_callback:
                    self.progress_callback(
                        current_step=f"Đang xử lý tháng {start_index}/{count}...",
                        processed=start_index,
                        total=count
                    )
                #----------------------------------------
                begin_day = self.arr_ed[self.i][0]
                end_day = self.arr_ed[self.i][1]
                spec = ""
                type_list_array = list(type_list)  # Convert set to list để có thể track index
                for type_idx, e in enumerate(type_list_array):
                    # ✅ Check cancelled flag trong vòng lặp type_list
                    if self._check_cancelled():
                        raise Exception("Job đã bị hủy (Ctrl+C)")
                    
                    # ✅ Thêm delay giữa các type để tránh rate limiting (trừ type đầu tiên)
                    if type_idx > 0:
                        delay_between_types = 2.0  # Delay 2 giây giữa các type
                        logger.info(f" ⏳ Delay {delay_between_types}s giữa các type để tránh rate limiting...")
                        time.sleep(delay_between_types)
                    
                    for i in range(br):
                        # ✅ Check cancelled flag trong vòng lặp br
                        if self._check_cancelled():
                            raise Exception("Job đã bị hủy (Ctrl+C)")
                        
                        if e == "8":
                            spec = "sco-"
                        else:
                            spec = ""
                        if i == 1:
                                    spec = "sco-"
                        j = 0
                        max_retries = 10  # ✅ Max retries cho mỗi request
                        retry_delay = 1.0  # ✅ Initial delay (seconds)
                        max_delay = 30.0  # ✅ Max delay (seconds)
                        
                        while j < max_retries:
                            # ✅ Check cancelled flag trong vòng lặp retry
                            if self._check_cancelled():
                                raise Exception("Job đã bị hủy (Ctrl+C)")
                            
                            try:
                                url = f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/{type_hoadon}?sort=tdlap:desc&size=50&search=tdlap=ge={begin_day}T00:00:00;tdlap=le={end_day}T23:59:59{d}{e}'
                                res = self.session.get(
                                    url,
                                    headers=headers,
                                    verify=False,timeout=tout
                                )
                                # ✅ Logging tiến trình
                                logger.info(f" Fetching invoices (initial) | Type: {type_hoadon} | Status: {res.status_code} | Response size: {len(res.content)} bytes | Period: {begin_day} to {end_day} | Retry: {j+1}/{max_retries}")
                                if res.status_code == 200:
                                    logger.info(f" Successfully fetched invoices for period: {begin_day} to {end_day}")
                                    # ✅ Thêm delay nhỏ sau khi fetch thành công để tránh rate limiting
                                    time.sleep(0.5)
                                    break
                                # ✅ Xử lý 429: Tạo session mới + rotate IP
                                elif res.status_code == 429:
                                    logger.warning(f" 429 Too Many Requests - Rotating IP and retrying... | Period: {begin_day} to {end_day} | Retry: {j+1}/{max_retries}")
                                    print(f"⚠️ 429 Too Many Requests - Creating new session with rotated IP...")
                                    self._recreate_session_with_new_proxy()
                                    
                                    # ✅ Check cancelled flag sau khi recreate session
                                    if self._check_cancelled():
                                        raise Exception("Job đã bị hủy (Ctrl+C)")
                                    
                                    # ✅ Delay nhỏ trước khi retry (0.5s)
                                    time.sleep(0.5)
                                    j += 1
                                    continue
                                # ✅ Xử lý 503: Service Unavailable - cần delay lâu hơn
                                elif res.status_code == 503:
                                    logger.warning(f" 503 Service Unavailable - Retrying with delay... | Period: {begin_day} to {end_day} | Retry: {j+1}/{max_retries}")
                                    print({
                                        "status": "error",
                                        "status_code": 503,
                                        "message": f"Service Unavailable (503) - Retrying in {retry_delay:.1f}s...",
                                        "data": {},
                                        "RetryAfter": retry_delay,
                                        "retry_count": j+1,
                                        "max_retries": max_retries
                                    })
                                    
                                    # ✅ Recreate session với proxy mới
                                    self._recreate_session_with_new_proxy()
                                    
                                    # ✅ Check cancelled flag sau khi recreate session
                                    if self._check_cancelled():
                                        raise Exception("Job đã bị hủy (Ctrl+C)")
                                    
                                    # ✅ Exponential backoff: delay tăng dần (1s, 2s, 4s, 8s, ... max 30s)
                                    time.sleep(min(retry_delay, max_delay))
                                    retry_delay = min(retry_delay * 2, max_delay)  # Double delay, max 30s
                                    j += 1
                                    continue
                                else:
                                    # Log response body khi loi de biet GDT tra ve gi
                                    try:
                                        body = res.text if getattr(res, "text", None) else (res.content or b"").decode("utf-8", errors="replace")
                                        logger.warning(" API %s body: %s", res.status_code, (body[:300] if body else ""))
                                    except Exception:
                                        pass
                                    logger.warning(f" Error {res.status_code} - Retrying... | Period: {begin_day} to {end_day} | Retry: {j+1}/{max_retries}")
                                    print({
                                        "status": "error",
                                        "status_code": res.status_code,
                                        "message": f"Lỗi khi gọi API: {res.status_code}",
                                        "data": {},
                                        "RetryAfter": retry_delay,
                                        "retry_count": j+1,
                                        "max_retries": max_retries
                                    })
                                    # ✅ Handle các error khác bằng recreate session
                                    self._recreate_session_with_new_proxy()
                                    
                                    # ✅ Check cancelled flag sau khi recreate session
                                    if self._check_cancelled():
                                        raise Exception("Job đã bị hủy (Ctrl+C)")
                                    
                                    if res.status_code == 401:
                                        return {
                                            "status": "error",
                                            "status_code": 401,
                                            "message": "Phiên đăng nhập đã hết hạn. Vui lòng đăng nhập lại.",
                                            "data": {}
                                        }
                                    
                                    # ✅ Delay trước khi retry (exponential backoff)
                                    time.sleep(min(retry_delay, max_delay))
                                    retry_delay = min(retry_delay * 1.5, max_delay)  # Increase delay
                                    j += 1
                                    continue

                            except Exception as ex:
                                # ✅ Check cancelled flag trong exception handler
                                if "Job đã bị hủy" in str(ex) or self._check_cancelled():
                                    raise Exception("Job đã bị hủy (Ctrl+C)")
                                
                                j += 1
                                logger.error(f"❌ Exception (retry {j}/{max_retries}): {ex}")
                                print(f"❌ Exception: {ex} | Retry: {j}/{max_retries}")
                                self._recreate_session_with_new_proxy()
                                
                                # ✅ Check cancelled flag sau khi recreate session
                                if self._check_cancelled():
                                    raise Exception("Job đã bị hủy (Ctrl+C)")
                                
                                # ✅ Delay trước khi retry exception
                                if j < max_retries:
                                    time.sleep(min(retry_delay, max_delay))
                                    retry_delay = min(retry_delay * 1.5, max_delay)
                        
                        # ✅ Nếu đã retry hết mà vẫn fail
                        if j >= max_retries:
                            error_msg = f"Không thể fetch invoices sau {max_retries} lần thử | Period: {begin_day} to {end_day}"
                            logger.error(error_msg)
                            print(f"❌ {error_msg}")
                            # ✅ Skip period này và tiếp tục period tiếp theo (không fail toàn bộ job)
                            # Return empty data để tiếp tục
                            data = {"datas": []}
                            if datas_first == "":
                                datas_first = data
                            else:
                                if "datas" not in datas_first:
                                    datas_first["datas"] = []
                                datas_first["datas"].extend(data["datas"])
                            print(f"⚠️ Đã skip period {begin_day} to {end_day} do lỗi liên tục")
                            continue
                        
                        try:
                            data = res.json()
                            if not isinstance(data, dict):
                                logger.error(f"❌ Response không phải dict: {type(data)} | Period: {begin_day} to {end_day}")
                                data = {"datas": []}
                                if datas_first == "":
                                    datas_first = data
                                else:
                                    if "datas" not in datas_first:
                                        datas_first["datas"] = []
                                    datas_first["datas"].extend(data["datas"])
                                continue
                        except Exception as json_error:
                            logger.error(f"❌ Lỗi parse JSON: {json_error} | Period: {begin_day} to {end_day}")
                            data = {"datas": []}
                            if datas_first == "":
                                datas_first = data
                            else:
                                if "datas" not in datas_first:
                                    datas_first["datas"] = []
                                datas_first["datas"].extend(data["datas"])
                            continue
                        
                        if not isinstance(data.get("datas"), list):
                            logger.warning(f"⚠️ Response không có 'datas' hoặc không phải list | Period: {begin_day} to {end_day}")
                            if "datas" not in data:
                                data["datas"] = []
                        
                        if datas_first == "":
                            if "datas" not in data:
                                data["datas"] = []
                            if not isinstance(data.get("datas"), list):
                                data["datas"] = []
                            datas_first = data
                        else:
                            if "datas" not in datas_first:
                                datas_first["datas"] = []
                            if isinstance(data.get("datas"), list):
                                datas_first["datas"].extend(data["datas"])
                        
                        time.sleep(0.3)
                        while True:
                            if self._check_cancelled():
                                raise Exception("Job đã bị hủy (Ctrl+C)")
                            
                            if data["state"] != None and data != "":
                                j = 0
                                max_retries_state = 10
                                retry_delay_state = 1.0
                                max_delay_state = 30.0
                                
                                while j < max_retries_state:
                                    if self._check_cancelled():
                                        raise Exception("Job đã bị hủy (Ctrl+C)")
                                    
                                    try:
                                        res = self.session.get(f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/{type_hoadon}?sort=tdlap:desc&size=50&state={data["state"]}&search=tdlap=ge={begin_day}T00:00:00;tdlap=le={end_day}T23:59:59',headers=headers,verify=False,timeout=tout)
                                        # ✅ Logging tiến trình
                                        logger.info(f" Fetching invoices with state={data['state']} | Status: {res.status_code} | Response size: {len(res.content)} bytes | Period: {begin_day} to {end_day} | Retry: {j+1}/{max_retries_state}")
                                        if res.status_code == 200:
                                            # ✅ Parse JSON sau khi break khỏi retry loop
                                            try:
                                                pagination_data = res.json()
                                                # ✅ Kiểm tra pagination_data có phải là dict không
                                                if not isinstance(pagination_data, dict):
                                                    logger.error(f"❌ Pagination response không phải dict: {type(pagination_data)} | Period: {begin_day} to {end_day}")
                                                    break
                                                if "state" in pagination_data:
                                                    # ✅ Đảm bảo pagination_data["datas"] là list
                                                    if "datas" not in pagination_data:
                                                        pagination_data["datas"] = []
                                                    if not isinstance(pagination_data.get("datas"), list):
                                                        pagination_data["datas"] = []
                                                    # ✅ Đảm bảo datas_first có "datas" key
                                                    if "datas" not in datas_first:
                                                        datas_first["datas"] = []
                                                    datas_first["datas"].extend(pagination_data["datas"])
                                                else:
                                                    break
                                            except Exception as json_error:
                                                logger.error(f"❌ Lỗi parse JSON (pagination): {json_error} | Period: {begin_day} to {end_day}")
                                                break
                                            break
                                        # ✅ Xử lý 429: Tạo session mới + rotate IP
                                        elif res.status_code == 429:
                                            logger.warning(f" 429 Too Many Requests (state pagination) - Rotating IP | State: {data['state']} | Period: {begin_day} to {end_day} | Retry: {j+1}/{max_retries_state}")
                                            print(f"⚠️ 429 Too Many Requests - Creating new session with rotated IP...")
                                            self._recreate_session_with_new_proxy()
                                            
                                            # ✅ Check cancelled flag sau khi recreate session
                                            if self._check_cancelled():
                                                raise Exception("Job đã bị hủy (Ctrl+C)")
                                            
                                            time.sleep(0.5)
                                            j += 1
                                            continue
                                        # ✅ Xử lý 503: Service Unavailable
                                        elif res.status_code == 503:
                                            logger.warning(f" 503 Service Unavailable (state pagination) - Retrying... | State: {data['state']} | Period: {begin_day} to {end_day} | Retry: {j+1}/{max_retries_state}")
                                            self._recreate_session_with_new_proxy()
                                            if self._check_cancelled():
                                                raise Exception("Job đã bị hủy (Ctrl+C)")
                                            time.sleep(min(retry_delay_state, max_delay_state))
                                            retry_delay_state = min(retry_delay_state * 2, max_delay_state)
                                            j += 1
                                            continue
                                        else:
                                            logger.warning(f" Error {res.status_code} (state pagination) - Retrying... | Retry: {j+1}/{max_retries_state}")
                                            self._recreate_session_with_new_proxy()
                                            
                                            # ✅ Check cancelled flag sau khi recreate session
                                            if self._check_cancelled():
                                                raise Exception("Job đã bị hủy (Ctrl+C)")
                                            
                                            time.sleep(min(retry_delay_state, max_delay_state))
                                            retry_delay_state = min(retry_delay_state * 1.5, max_delay_state)
                                            j += 1
                                            continue

                                    except Exception as e:
                                        # ✅ Check cancelled flag trong exception handler
                                        if "Job đã bị hủy" in str(e) or self._check_cancelled():
                                            raise Exception("Job đã bị hủy (Ctrl+C)")
                                        
                                        j += 1
                                        logger.error(f" Exception (state pagination, retry {j}/{max_retries_state}): {e}")
                                        
                                        # ✅ Check cancelled flag sau mỗi retry
                                        if self._check_cancelled():
                                            raise Exception("Job đã bị hủy (Ctrl+C)")
                                        
                                        if j < max_retries_state:
                                            self._recreate_session_with_new_proxy()
                                            time.sleep(min(retry_delay_state, max_delay_state))
                                            retry_delay_state = min(retry_delay_state * 1.5, max_delay_state)
                                
                                # ✅ Nếu retry hết mà vẫn fail, break khỏi state pagination
                                if j >= max_retries_state:
                                    logger.warning(f" Max retries reached for state pagination, breaking...")
                                    break
                                
                                    try:
                                        data = res.json()
                                        # ✅ Kiểm tra data có phải là dict không
                                        if not isinstance(data, dict):
                                            logger.error(f"❌ State pagination response không phải dict: {type(data)} | Period: {begin_day} to {end_day}")
                                            break
                                        if "state" in data:
                                            # ✅ Đảm bảo data["datas"] là list
                                            if "datas" not in data:
                                                data["datas"] = []
                                            if not isinstance(data.get("datas"), list):
                                                data["datas"] = []
                                            # ✅ Đảm bảo datas_first có "datas" key
                                            if "datas" not in datas_first:
                                                datas_first["datas"] = []
                                            datas_first["datas"].extend(data["datas"])
                                        else:
                                            break
                                    except Exception as json_error:
                                        logger.error(f"❌ Lỗi parse JSON (state pagination): {json_error} | Period: {begin_day} to {end_day}")
                                        break
                            else:
                                break
                        print(f"       [ ĐÃ XỬ LÝ XONG TỪ NGÀY {begin_day} ĐẾN NGÀY {end_day} ]")
                
                # ✅ Thêm delay giữa các period để tránh rate limiting (trừ period cuối)
                if self.i < len(self.arr_ed) - 1:
                    delay_between_periods = 1.0  # Delay 1 giây giữa các period
                    time.sleep(delay_between_periods)
            
            datas_first["datas"] = self.remove_duplicate_elements(datas_first["datas"])
            
            if type_invoice == 1:
                nm = "nmmst"
                nmten = "nmten"
            elif type_invoice == 2:
                nm = "nbmst"
                nmten = "nbten"
            headers_w = ["khmshdon", "khhdon", "shdon", "ntao", nm, nmten, "tgtcthue", "tgtthue", "ttcktmai","", "tgtttbso", "dvtte", "tthai", "ttxly"]
            self.a = 0
            try:
                n_range = 100/count
            except:
                print('Không có hóa đơn !')
                return {
                    "status": "error",
                    "status_code": 404,
                    "message": "Không có hóa đơn !",
                    "data": {}
                }
            hdon = {1: "Hóa đơn mới", 2: "Hóa đơn thay thế", 3: "Hóa đơn điều chỉnh", 4: "Hóa đơn đã bị thay thế", 5: "Hóa đơn đã bị điều chỉnh", 6: "Hóa đơn đã bị hủy"}
            ttxly = {0: "Tổng cục Thuế đã nhận", 1: "Đang tiến hành kiểm tra điều kiện cấp mã", 2: "CQT từ chối hóa đơn theo từng lần phát sinh", 3: "Hóa đơn đủ điều kiện cấp mã", 4: "Hóa đơn không đủ điều kiện cấp mã", 5: "Đã cấp mã hóa đơn", 6: "Tổng cục thuế đã nhận không mã", 7: "Đã kiểm tra định kỳ HĐĐT không có mã", 8: "Tổng cục thuế đã nhận hóa đơn có mã khởi tạo từ máy tính tiền"}
            spec = ""
            data_crawled = []  # Danh sách để lưu trữ dữ liệu JSON
            for data in datas_first["datas"]:
                # ✅ Check cancelled flag trước khi xử lý mỗi invoice
                if self._check_cancelled():
                    raise Exception("Job đã bị hủy (Ctrl+C)")
                
                last_row = sheet.max_row
                if data["ttxly"]== 8:
                    spec = "sco-"
                else:
                    spec = ""
                u = 0
                values = [data.get(header, "") for header in headers_w]
                values.insert(0, start_index)
                row_index = last_row + 1  # Hàng mới để ghi dữ liệu
                hdon_value = values[headers_w.index("tthai")+1]
                ttxly_value = values[headers_w.index("ttxly")+1]
                type_ = values[headers_w.index("khmshdon")+1]
                if hdon_value in hdon:
                    values[headers_w.index("tthai")+1] = hdon[hdon_value]
                if ttxly_value in ttxly:
                    values[headers_w.index("ttxly")+1] = ttxly[ttxly_value]
                if type_ == 2:
                    nbmst = data["nbmst"]
                    khhdon = data["khhdon"]
                    shd = data["shdon"]
                    detail_retry_count = 0
                    max_detail_retries = 5  # ✅ Max retries cho detail fetching (ít hơn vì chỉ là detail)
                    retry_delay_detail = 1.0
                    max_delay_detail = 15.0
                    
                    while detail_retry_count < max_detail_retries:
                        # ✅ Check cancelled flag trong vòng lặp detail fetching
                        if self._check_cancelled():
                            raise Exception("Job đã bị hủy (Ctrl+C)")
                        
                        try:    
                            link = f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/detail?nbmst={nbmst}&khhdon={khhdon}&shdon={shd}&khmshdon=2'          
                            res1 = self.session.get(link,headers=headers,verify=False,timeout=tout)
                            if res1.status_code == 200:
                                break
                            # ✅ Xử lý 429: Tạo session mới + rotate IP
                            elif res1.status_code == 429:
                                logger.warning(f" 429 Too Many Requests (detail) - Retrying... | Retry: {detail_retry_count+1}/{max_detail_retries}")
                                print(f"⚠️ 429 Too Many Requests - Creating new session with rotated IP...")
                                self._recreate_session_with_new_proxy()
                                
                                # ✅ Check cancelled flag sau khi recreate session
                                if self._check_cancelled():
                                    raise Exception("Job đã bị hủy (Ctrl+C)")
                                
                                time.sleep(0.5)
                                detail_retry_count += 1
                                continue
                            # ✅ Xử lý 503: Service Unavailable
                            elif res1.status_code == 503:
                                logger.warning(f" 503 Service Unavailable (detail) - Retrying... | Retry: {detail_retry_count+1}/{max_detail_retries}")
                                self._recreate_session_with_new_proxy()
                                if self._check_cancelled():
                                    raise Exception("Job đã bị hủy (Ctrl+C)")
                                time.sleep(min(retry_delay_detail, max_delay_detail))
                                retry_delay_detail = min(retry_delay_detail * 2, max_delay_detail)
                                detail_retry_count += 1
                                continue
                            else:
                                logger.warning(f" Error {res1.status_code} (detail) - Retrying... | Retry: {detail_retry_count+1}/{max_detail_retries}")
                                self._recreate_session_with_new_proxy()
                                
                                # ✅ Check cancelled flag sau khi recreate session
                                if self._check_cancelled():
                                    raise Exception("Job đã bị hủy (Ctrl+C)")
                                
                                print({
                                    "status": "error",
                                    "status_code": res1.status_code,
                                    "message": f"Lỗi khi gọi API: {res1.status_code}",
                                    "retry_count": detail_retry_count+1,
                                    "max_retries": max_detail_retries
                                })
                                time.sleep(min(retry_delay_detail, max_delay_detail))
                                retry_delay_detail = min(retry_delay_detail * 1.5, max_delay_detail)
                                detail_retry_count += 1
                                continue
                        except Exception as e:
                            # ✅ Check cancelled flag trong exception handler
                            if "Job đã bị hủy" in str(e) or self._check_cancelled():
                                raise Exception("Job đã bị hủy (Ctrl+C)")
                            
                            detail_retry_count += 1
                            logger.error(f" Exception (detail, retry {detail_retry_count}/{max_detail_retries}): {e}")
                            self._recreate_session_with_new_proxy()
                            
                            # ✅ Check cancelled flag sau khi recreate session
                            if self._check_cancelled():
                                raise Exception("Job đã bị hủy (Ctrl+C)")
                            
                            print({
                                "status": "error",
                                "message": str(e),
                                "message_detail": f"Lỗi khi lấy thêm chi tiết của hóa đơn khi tải tổng quát : {e}",
                                "retry_count": detail_retry_count,
                                "max_retries": max_detail_retries
                            })
                            
                            if detail_retry_count < max_detail_retries:
                                time.sleep(min(retry_delay_detail, max_delay_detail))
                                retry_delay_detail = min(retry_delay_detail * 1.5, max_delay_detail)
                    
                    # ✅ Nếu retry hết mà vẫn fail, skip detail này và tiếp tục
                    if detail_retry_count >= max_detail_retries:
                        logger.warning(f" Max retries reached for detail fetching, skipping this invoice detail...")
                        # Continue với invoice tiếp theo (không fail toàn bộ)
                        continue
                    
                    try:
                        data1 = res1.json()
                        # ✅ Kiểm tra data1 có phải là dict không
                        if not isinstance(data1, dict):
                            logger.error(f"❌ Detail response không phải dict: {type(data1)}")
                            print("lỗi nè")
                            u = 1
                            pass
                        else:
                            sum = 0
                            for dataz in data1.get('hdhhdvu', []):
                                if dataz["thtien"] == None :
                                    sum+=0
                                else:
                                    sum+=dataz["thtien"]
                            values[headers_w.index("tgtcthue")+1] = sum
                            print("3")
                    except:
                        print("lỗi nè")
                        u = 1
                        pass
                new_p = 0
                try:
                    for i in data["thttlphi"]:
                        new_p += i["tphi"]
                    values[10] = new_p
                except:
                    pass
                if values[6] == None:
                    try:
                        values[6]=data["nmtnmua"]
                    except:
                        pass
                    try:
                        values[6]=data["nbtnmua"]
                    except:
                        pass
                if u != 1:
                    for column_index, value in enumerate(values, start=1):

                        if column_index == 5:
                            try:
                                value = data["tdlap"]
                                value = value.split("T")[0]
                                new_value = value.split("-")
                                value = new_value[2] + "/" + new_value[1] + "/" + new_value[0]
                                value = self.increase_date(value)
                            except Exception as e:
                                print({
                                    "status": "error",
                                    "message": f"Lỗi định dạng ngày tháng: {e}",
                                    "data": {}
                                })
                                pass
                        sheet.cell(row=row_index, column=column_index, value=value)
                        #Ký hiệu mẫu số	Ký hiệu hóa đơn	Số hóa đơn	Ngày lập	MST người mua/MST người nhận hàng	Tên người mua/Tên người nhận hàng	Tổng tiền chưa thuế	Tổng tiền thuế	Tổng tiền chiết khấu thương mại	Tổng tiền phí	Tổng tiền thanh toán	Đơn vị tiền tệ	Trạng thái hóa đơn	Kết quả kiểm tra hóa đơn

                    # Thêm dữ liệu vào JSON
                    try:
                        nlap = values[4].split("-")
                        nlap = nlap[2] + "/" + nlap[1] + "/" + nlap[0]
                    except:
                        print("Can't format date")
                    json_record = {
                        "STT": values[0],
                        "Ký hiệu mẫu số": values[1],
                        "Ký hiệu hóa đơn": values[2],
                        "Số hóa đơn": values[3],
                        "Ngày lập": nlap,
                        "MST người mua/MST người nhận hàng": values[5],
                        "Tên người mua/Tên người nhận hàng": values[6],
                        "Tổng tiền chưa thuế": values[7],
                        "Tổng tiền thuế": values[8],
                        "Tổng tiền chiết khấu thương mại": values[9],
                        "Tổng tiền phí": values[10],
                        "Tổng tiền thanh toán": values[11],
                        "Đơn vị tiền tệ": values[12],
                        "Trạng thái hóa đơn": values[13],
                        "Kết quả kiểm tra hóa đơn": values[14]
                    }
                    data_crawled.append(json_record)
                    self.a += n_range
                    start_index += 1
                    last_row += 1
            columns = ['H','I','J','K','L']
            for column in columns:
                for cell in sheet[column]:
                    cell.number_format = '#,##0'
            alignment = Alignment(wrap_text=True)
            for row in sheet.iter_rows(min_row=7, max_row=last_row, min_col=1, max_col=16):
                for cell in row:
                    cell.alignment = alignment
            font = Font(size=12)
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
            columns = {"E"}
            for column in columns:
                    for cell in sheet[column]:
                        try:
                            new_value = cell.value
                            new_value = new_value.split("-")
                            new_value = new_value[2] + "/" + new_value[1] + "/" + new_value[0]
                            cell.value = new_value
                        except:
                            pass
            cell = sheet['A3']
            cell.font = Font(bold=True, size=16)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            excel_bytes_data = excel_buffer.getvalue()
            
            # ✅ Lưu file vào disk và tạo download_id (giống Go-Soft pattern)
            try:
                import sys
                import os as os_module
                sys.path.insert(0, os_module.path.dirname(os_module.path.dirname(os_module.path.dirname(os_module.path.abspath(__file__)))))
                from shared.download_service import save_file_to_disk
                
                download_id, file_path = save_file_to_disk(excel_bytes_data, 'xlsx')
                logger.info(f"✅ Đã lưu Excel file: {file_path} (download_id: {download_id})")
            except Exception as e:
                logger.error(f"❌ Lỗi khi lưu Excel file vào disk: {e}")
                # Fallback: vẫn tạo base64 nếu không lưu được
                download_id = None
                excel_bytes = base64.b64encode(excel_bytes_data).decode('utf-8')
            
            response = {
                "status": "success",
                "message": f"Hoàn tất tải thống kê tổng quát {count}/{count} hóa đơn",
                "data": {
                    "filename": "Thong_ke_tong_quat.xlsx",
                    "total_records": count,
                    "download_id": download_id,  # ✅ Trả về download_id thay vì excel_bytes
                    # ✅ Backward compatibility: vẫn có excel_bytes nếu không lưu được vào disk
                    "excel_bytes": base64.b64encode(excel_bytes_data).decode('utf-8') if download_id is None else None
                },
                # ✅ Cần trả datas để client gửi cho bước tải XML/HTML/PDF
                "data_crawled": data_crawled,
                "datas": datas_first["datas"],
            }

            print(f"       [ HOÀN TẤT TẢI THỐNG KÊ TỔNG QUÁT {count}/{count} HÓA ĐƠN ]")
            return response
    def chitiet_(self,datas_first = {},headers: dict = {},progress_callback=None):
            tout = 15
            self.progress_callback = progress_callback  # Lưu callback
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            excel_thongke_a = os.path.join(base_dir, '__pycache__', 'template', 'Thống kê chi tiết.xlsx')
            wb = load_workbook(excel_thongke_a)
            sheet = wb.active
            last_row = sheet.max_row
            start_index = 1
            data_crawled_detail = []  # Danh sách để lưu trữ dữ liệu JSON chi tiết
            
            total_invoices = len(datas_first["datas"])
            for data in datas_first["datas"]:
                # ✅ Check cancelled flag trước khi xử lý mỗi invoice
                if self._check_cancelled():
                    raise Exception("Job đã bị hủy (Ctrl+C)")
                
                if data["ttxly"] == 8:
                    spec = "sco-"
                else:
                    spec = ""
                start_index+=1
                
                # 📊 Báo tiến trình cho mỗi hóa đơn chi tiết
                if self.progress_callback:
                    self.progress_callback(
                        current_step=f"Đang lấy chi tiết hóa đơn {start_index}/{total_invoices}...",
                        processed=start_index,
                        total=total_invoices
                    )
                
                nbmst = data["nbmst"]
                khhdon = data["khhdon"]
                shd = data["shdon"]
                khmshdon = data["khmshdon"]
                f = 0
                import random
                while True:   
                    f+=1
                    try:                     
                        res1 = self.session.get(f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/detail?nbmst={nbmst}&khhdon={khhdon}&shdon={shd}&khmshdon={khmshdon}',headers=headers,verify=False,timeout =tout)
                        if res1.status_code == 200:
                            logger.info(f" Got invoice detail | Status: {res1.status_code} | Response size: {len(res1.content)} bytes | Attempt: {f} | | Invoice {start_index}/{total_invoices} |")
                            break
                        elif res1.status_code == 429:
                            logger.warning(f" 429 Too Many Requests detected | Invoice: {nbmst}-{khmshdon}-{shd} | Rotating IP...")
                            self._recreate_session_with_new_proxy()
                            continue
                    except Exception as ex:
                        logger.error(f" Request failed,change proxy now | Invoice: {nbmst}-{khmshdon}-{shd} | Error: {str(ex)}")
                        self._recreate_session_with_new_proxy()
                        print("ERROR")
                try:
                    data_ct = res1.json()
                    # ✅ Kiểm tra data_ct có phải là dict không
                    if not isinstance(data_ct, dict):
                        logger.error(f" Failed,change session,proxy now | Invoice: {nbmst} | Response không phải dict: {type(data_ct)}")
                        self._recreate_session_with_new_proxy()
                        continue
                except Exception as ex:
                    logger.error(f" Failed,change session,proxy now | Invoice: {nbmst} | Error: {str(ex)}")
                    self._recreate_session_with_new_proxy()
                    continue

                headers_w = ["khmshdon"	,"khhdon"	,"shdon","ntao"	,"nky"	,"mhdon"	,"nky"	,"dvtte"	,"tgia"	,"nbten"	, "nbmst"	,"nbdchi"	,"nmten"	,"nmmst"	,"nmdchi"	,"m_VT","ten","dvtinh","sluong","dgia","stckhau","tsuat","thtien","tthue","ttcktmai"	,"tgtphi"	,"tgtttbso"	,"tthai"	,"ttxly","url","mk","ghichu","thtttoan","tchat","dgiai"]
                s = 0
                n = 0
                url = ""
                mk = ""
                list_link = [['0100684378', 'https://0100684378-tt78.vnpt-invoice.com.vn'],['0314743623', 'https://ehoadondientu.com/Tra-cuu\n'],['0105987432', 'EASy'],['0106741551', 'https://tracuuhoadon.cyberbill.vn/#/tracuuhoadon/tracuu'],['0102721191-068', 'https://vat.ggg.com.vn/'],['0102516308', 'https://tracuuhoadon.mediamart.com.vn/'],['0107500414', 'https://tracuuhoadon.vetc.com.vn/'],['4600128263', 'https://hoadon.petrolimex.com.vn/SearchInvoicebycode/Index'],['0100107564-001', 'https://hoadon.petrolimex.com.vn/SearchInvoicebycode/Index'], ['0104128565', 'https://bit.ly/hdtrcuuFPT\n'], ['0302999571', 'https://tracuu.lcs-ca.vn\n'], ['0313963672', 'https://tracuuhoadon.kkvat.com.vn/\n'], ['0105232093', 'https://tracuu.cyberbill.vn/#/tracuuhoadon/tracuu\n'], ['0311942758', 
                    'http://www.ngogiaphat.vn ( website tra cứu không thể truy cập )\n'], ['0302712571', 'https://matbao.in/tra-cuu-hoa-don/\n'], ['0103930279', 'https://www.nacencomm.com.vn/dich-vu-chi-tiet/hoa-don-dien-tu\n'], ['0105844836', 
                    'https://tracuu.vininvoice.vn\n'], ['0312483391', 'https://azinvoice.com ( web tra cứu lỗi )\n'], ['0101243150', 'https://www.meinvoice.vn/tra-cuu/\n'], ['0106026495', 'https://tracuuhoadon.minvoice.com.vn/single/invoice\n'], ['0313906508', 'www.nguyenminhvat.vn\n'], ['0101300842', 'https://einvoice.vn/tra-cuu\n'], ['0306784030', 'https://ehoadon.online/einvoice/lookup\n'], ['0200638946', 'https://oinvoice.vn/tracuu/\n'], ['0312303803', 'https://tracuu.wininvoice.vn\n'], ['0100109106', 'https://vinvoice.viettel.vn/utilities/invoice-search\n'], ['0102454468', 'https://tax24.com.vn/thuedientu/xac-minh-hoa-don\n'], ['0105937449', 'https://newinvoice.com.vn/tra-cuu/\n'], ['0108516079', 'http://hddt.3asoft.vn/#tracuu\n'], ['0100686209', 'https://bit.ly/hdtrcuumobifone\n'], ['0101360697', 'https://bit.ly/hdtracuuVan\n'], ['0101162173', 'https://asiainvoice.vn/tra-cuu\n'], ['0401486901', 'https://tracuu.vin-hoadon.com/tracuuhoadon/tracuuxacthuc/tracuuhd\n'], ['0200784873', 'https://dinhvibachkhoa.vn\n'], ['0100684378', 'https://hoadon.petrolimex.com.vn/SearchInvoicebycode/Index?strFkey=\n'], 
                    ['0106713804', 'https://hiloinvoice.vn/tra-cuu/\n'], ['0314209362', 'https://hoadondientuvat.com/Tracuu.aspx\n'], ['0101352495', 'https://tracuu.v50.vninvoice.vn/\n'], ['0102182292', 
                    'https://hddt.vnpay.vn/Invoice/Index/\n'], ['0106870211', 'https://tracuu.vietinvoice.vn/#/\n'], ['0104614692', 'https://hoadontvan.com/TraCuu\n'], ['0309612872', 'https://ehd.smartvas.vn/HDDT/\n'], ['0309478306', 'https://tracuu.xuathoadon.vn/\n'], ['0315298333', 'https://tctinvoice.com/\n'], ['0303609305', 'https://ihoadondientu.com/Tra-cuu\n'], ['0100727825', 'https://invoice.fast.com.vn/lookup/tra-cuu-hoa-don-dien-tu.aspx\n'], ['0315467091', 'http://www.acconine.vn ( không tồn tại website )\n'], ['0315638251', 'https://laphoadon.htinvoice.vn/TraCuu\n'], ['0105958921', 'https://tracuu.cloudinvoice.vn/\n'], ['0302431595', 'https://tracuu.hoadon30s.vn/en/tin-tuc/\n'], ['0103018807', 'https://vnisc.com.vn\n'], ['0106820789', 'https://tracuu.hoadondientuvn.info/#/tracuuhoadon/tracuu\n'], ['0310151055', 'https://www.SAFEinvoice.vn ( website không thể truy cập )\n'], ['0303430876', 'www.spc-technology.com ( website không thể truy cập ) \n'], ['0301452923', 'https://tracuu.lienson.vn/#/tracuuhoadon/tracuu\n'], ['0314185087', 'https://hoadon.onlinevina.com.vn/invoice\n'], ['0100687474', 'https://hoadondientu-ptp.vn/tra-cuu/\n'], ['0400462489', 'https://e-invoicetuanchau.com/Tra-cuu\n'], ['3500456910', 'https://hoadonminhthuvungtau.com/Tra-cuu\n'], ['0104908371', 'https://hoadondientu.acman.vn/tra-cuu/hoa-don.html\n'], ['0315191291', 'https://hoadonsovn.evat.vn/\n'], ['0313844107', 'http://voice.hoadondientu.net.vn\n'], ['0311622035', 'http://voice.hoadondientu.net.vn/tra-cuu\n'], ['0106361479', 'https://tracuu.ahoadon.com/\n'], ['0312270160', 'https://ameinvoice.vn/invoice-inquiry/\n'], ['0104493085', 'https://fts.com.vn/phan-mem-hoa-don-dien-tu/\n'], ['0101289966', 'https://tracuu.e-hoadon.cloud/\n'], ['0303211948', 'https://vlc.evat.vn/\n'], ['0101622374', 'https://tamvietgroup.vn/hoa-don-dien-tu/\n'], ['0310768095', 'http://hoadondientu.link/tracuu\n'], ['0312961577', 'http://tracuuhoadon.benthanhinvoice.vn/\n'], ['0313950909', 'https://koffi.vn\n'], ['0311928954', 'https://tracuuhoadon.vietinfo.tech/\n'], ['0103770970', 'https://www.bitware.vn/tracuuhoadon/\n'], ['0305142231', 'https://www.rosysoft.vn/tin-cong-nghe/erp-rosy-giai-phap-hoa-don-dien-tu\n'], ['3702037020', 'https://trandinhtung.evat.vn/\n'], ['0101925883', 'http://tracuu.cmcsoft.com/\n'], ['0316642395', 'https://phuongnam.evat.vn/\n'], ['0315194912', 'https://ttltax.com/dich-vu/hoa-don-dien-tu-237.html\n'], ['0315983667', 'http://hoadondientuvietnam.vn/HDDT/\n'], ['0310926922', 'https://invoice.ehcm.vn/\n'], ['0101010702', 'https://www.thanglongsoft.com/index.php\n'], ['0102720409', 'http://tigtax.vn\n'], ['0314058603', 'https://portal.vdsg-invoice.vn/\n'], ['0301448733', 'https://accnet.vn/hoa-don-dien-tu\n'], ['0313253288', 'https://app.autoinvoice.vn/tracuu.zul\n'], ['0309889835', 'https://unit.com.vn/'], ['0202029650', 'https://hdbk.pmbk.vn/tra-cuu-hoa-don'], ['0108971656', 'https://tracuu.myinvoice.vn/#/'], ['0312942260', 'https://ihoadondientu.net/Tracuu.aspx\n'], ['1201496252', 'https://webcashvietnam.com/vn/e_invoice.html\n'], ['0303549303', 'https://e-invoices.vn/\n'], ['0311914694', 'https://brightbrain.vn/?s=tra+c%E1%BB%A9u\n'], ['0312617990', 'https://demo-eportal.cloudteam.vn/n'], ['0109282176', 'https://tracuu.vinvoice.vn/\n'], ['0102723181', 'http://hoadonct.gov.vn/'], ['0106858609', 'https://tracuuhoadon.vetc.com.vn/\n'], ['0315151651', 'https://ei.pvssolution.com/#/\n'], ['0310151739', 'https://news.yoinvoice.vn\n'], ['0312575123', 'https://www.ecount.com/vn/ecount/product/accounting_e-invoice'], ['0107732197', 'https://tracuuhoadon.atis.com.vn/\n'], ['0101659906', 'https://tracuu.kaike.vn/#/\n'], ['0103019524', 'https://einvoice.aits.vn/\n'], ['0316114998', 'https://bizzi.vn/\n'], ['0316636497', 'http://beetek.vn/\n'],]
                list_link2 = [['0106771637', 'https://newinvoice.com.vn/tra-cuu\n'],['0105987432', 'EASY'] ,
                    ['104918404', 'https://hoadon.winmart.vn/?branch=2&form=1&serial=K23TBP&seq=00023761\n'], ['0102519041', 'https://ihoadon.vn/kiem-tra/?lang=vn\n'], ['0310247046', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0301472278', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0107967449', 'https://tracuuhoadon.fpt.com.vn/\n'], ['0316016380', 'https://tracuu.wininvoice.vn\n'], ['0200662314', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0304043037', 'https://hoadon.247express.vn\n'], ['0303500749', 'https://viewhoadon78.nacencomm.vn/tracuuhd.aspx\n'], ['0100956381', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['3702058398', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['3600258976', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['1801210593', 
                    'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['3300854978-003', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0200662314-004', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['3300854978-004', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0313517445', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0104918404-067', 'https://hoadon.winmart.vn/?branch=2&form=1&serial=K23TBP&seq=00023761\n'], ['0301468144', 'https://tracuu.vietinvoice.vn\n'], ['0317993854', 'http://tctinvoice.com/hddt/sinv/sinv00101\n'], ['0104918404-007', 'https://hoadon.winmart.vn/?branch=2&form=1&serial=K23TBP&seq=00023761\n'], ['0200662314-001', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0312120895', 'https://hddt.centralretail.com.vn/vi-vn/tra-cuu-hoa-don-dien-tu\n'], ['0316521834', 'https://tracuuhd.smartsign.com.vn\n'], ['0104918404-002', 'https://hoadon.winmart.vn/?branch=4&form=1&serial=K23TBP&seq=00237857'],['4300763578', 'https://4300763578-tt78.vnpt-invoice.com.vn\n'],
                    ['2700560240', 'http://tracuu.ehoadon.vn\n'],
                    ['2900428497', 'https://petrolimexnghetinh-tt78.vnpt-invoice.com.vn\n'],
                    ['2801076188', 'https://2801076188-tt78.vnpt-invoice.com.vn\n'],
                    ['3300338405', 'https://ptshue-tt78.vnpt-invoice.com.vn\n'],
                    ['0101360697', 'https://van.ehoadon.vn/TCHD\n']]
                for i in list_link:
                    try:
                        if i[0] == data_ct["msttcgp"]:
                            try:
                                fkey = data['mhdon']
                            except:
                                pass
                            url = i[1]
                            if url == "EASy":
                                try:
                                    fkey = data['mhdon']
                                    url = f"https://{nbmst}hd.easyinvoice.com.vn/Search/?strFkey={fkey}"
                                except Exception as e:
                                    print(e)
                                    url = f"https://{nbmst}hd.easyinvoice.com.vn/Search/?strFkey="
                            mst_url_map = {
                                "0310471746": "https://hddt.bachhoaxanh.com",
                                "0303217354": "https://hddt.thegioididong.com",
                                "0301452948": "https://tracuuhoadondientu.acb.com.vn",
                                "0100107518": "https://einvoice.vietnamairlines.com",
                                "0309532909": "https://hoadon.tiki.com.vn",
                                "0312388363": "https://hoadon.tiki.com.vn",
                                "0304741634": "https://lottemart-nsg-tt78.vnpt-invoice.com.vn/Portal/Index",
                                "2500707753": "https://hoadon.petrolimex.com.vn/SearchInvoicebycode",
                                "4600391722": "https://4600391722-tt78.vnpt-invoice.com.vn",
                                "0302431595": "https://hoadon.pavietnam.vn",
                                "0316771954": "https://ehoadondientu.com/Tracuu.aspx",
                                "0100107437": "https://0100107437-tt78.vnpt-invoice.com.vn",
                                "2500233408": "https://2500233408-tt78.vnpt-invoice.com.vn/",
                                "0500387891": "https://petrolimexhatay-tt78.vnpt-invoice.com.vn/",
                                "2700941905": "https://2700941905-tt78.vnpt-invoice.com.vn/",
                                "0106494831": "https://dangkiemdongdo-tt78.vnpt-invoice.com.vn/",
                                "2700275317": "https://2700275317-tt78.vnpt-invoice.com.vn/",
                                "2700113651-006": "https://gpbank-tt78.vnpt-invoice.com.vn/",
                                "0100109120-024": "https://dangkiem2902v-tt78.vnpt-invoice.com.vn/",
                                "4601139074": "https://4601139074-tt78.vnpt-invoice.com.vn/",
                                "2902093542": "https://kdmhungphucan-tt78.vnpt-invoice.com.vn/",
                                "2500218618-001": "https://2500218618-001-tt78.vnpt-invoice.com.vn/",
                                "2700280162": "https://2700280162-tt78.vnpt-invoice.com.vn/",
                                "0500593012": "https://pattnhh-tt78.vnpt-invoice.com.vn/",
                                "2700871020": "https://xemhoadon.vnpt.vn/",
                                "2700714243": "https://2700714243-tt78.vnpt-invoice.com.vn/",
                                "0107466555": "https://tmdtcdt-tt78.vnpt-invoice.com.vn/",
                                "2802417722": "https://2802417722-tt78.vnpt-invoice.com.vn/",
                                "3600244645": "https://ajinomotodn-tt78.vnpt-invoice.com.vn",
                                "0302309845": "https://dienmaycholon-tt78.vnpt-invoice.com.vn",
                                "0314555531": "https://ctykingscross-tt78.vnpt-invoice.com.vn",
                                "0311945910": "https://0311945910-tt78.vnpt-invoice.com.vn"
                            }
                            nbmst = data_ct["nbmst"]
                            if nbmst in mst_url_map:
                                url = mst_url_map[nbmst]
                            else:
                                try:
                                    nbmst_prefix = nbmst.split("-")[0]
                                    if nbmst_prefix in mst_url_map:
                                        url = mst_url_map[nbmst_prefix]
                                except:
                                    pass
                            fkey = 0
                            try:
                                fkey = data['mhdon']
                            except Exception as e:
                                print("0000")
                                print(e)
                      
                            nbmst_prefix = data_ct["nbmst"].split("-")[0]
                            
                            if nbmst_prefix == "0105987432":
                                url = f"https://{nbmst}hd.easyinvoice.com.vn/Search/?strFkey={fkey}"
                            elif nbmst_prefix == "0100684378":
                                try:
                                    url = f"https://{nbmst}-tt78.vnpt-invoice.com.vn/?strFkey={fkey}"
                                except Exception as e:
                                    pass
                            elif nbmst_prefix == "0304741634":
                                url = "https://lottemart-nsg-tt78.vnpt-invoice.com.vn/Portal/Index"
                            elif nbmst_prefix == "3600244645":
                                url = "https://ajinomotosg-tt78.vnpt-invoice.com.vn/"
                            elif nbmst_prefix == "0104918404":
                                url = "https://hoadon.winmart.vn/"
                            
                            # Kiểm tra list_link2
                            for i in list_link2:
                                if data_ct["nbmst"] == i[0]:
                                    url = i[1]
                                    if url == "EASY":
                                        url = f"https://{nbmst}hd.easyinvoice.com.vn/Search/?strFkey={fkey}"
                            break

                    except Exception as e:
                        print("XXXXS")
                        pass
                if url == "":
                    for i in list_link:
                        try:
                            if i[0] == data_ct["tvandnkntt"]:
                                url = i[1]
                                if url == "EASY":
                                    url = f"https://{nbmst}hd.easyinvoice.com.vn/Search/?strFkey={fkey}"
                                break
                        except:
                            pass
                if mk == "":
                        try:
                            for i in data_ct["ttkhac"]:
                                if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "Mã tra cứu hóa đơn" or i["ttruong"] == "chungTuLienQuan"  or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice" or i["ttruong"] == "Mã tra cứu" :
                                    mk = i["dlieu"]
                            if mk == "":
                                for i in data_ct["cttkhac"]:
                                    if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "Mã tra cứu hóa đơn"  or i["ttruong"] == "chungTuLienQuan" or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice" or i["ttruong"] == "Mã tra cứu" :
                                        mk = i["dlieu"]
                            if mk == "":
                                for i in data_ct["ttttkhac"]:
                                    if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "chungTuLienQuan" or i["ttruong"] == "Mã tra cứu hóa đơn" or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice" or i["ttruong"] == "Mã tra cứu" :
                                        mk = i["dlieu"]
                            try:
                                if mk == "":
                                    for i in data_ct["TTKhac"]:
                                        if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "chungTuLienQuan" or i["ttruong"] == "Mã tra cứu hóa đơn" or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice":
                                            mk = i["dlieu"]
                            except:
                                pass
                        except:
                            pass
                if mk == "":
                    mk = "Không tìm thấy mã tra cứu trên file XML, vui lòng liên hệ người bán để được cung cấp file PDF gốc."
                if url == "https://tracuu.vietinvoice.vn/#/" or url == "https://tracuu.vininvoice.vn":
                    mk = data_ct["mhdon"]
                if url == "https://lottemart-nsg-tt78.vnpt-invoice.com.vn/Portal/Index":
                    mk = "Tìm kiếm theo thông tin hóa đơn"
                if url == "EASY":
                    try:
                        fkey = data['mhdon']
                    except Exception as e:
                                print(e)
                
                    url = f"https://{nbmst}hd.easyinvoice.com.vn/Search/?strFkey={fkey}"
                z = 0
                ttttoan = ""
                asss = 0 
                ds_san_pham = []  
                while True:
                        t = 0
                        temp = ""
                        tong_thue = 0 
                        try:    
                            len_ct = len(data_ct["hdhhdvu"])
                        except:
                            break
                        sp_n = 0
                        for sttt,sp in enumerate(data_ct["hdhhdvu"]):  
                            print(f"       [ THỐNG KÊ CHI TIẾT HÓA ĐƠN {start_index-1}/{len(datas_first['datas'])} - SẢN PHẨM {sp_n+1}/{len_ct} ]")
                            sp_n +=1
                            if z == 0:
                                    ttttoan = "tgtttbso"
                                    z +=1
                            else:
                                    ttttoan = ""
                                    z+=1  
                            row_index = last_row + 1 
                            values = []
                            n=0
                            for header in headers_w: 
                                if header in ["khmshdon", "khhdon", "shdon", "mhdon", "dvtte", "tgia", "nbten", "nbmst", "nbdchi", "nmten", "nmmst", "nmdchi", "ttcktmai", "tgtphi", ttttoan, "tthai", "ttxly","thtttoan"]:
                                    if header == "ttcktmai":
                                        if header == "ttcktmai" and n == len(headers_w)-1 :
                                            value = data["ttcktmai"]
                                        else:
                                            value = ""
                                    elif header == "thtien" and z == len(headers_w)-1 :
                                        value = data["tgtttbso"]-s
                                    else:
                                        if header == "tgtphi":
                                            if temp == data_ct.get("shdon", ""):
                                                value = ""
                                            else:
                                                value = data_ct.get(header, "") 
                                                temp = data_ct.get("shdon", "")
                                        else:
                                            value = data_ct.get(header, "") 
                                        if header == "thtien":
                                            s+=value
                                            asss +=value

                                    try:
                                        if header == 'nbten' and (data_ct.get("nbten", "") == 'null' or  data_ct.get("nbten", "") == None):
                                            value = data_ct.get("nmtnban", "")
                                        if header == 'nmten' and (data_ct.get("nmten", "") == 'null' or  data_ct.get("nmten", "") == None):
                                            value = data_ct.get("nmtnmua", "")
                                    except:
                                        pass
                                else:
                                    value= 0
                                    if header == "tsuat":
                                        value = data_ct["hdhhdvu"][sttt]["tsuat"]
                                        if data_ct["hdhhdvu"][sttt]["ltsuat"] == "KHAC":
                                            value = "KHAC"
                                            for item in data["thttltsuat"]:
                                                    tsuat = item.get("tsuat", "")
                                                    if tsuat.startswith("KHAC:") and tsuat.endswith("%"):
                                                        value = tsuat[5:-1] + "%"
                                        
                                        if data_ct["hdhhdvu"][sttt]["ltsuat"] == "KKKNT":
                                            value = "KKKNT"
                                        if data_ct["hdhhdvu"][sttt]["tsuat"] == 0.0:
                                            if data_ct["hdhhdvu"][sttt]["ltsuat"] == "KCT":
                                                value = "KCT"
                                            else:
                                                for item in data["thttltsuat"]:
                                                    if item.get("tsuat") == "KKKNT":
                                                        value = "KKKNT"
                                                        break
                                                    value = "0%"
                                        if value == "KCT":
                                                value = "KCT"
                                        if value == 121:
                                                value = "KKKNT"
                                        if value == 122:
                                                value = "KHAC"
                                        if value == "KHAC":
                                            for item in data["thttltsuat"]:
                                                    tsuat = item.get("tsuat", "")
                                                    if tsuat.startswith("KHAC:") and tsuat.endswith("%"):
                                                        value = tsuat[5:-1] + "%"
                                        if value == "KHAC":
                                            value = "KHAC"
                                        else:
                                            try:
                                                value = value
                                            except:
                                                pass
                                    elif header == "tthue":
                                            if len(data_ct["hdhhdvu"]) > 0:
                                                try:
                                                    cell = float(values[headers_w.index("thtien")])
                                                    if sp_n == len_ct:
                                                        value = data_ct["tgtthue"]-tong_thue
                                                    else:
                                                        try:
                                                            vl = values[headers_w.index("tsuat")]
                                                            if ":" in str(vl):
                                                                vl = str(vl).split(":")[1]
                                                            elif vl == 122:
                                                                vl = 5.263/100
                                                                for item in data["thttltsuat"]:
                                                                    tsuat = item.get("tsuat", "")
                                                                    if tsuat.startswith("KHAC:") and tsuat.endswith("%"):
                                                                        vl = float(int(tsuat[5:-1]))/100
                                                        except Exception as e:
                                                            print(e)
                                                        if data_ct["hdhhdvu"][sttt]["ltsuat"] == "KHAC":
                                                                vl = 5.263/100
                                                                for item in data["thttltsuat"]:
                                                                    tsuat = item.get("tsuat", "")
                                                                    if tsuat.startswith("KHAC:") and tsuat.endswith("%"):
                                                                        vl = float(int(tsuat[5:-1]))/100
                                                                
                                                        if vl != None:
                                                            try:
                                                                if "KKKNT" not in str(vl) or "KCT" not in str(vl) or str(vl) != "KCT":
                                                                    #value = math.ceil(float(vl)/100 * (round(round(float(cell),1),-1)))
                                                                    try:
                                                                        value = float(vl) * float(data_ct["hdhhdvu"][sttt]["thtien"])
                                                                    except:
                                                                        value = old_tthue
                                                                else:
                                                                    value = old_tthue          
                                                            except Exception as e:
                                                                print(e)
                                                        else:
                                                            value = " "
                                                        try:   
                                                            tong_thue +=value
                                                        except:
                                                            pass
                                                        old_tthue = data_ct["hdhhdvu"][sttt]["tthue"]
                                                        if old_tthue == None:
                                                            old_tthue = 0
                                                except Exception as e:
                                                    print(e)
                                            if value == " " :
                                                    try:
                                                        value = data_ct["tgtthue"]
                                                    except:
                                                        pass
                                    else: 
                                        if header == "ntao":
                                                from datetime import datetime, timedelta
                                                value = data["tdlap"]
                                                utc_time_str =  value
                                                utc_time = datetime.strptime(utc_time_str, "%Y-%m-%dT%H:%M:%SZ")
                                                vn_time = utc_time + timedelta(hours=7)
                                                vn_time_str = vn_time.strftime("%Y-%m-%dT%H:%M:%SZ")
                                                value = vn_time_str.split("T")[0]
                                                new_value = value.split("-")
                                                value = new_value[2] + "/" + new_value[1] + "/" + new_value[0]
                                        elif header == "url":
                                            value = url
                                            if t > 0:
                                                value = ""
                                        elif header == "mk":
                                            value = mk
                                            if t > 0:
                                                value = ""
                                            t +=1
                                        elif header == "nky":
                                            try:
                                                from datetime import datetime, timedelta
                                                utc_time_str = data["nky"]
                                                utc_time = datetime.strptime(utc_time_str, "%Y-%m-%dT%H:%M:%SZ")
                                                vn_time = utc_time + timedelta(hours=7)
                                                vn_time_str = vn_time.strftime("%Y-%m-%dT%H:%M:%SZ")
                                                value = vn_time_str.split("T")[0]
                                                new_value = value.split("-")
                                                value = new_value[2] + "/" + new_value[1] + "/" + new_value[0]
                                            except:
                                                value = values[headers_w.index("ntao")]
                                        elif header == "ghichu":
                                            d1 = values[3].split("/")
                                            d2 = values[4].split("/")
                                            try:
                                                if (d2[1] != d1[1]) or (d2[0] != d1[0]) or (d2[2] != d1[2]):
                                                    value = "  "
                                                else:
                                                    value = " "
                                            except:
                                                value = " "
                                        elif header == "tchat":
                                            value = sp.get(header,  "")
                                            if value == 1:
                                                value = "Hàng hóa, dịch vụ"
                                            if value == 2:
                                                value = "Khuyến mại"
                                            if value == 3:
                                                value = "Chiết khấu"  
                                                values[22] = values[22]*(-1)  
                                            if value == 4:
                                                value = "Ghi chú, diễn giải"
                                        elif header =="dgiai":
                                            value = " "
                                            if value == " ":
                                                khmshd = data_ct["khmshdgoc"]
                                                khhd = data_ct["khhdgoc"]
                                                shd_ = data_ct["shdgoc"]
                                                if data_ct["tdlhdgoc"] != None :
                                                    nlap = str(data_ct["tdlhdgoc"]).split("T")[0]
                                                    value = f"Điều chỉnh cho ký hiệu mẫu số hóa đơn {khmshd}, ký hiệu hóa đơn {khhd}, số hóa đơn {shd_}, ngày lập {nlap}"
                                            if value == " ":
                                                for i in data_ct["ttkhac"]:
                                                    if i["ttruong"] == "Ghi chú hóa đơn":
                                                        try:
                                                            if len(i["dlieu"])>20:
                                                                value = i["dlieu"]
                                                        except:
                                                            pass
                                        else:
                                            if header == "m_VT":
                                                n_product = sp.get("ten", "")
                                                for i in ds_san_pham:
                                                    if i[1] == n_product:
                                                        import string
                                                        value = i[0]
                                            else:
                                                    value = sp.get(header, "")
                                            
                                n+=1
                                values.append(value)
                            if values[21]=="KCT" :
                                values[23] = 0
                            if values[21]=="0%" :
                                values[23] = 0
                            if values[21]=="KKKNT":
                                values[23] = 0
                            if values[21]=="KHAC":
                                try:    
                                    values[23] = float(5.263/100) * float(data_ct["hdhhdvu"][sttt]["thtien"])
                                except:
                                    values[23] = 0

                            if values[21]=="" or values[21]== None or values[21]== " ":
                                values[23]=0
                            if values[16] == "Chiết khấu":
                                values[22]=float(values[22])*-1
                            values[2] = " "+str(values[2])
                            hdon_value = values[headers_w.index("tthai")]
                            ttxly_value = values[headers_w.index("ttxly")]
                            hdon = {1: "Hóa đơn mới", 2: "Hóa đơn thay thế", 3: "Hóa đơn điều chỉnh", 4: "Hóa đơn đã bị thay thế", 5: "Hóa đơn đã bị điều chỉnh", 6: "Hóa đơn đã bị hủy"}
                            ttxly = {0: "Tổng cục Thuế đã nhận", 1: "Đang tiến hành kiểm tra điều kiện cấp mã", 2: "CQT từ chối hóa đơn theo từng lần phát sinh", 3: "Hóa đơn đủ điều kiện cấp mã", 4: "Hóa đơn không đủ điều kiện cấp mã", 5: "Đã cấp mã hóa đơn", 6: "Tổng cục thuế đã nhận không mã", 7: "Đã kiểm tra định kỳ HĐĐT không có mã", 8: "Tổng cục thuế đã nhận hóa đơn có mã khởi tạo từ máy tính tiền"}
                            if hdon_value in hdon:
                                values[headers_w.index("tthai")] = hdon[hdon_value]
                            if ttxly_value in ttxly:
                                values[headers_w.index("ttxly")] = ttxly[ttxly_value]
                            for i in sp["ttkhac"]:
                                if i["ttruong"] == "Lot" or i["ttruong"] == "Extra1" or i["ttruong"] == "BatchNo" or i["ttruong"] == "SoLo":
                                    values.append(i["dlieu"])
                                if i["ttruong"] == "ExpireDate" or i["ttruong"] == "Extra2" or i["ttruong"] == "Expiry" or i["ttruong"] == "HanDung":
                                    values.append(i["dlieu"])
                            #Mẫu số HD	Ký hiệu hóa  đơn	Số hóa đơn	Ngày lập hóa đơn	Ngày người bán ký số	MCCQT	Ngày CQT ký số	Đơn vị tiền tệ	Tỷ giá	Tên người bán	MST người bán	Địa chỉ người bán	Tên người mua	MST người mua	Địa chỉ người mua	Mã VT	Tên hàng hóa, dịch vụ	Đơn vị tính	Số lượng	Đơn giá	Chiết khấu	Thuế suất	Thành tiền chưa thuế	Tiền thuế	Tổng tiền CKTM	Tổng tiền phí	Tổng tiền thanh toán	Trạng thái hóa đơn	Kết quả kiểm tra hóa đơn	url  tra cứu hóa đơn	Mã tra cứu	Ghi chú 1	Hình  thức thanh toán	Tính chất	Ghi chú 2	Số lô 	Hạn dùng 

                            for column_index, value in enumerate(values, start=1):
                                sheet.cell(row=row_index, column=column_index, value=value)
                            if values[12]==None:
                                try:
                                    values[12]=data["nmtnmua"]
                                except:
                                    pass
                                try:
                                    values[12]=data["nbtnmua"]
                                except:
                                    pass
                            # Thêm dữ liệu vào JSON chi tiết
                            headers_detail = ["Mẫu số HD", "Ký hiệu hóa đơn", "Số hóa đơn", "Ngày lập hóa đơn", "Ngày người bán ký số", 
                                            "MCCQT", "Ngày CQT ký số", "Đơn vị tiền tệ", "Tỷ giá", "Tên người bán", "MST người bán", 
                                            "Địa chỉ người bán", "Tên người mua", "MST người mua", "Địa chỉ người mua", "Mã VT", 
                                            "Tên hàng hóa, dịch vụ", "Đơn vị tính", "Số lượng", "Đơn giá", "Chiết khấu", "Thuế suất", 
                                            "Thành tiền chưa thuế", "Tiền thuế", "Tổng tiền CKTM", "Tổng tiền phí", "Tổng tiền thanh toán", 
                                            "Trạng thái hóa đơn", "Kết quả kiểm tra hóa đơn", "url tra cứu hóa đơn", "Mã tra cứu", 
                                            "Ghi chú 1", "Hình thức thanh toán", "Tính chất", "Ghi chú 2", "Số lô", "Hạn dùng"]
                            
                            json_record = {}
                            for idx, header in enumerate(headers_detail):
                                if idx < len(values):
                                    json_record[header] = values[idx]
                            
                            data_crawled_detail.append(json_record)
                            last_row += 1
                        break
                if values[12]==None:
                    try:
                        values[12]=data["nmtnmua"]
                    except:
                        pass
                    try:
                        values[12]=data["nbtnmua"]
                    except:
                        pass
            border = Border(left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=37):
                for cell in row:
                    cell.border = border
            columns = ['T', 'U',"W", 'X', 'AA']
            for column in columns:
                for cell in sheet[column]:
                    cell.number_format = '#,##0'
            alignment = Alignment(wrap_text=True)
            column = 'AD'
            for row in range(2, sheet.max_row + 1):
                cell = sheet[column + str(row)]
                font = Font(underline="single", color="0563C1")
                cell.font = font
                cell_value = cell.value
                sheet['AD' + str(row)].value = f'=HYPERLINK("{cell_value}","{cell_value}")'
            
            font = Font(size=12)
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
            
            columns = {'D','E','G'}
            for column in columns:
                    for cell in sheet[column]:
                        try:
                            new_value = cell.value
                            new_value = new_value.split("-")
                            new_value = new_value[2] + "/" + new_value[1] + "/" + new_value[0]
                            cell.value = new_value
                        except:
                            pass
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            excel_bytes_data = excel_buffer.getvalue()
            
            # ✅ Lưu file vào disk và tạo download_id (giống Go-Soft pattern)
            try:
                import sys
                import os as os_module
                sys.path.insert(0, os_module.path.dirname(os_module.path.dirname(os_module.path.dirname(os_module.path.abspath(__file__)))))
                from shared.download_service import save_file_to_disk
                
                download_id, file_path = save_file_to_disk(excel_bytes_data, 'xlsx')
                logger.info(f"✅ Đã lưu Excel file (chitiet): {file_path} (download_id: {download_id})")
            except Exception as e:
                logger.error(f"❌ Lỗi khi lưu Excel file (chitiet) vào disk: {e}")
                # Fallback: vẫn tạo base64 nếu không lưu được
                download_id = None
                excel_bytes = base64.b64encode(excel_bytes_data).decode('utf-8')
            
            response = {
                "status": "success",
                "message": f"Hoàn tất tải chi tiết {len(datas_first['datas'])} hóa đơn",
                "data": {
                    "filename": "Chi_tiet_hoa_don.xlsx",
                    "total_records": len(datas_first['datas']),
                    "download_id": download_id,  # ✅ Trả về download_id thay vì excel_bytes
                    # ✅ Backward compatibility: vẫn có excel_bytes nếu không lưu được vào disk
                    "excel_bytes": base64.b64encode(excel_bytes_data).decode('utf-8') if download_id is None else None
                },
                "data_crawled_detail": data_crawled_detail,
                # ✅ KHÔNG trả về raw_data nữa (quá lớn, có thể gây tràn RAM)
                # "raw_data": datas_first["datas"]  # ❌ Comment out để tránh tràn RAM
            }
            return response
    
    def xmlahtml(self,datas_first = {},headers: dict = {},type_export:dict = {},progress_callback=None):
        tout = 15 
        self.progress_callback = progress_callback  # Lưu callback
        xml_list = []
        html_list = []
        xml_buffer_dict = {}  
        html_buffer_dict = {} 
        
        i = 0
        total_invoices = len(datas_first["datas"])
        for data in datas_first["datas"]:
                # ✅ Check cancelled flag trước khi xử lý mỗi invoice
                if self._check_cancelled():
                    raise Exception("Job đã bị hủy (Ctrl+C)")
                
                i+=1
                
                # ✅ Cập nhật progress cho mỗi invoice đang xử lý
                # Hiển thị đúng message dựa trên type_export và context (PDF hay không)
                if self.progress_callback:
                    # ✅ Kiểm tra xem có phải đang chạy PDF không (từ raw_data trong datas_first)
                    is_pdf_context = False
                    if isinstance(datas_first, dict) and datas_first.get("_is_pdf_context") == True:
                        is_pdf_context = True
                    
                    if is_pdf_context and type_export.get("html") == True:
                        # ✅ Khi chạy PDF, hiển thị message rõ ràng là đang lấy HTML để chuyển PDF
                        step_message = f"Đang lấy HTML để chuyển PDF {i}/{total_invoices}..."
                    elif type_export.get("xml") == True and type_export.get("html") == True:
                        step_message = f"Đang xuất XML/HTML {i}/{total_invoices}..."
                    elif type_export.get("xml") == True:
                        step_message = f"Đang xuất XML {i}/{total_invoices}..."
                    elif type_export.get("html") == True:
                        step_message = f"Đang xuất HTML {i}/{total_invoices}..."
                    else:
                        step_message = f"Đang xử lý {i}/{total_invoices}..."
                    
                    self.progress_callback(
                        current_step=step_message,
                        processed=i,
                        total=total_invoices
                    )
                
                if data["ttxly"] == 8:
                    spec = "sco-"
                else:
                    spec = ""
                nbmst = data["nbmst"]
                khhdon = data["khhdon"]
                shd = data["shdon"]
                khmshdon = data["khmshdon"]
                time_delay = 1
                check_timelimit = 0 
                check_show = 1
                while True:
                    # ✅ Check cancelled flag trong vòng lặp retry
                    if self._check_cancelled():
                        raise Exception("Job đã bị hủy (Ctrl+C)")
                    
                    print(f"[{i}]")  
                    try:    
                        # ✅ Sử dụng _safe_get để tự động check cancelled flag
                        response = self._safe_get(
                            f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/export-xml?nbmst={nbmst}&khhdon={khhdon}&shdon={shd}&khmshdon={khmshdon}',
                            headers=headers,
                            verify=False,
                            timeout=3
                        )
                        
                        logger.info(f"Processing invoice {i}/{total_invoices} - Status Code: {response.status_code} - SIZE: {len(response.content)} bytes")
                        if response.status_code == 429:
                            # ✅ Xử lý 429: Tạo session mới + rotate IP
                            print(f"⚠️ 429 Too Many Requests - Creating new session with rotated IP...")
                            self._recreate_session_with_new_proxy()
                            
                            # ✅ Check cancelled flag trước khi continue
                            if self._check_cancelled():
                                raise Exception("Job đã bị hủy (Ctrl+C)")
                            
                            continue
                        elif response.status_code:
                            zip_data = io.BytesIO(response.content)
                            zip_data.seek(0)
                            file_header = zip_data.read(4)
                            zip_data.seek(0)
                            is_zip = file_header[:4] == b'PK\x03\x04'
                            if not is_zip:
                                check_timelimit+=1
                                if check_timelimit %2 == 0:
                                    time_delay+=2
                                if time_delay > 20:
                                    time_delay = 10
                                print(f"Too Many Requests,Retry after {time_delay} {response.status_code} {response.text}")
                                
                                # ✅ Check cancelled flag trước khi retry
                                if self._check_cancelled():
                                    raise Exception("Job đã bị hủy (Ctrl+C)")
                                
                                if response.status_code == 500:
                                    check_show = 0
                                    break
                            else:
                                break
                    except Exception as e:
                        error_str = str(e)
                        
                        # ✅ Check cancelled flag ngay khi có exception (có thể là cancelled exception)
                        if "Job đã bị hủy" in error_str or self._check_cancelled():
                            raise Exception("Job đã bị hủy (Ctrl+C)")
                        
                        print({"status": "error", "message": error_str})
                        self._recreate_session_with_new_proxy()
                        
                        # ✅ Check cancelled flag sau khi recreate session
                        if self._check_cancelled():
                            raise Exception("Job đã bị hủy (Ctrl+C)")
                        
                        # ✅ Check cancelled flag trước khi continue retry
                        if self._check_cancelled():
                            raise Exception("Job đã bị hủy (Ctrl+C)")
                        
                        if "ReadTimeoutError" in error_str or "ConnectionError" in error_str or "Read timed out" in error_str:
                            pass
                if check_show == 0:
                    print(f"Bỏ qua hóa đơn do {response.text}")
                    continue
                if type_export.get("xml") == True:
                    try:
                        with zipfile.ZipFile(zip_data, "r") as zip_file:
                            for filename in zip_file.namelist():
                                if filename == "invoice.xml":
                                    file_content = zip_file.read('invoice.xml').decode('utf-8')
                                    xml_list.append({
                                        "khhdon": khhdon,
                                        "shdon": shd,
                                        "khmshdon": khmshdon,
                                        "xml_content": file_content
                                    })
                                    # Lưu vào bộ đệm XML
                                    file_name = f"{khhdon}_{shd}.xml"
                                    xml_buffer_dict[file_name] = file_content
                                    break
                    except Exception as e:
                        return {"status": "error", "message": f"❌ Lỗi khi xử lý XML: {e}"}

                if type_export.get("html") == True:
                    try:
                        with zipfile.ZipFile(zip_data, "r") as zip_file:
                            for filename in zip_file.namelist():
                                if filename == "invoice.html":
                                    file_content = zip_file.read('invoice.html').decode('utf-8')
                                    html_list.append({
                                        "khhdon": khhdon,
                                        "shdon": shd,
                                        "khmshdon": khmshdon,
                                        "xml_content": file_content
                                    })
                                    # Lưu vào bộ đệm HTML
                                    file_name = f"{khhdon}_{shd}.html"
                                    html_buffer_dict[file_name] = file_content
                                    break
                    except Exception as e:
                        return {"status": "error", "message": f"❌ Lỗi khi xử lý HTML: {e}"}
        
        # ✅ Tạo 2 ZIP riêng biệt cho XML và HTML để cache riêng
        xml_zip_bytes = None
        html_zip_bytes = None
        xml_filename = None
        html_filename = None
        
        # Tạo ZIP cho XML (nếu có)
        if type_export.get("xml") == True and len(xml_buffer_dict) > 0:
            xml_zip_buffer = io.BytesIO()
            with zipfile.ZipFile(xml_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as xml_zip_archive:
                for filename, content in xml_buffer_dict.items():
                    xml_zip_archive.writestr(filename, content)  # Không có folder xml/ nữa
            xml_zip_buffer.seek(0)
            xml_zip_bytes = base64.b64encode(xml_zip_buffer.getvalue()).decode('utf-8')
            xml_filename = "invoices_xml.zip"
        
        # Tạo ZIP cho HTML (nếu có)
        if type_export.get("html") == True and len(html_buffer_dict) > 0:
            html_zip_buffer = io.BytesIO()
            with zipfile.ZipFile(html_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as html_zip_archive:
                for filename, content in html_buffer_dict.items():
                    html_zip_archive.writestr(filename, content)  # Không có folder html/ nữa
            html_zip_buffer.seek(0)
            html_zip_bytes = base64.b64encode(html_zip_buffer.getvalue()).decode('utf-8')
            html_filename = "invoices_html.zip"
        
        # ✅ Lưu file vào disk và tạo download_id (giống Go-Soft pattern)
        xml_download_id = None
        html_download_id = None
        combined_download_id = None
        
        try:
            import sys
            import os as os_module
            sys.path.insert(0, os_module.path.dirname(os_module.path.dirname(os_module.path.dirname(os_module.path.abspath(__file__)))))
            from shared.download_service import save_file_to_disk
            
            # ✅ Lưu XML ZIP vào disk
            if xml_zip_bytes:
                xml_zip_bytes_data = base64.b64decode(xml_zip_bytes)
                xml_download_id, xml_file_path = save_file_to_disk(xml_zip_bytes_data, 'zip')
                logger.info(f"✅ Đã lưu XML ZIP file: {xml_file_path} (download_id: {xml_download_id})")
            
            # ✅ Lưu HTML ZIP vào disk
            if html_zip_bytes:
                html_zip_bytes_data = base64.b64decode(html_zip_bytes)
                html_download_id, html_file_path = save_file_to_disk(html_zip_bytes_data, 'zip')
                logger.info(f"✅ Đã lưu HTML ZIP file: {html_file_path} (download_id: {html_download_id})")
            
            # ✅ Lưu combined ZIP vào disk (nếu cả 2 đều có)
            if xml_zip_bytes and html_zip_bytes:
                combined_zip_buffer = io.BytesIO()
                with zipfile.ZipFile(combined_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as combined_zip_archive:
                    # Thêm XML files vào folder xml/
                    for filename, content in xml_buffer_dict.items():
                        combined_zip_archive.writestr(f"xml/{filename}", content)
                    # Thêm HTML files vào folder html/
                    for filename, content in html_buffer_dict.items():
                        combined_zip_archive.writestr(f"html/{filename}", content)
                combined_zip_buffer.seek(0)
                combined_zip_bytes_data = combined_zip_buffer.getvalue()
                combined_download_id, combined_file_path = save_file_to_disk(combined_zip_bytes_data, 'zip')
                logger.info(f"✅ Đã lưu Combined ZIP file: {combined_file_path} (download_id: {combined_download_id})")
        except Exception as e:
            logger.error(f"❌ Lỗi khi lưu ZIP file vào disk: {e}")
            # Fallback: vẫn tạo base64 nếu không lưu được
        
        # ✅ Tạo response với download_id thay vì base64
        data_obj = {
            "total_xml": len(xml_buffer_dict),
            "total_html": len(html_buffer_dict)
        }
        
        # ✅ Chỉ thêm download_id nếu có data (tránh lưu empty)
        if xml_download_id:
            data_obj["xml_download_id"] = xml_download_id
            data_obj["xml_filename"] = xml_filename
            # ✅ Backward compatibility: vẫn có base64 nếu không lưu được vào disk
            if not xml_download_id:
                data_obj["xml_base64"] = xml_zip_bytes
        
        if html_download_id:
            data_obj["html_download_id"] = html_download_id
            data_obj["html_filename"] = html_filename
            # ✅ Backward compatibility: vẫn có base64 nếu không lưu được vào disk
            if not html_download_id:
                data_obj["html_base64"] = html_zip_bytes
        
        # ✅ Backward compatibility: vẫn có zip_bytes và download_id nếu cả 2 đều có
        if combined_download_id:
            data_obj["download_id"] = combined_download_id
            data_obj["zip_filename"] = "invoices_xmlhtml.zip"
        elif xml_download_id:
            # Chỉ có XML
            data_obj["download_id"] = xml_download_id
            data_obj["zip_filename"] = xml_filename
        elif html_download_id:
            # Chỉ có HTML
            data_obj["download_id"] = html_download_id
            data_obj["zip_filename"] = html_filename
        
        response = {
            "status": "success",
            "message": f"Hoàn tất tải xml/html {len(datas_first['datas'])} hóa đơn",
            "xml_list": xml_list,
            "html_list": html_list,
            "data": data_obj
        }
        return response
    
    def html2pdf(self, html_list=[],progress_callback=None):
        try:
            from playwright.sync_api import sync_playwright
            import tempfile
            self.progress_callback = progress_callback  # Lưu callback
            
            pdf_buffer_list = []
            total_pdfs = len(html_list)
            for idx, item in enumerate(html_list, 1):
                # ✅ Check cancelled flag trước khi xử lý mỗi PDF
                if self._check_cancelled():
                    raise Exception("Job đã bị hủy (Ctrl+C)")
                
                # 📊 Báo tiến trình cho mỗi PDF được convert
                if self.progress_callback:
                    self.progress_callback(
                        current_step=f"Đang chuyển đổi sang PDF {idx}/{total_pdfs}...",
                        processed=idx,
                        total=total_pdfs
                    )
                
                if isinstance(item, dict):
                    khhdon = item.get('khhdon', 'unknown')
                    shdon = item.get('shdon', 'unknown')
                    khmshdon = item.get('khmshdon', 'unknown')
                    xml_content = item.get('xml_content', '')
                else:
                    khhdon = 'unknown'
                    shdon = idx
                    khmshdon = 'unknown'
                    xml_content = item
                
                try:
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as tmp_html:
                        tmp_html.write(xml_content)
                        tmp_html_path = tmp_html.name
                    
                    with sync_playwright() as p:
                        browser = p.chromium.launch(headless=True)
                        page = browser.new_page()
                        page.goto(f'file:///{os.path.abspath(tmp_html_path)}')
                        pdf_bytes = page.pdf(format='A4', landscape=False)
                        
                        browser.close()
                    os.unlink(tmp_html_path)
                    
                    pdf_buffer_list.append({
                        "khhdon": khhdon,
                        "shdon": shdon,
                        "khmshdon": khmshdon,
                        "index": idx,
                        "pdf_bytes": pdf_bytes,
                        "filename": f"{khhdon}_{shdon}.pdf"
                    })
                    print(f"✓ [{idx}] Converted: {khhdon}_{shdon}.pdf")
                except Exception as e:
                    print(f"❌ [{idx}] Error converting {khhdon}_{shdon}: {e}")
                    continue
            
            # Nén tất cả PDF vào ZIP buffer (in-memory)
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for item in pdf_buffer_list:
                    filename = item["filename"]
                    pdf_bytes = item["pdf_bytes"]
                    zip_file.writestr(filename, pdf_bytes)
            
            zip_buffer.seek(0)
            zip_bytes_data = zip_buffer.getvalue()
            
            # ✅ Lưu file vào disk và tạo download_id (giống Go-Soft pattern)
            pdf_download_id = None
            try:
                import sys
                import os as os_module
                sys.path.insert(0, os_module.path.dirname(os_module.path.dirname(os_module.path.dirname(os_module.path.abspath(__file__)))))
                from shared.download_service import save_file_to_disk
                
                pdf_download_id, pdf_file_path = save_file_to_disk(zip_bytes_data, 'zip')
                logger.info(f"✅ Đã lưu PDF ZIP file: {pdf_file_path} (download_id: {pdf_download_id})")
            except Exception as e:
                logger.error(f"❌ Lỗi khi lưu PDF ZIP file vào disk: {e}")
                # Fallback: vẫn tạo base64 nếu không lưu được
                zip_bytes = base64.b64encode(zip_bytes_data).decode('utf-8')
            
            response = {
                "status": "success",
                "message": f"Hoàn tất chuyển {len(pdf_buffer_list)}/{len(html_list)} PDF",
                "data": {
                    "filename": "invoices_pdf.zip",
                    "total_pdf": len(pdf_buffer_list),
                    "download_id": pdf_download_id,  # ✅ Trả về download_id thay vì zip_bytes
                    # ✅ Backward compatibility: vẫn có zip_bytes nếu không lưu được vào disk
                    "zip_bytes": base64.b64encode(zip_bytes_data).decode('utf-8') if pdf_download_id is None else None
                },
                # ✅ KHÔNG trả về pdf_list nữa (quá lớn, có thể gây tràn RAM)
                # "pdf_list": pdf_buffer_list  # ❌ Comment out để tránh tràn RAM
            }
            print(f"✓ ZIP created with {len(pdf_buffer_list)} PDFs")
            return response
            
        except Exception as e:
            print(f"❌ Error in html2pdf: {e}")
            return {
                "status": "error",
                "message": f"Lỗi khi chuyển PDF: {e}",
                "data": {}
            }
        