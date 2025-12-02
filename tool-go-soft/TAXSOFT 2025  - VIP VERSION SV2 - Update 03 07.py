try:
    from PyQt5.QtWidgets import QApplication,QMainWindow,QMessageBox
    from PyQt5 import QtWidgets
    import shutil
    from selenium.webdriver.support.select import Select
    from PyQt5.QtCore import QDate
    import unidecode
    from PyQt5.QtGui import QPixmap
    from datetime import  timedelta,datetime
    from PyQt5.QtWidgets import QFileDialog
    import os,re
    import xml.etree.ElementTree as ET
    from openpyxl import Workbook
    import sys
    import time
    import base64
    import selenium
    import pprint
    from PyQt5 import QtCore, QtGui, QtWidgets
    from PyQt5.QtGui import QPixmap, QImage
    from PyQt5.QtCore import QByteArray, QBuffer
    from PyQt5.QtWidgets import QMainWindow
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from PyQt5.QtCore import QByteArray
    from PyQt5.QtGui import QPixmap
    import hashlib
    import math
    import socket
    import openpyxl
    import json
    from openpyxl.styles import Font, Border, Side, Alignment
    import os
    import io
    from requests import get
    from openpyxl import load_workbook
    from __pycache__.ui import TOOL1,KEY_ERROR,THONGBAO1,LOGIN,captcha_CHECK,choose_tk
    from PyQt5.QtWidgets import QMessageBox, QPushButton
    from requests import adapters
    from urllib3 import poolmanager
    from ssl import create_default_context, Purpose, CERT_NONE
    from requests import Session
    from openpyxl.utils import get_column_letter
    from PyQt5.QtGui import QDesktopServices
    from PyQt5.QtCore import Qt, QUrl
    import zipfile
    from pyhtml2pdf import converter
    import uuid,wmi
except Exception as e:
    print(e)
    import os
    os.system('py -m pip install -r requirements.txt')
download_path = '.\\__pycache__\\cache_' 
pb_t = "v3.5.9 (03/07/2025)"
full_path = os.path.abspath(download_path)

'''class CustomHttpAdapter (adapters.HTTPAdapter):
    def __init__(self, ssl_context=None, **kwargs):
        self.ssl_context = ssl_context
        super().__init__(**kwargs)
    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = poolmanager.PoolManager(
            num_pools=connections, maxsize=maxsize,
            block=block, ssl_context=self.ssl_context)
def log_request(**kwargs):
                                            if 'request' in kwargs:
                                                request = kwargs['request']
                                                print(f"Request URL: {request['url']}")    '''        
'''def ssl_supressed_session():
    ctx = create_default_context(Purpose.SERVER_AUTH)
    # to bypass verification after accepting Legacy connections
    ctx.check_hostname = False
    ctx.verify_mode = CERT_NONE
    # accepting legacy connections
    ctx.options |= 0x4    
    session = Session()
    session.mount('https://', CustomHttpAdapter(ctx))
    return session'''
def get_disk_info():
    disk_info = {}
    import wmi
    c = wmi.WMI()
    for disk in c.Win32_DiskDrive():
        serial_number = disk.SerialNumber.strip()
    disk_info['serial_number'] = serial_number if serial_number else 'N/A'

    return disk_info

proxies = {
    "http":'http://103.82.138.76:10906:muaxu10906:PEKer'
}

#////////////////////////////////////////////////////////////////////////////////////////
class MAIN():
    def __init__(self):
        self.ch = 0
        self.thongbao1_ = QMainWindow()
        self.thongbao1 = THONGBAO1()
        self.thongbao1.setupUi(self.thongbao1_)
        self.thongbao1_.show()
        self.check_key = self.load_key()
        if self.check_key:#
            self.check_user()
            self.thongbao1_.close()
        else:
            self.error_ = QMainWindow()
            self.error = KEY_ERROR()
            self.error.setupUi(self.error_)
            self.error.textBrowser.setText(self.pr_key)
            self.thongbao1_.close()
            self.error_.show()
    def get_current_date(self):
        ngay_hien_tai = datetime.now()
        ngay_thang_nam = ngay_hien_tai.strftime('%d/%m/%Y')
        return ngay_thang_nam
    def compare_dates(self,date1, date2):
        format = "%d/%m/%Y"
        datetime1 = datetime.strptime(date1, format)
        datetime2 = datetime.strptime(date2, format)
        if datetime1 < datetime2:
            print(0)
            return 0
        return 1
    def open_link(self):
        QDesktopServices.openUrl(QUrl(self.link_update))
    def load_key(self):
        #=================================
        disk_info = get_disk_info()
        mac_address =  disk_info['serial_number']
        data = f"{mac_address}"
        hash_object = hashlib.sha256(data.encode())
        authentication_key = hash_object.hexdigest()
        self.key = 'key' + authentication_key
        self.key = self.key[:-35]
        self.pr_key = self.key
        while True:
            try: 
                r = get('https://severtmoclan.click/SERVER2/TS/vip.txt',timeout=5)
                self.keys = r.text.split('\n')
                if self.keys[0].split("|")[0] == "update":
                    self.tbup = self.keys[0].split("|")[2]
                    print(len(self.tbup))
                    print(len(pb_t))
                    if pb_t == self.tbup:
                        print("TRÙNG")
                        pass
                    else:
                        self.link_update = self.keys[0].split("|")[1]
                        msg_box = QMessageBox()
                        msg_box.setWindowTitle('THÔNG BÁO UPDATE')
                        msg_box.setIcon(QMessageBox.Information)
                        self.get_current_date()
                        msg_box.setText(f'Đã có bản cập nhật {self.tbup}. Nhấn OK để cập nhật.')
                        ok_button = QPushButton('OK')
                        cancel_button = QPushButton('Cancel')
                        msg_box.addButton(ok_button, QMessageBox.AcceptRole)
                        msg_box.addButton(cancel_button, QMessageBox.RejectRole)
                        msg_box.setDefaultButton(ok_button)
                        user_choice = msg_box.exec()
                        if user_choice == QMessageBox.AcceptRole:
                            self.open_link()
                        elif user_choice == QMessageBox.RejectRole:
                            # Handle the cancel action here
                            pass
                for i in range(len(self.keys)):
                    self.keys[i] = self.keys[i].split('|')
                for i in range(len(self.keys)):
                    if self.key == self.keys[i][0]:
                        self.thongbao1_.close()
                        self.mst_gh = self.keys[i][4]
                        self.nhh = self.keys[i][2]
                        nht = self.get_current_date()
                        print(self.nhh,nht)
                        if self.compare_dates(self.nhh,nht) == 0:
                            print("thõa")
                            QMessageBox.critical(None, 'Error', 'Hết hạn sử dụng , vui lòng gia hạn !' )
                            time.sleep(99999999)
                            time.sleep(99999999)
                            return False
                        return True
                self.thongbao1_.close()
                return False
            except:
                time.sleep(3)
                print("try")
                pass
        #=================================   
    def check_user(self):
        #login
        if self.ch == 1:
            self.tool.close()
        if self.check_key: #
            self.login = QMainWindow()
            self.loginweb = LOGIN() 
            self.loginweb.setupUi(self.login)
            for i in range(len(self.keys)):
                if self.key == self.keys[i][0]:
                    self.key_ = self.keys[i]
                    if self.key_[2] == "l":
                        self.color = 'blue'
                        self.type_v_l = 'Hạn chế'
                    else:
                        print("VIP ")
                        self.color = 'green'
                        self.type_v_l = 'VIP'
            #====================
            with open(".\\__pycache__\\cache_\\user.txt", "r") as file:
                # Đọc nội dung tệp tin
                line = file.readline()
                line = line.strip()
                self.user, self.password = line.strip().split("|")
            if self.key_[2] == 'l':
                self.loginweb.lineEdit.setText(self.key_[3])
                self.loginweb.lineEdit_2.setText(self.password)
            else:
                self.loginweb.lineEdit.setText(self.user)
                self.loginweb.lineEdit_2.setText(self.password)
                #--=-=--=-=--==-=-=--=-=-=-=-=-=-=-=
            self.check_h = 0
            if self.check_h == 0:
                self.loginweb.lineEdit_2.setEchoMode(QtWidgets.QLineEdit.Password)
            self.loginweb.password_toggle_btn.clicked.connect(self.show_hide)
            self.loginweb.close_btn.clicked.connect(QtWidgets.QApplication.instance().quit)
            self.loginweb.pushButton.clicked.connect(self.show_load)
            self.login.show()
        else:
            self.error_ = QMainWindow()
            self.errorKey = KEY_ERROR()
            self.errorKey.setupUi(self.error_)
            self.errorKey.textBrowser.setText(self.pr_key)
            self.error_.show()
    def show_hide(self):
        if self.check_h == 1:
            self.loginweb.lineEdit_2.setEchoMode(QtWidgets.QLineEdit.Normal)
            self.check_h = 0
        else:
            self.loginweb.lineEdit_2.setEchoMode(QtWidgets.QLineEdit.Password)
            self.check_h = 1
        
    def show_load(self):
        self.user = self.loginweb.lineEdit.text()  
        self.password = self.loginweb.lineEdit_2.text()

        self.in_info()
    def in_info(self):
         
        if self.user not in self.mst_gh.split(","):
            if self.mst_gh == 'o':
                pass
            else:
                QMessageBox.critical(None, 'Error', 'Tài khoản bị hạn chế !' )
                self.load_key()
                return
        options = webdriver.ChromeOptions()
        options.add_argument('ignore-certificate-errors')
        options.add_argument("--disable-animations")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-features=VideoPlayback,Images")
        options.set_capability("goog:loggingPrefs", {"performance": "ALL"})
        path_f = os.path.abspath(r".\__pycache__\cache\cache_tk")
        folder_path = path_f  
        # Kiểm tra xem thư mục tồn tại hay không
        if os.path.exists(folder_path):
            # Lặp qua tất cả các file trong thư mục
            for filename in os.listdir(folder_path):
                file_path = os.path.join(folder_path, filename)
                # Kiểm tra xem đường dẫn có phải là file hay không
                if os.path.isfile(file_path):
                    # Xóa file
                    os.remove(file_path)
            print("Đã xóa tất cả các file trong thư mục.")
        else:
            print("Thư mục không tồn tại.")
        self.path_f1 = path_f
        options.add_experimental_option("prefs", {
            "download.default_directory": path_f,
            "safebrowsing.enabled": True
        })
        from selenium.webdriver.chrome.service import Service
        options.add_argument("--headless") #Hiện chrome/Ẩn Chrome
        self.driver = webdriver.Chrome(options=options)
        self.driver.get('https://thuedientu.gdt.gov.vn/')
        wait = WebDriverWait(self.driver,3)
        while True:
            try:
                DN = wait.until(EC.presence_of_element_located((By.XPATH, '//a[span[text()="DOANH NGHIỆP"]]')))
                DN.click()
                break
            except:
                self.driver.get('https://thuedientu.gdt.gov.vn/')
                time.sleep(1)
        from selenium.webdriver.common.alert import Alert
        time.sleep(4)
        alert = Alert(self.driver)
        alert.accept() 
        login_bt = self.driver.find_element(By.XPATH, '//div[@class="dangnhap"]')
        login_bt = login_bt .find_elements(By.TAG_NAME, 'span')
        if len(login_bt) >= 2:
            login_bt[1].click()
        time.sleep(3)
        try:
            self.driver.execute_script("popupThungo();")
        except:
            self.driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(3)
            self.driver.execute_script("popupThungo();")
        try:
            element = WebDriverWait(self.driver, 10).until( EC.element_to_be_clickable((By.XPATH, "//*[text()='Đăng nhập bằng tài khoản Thuế điện tử']")))
            element.click()
        except:
            pass
        time.sleep(2)
        user_ip =  self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/div/div[2]/form/table/tbody/tr[1]/td/input')
        user_ip.send_keys(self.user)
        pass_ip =  self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/div/div[2]/form/table/tbody/tr[2]/td/input')
        pass_ip.send_keys(self.password)
        img_element = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/div/div[2]/form/table/tbody/tr[4]/td/div/div[2]/img')
        resources = self.driver.execute_script("return performance.getEntriesByType('resource');")
        # Duyệt qua các requests mạng
        for request in resources:
            # Kiểm tra xem request có phải là ImageServlet
            if 'ImageServlet' in request['name']:
                resource_name = request['name']
                script = f"""
                    var img = new Image();
                    img.src = '{resource_name}';
                    var canvas = document.createElement('canvas');
                    var ctx = canvas.getContext('2d');
                    ctx.drawImage(img, 0, 0);
                    var base64Data = canvas.toDataURL('image/png').replace(/^data:image\/(png|jpg|jpeg);base64,/, '');
                    return base64Data;
                """
                base64_data = self.driver.execute_script(script)
                # In base64 ra console để kiểm tra
                print(base64_data)
        
        self.check_click = 0
        self.captcha_(base64_data)
    def captcha_(self, imagen64):
        # Tạo QImage
        imagen = QtGui.QImage()
        bytearr = QtCore.QByteArray.fromBase64(imagen64.encode())
        pprint.pprint(bytearr.length())
        imagen.loadFromData(bytearr, 'PNG')
        pixmap = QtGui.QPixmap.fromImage(imagen)
        self.captcha = QMainWindow()
        self.captcha_check = captcha_CHECK()
        self.captcha_check.setupUi(self.captcha)
        self.captcha_check.image.setPixmap(pixmap)
        self.captcha_check.pushButton.clicked.connect(self.set_captcha_inp)
        self.captcha.setWindowModality(1)
        self.captcha.show()
        '''self.TBLOAD.close()'''

    def set_captcha_inp(self):
        self.captcha_inp = self.captcha_check.lineEdit.text()
        self.login_web()
    def remove_accents(self,text):
        
        return unidecode.unidecode(text)
    def login_web(self):
            try:
                captcha_ip = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/div/div[2]/form/table/tbody/tr[4]/td/div/div[1]/input')
                captcha_ip.send_keys(self.captcha_inp)
            except:
                pass
            time.sleep(0.5)                             
            try:
                begin_lg = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/div/div[2]/form/table/tbody/tr[5]/td/input')
            except:
                try:
                    begin_lg = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/div/div[2]/form/table/tbody/tr[4]/td/input')
                except:
                    pass
            begin_lg.click()
            d = 0
            while True:
                    self.token_ = 0
                    self.true_acc = True
                    if self.key_[2] == 'l':  
                        if self.key_[3] != self.user:
                            self.true_acc = False
                    try:
                        self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/div/div[2]/form/table/tbody/tr[5]/td/input')
                        self.true_acc = 0
                    except:
                        self.true_acc = 1
                    if self.true_acc:
                        self.captcha.close()
                        print(self.user)

                        with open(".\\__pycache__\\cache_\\user.txt", "w") as file:
                            file.write(f"{self.user.rstrip()}|{self.password.rstrip()}")
                        self.tool_RUN()
                        print("ok1")
                        break
                    else:
                        print("err1")
                        self.captcha.close()
                        QMessageBox.critical(None, 'Error', 'Nhập sai tài khoản, mật khẩu hoặc captcha !' )
                        break
                    break

                    
    #=========================== TOOL ===========================
    def tool_RUN(self):
        logs = self.driver.get_log("performance")
        urls = []
        # Duyệt log để trích URL từ request
        for entry in logs:
            try:
                message = json.loads(entry["message"])["message"]
                if message["method"] == "Network.requestWillBeSent":
                    url = message["params"]["request"]["url"]
                    print(url)
                    urls.append(url)
            except Exception:
                pass
        ssid = "NotFound"
        for url in urls:
            match = re.search(r"&dse_sessionId=([^&]+)&", url)
            if match:
                ssid = match.group(1)
                print(f"✅ Found dse_sessionId: {ssid}")
        self.ss = ssid
        wait = WebDriverWait(self.driver, 7)
        try:    
            wait.until(EC.presence_of_element_located((By.XPATH,'/html/body/div/div[2]/ul/li[3]')))
        except:
            QMessageBox.critical(None, 'Error', 'Nhập sai tài khoản, mật khẩu hoặc captcha !' )
            return
        kthue = self.driver.find_element(By.XPATH,'/html/body/div/div[2]/ul/li[3]' )
        kthue.click()
        wait.until(EC.presence_of_element_located((By.XPATH,'/html/body/div/div[3]/div/div[3]/ul/li[8]')))
        tcuutkhai = self.driver.find_element(By.XPATH, '/html/body/div/div[3]/div/div[3]/ul/li[8]')
        tcuutkhai.click()
        print("GO GO")
        
        self.login.close()
        self.captcha.close()
        self.tool = QMainWindow()
        self.tool_run = TOOL1()
        self.tool_run.setupUi(self.tool,0)
        default_year_begin = 2023  
        default_month_begin = 1  
        default_day_begin = 1
        default_date_begin = QDate(default_year_begin, default_month_begin, default_day_begin)
        self.tool_run.begin.setDate(default_date_begin)
        default_year_begin = 2023 
        default_month_begin = 2
        default_day_begin = 2 
        default_date_begin = QDate(default_year_begin, default_month_begin, default_day_begin)
        self.tool_run.end.setDate(default_date_begin)
        #--------------folder lưu trữ--------------------
        self.selected_directory = ""
        self.tool_run.pathfolder.clicked.connect(self.openDirectoryDialog)
        self.tool_run.c_xml_p.clicked.connect(self.openDirectoryDialog1)
        self.tool_run.type_tk_3.clicked.connect(self.list_tk)
        self.tool_run.tracuu.clicked.connect(self.checktool)
        self.tool_run.ok_bd.clicked.connect(self.xml2xlsx)
        if self.tool_run.change_acc.clicked:
            self.ch = 1
            self.tool_run.change_acc.clicked.connect(self.check_user)
        self.tool.show()
    def format_number(self,num_str):
        num = int(num_str)
        num_str = str(num)
        result = ""
        for i, digit in enumerate(reversed(num_str)):
            if i > 0 and i % 3 == 0:
                result = "." + result
            result = digit + result
        return result
    def xml2xlsx(self):
        print("DÔ")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Tên',  'Kỳ tính thuế Tháng/Quý', 'Lần','Năm', 'VAT đầu kỳ', 'Giá trị HH mua vào', 'VAT mua vào', 'VAT được khấu trừ kỳ này','Giá trị HH bán ra', 'VAT bán ra','Điều chỉnh tăng','Điều chỉnh giảm','Thuế vãng lai ngoại tỉnh', 'VAT còn phải nộp', 'VAT còn được khấu trừ chuyển kỳ sau'])
        len_n = len(os.listdir(self.selected_directoryz))
        n_b = 0
        div_n = 100/len_n

        for filename in os.listdir(self.selected_directoryz):
            
            n_b +=div_n
            if n_b>90:
                n_b = 99
            QtWidgets.QApplication.processEvents()
            self.tool_run.progressBar.setValue(int(n_b))
            print(".")
            if filename.endswith('.xml'):
                tree = ET.parse(os.path.join(self.selected_directoryz, filename))
                root = tree.getroot()
                namespace = {'ns0': root.tag.split('}')[0][1:]}
                print(ET.tostring(root, encoding='unicode'))
                kyKKhai_element = root.find('.//ns0:kyKKhai', namespace)
                kytinhthue = kyKKhai_element.text if kyKKhai_element is not None else ''
                ky = kytinhthue.split("/")[0] if kytinhthue else ''
                nam = kytinhthue.split("/")[1] if kytinhthue else ''
                try:
                    soLan_text = filename.split("-")[2] + " " + filename.split("-")[3]
                except:
                    soLan_text = " "
                vat_dk = root.findall('.//ns0:ct22', namespace)
                vat_dk_text = vat_dk[0].text if vat_dk else ''
                hh_mv = root.findall('.//ns0:ct23', namespace)
                hh_mv_text = hh_mv[0].text if hh_mv else ''
                vat_mv = root.findall('.//ns0:ct24', namespace)
                vat_mv_text = vat_mv[0].text if vat_mv else ''
                vat_ktkn = root.findall('.//ns0:ct25', namespace)
                vat_ktkn_text = vat_ktkn[0].text if vat_ktkn else ''
                hh_br = root.findall('.//ns0:ct34', namespace)
                hh_br_text = hh_br[0].text if hh_br else ''
                vat_br = root.findall('.//ns0:ct35', namespace)
                vat_br_text = vat_br[0].text if vat_br else ''
                vat_cpn = root.findall('.//ns0:ct40', namespace)
                vat_cpn_text = vat_cpn[0].text if vat_cpn else ''
                vat_ktks = root.findall('.//ns0:ct43', namespace)
                vat_ktks_text = vat_ktks[0].text if vat_ktks else ''
                dct = root.findall('.//ns0:ct38', namespace)
                dct = dct[0].text if dct else ''
                dcg = root.findall('.//ns0:ct37', namespace)
                dcg = dcg[0].text if dcg else ''
                tvl = root.findall('.//ns0:ct39', namespace)
                tvl = tvl[0].text if tvl else ''
                print([filename, ky,soLan_text,  nam, vat_dk_text, hh_mv_text, vat_mv_text, vat_ktkn_text, hh_br_text,vat_br_text,dct,dcg,tvl, vat_cpn_text, vat_ktks_text])
                worksheet.append([
                    filename,  ky,soLan_text, nam, vat_dk_text, hh_mv_text, vat_mv_text, vat_ktkn_text, hh_br_text,
                    vat_br_text,dct,dcg,tvl, vat_cpn_text, vat_ktks_text
                ])
        name_ = self.selected_directoryz.split("/")[-1]
        path_save = os.path.join(self.selected_directory1,f"{name_}.xlsx")
        print(path_save)
        header_font = Font(bold=True)
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in range(1, worksheet.max_column+1):
            worksheet.cell(row=1, column=col).font = header_font
            worksheet.cell(row=1, column=col).border = header_border

        # Định dạng các ô còn lại
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in range(2, worksheet.max_row+1):
            for col in range(1, worksheet.max_column+1):
                worksheet.cell(row=row, column=col).border = cell_border
        for col in range(1, worksheet.max_column+1):
            # Get the header text for the current column
            header_text = worksheet.cell(row=1, column=col).value
            # Set the column width based on the length of the header text
            worksheet.column_dimensions[get_column_letter(col)].width = len(str(header_text)) + 10
        from openpyxl .styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
        for row in range(2, worksheet.max_row + 1):
            for col in range(5, 16): 
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                    except ValueError:
                        pass
        workbook.save(path_save)
        print("v")
        QtWidgets.QApplication.processEvents()
        self.tool_run.progressBar.setValue(100)
    def list_tk(self):
        
        self.choose_tkk = QMainWindow()
        self.choose = choose_tk()
        self.choose.setupUi(self.choose_tkk )
        self.choose_tkk.show()
        self.choose.optionSelected.connect(self.handle_option_selected)  # Kết nối tín hiệu optionSelected với khe handle_option_selected
        self.choose.pushButton.clicked.connect(self.close_tp)

    def handle_option_selected(self, option):
        print("Lựa chọn:", option)  # In ra lựa chọn được chọn từ choose_tk
        self.option_ = option
    def close_tp(self):
        self.choose_tkk.close()
        try:
            self.option_tk = self.option_.lstrip()
            self.tool_run.type_tk.setText(self.option_tk)
        except:
            QMessageBox.critical(None, 'Error', 'Chưa chọn loại tờ khai !')
    def checktool(self):
        type_ = 0
        self.list_new_name = []
        #============================
        date = self.tool_run.begin.date()
        self.day = date.day()
        self.month = date.month()
        self.year = date.year()
        self.begin_ = f'{self.day}/{self.month}/{self.year}'
        self.begin_CONST = self.begin_
        date2 = self.tool_run.end.date()
        self.day1 = date2.day()
        self.month1 = date2.month()
        self.year1 = date2.year()
        self.end_ = f'{self.day1}/{self.month1}/{self.year1}'
        self.end_CONST = self.end_
        self.be = f" {self.begin_CONST}_{self.end_CONST}"
        self.range_DAY_const = f' Thời gian : {self.begin_CONST}=>{self.end_CONST}'
        self.arr_ed = self.day_dow(self.begin_,self.end_) 
        print(self.arr_ed)
        self.path_ketqua = self.tool_run.prpath.text()
        #============================
        self.checkspc = " "
        #type
        self.tkk__ = self.tool_run.tkk_.isChecked()
        self.tbb__ = self.tool_run.tbb_.isChecked()
        self.gntt__ = self.tool_run.gntt_.isChecked()
        self.filter()
    def openDirectoryDialog(self):
        file_dialog = QFileDialog()
        file_dialog.setWindowTitle("Chọn thư mục")
        file_dialog.setFileMode(QFileDialog.Directory)
        file_dialog.setViewMode(QFileDialog.Detail)
        if file_dialog.exec():
            selected_directory = file_dialog.selectedFiles()[0]
            self.selected_directory = selected_directory
            print("Đường dẫn thư mục đã chọn:", selected_directory)
        try:
            self.selected_directory = selected_directory
            self.selected_directory1 = self.selected_directory 
            self.tool_run.prpath.setText(self.selected_directory)
        except:
            pass
    def openDirectoryDialog1(self):
        file_dialog = QFileDialog()
        file_dialog.setWindowTitle("Chọn thư mục")
        file_dialog.setFileMode(QFileDialog.Directory)
        file_dialog.setViewMode(QFileDialog.Detail)
        if file_dialog.exec():
            selected_directory = file_dialog.selectedFiles()[0]
            self.selected_directory = selected_directory
            print("Đường dẫn thư mục đã chọn:", selected_directory)
        try:
            self.selected_directoryz = selected_directory
            self.tool_run.path_xml.setText(self.selected_directoryz)
        except:
            pass
    def increase_date(self,date_string):
        try:
            # Chuyển đổi chuỗi ngày thành đối tượng datetime
            date = datetime.strptime(date_string, "%d/%m/%Y")
            
            # Tăng thêm một ngày
            increased_date = date + timedelta(days=1)
            
            # Chuyển đổi ngày mới thành chuỗi ngày
            increased_date_string = increased_date.strftime("%d/%m/%Y")
            
            return increased_date_string
        except ValueError:
            return "Định dạng ngày không hợp lệ!"
#//////////////////////////////////////////////////////////////////////////
    def day_dow(self,start_date, end_date):
        print("=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=")
        print(start_date)
        print(end_date)
        date_format = "%d/%m/%Y"  # Định dạng ngày tháng
        date1 = datetime.strptime(start_date, date_format)  # Chuyển đổi chuỗi thành đối tượng datetime
        date2 = datetime.strptime(end_date, date_format)  # Chuyển đổi chuỗi thành đối tượng datetime
        one_month = timedelta(days=350)  # Khoảng thời gian 29 ngày
        date_ranges = []  # Mảng chứa các mảng con
        print(date_ranges )
        while date1 <= date2:
            sub_array = []  # Mảng con
            sub_array.append(date1.strftime(date_format))  # Thêm giá trị date1 vào mảng con
            date1 += one_month  # Tăng date1 lên 29 ngày
            if date1 > date2:
                date1 = date2  # Nếu date1 vượt quá date2, gán date1 bằng date2
            sub_array.append(date1.strftime(date_format))  # Thêm giá trị date1 mới vào mảng con
            date_ranges.append(sub_array)  # Thêm mảng con vào mảng chứa các mảng con
            date1 += timedelta(days=1)  # Tăng date1 lên 1 ngày để tạo khoảng thời gian tiếp theo
        return date_ranges
    def day_dow2(self,start_date, end_date):
        print("=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=")
        print(start_date)
        print(end_date)
        date_format = "%d/%m/%Y"  # Định dạng ngày tháng
        date1 = datetime.strptime(start_date, date_format)  # Chuyển đổi chuỗi thành đối tượng datetime
        date2 = datetime.strptime(end_date, date_format)  # Chuyển đổi chuỗi thành đối tượng datetime
        one_month = timedelta(days=360)  # Khoảng thời gian 29 ngày
        date_ranges = []  # Mảng chứa các mảng con
        print(date_ranges )
        while date1 <= date2:
            sub_array = []  # Mảng con
            sub_array.append(date1.strftime(date_format))  # Thêm giá trị date1 vào mảng con
            date1 += one_month  # Tăng date1 lên 29 ngày
            if date1 > date2:
                date1 = date2  # Nếu date1 vượt quá date2, gán date1 bằng date2
            sub_array.append(date1.strftime(date_format))  # Thêm giá trị date1 mới vào mảng con
            date_ranges.append(sub_array)  # Thêm mảng con vào mảng chứa các mảng con
            date1 += timedelta(days=1)  # Tăng date1 lên 1 ngày để tạo khoảng thời gian tiếp theo
        return date_ranges
    def filter(self):
        print("CHECC")
        if self.tkk__:
            self.tt_khai()
        if self.tbb__:
            self.tt_bao()
        if self.gntt__:
            self.tgnt()
    def tgnt(self):
        print("TBB")
        QtWidgets.QApplication.processEvents()
        self.tool_run.progressBar.setValue(0)
        new_directory_name = self.user
        folder_son = os.path.join(self.path_ketqua, new_directory_name)
        if os.path.exists(self.path_ketqua):
            if os.path.exists(folder_son):
                pass
            else:
                os.makedirs(folder_son)
                print("Thư mục con đã được tạo thành công.")
        a = str("GNT "+self.be)
        a = a.replace("/","_")
        folder_tk = os.path.join(folder_son,a )
        self.folder_tk1 = folder_tk
        if os.path.exists(folder_son):
            if os.path.exists(folder_tk):
                folder_path = folder_tk  # Thay đổi đường dẫn này thành đường dẫn thực tế của thư mục bạn muốn xóa
                # Kiểm tra xem thư mục tồn tại hay không
                if os.path.exists(folder_path):
                    # Lặp qua tất cả các file trong thư mục
                    for filename in os.listdir(folder_path):
                        file_path = os.path.join(folder_path, filename)
                        # Kiểm tra xem đường dẫn có phải là file hay không
                        if os.path.isfile(file_path):
                            # Xóa file
                            os.remove(file_path)
                    print("Đã xóa tất cả các file trong thư mục.")
                else:
                    print("Thư mục không tồn tại.")
            else:
                os.makedirs(folder_tk)
                print(f"Thư mục con đã được tạo thành công: {folder_tk}")
        else:
            QMessageBox.critical(None, 'Error', 'Thư mục không tồn tại !')
        self.driver.switch_to.default_content()
        self.driver.find_element(By.XPATH, f"/html/body/div[1]/div[2]/ul/li[4]").click() 
        self.driver.find_element(By.XPATH, f"/html/body/div[1]/div[3]/div/div[4]/ul/li[4]").click() 
        try:    
            self.driver.switch_to.frame("mainframe")
        except:
            pass
        #----------------------------
        
        x = 0 
        n_tkk = 0
        self.list_new_name = []
        self.arr_ed = self.day_dow2(self.begin_,self.end_) 
        name_r = 0
        for i in self.arr_ed:
            for k in range(0,2):
                if k == 0:                              
                    try:                                                                                             
                        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div/div/table[1]/tbody/tr[4]/td[2]/input')))                                                    
                        self.driver.find_element(By.XPATH, "/html/body/form/div/div/table[1]/tbody/tr[4]/td[2]/input").clear()
                        self.driver.find_element(By.XPATH, "/html/body/form/div/div/table[1]/tbody/tr[4]/td[2]/input").send_keys(i[k])
                        time.sleep(1)
                    except:
                        pass
                else:                                                                                            
                    select_element = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div/div/table[1]/tbody/tr[4]/td[4]/input')))                        
                    self.driver.find_element(By.XPATH, "/html/body/form/div/div/table[1]/tbody/tr[4]/td[4]/input").click()
                    from selenium.webdriver.common.keys import Keys
                    self.driver.find_element(By.XPATH, "/html/body/form/div/div/table[1]/tbody/tr[4]/td[4]/input").send_keys(Keys.CONTROL, "a")
                    self.driver.find_element(By.XPATH, "/html/body/form/div/div/table[1]/tbody/tr[4]/td[4]/input").send_keys(i[k]) 
            QtWidgets.QApplication.processEvents()
            #crawl                             
            self.driver.find_element(By.XPATH, "/html/body/form/div/div/table[2]/tbody/tr/td/div/input").click()
            check = 1
            check = 1
            check_pages = 1
            while check_pages:
                try:   
                    time.sleep(2)  
                    WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div/div[2]/div/div[2]/div/table/tbody[1]')))                                       
                    tb_tk = self.driver.find_element(By.XPATH, "/html/body/form/div/div[2]/div/div[2]/div/table/tbody[1]")
                except:
                    check = 0
                
                if check == 0:
                    print("KHÔNG CÓ HÓA ĐƠN TGIAN NÀY")
                else:
                    time.sleep(0.5)
                    #XL                                         /html/body/form/div/div[2]/div/div[2]/div/table/tbody[1]
                    tb_tk = self.driver.find_element(By.XPATH, "/html/body/form/div/div[2]/div/div[2]/div/table/tbody[1]")
                    trs = tb_tk.find_elements(By.TAG_NAME, "tr")
                    x+= len(trs)
                    q = 0
                    n_count = 0
                    n_0 = 0
                    n_load = len(trs)
                    n_tkk += n_load
                    for tr in trs: 
                        div_pt = 100/n_load
                        n_count+=1
                        QtWidgets.QApplication.processEvents()
                        n_0+=div_pt
                        self.tool_run.progressBar.setValue(int(n_0))
                        QtWidgets.QApplication.processEvents()
                        self.tool_run.labeltb1.setText(f"[{n_count}/{n_load} | Tổng {n_tkk}]")
                        try:
                            tds = tr.find_elements(By.XPATH, "./td")
                            id_tk = tds[2].text
                            for i in range(14,21):
                                try:
                                    if i == 17 or i == 18 or i == 19 or i == 20: 
                                            dd = tds[i].find_elements(By.TAG_NAME, "a")
                                            if len(dd) == 2:
                                                dd[1].click()
                                            else:
                                                tds[i].find_element(By.TAG_NAME, "a").click()
                                            q+=1
                                            new_path = os.path.join(self.path_f1, f'chungtu{name_r}.xml')
                                            name_r+=1
                                            old = os.path.join(self.path_f1, f'chungtu.xml')
                                            time.sleep(0.5)
                                            os.rename(old , new_path)
                                            time.sleep(0.5)
                                except:
                                    pass
                        except:
                            pass
                        QtWidgets.QApplication.processEvents()
                        self.tool_run.progressBar.setValue(99)
                wait = WebDriverWait(self.driver, 10)
                try:
                    pagination_button = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "img[src='/etaxnnt/static/images/pagination_right.gif']")))
                    pagination_button.click()
                    print("Đã nhấn vào nút phân trang")
                except:
                    check_pages = 0
        time.sleep(2)
        file_list = os.listdir(self.path_f1)
        print(file_list)
        nnn = 0
        for file_name in file_list:
                        nnn+=1
                        print(file_name)
                        file_path = os.path.join(self.path_f1, file_name)
                        # Trích xuất nội dung yêu cầu
                        from bs4 import BeautifulSoup
                        with open(file_path, 'r', encoding='utf-8') as f:
                            soup = BeautifulSoup(f, 'html.parser')
                        mgd = soup.find('ma_ndkt').text
                        n_lap = soup.find('ngay_lap').text
                        n_lap = n_lap.replace("/","-")
                        ttb = soup.find('ma_chuong').text
                        ttb_2 = soup.find('ky_thue').text
                        ttb_2 = ttb_2.replace("/","-")
                        file_path = str(file_path.replace("\\", "\**").replace("**",""))
                        new_file_name = f"{mgd} - {ttb} - Kynopthue - {ttb_2} - Ngaynopthue - {n_lap} [{nnn}].xml" 
                        print(new_file_name)
                        new_file_name  = self.remove_accents(new_file_name)
                        print(new_file_name)
                        # Đường dẫn và tên tệp tin mới
                        new_file_path = os.path.join(self.folder_tk1, new_file_name)
                        # Di chuyển và đổi tên tệp tin
                        new_file_path = new_file_path.replace("\\", "\**").replace("**","")
                        file_path_c = file_path.split(".")
                        if len(file_path_c)>2:
                            file_path = str(file_path_c[0]+"."+file_path_c[1])
                        print(f"[{file_path}] ==> [{new_file_path}]")
                        shutil.move(file_path, new_file_path)
        time.sleep(0.5)
        for file_name in file_list:
            file_path = os.path.join(self.path_f1, file_name)
            file_path = str(file_path.replace("\\", "\**").replace("**",""))
            # Duyệt qua mảng lớn
            new_file_name = ""   
            new_file_path = os.path.join(self.folder_tk1, file_name)
            # Di chuyển và đổi tên tệp tin
            new_file_path = new_file_path.replace("\\", "\**").replace("**","") 
            try:    
                shutil.move(file_path, new_file_path)
            except:
                pass
        QtWidgets.QApplication.processEvents()
        self.tool_run.progressBar.setValue(100)
        self.tool_run.labeltb1.setText(f"HOÀN THÀNH [{n_tkk}]")
    def tt_bao(self):
        print("TBB")
        QtWidgets.QApplication.processEvents()
        self.tool_run.progressBar.setValue(0)
        new_directory_name = self.user
        folder_son = os.path.join(self.path_ketqua, new_directory_name)
        if os.path.exists(self.path_ketqua):
            if os.path.exists(folder_son):
                pass
            else:
                os.makedirs(folder_son)
                print("Thư mục con đã được tạo thành công.")
        a = str("TB "+self.be)
        a = a.replace("/","_")
        folder_tk = os.path.join(folder_son,a )
        self.folder_tk1 = folder_tk
        if os.path.exists(folder_son):
            if os.path.exists(folder_tk):
                folder_path = folder_tk  # Thay đổi đường dẫn này thành đường dẫn thực tế của thư mục bạn muốn xóa
                # Kiểm tra xem thư mục tồn tại hay không
                if os.path.exists(folder_path):
                    # Lặp qua tất cả các file trong thư mục
                    for filename in os.listdir(folder_path):
                        file_path = os.path.join(folder_path, filename)
                        # Kiểm tra xem đường dẫn có phải là file hay không
                        if os.path.isfile(file_path):
                            # Xóa file
                            os.remove(file_path)
                    print("Đã xóa tất cả các file trong thư mục.")
                else:
                    print("Thư mục không tồn tại.")
            else:
                os.makedirs(folder_tk)
                print(f"Thư mục con đã được tạo thành công: {folder_tk}")
        else:
            QMessageBox.critical(None, 'Error', 'Thư mục không tồn tại !')
        self.driver.switch_to.default_content()
        time.sleep(1)
        self.driver.find_element(By.XPATH, f"/html/body/div[1]/div[3]/div/div[3]/ul/li[9]").click() 
        try:    
            self.driver.switch_to.frame("mainframe")
        except:
            pass
        #----------------------------
        
        x = 0 
        n_tkk = 0
        self.list_new_name = []
        for i in self.arr_ed:
            for k in range(0,2):
                if k == 0:      
                    while True: 
                        try:
                            select_element = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[2]/div[1]/table[1]/tbody/tr[3]/td[2]/input')))                       
                            self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[1]/table[1]/tbody/tr[3]/td[2]/input").clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[1]/table[1]/tbody/tr[3]/td[2]/input").send_keys(i[k])
                            break
                        except Exception as e:
                            print("Thử lại ")
                            time.sleep(2)
                            print(e)
                            pass
                else:   
                    self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[1]/table[1]/tbody/tr[3]/td[3]/input").click()
                    self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[1]/table[1]/tbody/tr[3]/td[3]/input").clear()
                    self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[1]/table[1]/tbody/tr[3]/td[3]/input").send_keys(i[k]) 
            QtWidgets.QApplication.processEvents()
            #crawl
            self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[1]/table[2]/tbody/tr/td/div/input").click()
            check = 1
            time.sleep(2)
            try:                                            
                tb_tk = self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[3]/div[2]/div[2]/div/div/table/tbody")
            except:
                check = 0
            
            if check == 0:
                print("KHÔNG CÓ HÓA ĐƠN TGIAN NÀY")
            else:
                time.sleep(0.5)
                #XL
                tb_tk = self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div[3]/div[2]/div[2]/div/div/table/tbody")
                trs = tb_tk.find_elements(By.TAG_NAME, "tr")
                x+= len(trs)
                q = 0
                n_count = 0
                n_0 = 0
                n_load = len(trs)
                n_tkk += n_load
                for tr in trs: 
                    time.sleep(0.8)
                    div_pt = 100/n_load
                    n_count+=1
                    QtWidgets.QApplication.processEvents()
                    n_0+=div_pt
                    self.tool_run.progressBar.setValue(int(n_0))
                    QtWidgets.QApplication.processEvents()
                    self.tool_run.labeltb1.setText(f"[{n_count}/{n_load} | Tổng {n_tkk}]")
                    tds = tr.find_elements(By.XPATH, "./td")
                    id_tk = tds[2].text
                    for i in range(3,6):
                        
                        if i == 5:
                                tds[i].find_element(By.TAG_NAME, "a").click()
                                q+=1
                                time.sleep(0.2)
                    QtWidgets.QApplication.processEvents()
                    self.tool_run.progressBar.setValue(99)
        time.sleep(2)
        file_list = os.listdir(self.path_f1)
        print(len(file_list))
        n = 0
        for file_name in file_list:
                        n+=1
                        print(file_name)
                        file_path = os.path.join(self.path_f1, file_name)
                        # Trích xuất nội dung yêu cầu
                        from bs4 import BeautifulSoup
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                soup = BeautifulSoup(f, 'html.parser')
                            mgd = soup.find('magiaodichdtu').text
                            ttb = soup.find('tentbao').text
                            maTBao = soup.find('matbao').text
                            if "Tiếp nhận" in ttb:
                                ttb = "Tiếp nhận"
                            if "Xác nhận" in ttb:
                                ttb = "Xác nhận"
                            if maTBao == 844:
                                ttb = "Không chấp nhận"
                            else:
                                if maTBao == 451:
                                    ttb = "Chấp nhận"
                            try:
                                if ttb == "Tiếp nhận":
                                    try:
                                        ttb_2 = soup.find('ngaytbao').text
                                    except:
                                        print(soup)
                                else:
                                    ttb_2 = soup.find('ngaychapnhan').text
                                    ttb_2 = ttb_2.split("T")[0]
                            except:
                                ttb_2 = "X"
                            file_path = str(file_path.replace("\\", "\**").replace("**",""))
                            ttb_2 = ttb_2.replace("/","-")
                            new_file_name = f"{mgd} - {ttb} - {ttb_2}.xml"
                            new_file_name  = self.remove_accents(new_file_name)
                            print(new_file_name)
                            # Đường dẫn và tên tệp tin mới
                            new_file_path = os.path.join(self.folder_tk1, new_file_name)
                            # Di chuyển và đổi tên tệp tin
                            new_file_path = new_file_path.replace("\\", "\**").replace("**","")
                            file_path_c = file_path.split(".")
                            if len(file_path_c)>2:
                                file_path = str(file_path_c[0]+"."+file_path_c[1])
                            print(f"[{file_path}] ==> [{new_file_path}]")
                            shutil.move(file_path, new_file_path)
                        except:
                            pass
        time.sleep(0.5)
        for file_name in file_list:
            file_path = os.path.join(self.path_f1, file_name)
            file_path = str(file_path.replace("\\", "\**").replace("**",""))
            # Duyệt qua mảng lớn
            new_file_name = ""   
            new_file_path = os.path.join(self.folder_tk1, file_name)
            # Di chuyển và đổi tên tệp tin
            new_file_path = new_file_path.replace("\\", "\**").replace("**","") 
            try:    
                shutil.move(file_path, new_file_path)
            except:
                pass
        QtWidgets.QApplication.processEvents()
        self.tool_run.progressBar.setValue(100)
        self.tool_run.labeltb1.setText(f"HOÀN THÀNH [{n_tkk}]")
    # Lắng nghe sự kiện mạng
    

    def tt_khai(self):  

        QtWidgets.QApplication.processEvents()
        self.tool_run.progressBar.setValue(0)
        new_directory_name = self.user
        folder_son = os.path.join(self.path_ketqua, new_directory_name)
        if os.path.exists(self.path_ketqua):
            if os.path.exists(folder_son):
                pass
            else:
                os.makedirs(folder_son)
                print("Thư mục con đã được tạo thành công.")
        a = str(self.option_tk + self.be)
        a = a.replace("/","_")
        a  = self.remove_accents(a)
        folder_tk = os.path.join(folder_son,a )
        self.folder_tk1 = folder_tk
        if os.path.exists(folder_son):
            if os.path.exists(folder_tk):
                folder_path = folder_tk  # Thay đổi đường dẫn này thành đường dẫn thực tế của thư mục bạn muốn xóa
                # Kiểm tra xem thư mục tồn tại hay không
                if os.path.exists(folder_path):
                    # Lặp qua tất cả các file trong thư mục
                    for filename in os.listdir(folder_path):
                        file_path = os.path.join(folder_path, filename)
                        # Kiểm tra xem đường dẫn có phải là file hay không
                        if os.path.isfile(file_path):
                            # Xóa file
                            os.remove(file_path)
                    print("Đã xóa tất cả các file trong thư mục.")
                else:
                    print("Thư mục không tồn tại.")
            else:
                os.makedirs(folder_tk)
                print(f"Thư mục con đã được tạo thành công: {folder_tk}")
        else:
            QMessageBox.critical(None, 'Error', 'Thư mục không tồn tại !')
        self.driver.switch_to.default_content()
        time.sleep(1)                    
        self.driver.find_element(By.XPATH, f"/html/body/div[1]/div[3]/div/div[3]/ul/li[8]").click() 
        self.driver.switch_to.frame("mainframe")
        select_element = WebDriverWait(self.driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[2]/div/table[1]/tbody/tr[1]/td[2]/select'))
        )
        select_element.click()
        time.sleep(3)
        try:
            print(self.option_tk)
            self.driver.find_element(By.XPATH, f"//select/option[contains(text(), '{self.option_tk}')]").click()
        except:
            print("LỖI ") 
            self.driver.find_element(By.XPATH, f"//select/option[contains(text(), '{self.option_tk}')]").click()
        x = 0 
        n_tkk = 0
        self.list_new_name = []
        for i in self.arr_ed:
            for k in range(0,2):
                if k == 0:    
                    while True:
                        try:                 
                            self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div/table[1]/tbody/tr[3]/td[2]/input").clear()
                            self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div/table[1]/tbody/tr[3]/td[2]/input").send_keys(i[k])
                            break
                        except:
                            print("Thử lại")
                            pass

                else:
                    self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div/table[1]/tbody/tr[3]/td[4]/input").click()
                    self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div/table[1]/tbody/tr[3]/td[4]/input").clear()
                    self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div/table[1]/tbody/tr[3]/td[4]/input").send_keys(i[k]) 
            QtWidgets.QApplication.processEvents()
            #crawl
            self.driver.find_element(By.XPATH, "/html/body/form/div[2]/div/table[2]/tbody/tr/td/div/input").click()
            check = 1
            check_pages = 1
            cache_id = []
            ssid = self.ss
            while check_pages:
                try:
                    tb_tk = self.driver.find_element(By.XPATH, "/html/body/form/div[4]/div[2]/div[2]/div/div/table/tbody")
                except:
                    check +=1
                
                if check == 15:
                    print("KHÔNG CÓ HÓA ĐƠN TGIAN NÀYY")
                else:
                    time.sleep(3)
                    #XL
                    try:
                        tb_tk = self.driver.find_element(By.XPATH, "/html/body/form/div[4]/div[2]/div[2]/div/div/table/tbody")
                    except:
                        print("KHÔNG CÓ HÓA ĐƠN TGIAN NÀYYYY")
                        continue
                    trs = tb_tk.find_elements(By.TAG_NAME, "tr")
                    x+= len(trs)
                    q = 0
                    n_count = 0
                    n_0 = 0
                    n_load = len(trs)
                    n_tkk += n_load
                    
                                        
                                        # Lọc và in ra URL từ các entries
                    for tr in trs: 
                        div_pt = 100/n_load
                        n_count+=1
                        QtWidgets.QApplication.processEvents()
                        n_0+=div_pt
                        self.tool_run.progressBar.setValue(int(n_0))
                        QtWidgets.QApplication.processEvents()
                        self.tool_run.labeltb1.setText(f"[{n_count}/{n_load} | Tổng {n_tkk}]")
                        b = 0
                        try:
                            tds = tr.find_elements(By.XPATH, "./td")
                            id_tk = tds[1].text
                        except:
                            b = 1
                        print("***")
                        print(id_tk)
                        if len(id_tk)  >= 4 or b == 1: 
                            '''id_tk = id_tk.replace("/","-")'''
                            for i in range(0,3):
                                
                                if i == 2:
                                    try:
                                        tds[i].find_element(By.TAG_NAME, "a").click()
                                        q+=1
                                        time.sleep(0.2)
                                    except Exception as e:
                                        print("HÓA ĐƠN ĐẶC BIỆT")
                                        
                                        
                                        # Lặp qua các phần để tìm giá trị cần lấy
                                        '''for part in parts:
                                            print(part)
                                            print("-1-1-1-1-")
                                            if '&messageId=' in part:
                                                print(part)
                                                time.sleep(10000)
                                                ssid = part.split('=')[1]
                                                break'''
                                        entries = self.driver.execute_script("return window.performance.getEntries();")
                                        
                                        # Lọc và in ra URL từ các entries
                                        for entry in entries:
                                            print("-=-=11")
                                        
                                            parts = entry['name'].split('&')
                                            for part in parts:
                                                print(part)
                                                print("-1-1-1-1-")
                                                if 'dse_sessionId' in part:
                                                    print(part)
                                                    try:
                                                        ssid = part.split('%3D')[1]
                                                        ssid = ssid.split("%26")[0]
                                                    except:
                                                        ssid = part.split('=')[1]
                                                    print(ssid)
                                                    break
                                        print("Ssid=",ssid)
                                        from selenium.webdriver.common.keys import Keys
                                        if ssid == "NOTFOUND":
                                            cache_id.append(id_tk)
                                        else:
                                            try:
                                                link_n_tk = f"https://thuedientu.gdt.gov.vn/etaxnnt/Request?dse_sessionId={ssid}&dse_applicationId=-1&dse_operationName=traCuuToKhaiProc&dse_pageId=14&dse_processorState=viewTraCuuTkhai&dse_nextEventName=downTkhai&messageId={id_tk}"
                                            except:
                                                break
                                            print("LINK 05 QTT")
                                            print(link_n_tk)
                                            # Mở liên kết trong tab mới
                                            script = "window.open(arguments[0], '_blank');"
                                            self.driver.execute_script(script, link_n_tk)
                                            time.sleep(0.5)
                                            # Lấy handle của tab mới
                                            new_tab_handle = self.driver.window_handles[-1]
                                            
                                            # Chuyển sang tab mới
                                            self.driver.switch_to.window(new_tab_handle)
                                            time.sleep(0.5)
                                            # Đóng tab mới

                                            # Quay lại tab ban đầu
                                            self.driver.switch_to.window(self.driver.window_handles[0])
                                            
                                            try:    
                                                self.driver.switch_to.frame("mainframe")
                                            except:
                                                pass
                                            #Hóa đơn đặc biệt
                                    break
            
                                
                            try:
                                name_tk_w = tds[2].text 
                            except:
                                name_tk_w = " X "
                                pass
                            try:
                                check_cn = tds[10].text 
                                print(f"^^^^^^^^^^^^^^^^{check_cn}")
                                if "không chấp nhận" in check_cn or "Không chấp nhận" in check_cn:
                                    check_cn = "[Không chấp nhận]"
                                elif "Chấp nhận" in check_cn or "chấp nhận" in check_cn:
                                    check_cn = "[Chấp nhận]"
                            except:
                                print("ERROR333")
                            tgiant = 0
                            try:
                                tgiant = tds[7].text 
                            except:
                                pass
                            print(name_tk_w)
                            print("=-=-=-=-=-=-=-=")
                            if "TỜ KHAI QUYẾT TOÁN THUẾ THU NHẬP CÁ NHÂN" in name_tk_w :
                                if "(TT92/2015)" in name_tk_w:
                                    name_tk_w = "05/QTT-TNCN (TT92/2015)"
                                elif "TT80/2021" in name_tk_w:
                                    name_tk_w = "05/QTT-TNCN (TT80/2021)"
                            elif "03/TNDN" in name_tk_w :
                                if "(TT80/2021)" in name_tk_w:
                                    name_tk_w = "03/TNDN (TT80/2021)"
                            elif "01A/TNDN" in name_tk_w :
                                name_tk_w = "01A/TNDN"
                            elif ("GH_TNDN_TT83"  in name_tk_w and "TT83/2012/TT-BTC" in name_tk_w ) :
                                name_tk_w = "GH_TNDN_TT83 (TT83/2012/TT-BTC)"
                            elif "01B/TNDN" in name_tk_w :
                                name_tk_w = "01B/TNDN"
                            elif "02/TNDN" in name_tk_w :
                                name_tk_w = "02/TNDN"
                            elif ("06/KK-TNCN"  in name_tk_w and "(TT156/2013)" in name_tk_w ):
                                name_tk_w = "06/KK-TNCN (TT156/2013)"
                            elif ("05/KK-TNCN" in name_tk_w  and "(TT92/2015)" in name_tk_w ):
                                name_tk_w = "05/KK-TNCN (TT92/2015)"
                            elif ("05/KK-TNCN"  in name_tk_w and "(TT156/2013)" in name_tk_w ) :
                                name_tk_w = "05/KK-TNCN (TT156/2013)"
                            elif ("02/KK-TNCN"  in name_tk_w and "(TT156/2013)" in name_tk_w ) :
                                name_tk_w = " 02/KK-TNCN (TT156/2013)"
                            elif ("05/KK-TNCN" in name_tk_w  and "(TT80)" in name_tk_w ):
                                name_tk_w = "05/KK-TNCN - KTTTNCN (TT80)"
                            elif "QĐ 48/2006-BTC" in name_tk_w :
                                name_tk_w = "QĐ 48/2006-BTC"
                            elif "TT133_VuaVaNho_LT_B01a" in name_tk_w :
                                name_tk_w = "TT133_VuaVaNho_LT_B01a"
                            elif "TT133_VuaVaNho_LT_B01b" in name_tk_w :
                                name_tk_w = "TT133_VuaVaNho_LT_B01b"
                            elif ("BCTC_TT24_B01a" in name_tk_w  and "(TT24/2017/TT-BTC - mẫu B01a theo TT133)" in name_tk_w ) :
                                name_tk_w = "BCTC_TT24_B01a (TT24/2017/TT-BTC - mẫu B01a theo TT133)"
                            elif ("01/MBAI"  in name_tk_w and"(TT156/2013)" in name_tk_w):
                                name_tk_w = "01/MBAI (TT156/2013)"
                            elif ( "01/LPMB" in name_tk_w and "(TT80/2021)" in name_tk_w):
                                name_tk_w = "01/LPMB (TT80/2021)"
                            elif "GDNGHNT" in name_tk_w :
                                name_tk_w = "GDNGHNT"
                            elif "TB01/AC" in name_tk_w :
                                name_tk_w = "TB01/AC"
                            elif "TB03/AC" in name_tk_w :
                                name_tk_w = "TB03/AC"
                            elif "BC26/AC" in name_tk_w :
                                name_tk_w = "BC26/AC"
                            elif  ("01/GTGT" in name_tk_w and "(GTGT)" in name_tk_w)  :
                                name_tk_w = "01/GTGT (GTGT)"
                            elif ("01/GTGT" in name_tk_w and "(TT80/2021)" in name_tk_w) :
                                name_tk_w = "01/GTGT (TT80/2021)"
                            elif ("BẢNG TỔNG HỢP ĐĂNG KÝ NGƯỜI PHỤ THUỘC GIẢM TRỪ GIA CẢNH" in name_tk_w ) :
                                name_tk_w = "DK_NPT (TT156/2013)"
                            elif ("Tờ khai thuế nhà thầu nước ngoài - MẪU SỐ 01/NTNN" in name_tk_w ) :
                                name_tk_w = "DK_NPT (TT156/2013)"
                            elif ("BẢNG TỔNG HỢP ĐĂNG KÝ NGƯỜI PHỤ THUỘC GIẢM TRỪ GIA CẢNH" in name_tk_w ) :
                                name_tk_w = "DK_NPT (TT156/2013)"
                            print(name_tk_w)
                            try:
                                ktt = tds[3].text
                            except:
                                ktt = " "
                            try:    
                                ltk = tds[4].text
                            except:
                                ltk = " "
                            try:
                                ln = tds[5].text
                            except:
                                ln = " "
                            try:
                                nn = tds[7].text
                                c=id_tk.replace("/","-")
                                name_fl = f"{name_tk_w} -{ktt} -L{ln} -{ltk} -({c}) -[{nn}] [{check_cn}]"
                                name_fl  = self.remove_accents(name_fl)
                                name_fl = name_fl.replace(":","_")
                                self.list_new_name.append([id_tk,name_fl])
                            except:
                                pass
                                
                        QtWidgets.QApplication.processEvents()
                        self.tool_run.progressBar.setValue(99)
                wait = WebDriverWait(self.driver, 2)
                try:
                    pagination_button = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "img[src='/etaxnnt/static/images/pagination_right.gif']")))
                    pagination_button.click()
                    print("Đã nhấn vào nút phân trang")
                except:
                    check_pages = 0
            if len(cache_id) > 0 :
                print("TRY AGAIN")
                print("SSID LASTED = ", ssid)         
                for id_tk in cache_id:
                                            
                                            try:
                                                                            link_n_tk = f"https://thuedientu.gdt.gov.vn/etaxnnt/Request?dse_sessionId={ssid}&dse_applicationId=-1&dse_operationName=traCuuToKhaiProc&dse_pageId=14&dse_processorState=viewTraCuuTkhai&dse_nextEventName=downTkhai&messageId={id_tk}"
                                            except:
                                                                            continue
                                            print("LINK 05 QTT")
                                            print(link_n_tk)
                                            # Mở liên kết trong tab mới
                                            script = "window.open(arguments[0], '_blank');"
                                            self.driver.execute_script(script, link_n_tk)
                                            time.sleep(0.5)
                                            # Lấy handle của tab mới
                                            new_tab_handle = self.driver.window_handles[-1]
                                            
                                            # Chuyển sang tab mới
                                            self.driver.switch_to.window(new_tab_handle)
                                            time.sleep(0.5)
                                            # Đóng tab mới

                                            # Quay lại tab ban đầu
                                            self.driver.switch_to.window(self.driver.window_handles[0])
                                            
                                            try:    
                                                self.driver.switch_to.frame("mainframe")
                                            except:
                                                pass
        file_list = os.listdir(self.path_f1)
        time.sleep(1)
        cache_p = []
        for file_name in file_list:
            file_path = os.path.join(self.path_f1, file_name)
            file_path = str(file_path.replace("\\", "\**").replace("**",""))
            # Duyệt qua mảng lớn
            new_file_name = ""
            for sub_array in self.list_new_name:
                # Duyệt qua mảng con
                    id_tk = sub_array[0]
                    cache_p.append(id_tk)
                    value = sub_array[1]
                    value = value.replace("/","_")
                    new_file_name = ""
                    if id_tk in file_name and id_tk != "":
                        new_file_name = file_name.replace(id_tk, value)
                        new_file_name = new_file_name.replace("ETAX","")
                        if len(new_file_name) > 70:
                            i = file_path.split(".")[1]
                            new_file_name = value + f".{i}"
                        # Đường dẫn và tên tệp tin mới
                        new_file_path = os.path.join(self.folder_tk1, new_file_name)
                        # Di chuyển và đổi tên tệp tin
                        new_file_path = new_file_path.replace("\\", "\**").replace("**","")
                        file_path_c = file_path.split(".")
                        if len(file_path_c)>2:
                            file_path = str(file_path_c[0]+"."+file_path_c[1])
                        if ".xml" not in new_file_path and "." not in new_file_path:
                            new_file_path = new_file_path + ".xml"
                        print(f"[{file_path}] ==> [{new_file_path}]")
                        try:    
                            shutil.move(file_path, new_file_path)
                        except Exception as e:
                            print(e)
                            time.sleep(100)
                            pass
                        break
        time.sleep(0.5)
        for file_name in file_list:
            file_path = os.path.join(self.path_f1, file_name)
            file_path = str(file_path.replace("\\", "\**").replace("**",""))
            # Duyệt qua mảng lớn
            new_file_name = ""   
            new_file_path = os.path.join(self.folder_tk1, file_name)
            # Di chuyển và đổi tên tệp tin
            new_file_path = new_file_path.replace("\\", "\**").replace("**","") 
            try:    
                shutil.move(file_path, new_file_path)
            except:
                pass
        path_ce = ".\\__pycache__\\cache\\cache_tk"
        file_list = os.listdir(path_ce)
        for file_name in file_list:
            file_path = os.path.join(path_ce, file_name)
            new_file_name = ""   
            new_file_path = os.path.join(self.folder_tk1, file_name)
            # Di chuyển và đổi tên tệp tin
            new_file_path = new_file_path.replace("\\", "\**").replace("**","") 
            try:    
                shutil.move(file_path, new_file_path)
            except:
                pass
        QtWidgets.QApplication.processEvents()
        self.tool_run.progressBar.setValue(100)
        self.tool_run.labeltb1.setText(f"HOÀN THÀNH [{n_tkk}]")
if __name__ == "__main__":
    import os
    from PyQt5.QtCore import Qt
    from PyQt5.QtWidgets import QApplication

    # Xử lý DPI cho các máy độ phân giải cao (4K, scaling 125%, 150%...)
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication([])
    ui = MAIN()
    sys.exit(app.exec_())

        