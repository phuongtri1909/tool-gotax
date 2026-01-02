import os
import asyncio
import base64
import logging
import tempfile
import shutil
import re
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, AsyncGenerator
from io import BytesIO
import zipfile
import uuid

import httpx
from bs4 import BeautifulSoup, XMLParsedAsHTMLWarning
import warnings
from openpyxl import Workbook

# Suppress XMLParsedAsHTMLWarning khi parse XML với html.parser
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from .session_manager import SessionManager, SessionData

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# Mapping loại tờ khai
TOKHAI_TYPES = {
    "01/GTGT": "01/GTGT",
    "01/GTGT (TT80/2021)": "01/GTGT (TT80/2021)",
    "05/QTT-TNCN": "05/QTT-TNCN",
    "03/TNDN": "03/TNDN",
    "01A/TNDN": "01A/TNDN",
    "01B/TNDN": "01B/TNDN",
    "02/TNDN": "02/TNDN",
    "05/KK-TNCN": "05/KK-TNCN",
    "06/KK-TNCN": "06/KK-TNCN",
    "01/MBAI": "01/MBAI",
    "01/LPMB": "01/LPMB",
}

# Base URL
BASE_URL = "https://thuedientu.gdt.gov.vn"


class TaxCrawlerService:
    ZIP_STORAGE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'temp')
    
    def __init__(self, session_manager: SessionManager):
        self.session_manager = session_manager
        os.makedirs(self.ZIP_STORAGE_DIR, exist_ok=True)
        self._http_clients: Dict[str, httpx.AsyncClient] = {}
    
    async def _get_http_client(self, session_id: str) -> Optional[httpx.AsyncClient]:
        """
        Lấy hoặc tạo httpx client với cookies từ session
        Dùng để crawl nhanh sau khi login
        """
        session = self.session_manager.get_session(session_id)
        if not session or not session.is_logged_in:
            return None
        
        if session_id not in self._http_clients:
            cookies = await self.session_manager.get_cookies_for_httpx(session_id)
            if not cookies:
                return None
            
            self._http_clients[session_id] = httpx.AsyncClient(
                cookies=cookies,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'vi-VN,vi;q=0.9,en;q=0.8',
                },
                timeout=30.0,
                verify=False,
                follow_redirects=True
            )
        
        return self._http_clients[session_id]
    
    async def close_http_client(self, session_id: str):
        """Đóng httpx client khi session kết thúc"""
        if session_id in self._http_clients:
            await self._http_clients[session_id].aclose()
            del self._http_clients[session_id]
    
    async def _check_cancelled(self, job_id: str) -> bool:
        """
        Kiểm tra xem job có bị cancel không
        
        Returns:
            True nếu job bị cancel, False nếu không
        """
        try:
            from shared.redis_client import get_redis_client
            redis_client = get_redis_client()
            cancelled = redis_client.get(f"job:{job_id}:cancelled")
            if cancelled:
                cancelled = cancelled.decode('utf-8') if isinstance(cancelled, bytes) else str(cancelled).strip()
                if cancelled == '1':
                    return True
            
            # Check status
            status = redis_client.get(f"job:{job_id}:status")
            if status:
                status = status.decode('utf-8') if isinstance(status, bytes) else str(status).strip()
                if status == 'cancelled':
                    return True
            return False
        except Exception as e:
            logger.warning(f"Error checking cancelled flag: {e}")
            return False
    
    async def _check_session_timeout(self, page) -> bool:
        """
        Kiểm tra xem session có bị timeout không
        
        Returns:
            True nếu session timeout, False nếu không
        """
        try:
            current_url = page.url
            
            # Check URL timeout
            if 'timeout.jsp' in current_url:
                logger.warning("Session timeout detected from URL")
                return True
            
            # Check content timeout
            try:
                # Tìm text "Phiên giao dịch hết hạn"
                timeout_text = page.locator('text=Phiên giao dịch hết hạn')
                if await timeout_text.count() > 0:
                    logger.warning("Session timeout detected from content")
                    return True
                
                # Tìm nút "Trở lại" với onclick chứa corpIndexProc
                back_button = page.locator('input[type="button"][onclick*="corpIndexProc"]')
                if await back_button.count() > 0:
                    logger.warning("Session timeout detected from back button")
                    return True
            except Exception as e:
                logger.debug(f"Error checking timeout content: {e}")
            
            return False
        except Exception as e:
            logger.error(f"Error in _check_session_timeout: {e}")
            return False
    
    def _get_date_ranges(self, start_date: str, end_date: str, days_interval: int = 350) -> List[List[str]]:
        date_format = "%d/%m/%Y"
        date1 = datetime.strptime(start_date, date_format)
        date2 = datetime.strptime(end_date, date_format)
        interval = timedelta(days=days_interval)
        
        date_ranges = []
        while date1 <= date2:
            sub_array = [date1.strftime(date_format)]
            date1 += interval
            if date1 > date2:
                date1 = date2
            sub_array.append(date1.strftime(date_format))
            date_ranges.append(sub_array)
            date1 += timedelta(days=1)
        
        return date_ranges
    
    def _calculate_days_between(self, start_date: str, end_date: str) -> int:
        """Tính số ngày giữa 2 ngày (format: DD/MM/YYYY)"""
        date_format = "%d/%m/%Y"
        try:
            date1 = datetime.strptime(start_date, date_format)
            date2 = datetime.strptime(end_date, date_format)
            return (date2 - date1).days + 1  # +1 để tính cả ngày cuối
        except Exception as e:
            logger.warning(f"Error calculating days between {start_date} and {end_date}: {e}")
            return 0
    
    def _normalize_tokhai_name(self, name_tk: str) -> str:
        """Chuẩn hóa tên tờ khai"""
        if "TỜ KHAI QUYẾT TOÁN THUẾ THU NHẬP CÁ NHÂN" in name_tk:
            if "(TT92/2015)" in name_tk:
                return "05/QTT-TNCN (TT92/2015)"
            elif "TT80/2021" in name_tk:
                return "05/QTT-TNCN (TT80/2021)"
        elif "03/TNDN" in name_tk and "(TT80/2021)" in name_tk:
            return "03/TNDN (TT80/2021)"
        elif "01A/TNDN" in name_tk:
            return "01A/TNDN"
        elif "01B/TNDN" in name_tk:
            return "01B/TNDN"
        elif "02/TNDN" in name_tk:
            return "02/TNDN"
        elif "06/KK-TNCN" in name_tk and "(TT156/2013)" in name_tk:
            return "06/KK-TNCN (TT156/2013)"
        elif "05/KK-TNCN" in name_tk and "(TT92/2015)" in name_tk:
            return "05/KK-TNCN (TT92/2015)"
        elif "05/KK-TNCN" in name_tk and "(TT80)" in name_tk:
            return "05/KK-TNCN (TT80)"
        elif "01/GTGT" in name_tk and "(GTGT)" in name_tk:
            return "01/GTGT (GTGT)"
        elif "01/GTGT" in name_tk and "(TT80/2021)" in name_tk:
            return "01/GTGT (TT80/2021)"
        elif "01/MBAI" in name_tk and "(TT156/2013)" in name_tk:
            return "01/MBAI (TT156/2013)"
        elif "01/LPMB" in name_tk and "(TT80/2021)" in name_tk:
            return "01/LPMB (TT80/2021)"
        
        return name_tk
    
    async def _navigate_to_tokhai_page(self, page, dse_session_id: str) -> bool:
        success = False
        frame = None
        
        try:
            # Bước 1: Navigate đến trang dich-vu-khac
            # QUAN TRỌNG: Page mới cần navigate đến đúng URL, không dùng current_url
            # Vì page mới có thể chưa có URL hoặc URL không đúng
            logger.info("Navigating to /tthc/dich-vu-khac...")
            try:
                # Đảm bảo navigate đến đúng URL (không phải homelogin)
                target_url = 'https://dichvucong.gdt.gov.vn/tthc/dich-vu-khac'
                current_url = page.url
                
                # Nếu đang ở homelogin hoặc URL khác, navigate lại
                if '/tthc/dich-vu-khac' not in current_url:
                    await page.goto(target_url, wait_until='domcontentloaded', timeout=30000)
                    await asyncio.sleep(2)
                    logger.info(f"Successfully navigated to dich-vu-khac, current URL: {page.url}")
                else:
                    logger.info(f"Already on dich-vu-khac page: {current_url}")
            except Exception as nav_err:
                logger.error(f"Error navigating to dich-vu-khac: {nav_err}")
                return False
            
            # Bước 2: Gọi trực tiếp hàm JavaScript connectSSO('360103', '', '', '')
            logger.info("Calling connectSSO('360103', '', '', '') via JavaScript...")
            
            try:
                # Gọi hàm connectSSO trực tiếp bằng JavaScript
                await page.evaluate("""
                    async () => {
                        // Kiểm tra xem hàm connectSSO có tồn tại không
                        if (typeof connectSSO === 'function') {
                            await connectSSO('360103', '', '', '');
                            return { success: true, message: 'connectSSO called' };
                        } else {
                            return { success: false, message: 'connectSSO function not found' };
                        }
                    }
                """)
                logger.info("connectSSO('360103', '', '', '') called successfully")
                # Đợi AJAX hoàn tất và iframe được set src
                await asyncio.sleep(3)
            except Exception as e:
                logger.error(f"Error calling connectSSO: {e}")
                return False
            
            # Bước 3: Đợi iframe load với src từ thuedientu.gdt.gov.vn
            logger.info("Waiting for iframe to load with thuedientu.gdt.gov.vn...")
            
            # Tìm iframe trong #iframeRenderSSO
            max_wait = 20  # Đợi tối đa 10 giây (20 * 0.5)
            for i in range(max_wait):
                try:
                    # Tìm iframe trong modal #iframeRenderSSO
                    iframe_elem = page.locator('#iframeRenderSSO iframe').first
                    if await iframe_elem.count() > 0:
                        # Lấy src của iframe
                        iframe_src = await iframe_elem.get_attribute('src')
                        if iframe_src and 'thuedientu.gdt.gov.vn' in iframe_src:
                            logger.info(f"Found iframe with src: {iframe_src[:100]}...")
                            
                            # Tìm frame từ page.frames
                            frames = page.frames
                            for f in frames:
                                if 'thuedientu.gdt.gov.vn' in f.url:
                                    frame = f
                                    logger.info(f"Found frame: {frame.url[:100]}...")
                                    break
                            
                            if frame:
                                break
                except Exception as e:
                    logger.debug(f"Waiting for iframe (attempt {i + 1}/{max_wait}): {e}")
                
                await asyncio.sleep(0.5)
            
            # Bước 4: Switch vào iframe và đợi #maTKhai xuất hiện
            if frame:
                try:
                    logger.info("Waiting for #maTKhai in iframe...")
                    await frame.wait_for_load_state('domcontentloaded', timeout=15000)
                    await asyncio.sleep(1)
                    await frame.wait_for_selector('#maTKhai', timeout=15000)
                    success = True
                    logger.info("Tra cuu tokhai page loaded successfully via SSO iframe")
                except Exception as e:
                    logger.warning(f"Frame found but #maTKhai not found: {e}")
                    # Thử đợi thêm một chút
                    try:
                        await asyncio.sleep(2)
                        await frame.wait_for_selector('#maTKhai', timeout=10000)
                        success = True
                        logger.info("Tra cuu tokhai page loaded after additional wait")
                    except:
                        logger.error("Still cannot find #maTKhai after additional wait")
            else:
                logger.error("Iframe not found after clicking connectSSO link")
            
            return success
            
        except Exception as e:
            logger.error(f"Error navigating to tokhai page: {e}")
            return False
    
    async def _navigate_to_tokhai_search(self, session: SessionData) -> bool:
        return await self._navigate_to_tokhai_page(session.page, session.dse_session_id)
    
    async def _navigate_to_thongbao_page(self, page, dse_session_id: str) -> bool:
        success = False
        frame = None
        
        try:
            # Bước 1: Navigate đến trang dich-vu-khac
            current_url = page.url
            if '/tthc/dich-vu-khac' not in current_url:
                logger.info("Navigating to /tthc/dich-vu-khac for thongbao...")
                await page.goto('https://dichvucong.gdt.gov.vn/tthc/dich-vu-khac', wait_until='domcontentloaded', timeout=30000)
                await asyncio.sleep(2)
            else:
                logger.info("Already on /tthc/dich-vu-khac page")
            
            # Bước 2: Gọi trực tiếp hàm JavaScript connectSSO('360102', '', '', '')
            logger.info("Calling connectSSO('360102', '', '', '') via JavaScript...")
            
            try:
                # Gọi hàm connectSSO trực tiếp bằng JavaScript
                await page.evaluate("""
                    async () => {
                        // Kiểm tra xem hàm connectSSO có tồn tại không
                        if (typeof connectSSO === 'function') {
                            await connectSSO('360102', '', '', '');
                            return { success: true, message: 'connectSSO called' };
                        } else {
                            return { success: false, message: 'connectSSO function not found' };
                        }
                    }
                """)
                logger.info("connectSSO('360102', '', '', '') called successfully")
                # Đợi AJAX hoàn tất và iframe được set src
                await asyncio.sleep(3)
            except Exception as e:
                logger.error(f"Error calling connectSSO for thongbao: {e}")
                return False
            
            # Bước 3: Đợi iframe load với src từ thuedientu.gdt.gov.vn
            logger.info("Waiting for iframe to load with thuedientu.gdt.gov.vn for thongbao...")
            
            # Tìm iframe trong #iframeRenderSSO
            max_wait = 20  # Đợi tối đa 10 giây (20 * 0.5)
            for i in range(max_wait):
                try:
                    # Tìm iframe trong modal #iframeRenderSSO
                    iframe_elem = page.locator('#iframeRenderSSO iframe').first
                    if await iframe_elem.count() > 0:
                        # Lấy src của iframe
                        iframe_src = await iframe_elem.get_attribute('src')
                        if iframe_src and 'thuedientu.gdt.gov.vn' in iframe_src:
                            logger.info(f"Found iframe with src: {iframe_src[:100]}...")
                            
                            # Tìm frame từ page.frames
                            frames = page.frames
                            for f in frames:
                                if 'thuedientu.gdt.gov.vn' in f.url:
                                    frame = f
                                    logger.info(f"Found frame: {frame.url[:100]}...")
                                    break
                            
                            if frame:
                                break
                except Exception as e:
                    logger.debug(f"Waiting for iframe (attempt {i + 1}/{max_wait}): {e}")
                
                await asyncio.sleep(0.5)
            
            # Bước 4: Switch vào iframe và đợi form thông báo xuất hiện
            if frame:
                try:
                    logger.info("Waiting for thong bao form in iframe...")
                    await frame.wait_for_load_state('domcontentloaded', timeout=15000)
                    await asyncio.sleep(1)
                    # Đợi form thông báo load - kiểm tra input qryFromDate
                    await frame.wait_for_selector('#qryFromDate', timeout=15000)
                    success = True
                    logger.info("Tra cuu thong bao page loaded successfully via SSO iframe")
                except Exception as e:
                    logger.warning(f"Frame found but form not found: {e}")
                    # Thử đợi thêm một chút
                    try:
                        await asyncio.sleep(2)
                        await frame.wait_for_selector('#qryFromDate', timeout=10000)
                        success = True
                        logger.info("Tra cuu thong bao page loaded after additional wait")
                    except:
                        logger.error("Still cannot find form after additional wait")
            else:
                logger.error("Iframe not found after calling connectSSO for thongbao")
            
            return success
            
        except Exception as e:
            logger.error(f"Error navigating to thongbao page: {e}")
            return False
    
    async def _navigate_to_giaynoptien_page(self, page, dse_session_id: str) -> bool:
        """
        Navigate đến trang tra cứu giấy nộp tiền qua dichvucong.gdt.gov.vn
        Giống như tờ khai nhưng dùng connectSSO('330410')
        
        Flow:
        1. Navigate đến /tthc/dich-vu-khac
        2. Gọi connectSSO('330410', '', '', '')
        3. Đợi iframe load với src từ thuedientu.gdt.gov.vn
        4. Switch vào iframe và đợi form giấy nộp tiền xuất hiện
        
        Returns: True nếu thành công
        """
        success = False
        frame = None
        
        try:
            # Bước 1: Navigate đến trang dich-vu-khac
            current_url = page.url
            if '/tthc/dich-vu-khac' not in current_url:
                logger.info("Navigating to /tthc/dich-vu-khac for giaynoptien...")
                await page.goto('https://dichvucong.gdt.gov.vn/tthc/dich-vu-khac', wait_until='domcontentloaded', timeout=30000)
                await asyncio.sleep(2)
            else:
                logger.info("Already on /tthc/dich-vu-khac page")
            
            # Bước 2: Gọi trực tiếp hàm JavaScript connectSSO('330410', '', '', '')
            logger.info("Calling connectSSO('330410', '', '', '') via JavaScript...")
            
            try:
                # Gọi hàm connectSSO trực tiếp bằng JavaScript
                await page.evaluate("""
                    async () => {
                        // Kiểm tra xem hàm connectSSO có tồn tại không
                        if (typeof connectSSO === 'function') {
                            await connectSSO('330410', '', '', '');
                            return { success: true, message: 'connectSSO called' };
                        } else {
                            return { success: false, message: 'connectSSO function not found' };
                        }
                    }
                """)
                logger.info("connectSSO('330410', '', '', '') called successfully")
                # Đợi AJAX hoàn tất và iframe được set src
                await asyncio.sleep(3)
            except Exception as e:
                logger.error(f"Error calling connectSSO for giaynoptien: {e}")
                return False
            
            # Bước 3: Đợi iframe load với src từ thuedientu.gdt.gov.vn
            logger.info("Waiting for iframe to load with thuedientu.gdt.gov.vn for giaynoptien...")
            
            # Tìm frame trực tiếp từ page.frames (đáng tin cậy hơn)
            max_wait = 30  # Đợi tối đa 15 giây (30 * 0.5)
            frame = None
            for i in range(max_wait):
                try:
                    # Tìm frame từ page.frames trực tiếp
                    frames = page.frames
                    for f in frames:
                        if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                            frame = f
                            logger.info(f"Found frame: {frame.url[:100]}...")
                            break
                    
                    if frame:
                        # Kiểm tra xem frame đã load chưa
                        try:
                            await frame.wait_for_load_state('domcontentloaded', timeout=2000)
                            break
                        except:
                            # Frame chưa load xong, tiếp tục đợi
                            frame = None
                            pass
                except Exception as e:
                    logger.debug(f"Waiting for frame (attempt {i + 1}/{max_wait}): {e}")
                
                await asyncio.sleep(0.5)
            
            # Bước 4: Switch vào iframe và đợi form giấy nộp tiền xuất hiện
            if frame:
                try:
                    logger.info("Waiting for giay nop tien form in iframe...")
                    await frame.wait_for_load_state('domcontentloaded', timeout=15000)
                    await asyncio.sleep(1)
                    # Đợi form giấy nộp tiền load
                    await frame.wait_for_selector('input[name="ngay_lap_tu_ngay"], #ngay_lap_tu_ngay', timeout=15000)
                    success = True
                    logger.info("Tra cuu giay nop tien page loaded successfully via SSO iframe")
                except Exception as e:
                    logger.warning(f"Frame found but form not found: {e}")
                    # Thử đợi thêm một chút
                    try:
                        await asyncio.sleep(2)
                        await frame.wait_for_selector('input[name="ngay_lap_tu_ngay"], #ngay_lap_tu_ngay', timeout=10000)
                        success = True
                        logger.info("Tra cuu giay nop tien page loaded after additional wait")
                    except:
                        logger.error("Still cannot find form after additional wait")
            else:
                logger.error("Iframe not found after calling connectSSO for giaynoptien")
            
            return success
            
        except Exception as e:
            logger.error(f"Error navigating to giaynoptien page: {e}")
            return False
    
    async def crawl_tokhai_info(
        self,
        session_id: str,
        tokhai_type: str,
        start_date: str,
        end_date: str,
    ) -> AsyncGenerator[Dict[str, Any], None]:
        """
        Chỉ lấy thông tin tờ khai (KHÔNG download file)
        Dùng để hiển thị danh sách trước, user chọn tải sau
        
        Yields:
            Dict với các key: type, data, progress, error
        """
        session = self.session_manager.get_session(session_id)
        if not session:
            yield {"type": "error", "error": "Session không tồn tại hoặc đã hết hạn", "error_code": "SESSION_NOT_FOUND"}
            return
        
        if not session.is_logged_in:
            yield {"type": "error", "error": "Chưa đăng nhập. Vui lòng đăng nhập lại.", "error_code": "NOT_LOGGED_IN"}
            return
        
        page = session.page
        
        try:
            yield {"type": "info", "message": "Đang xử lý tờ khai..."}
            
            # Navigate đến trang tra cứu tờ khai bằng JavaScript (nhanh hơn click menu)
            success = await self._navigate_to_tokhai_page(page, session.dse_session_id)
            
            if not success:
                yield {"type": "error", "error": "Không thể navigate đến trang tra cứu. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            # Switch to mainframe
            frame = page.frame('mainframe')
            if not frame:
                yield {"type": "error", "error": "Không tìm thấy mainframe", "error_code": "NAVIGATION_ERROR"}
                return
            
            yield {"type": "info", "message": "Đang chọn loại tờ khai..."}
            
            # Chọn loại tờ khai
            try:
                select_element = frame.locator('#maTKhai')
                await select_element.wait_for(timeout=10000)
                
                if tokhai_type in ["00", "Tất cả", "tat_ca", None, ""]:
                    await select_element.select_option(value="00")
                    is_all_types = True
                else:
                    try:
                        await select_element.select_option(value=tokhai_type)
                        is_all_types = False
                    except:
                        option = frame.locator(f'#maTKhai option:has-text("{tokhai_type}")')
                        if await option.count() > 0:
                            option_value = await option.first.get_attribute('value')
                            await select_element.select_option(value=option_value)
                            is_all_types = (option_value == "00")
                        else:
                            raise Exception(f"Option not found: {tokhai_type}")
            except Exception as e:
                yield {"type": "error", "error": f"Không tìm thấy loại tờ khai: {tokhai_type}", "error_code": "INVALID_TOKHAI_TYPE"}
                return
            
            await asyncio.sleep(0.5)
            
            # Chia khoảng thời gian
            date_ranges = self._get_date_ranges(start_date, end_date)
            
            total_count = 0
            results = []
            
            yield {"type": "info", "message": f"Bắt đầu crawl {len(date_ranges)} khoảng thời gian..."}
            
            for range_idx, date_range in enumerate(date_ranges):
                yield {
                    "type": "progress", 
                    "current": range_idx + 1, 
                    "total": len(date_ranges),
                    "message": f"Đang xử lý khoảng {date_range[0]} - {date_range[1]}..."
                }
                
                try:
                    # Nhập ngày
                    start_input = frame.locator('#qryFromDate')
                    await start_input.fill('')
                    await start_input.fill(date_range[0])
                    
                    end_input = frame.locator('#qryToDate')
                    await end_input.click()
                    await end_input.fill('')
                    await end_input.fill(date_range[1])
                    
                    # Click Tra cứu
                    search_btn = frame.locator('input[value="Tra cứu"]')
                    await search_btn.click()
                    
                    await asyncio.sleep(2)
                    
                    # Xử lý pagination
                    check_pages = True
                    while check_pages:
                        try:
                            table_body = frame.locator('#allResultTableBody, table.md_list2 tbody, table#data_content_onday tbody').first
                            await table_body.wait_for(timeout=5000)
                        except:
                            yield {"type": "info", "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}"}
                            break
                        
                        rows = table_body.locator('tr')
                        row_count = await rows.count()
                        
                        yield {"type": "progress", "current": total_count, "message": f"Đang parse {row_count} tờ khai (trang hiện tại)..."}
                        
                        for i in range(row_count):
                            try:
                                row = rows.nth(i)
                                cols = row.locator('td')
                                col_count = await cols.count()
                                
                                if col_count < 3:
                                    continue
                                
                                # Cột 1: Mã giao dịch (id_tk)
                                id_tk = await cols.nth(1).text_content()
                                id_tk = id_tk.strip() if id_tk else ""
                                
                                if len(id_tk) < 4:
                                    continue
                                
                                # Extract thông tin
                                name_tk = await cols.nth(2).text_content() if col_count > 2 else ""
                                ky_tinh_thue = await cols.nth(3).text_content() if col_count > 3 else ""
                                loai_tk = await cols.nth(4).text_content() if col_count > 4 else ""
                                lan_nop = await cols.nth(5).text_content() if col_count > 5 else ""
                                lan_bs = await cols.nth(6).text_content() if col_count > 6 else ""
                                ngay_nop = await cols.nth(7).text_content() if col_count > 7 else ""
                                noi_nop = await cols.nth(9).text_content() if col_count > 9 else ""
                                trang_thai = await cols.nth(10).text_content() if col_count > 10 else ""
                                
                                # Chuẩn hóa tên tờ khai
                                name_tk_normalized = self._normalize_tokhai_name(name_tk.strip() if name_tk else "")
                                
                                # Xác định trạng thái
                                status = "unknown"
                                status_text = ""
                                trang_thai_lower = trang_thai.lower() if trang_thai else ""
                                if "không chấp nhận" in trang_thai_lower:
                                    status = "rejected"
                                    status_text = "[Khong chap nhan]"
                                elif "chấp nhận" in trang_thai_lower:
                                    status = "accepted"
                                    status_text = "[Chap nhan]"
                                
                                # Tạo tên file (để user biết tên file sẽ được tải)
                                ngay_nop_clean = ngay_nop.strip().replace("/", "-").replace(":", "-") if ngay_nop else ""
                                file_name = f"{name_tk_normalized} -{ky_tinh_thue.strip()} -L{lan_nop.strip()} -{loai_tk.strip()} -({id_tk}) -[{ngay_nop_clean}] {status_text}"
                                file_name = self._remove_accents(file_name)
                                file_name = file_name.replace("/", "_").replace(":", "_").replace("\\", "_")
                                
                                # Check xem có link download không
                                has_link = False
                                try:
                                    col2 = cols.nth(2)
                                    download_link = col2.locator('a')
                                    link_count = await download_link.count()
                                    
                                    if link_count > 0:
                                        first_link = download_link.first
                                        onclick = await first_link.get_attribute('onclick')
                                        title = await first_link.get_attribute('title')
                                        
                                        if onclick and 'downloadTkhai' in onclick:
                                            has_link = True
                                        elif title and 'Tải tệp' in title:
                                            has_link = True
                                except:
                                    has_link = False
                                
                                result = {
                                    "id": id_tk,
                                    "name": name_tk_normalized,
                                    "ky_tinh_thue": ky_tinh_thue.strip() if ky_tinh_thue else "",
                                    "loai": loai_tk.strip() if loai_tk else "",
                                    "lan_nop": lan_nop.strip() if lan_nop else "",
                                    "lan_bo_sung": lan_bs.strip() if lan_bs else "",
                                    "ngay_nop": ngay_nop.strip() if ngay_nop else "",
                                    "noi_nop": noi_nop.strip() if noi_nop else "",
                                    "trang_thai": status,
                                    "trang_thai_text": status_text,
                                    "file_name": file_name + ".xml",
                                    "has_download_link": has_link  # Có link download sẵn hay không
                                }
                                
                                results.append(result)
                                total_count += 1
                                
                                yield {"type": "item", "data": result}
                                
                            except Exception as e:
                                logger.error(f"Error processing row: {e}")
                                continue
                        
                        # Check pagination
                        try:
                            next_btn = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                            if await next_btn.count() > 0:
                                await next_btn.click()
                                await asyncio.sleep(1)
                            else:
                                check_pages = False
                        except:
                            check_pages = False
                
                except Exception as e:
                    logger.error(f"Error processing date range {date_range}: {e}")
                    yield {"type": "warning", "message": f"Lỗi xử lý khoảng {date_range}: {str(e)}"}
                    continue
            
            yield {
                "type": "complete",
                "total": total_count,
                "results": results
            }
            
        except Exception as e:
            logger.error(f"❌ Lỗi trong crawl_tokhai_info: {e}")
            yield {"type": "error", "error": f"Lỗi khi tra cứu thông tin tờ khai: {str(e)}", "error_code": "CRAWL_ERROR"}
    
    async def crawl_tokhai(
        self,
        session_id: str,
        tokhai_type: str,
        start_date: str,
        end_date: str,
        job_id: Optional[str] = None,  # ✅ Thêm job_id để check cancelled
    ) -> AsyncGenerator[Dict[str, Any], None]:
        session = self.session_manager.get_session(session_id)
        if not session:
            yield {"type": "error", "error": "Session không tồn tại hoặc đã hết hạn", "error_code": "SESSION_NOT_FOUND"}
            return
        
        if not session.is_logged_in:
            yield {"type": "error", "error": "Chưa đăng nhập. Vui lòng đăng nhập lại.", "error_code": "NOT_LOGGED_IN"}
            return
        
        page = session.page
        
        # ✅ FIX: Tạo temp directory trong source code thay vì system temp
        # Lấy đường dẫn project (tool-go-soft)
        current_dir = os.path.dirname(os.path.abspath(__file__))  # .../services/
        services_dir = os.path.dirname(current_dir)  # .../tool-go-soft/
        temp_base_dir = os.path.join(services_dir, "temp")  # .../tool-go-soft/temp/
        os.makedirs(temp_base_dir, exist_ok=True)
        
        # Tạo temp directory với timestamp để tránh conflict
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_dir = os.path.join(temp_base_dir, f"tokhai_{timestamp}")
        os.makedirs(temp_dir, exist_ok=True)
        
        logger.info(f"📁 Temp directory for debug files: {temp_dir}")  # ✅ Log temp_dir path để dễ tìm file debug
        ssid = session.dse_session_id
        
        try:
            yield {"type": "info", "message": "Đang xử lý tờ khai..."}
            
            success = await self._navigate_to_tokhai_page(page, ssid)
            
            if not success:
                yield {"type": "error", "error": "Không thể navigate đến trang tra cứu. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            frame = None
            try:
                frames = page.frames
                for f in frames:
                    if 'thuedientu.gdt.gov.vn' in f.url:
                        frame = f
                        break
            except Exception as e:
                logger.warning(f"Lỗi khi tìm frame: {e}")
            
            if not frame:
                yield {"type": "error", "error": "Không tìm thấy iframe sau khi navigate. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            try:
                await frame.wait_for_load_state('domcontentloaded', timeout=15000)
                await asyncio.sleep(1)
                await frame.wait_for_selector('#maTKhai', timeout=15000)
            except Exception as e:
                yield {"type": "error", "error": "Không tìm thấy form tra cứu. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            if await self._check_session_timeout(page):
                yield {
                    "type": "error",
                    "error": "Phiên giao dịch hết hạn. Vui lòng đăng nhập lại.",
                    "error_code": "SESSION_EXPIRED"
                }
                return
            
            yield {"type": "info", "message": "Đang chọn loại tờ khai..."}
            
            try:
                select_element = frame.locator('#maTKhai')
                await select_element.wait_for(timeout=10000)
                
                if tokhai_type in ["00", "Tất cả", "tat_ca", None, ""]:
                    await select_element.select_option(value="00")
                    logger.info("Selected tokhai: Tất cả")
                    is_all_types = True
                else:
                    try:
                        await select_element.select_option(value=tokhai_type)
                        logger.info(f"Selected tokhai by value: {tokhai_type}")
                        is_all_types = False
                    except:
                        option = frame.locator(f'#maTKhai option:has-text("{tokhai_type}")')
                        if await option.count() > 0:
                            option_value = await option.first.get_attribute('value')
                            await select_element.select_option(value=option_value)
                            is_all_types = (option_value == "00")
                        else:
                            raise Exception(f"Option not found: {tokhai_type}")
                        
            except Exception as e:
                logger.error(f"❌ Lỗi khi chọn loại tờ khai: {e}")
                yield {"type": "error", "error": f"Không tìm thấy loại tờ khai: {tokhai_type}. Hãy dùng value như '842', '00' (Tất cả), hoặc text như '01/GTGT'", "error_code": "INVALID_TOKHAI_TYPE"}
                return
            
            await asyncio.sleep(0.5)
            
            date_ranges = self._get_date_ranges(start_date, end_date)
            
            total_days = 0
            range_days = []
            for date_range in date_ranges:
                days = self._calculate_days_between(date_range[0], date_range[1])
                range_days.append(days)
                total_days += days
            
            range_percentages = []
            for days in range_days:
                if total_days > 0:
                    percent = (days / total_days) * 100
                else:
                    percent = 100.0 if len(date_ranges) == 1 else 0.0
                range_percentages.append(percent)
            
            total_count = 0
            results = []
            accumulated_total_so_far = 0
            accumulated_percent_so_far = 0.0 
            all_special_items = []
            thuyet_minh_total = 0
            thuyet_minh_downloaded = 0
            
            yield {"type": "info", "message": f"Bắt đầu crawl {len(date_ranges)} khoảng thời gian..."}
            
            http_client = await self._get_http_client(session_id)
            
            for range_idx, date_range in enumerate(date_ranges):
                if job_id and await self._check_cancelled(job_id):
                    logger.info(f"Job {job_id} đã bị cancel, dừng crawl")
                    yield {"type": "error", "error": "Job đã bị hủy", "error_code": "JOB_CANCELLED"}
                    return
                
                yield {
                    "type": "progress", 
                    "current": range_idx + 1, 
                    "total": len(date_ranges),
                    "message": f"Đang xử lý khoảng {date_range[0]} - {date_range[1]}...",
                    "percent": int(round(accumulated_percent_so_far)),
                    "accumulated_percent": int(round(accumulated_percent_so_far)),
                    "accumulated_total": accumulated_total_so_far,
                    "accumulated_downloaded": total_count,
                    "thuyet_minh_downloaded": thuyet_minh_downloaded,
                    "thuyet_minh_total": thuyet_minh_total
                }
                
                try:
                    # Nhập ngày bắt đầu (id="qryFromDate")
                    start_input = frame.locator('#qryFromDate')
                    await start_input.fill('')
                    await start_input.fill(date_range[0])
                    
                    # Nhập ngày kết thúc (id="qryToDate")
                    end_input = frame.locator('#qryToDate')
                    await end_input.click()
                    await end_input.fill('')
                    await end_input.fill(date_range[1])
                    
                    # Click button Tra cứu
                    search_btn = frame.locator('input[value="Tra cứu"]')
                    await search_btn.click()
                    
                    await asyncio.sleep(1)
                    
                    try:
                        frames = page.frames
                        for f in frames:
                            if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                frame = f
                                break
                    except Exception as refind_frame_e:
                        pass
                    
                    try:
                        await frame.wait_for_load_state('networkidle', timeout=5000)
                    except Exception as frame_load_e:
                        pass
                    
                    try:
                        table_body = frame.locator('#allResultTableBody, table.md_list2 tbody, table#data_content_onday tbody').first
                        await table_body.wait_for(timeout=10000, state='visible')
                        await asyncio.sleep(1.5)
                    except Exception as e:
                        pass
                        accumulated_percent_so_far = min(100.0, accumulated_percent_so_far)
                        yield {
                            "type": "info", 
                            "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}",
                            "percent": int(round(accumulated_percent_so_far)),
                            "accumulated_percent": int(round(accumulated_percent_so_far)),
                            "accumulated_total": accumulated_total_so_far,
                            "accumulated_downloaded": total_count,
                            "thuyet_minh_downloaded": thuyet_minh_downloaded,
                            "thuyet_minh_total": thuyet_minh_total
                        }
                        continue
                    
                    await asyncio.sleep(1)
                    
                    pagination_info = await self._extract_pagination_info(frame)
                    if not pagination_info:
                        rows = table_body.locator('tr')
                        row_count = await rows.count()
                        if row_count == 0:
                            yield {"type": "info", "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}"}
                            continue
                        else:
                            pagination_info = {
                                "current_page": 1,
                                "total_pages": 1,
                                "total_records": row_count
                            }
                    
                    total_pages = pagination_info["total_pages"]
                    total_records_estimated = pagination_info["total_records"]
                    
                    range_percent = range_percentages[range_idx]
                    
                    yield {
                        "type": "info",
                        "message": f"Tìm thấy {total_records_estimated} bản ghi trong {total_pages} trang. Bắt đầu tải..."
                    }
                    
                    yield {
                        "type": "download_start",
                        "total_to_download": total_records_estimated,
                        "date_range": f"{date_range[0]} - {date_range[1]}",
                        "range_index": range_idx + 1,
                        "total_ranges": len(date_ranges),
                        "accumulated_total": accumulated_total_so_far + total_records_estimated,
                        "accumulated_downloaded": total_count,
                        "range_percent": range_percent,
                        "accumulated_percent": accumulated_percent_so_far
                    }
                    
                    downloaded_count = 0
                    actual_downloaded = 0
                    actual_thuyet_minh_downloaded = 0
                    special_items = []
                    range_thuyet_minh_total = 0
                    
                    tokhai_count = total_records_estimated
                    if tokhai_count > 0:
                        percent_per_tokhai = range_percent / tokhai_count
                    else:
                        percent_per_tokhai = 0.0
                    
                    previous_first_row_id = None
                    
                    for page_num in range(1, total_pages + 1):
                        if job_id and await self._check_cancelled(job_id):
                            logger.info(f"Job {job_id} đã bị cancel, dừng crawl")
                            yield {"type": "error", "error": "Job đã bị hủy", "error_code": "JOB_CANCELLED"}
                            return
                        
                        if page_num > 1:
                            try:
                                next_btn = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                                next_btn_count = await next_btn.count()
                                if next_btn_count > 0:
                                    await asyncio.wait_for(next_btn.click(), timeout=10.0)
                                else:
                                    break
                            except asyncio.TimeoutError:
                                break
                            except Exception as click_e:
                                break
                                break
                            
                            logger.info(f"⏳ [TOKHAI] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num - 1}: Đợi 2 giây sau khi click...")
                            await asyncio.sleep(2)
                            logger.info(f"✅ [TOKHAI] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num - 1}: Đã đợi xong 2 giây, bắt đầu đợi table load...")
                            
                            try:
                                frames = page.frames
                                for f in frames:
                                    if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                        frame = f
                                        logger.info(f"🔄 [TOKHAI] [{range_idx + 1}/{len(date_ranges)}] Đã tìm lại frame mới sau khi click next: {frame.url[:100]}...")
                                        break
                            except Exception as refind_frame_e:
                                logger.warning(f"⚠️ [TOKHAI] [{range_idx + 1}/{len(date_ranges)}] Không thể tìm lại frame mới sau khi click next: {refind_frame_e}")
                        
                        try:
                            table_body = frame.locator('#allResultTableBody, table.md_list2 tbody, table#data_content_onday tbody').first
                            await asyncio.wait_for(
                                table_body.wait_for(timeout=15000, state='visible'),
                                timeout=20.0
                            )
                            
                            try:
                                await frame.wait_for_load_state('networkidle', timeout=5000)
                            except Exception as frame_load_e:
                                pass
                            
                            await asyncio.sleep(1.5)
                            
                            if page_num > 1:
                                try:
                                    rows_check = table_body.locator('tr')
                                    row_count_check = await rows_check.count()
                                    
                                    first_row_id = None
                                    if row_count_check > 0:
                                        try:
                                            first_row = rows_check.first
                                            first_cols = first_row.locator('td')
                                            col_count = await first_cols.count()
                                            if col_count > 1:
                                                first_row_id = await first_cols.nth(1).text_content()
                                                first_row_id = first_row_id.strip() if first_row_id else None
                                        except Exception as get_id_e:
                                            pass
                                    
                                    if previous_first_row_id and first_row_id:
                                        if previous_first_row_id == first_row_id:
                                            await asyncio.sleep(2)
                                            first_row_id_after_wait = None
                                            if row_count_check > 0:
                                                try:
                                                    first_row_after = rows_check.first
                                                    first_cols_after = first_row_after.locator('td')
                                                    col_count_after = await first_cols_after.count()
                                                    if col_count_after > 1:
                                                        first_row_id_after_wait = await first_cols_after.nth(1).text_content()
                                                        first_row_id_after_wait = first_row_id_after_wait.strip() if first_row_id_after_wait else None
                                                except Exception as get_id_e2:
                                                    logger.debug(f"⚠️ [TOKHAI] [{range_idx + 1}/{len(date_ranges)}] Không thể lấy mã giao dịch sau khi đợi: {get_id_e2}")
                                            
                                            if first_row_id_after_wait and previous_first_row_id == first_row_id_after_wait:
                                                logger.error(f"❌ [TOKHAI] [{range_idx + 1}/{len(date_ranges)}] Table vẫn chưa chuyển trang sau khi đợi thêm!")
                                                break
                                            elif first_row_id_after_wait:
                                                first_row_id = first_row_id_after_wait
                                        else:
                                            pass
                                except Exception as verify_e:
                                    pass
                        except asyncio.TimeoutError:
                            break
                        except Exception as wait_table_e:
                            break
                        
                        page_params = await self._extract_download_params(frame)
                        if not page_params:
                            continue
                        
                        rows = table_body.locator('tr')
                        row_count = await rows.count()
                        
                        first_row_id_current = None
                        if row_count > 0:
                            try:
                                first_row = rows.first
                                first_cols = first_row.locator('td')
                                col_count_first = await first_cols.count()
                                if col_count_first > 1:
                                    first_row_id_current = await first_cols.nth(1).text_content()
                                    first_row_id_current = first_row_id_current.strip() if first_row_id_current else None
                            except Exception as get_first_id_e:
                                pass
                        
                        page_items_to_download = []
                        
                        for i in range(row_count):
                            try:
                                row = rows.nth(i)
                                cols = row.locator('td')
                                col_count = await cols.count()
                                
                                if col_count < 3:
                                    continue
                                
                                # Cột 1: Mã giao dịch (id_tk)
                                id_tk = await cols.nth(1).text_content()
                                id_tk = id_tk.strip() if id_tk else ""
                                
                                # Cột 2: Tờ khai/Phụ lục
                                name_tk = await cols.nth(2).text_content() if col_count > 2 else ""
                                
                                # Check xem có link download không
                                download_type = None
                                has_link = False
                                extracted_id = None
                                
                                try:
                                    col2 = cols.nth(2)
                                    download_link = col2.locator('a')
                                    link_count = await download_link.count()
                                    
                                    if link_count > 0:
                                        first_link = download_link.first
                                        onclick = await first_link.get_attribute('onclick')
                                        title = await first_link.get_attribute('title')
                                        
                                        if onclick and 'downloadBke' in onclick:
                                            download_type = "downloadBke"
                                            has_link = True
                                            range_thuyet_minh_total += 1
                                            match = re.search(r"downloadBke\(['\"]?(\d+)['\"]?\)", onclick)
                                            if match:
                                                extracted_id = match.group(1)
                                                if not id_tk or len(id_tk) < 4:
                                                    id_tk = extracted_id
                                        elif onclick and 'downloadTkhai' in onclick:
                                            download_type = "downloadTkhai"
                                            has_link = True
                                            match = re.search(r"downloadTkhai\(['\"]?(\d+)['\"]?\)", onclick)
                                            if match:
                                                extracted_id = match.group(1)
                                        elif title and 'Tải tệp' in title:
                                            has_link = True
                                            download_type = "downloadTkhai"
                                except:
                                    has_link = False
                                
                                if not id_tk or len(id_tk) < 4:
                                    if extracted_id:
                                        id_tk = extracted_id
                                    else:
                                        continue
                                
                                name_tk_normalized = self._normalize_tokhai_name(name_tk.strip() if name_tk else "")
                                
                                if has_link:
                                    ky_tinh_thue = await cols.nth(3).text_content() if col_count > 3 else ""
                                    loai_tk = await cols.nth(4).text_content() if col_count > 4 else ""
                                    lan_nop = await cols.nth(5).text_content() if col_count > 5 else ""
                                    lan_bs = await cols.nth(6).text_content() if col_count > 6 else ""
                                    ngay_nop = await cols.nth(7).text_content() if col_count > 7 else ""
                                    noi_nop = await cols.nth(9).text_content() if col_count > 9 else ""
                                    trang_thai = await cols.nth(10).text_content() if col_count > 10 else ""
                                    
                                    status = "unknown"
                                    status_text = ""
                                    trang_thai_lower = trang_thai.lower() if trang_thai else ""
                                    
                                    if "tiếp nhận" in trang_thai_lower or "tiep nhan" in trang_thai_lower:
                                        status = "received"
                                        status_text = "Tiếp nhận"
                                    elif "xác nhận" in trang_thai_lower or "xac nhan" in trang_thai_lower:
                                        status = "confirmed"
                                        status_text = "Xác nhận"
                                    elif "không chấp nhận" in trang_thai_lower or "khong chap nhan" in trang_thai_lower:
                                        status = "rejected"
                                        status_text = "Không chấp nhận"
                                    elif "chấp nhận" in trang_thai_lower or "chap nhan" in trang_thai_lower:
                                        status = "accepted"
                                        status_text = "Chấp nhận"
                                    else:
                                        status_text = trang_thai.strip()[:20] if trang_thai else "Unknown"
                                    
                                    status_text_clean = self._remove_accents(status_text)
                                    
                                    ngay_clean = ""
                                    if ngay_nop:
                                        ngay_parts = ngay_nop.strip().split(" ")
                                        if ngay_parts:
                                            ngay_only = ngay_parts[0]
                                            date_parts = ngay_only.split("/")
                                            if len(date_parts) == 3:
                                                ngay_clean = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                                            else:
                                                ngay_clean = ngay_only.replace("/", "-")
                                    
                                    if ngay_clean:
                                        file_name = f"{id_tk} - {status_text_clean} - {ngay_clean}"
                                    else:
                                        file_name = f"{id_tk} - {status_text_clean}"
                                    
                                    max_filename_length = 150
                                    if len(file_name) > max_filename_length:
                                        status_short = {
                                            "Tiep nhan": "TN",
                                            "Chap nhan": "CN",
                                            "Khong chap nhan": "KCN",
                                            "Xac nhan": "XN"
                                        }.get(status_text_clean, status_text_clean[:3])
                                        
                                        if ngay_clean:
                                            file_name = f"{id_tk} - {status_short} - {ngay_clean}"
                                        else:
                                            file_name = f"{id_tk} - {status_short}"
                                        
                                        if len(file_name) > max_filename_length:
                                            file_name = file_name[:max_filename_length]
                                    
                                    # Lưu item để download batch
                                    item = {
                                    "id": id_tk,
                                    "name": name_tk_normalized,
                                    "ky_tinh_thue": ky_tinh_thue.strip() if ky_tinh_thue else "",
                                    "loai": loai_tk.strip() if loai_tk else "",
                                    "lan_nop": lan_nop.strip() if lan_nop else "",
                                    "lan_bo_sung": lan_bs.strip() if lan_bs else "",
                                    "ngay_nop": ngay_nop.strip() if ngay_nop else "",
                                    "noi_nop": noi_nop.strip() if noi_nop else "",
                                    "trang_thai": status,
                                    "file_name": file_name,
                                        "has_link": True,
                                        "download_type": download_type,
                                        "page_number": page_num
                                    }
                                    page_items_to_download.append(item)
                                else:
                                    special_item = {
                                        "id": id_tk,
                                        "name": name_tk_normalized,
                                        "page_number": page_num,
                                        "has_link": False,
                                        "date_range": f"{date_range[0]} - {date_range[1]}"
                                    }
                                    special_items.append(special_item)
                                    logger.info(f"📝 Collected special item (no link): {id_tk} - {name_tk_normalized} (page {page_num})")
                                
                            except Exception as e:
                                logger.error(f"Error processing row: {e}")
                                continue
                        
                        if page_items_to_download:
                            page_params_map = {page_num: page_params}
                            
                            tokhai_count = total_records_estimated - range_thuyet_minh_total
                            
                            if tokhai_count > 0:
                                percent_per_tokhai = range_percent / tokhai_count
                            else:
                                percent_per_tokhai = 0.0
                            
                            current_percent_for_page = accumulated_percent_so_far + (actual_downloaded * percent_per_tokhai)
                            current_percent_for_page = min(100.0, current_percent_for_page)
                            accumulated_total_current = accumulated_total_so_far + total_records_estimated
                            yield {
                                "type": "progress",
                                "current": actual_downloaded,
                                "message": f"Đang xử lý trang {page_num}/{total_pages} ({len(page_items_to_download)} tờ khai)...",
                                "percent": int(round(current_percent_for_page)),
                                "accumulated_percent": int(round(current_percent_for_page)),
                                "accumulated_total": accumulated_total_current,
                                "accumulated_downloaded": total_count + actual_downloaded,
                                "thuyet_minh_downloaded": thuyet_minh_downloaded + actual_thuyet_minh_downloaded,
                                "thuyet_minh_total": thuyet_minh_total + range_thuyet_minh_total
                            }
                            
                            successful_downloads = []
                            base_params = page_params
                            
                            for idx, item in enumerate(page_items_to_download):
                                file_num = idx + 1
                                
                                try:
                                    fresh_params = await self._extract_download_params(frame)
                                    if fresh_params:
                                        base_params = fresh_params
                                except:
                                    pass
                                
                                try:
                                    result = await self._download_one_via_url(
                                        session_id,
                                        item["id"],
                                        item,
                                        base_params,
                                        temp_dir,
                                        frame=frame
                                    )
                                    
                                    if result and not isinstance(result, Exception):
                                        successful_downloads.append(result)
                                        
                                        if result.get("download_type") == "downloadBke":
                                            actual_thuyet_minh_downloaded += 1
                                        else:
                                            actual_downloaded += 1
                                        downloaded_count += 1
                                        
                                        current_percent = accumulated_percent_so_far + (actual_downloaded * percent_per_tokhai)
                                        current_percent = min(100.0, current_percent)
                                        
                                        accumulated_total_current = accumulated_total_so_far + total_records_estimated
                                        
                                        yield {
                                            "type": "download_progress",
                                            "downloaded": actual_downloaded,
                                            "total": total_records_estimated,
                                            "percent": int(round(current_percent)),
                                            "date_range": f"{date_range[0]} - {date_range[1]}",
                                            "range_index": range_idx + 1,
                                            "total_ranges": len(date_ranges),
                                            "accumulated_downloaded": total_count + actual_downloaded,
                                            "accumulated_total": accumulated_total_current,
                                            "accumulated_percent": int(round(current_percent)),
                                            "thuyet_minh_downloaded": thuyet_minh_downloaded + actual_thuyet_minh_downloaded,
                                            "thuyet_minh_total": thuyet_minh_total + range_thuyet_minh_total
                                        }
                                        
                                        result_data = {
                                            "id": result["id"],
                                            "name": result["name"],
                                            "ky_tinh_thue": result["ky_tinh_thue"],
                                            "loai": result["loai"],
                                            "lan_nop": result["lan_nop"],
                                            "lan_bo_sung": result["lan_bo_sung"],
                                            "ngay_nop": result["ngay_nop"],
                                            "noi_nop": result["noi_nop"],
                                            "trang_thai": result["trang_thai"],
                                            "file_name": result["file_name"] + ".xml"
                                        }
                                        results.append(result_data)
                                        yield {"type": "item", "data": result_data}
                                    
                                    await asyncio.sleep(0.1)
                                except Exception as e:
                                    logger.error(f"❌ Error downloading {item.get('id', 'unknown')}: {e}")
                        
                        previous_first_row_id = first_row_id_current
                    
                    accumulated_total_so_far += total_records_estimated
                    total_count += actual_downloaded
                    
                    accumulated_percent_so_far += range_percent
                    accumulated_percent_so_far = min(100.0, accumulated_percent_so_far)
                    
                    thuyet_minh_total += range_thuyet_minh_total
                    thuyet_minh_downloaded += actual_thuyet_minh_downloaded
                    
                    if special_items:
                        all_special_items.extend(special_items)
                        yield {
                            "type": "special_items",
                            "count": len(special_items),
                            "items": special_items,
                            "date_range": f"{date_range[0]} - {date_range[1]}",
                            "message": f"Có {len(special_items)} tờ khai đặc biệt trong khoảng {date_range[0]} - {date_range[1]} (chưa có cách tải, đã lưu metadata để sau này)",
                            "percent": int(round(accumulated_percent_so_far)),
                            "accumulated_percent": int(round(accumulated_percent_so_far)),
                            "accumulated_total": accumulated_total_so_far,
                            "accumulated_downloaded": total_count,
                            "thuyet_minh_downloaded": thuyet_minh_downloaded,
                            "thuyet_minh_total": thuyet_minh_total
                        }
                        logger.info(f"📋 Found {len(special_items)} special items (no download link) in date range {date_range[0]} - {date_range[1]}")
                    
                    accumulated_percent_so_far = min(100.0, accumulated_percent_so_far)
                    yield {
                        "type": "info",
                        "message": f"Đã tải {actual_downloaded} file từ {total_pages} trang (ước tính {total_records_estimated} bản ghi). Có {len(special_items)} tờ khai đặc biệt chưa tải.",
                        "percent": int(round(accumulated_percent_so_far)),
                        "accumulated_percent": int(round(accumulated_percent_so_far)),
                        "accumulated_total": accumulated_total_so_far,
                        "accumulated_downloaded": total_count,
                        "thuyet_minh_downloaded": thuyet_minh_downloaded,
                        "thuyet_minh_total": thuyet_minh_total
                    }
                
                except Exception as e:
                    logger.error(f"Error processing date range {date_range}: {e}")
                    yield {"type": "warning", "message": f"Lỗi xử lý khoảng {date_range}: {str(e)}"}
                    continue
            
            # Tạo ZIP file từ các file đã download
            zip_base64 = None
            download_id = None
            files_info = []
            total_size = 0
            
            if os.listdir(temp_dir):
                # Tạo tên file ZIP
                if is_all_types:
                    zip_filename = f"tokhai_TAT_CA_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                    tokhai_type_label = "Tất cả"
                else:
                    zip_filename = f"tokhai_{tokhai_type}_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                    tokhai_type_label = tokhai_type
                
                # Tạo download_id (UUID) để worker có thể download sau
                download_id = str(uuid.uuid4())
                zip_file_path = os.path.join(self.ZIP_STORAGE_DIR, f"{download_id}.zip")
                
                # Lưu zip vào disk thay vì chỉ tạo base64
                with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for file_name in os.listdir(temp_dir):
                        file_path = os.path.join(temp_dir, file_name)
                        if os.path.isfile(file_path):
                            file_size = os.path.getsize(file_path)
                            total_size += file_size
                            zf.write(file_path, file_name)
                            files_info.append({
                                "name": file_name,
                                "size": file_size
                            })
                
                # Đọc file để tạo base64 (vẫn cần cho Redis)
                with open(zip_file_path, 'rb') as f:
                    zip_base64 = base64.b64encode(f.read()).decode('utf-8')
                
                logger.info(f"✅ Đã tạo file ZIP: {zip_filename} (download_id: {download_id})")
                
                # Lưu download_id vào Redis
                try:
                    from shared.redis_client import get_redis_client
                    redis_client = get_redis_client()
                    redis_key = f"session:{session_id}:download_id"
                    redis_client.setex(redis_key, 3600, download_id.encode('utf-8'))
                except Exception as redis_err:
                    logger.warning(f"⚠️ Không thể lưu download_id vào Redis: {redis_err}")
            else:
                # Không có files
                pass
            
            # Tạo zip_filename nếu chưa có (cho trường hợp không có files)
            if not download_id:
                if is_all_types:
                    zip_filename = f"tokhai_TAT_CA_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                    tokhai_type_label = "Tất cả"
            else:
                zip_filename = f"tokhai_{tokhai_type}_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                tokhai_type_label = tokhai_type
            
            # Đếm lại số file thực tế đã download (tờ khai + tờ thuyết minh)
            actual_files_count = len(files_info)
            # Đếm số results thực tế
            actual_results_count = len(results)
            
            # ✅ Tính số tờ khai đã tải (không bao gồm tờ thuyết minh)
            tokhai_downloaded = total_count  # Số tờ khai đã tải (không tính tờ thuyết minh)
            # ✅ Tổng số file đã tải = tờ khai + tờ thuyết minh
            total_files_downloaded = tokhai_downloaded + thuyet_minh_downloaded
            
            # ✅ LOG để debug
            logger.info(f"📊 Complete event - tokhai_downloaded: {tokhai_downloaded}, thuyet_minh_downloaded: {thuyet_minh_downloaded}, total_files_downloaded: {total_files_downloaded}, accumulated_total_so_far: {accumulated_total_so_far}")
            
            # ✅ Message hiển thị khi hoàn thành
            completion_message = f"Hoàn thành! Đã tải {tokhai_downloaded}/{accumulated_total_so_far} tờ khai"
            if thuyet_minh_total > 0:
                completion_message += f" - {thuyet_minh_downloaded}/{thuyet_minh_total} tờ thuyết minh"
            if len(all_special_items) > 0:
                completion_message += f". Có {len(all_special_items)} tờ khai đặc biệt không tải được"
            
            # Total = số file thực tế đã download (tờ khai + tờ thuyết minh) - đây là số hiển thị trên button
            # Nếu muốn biết số items đã tìm thấy, dùng actual_results_count
            # Complete event KHÔNG gửi zip_base64 và results (quá lớn), chỉ gửi metadata
            yield {
                "type": "complete",
                "total": total_files_downloaded,  # ✅ Số file đã tải (tờ khai + tờ thuyết minh) - để hiển thị trên button
                "tokhai_downloaded": tokhai_downloaded,  # Số tờ khai đã tải
                "tokhai_total": accumulated_total_so_far,  # Tổng số tờ khai tìm thấy
                "thuyet_minh_downloaded": thuyet_minh_downloaded,  # Số tờ thuyết minh đã tải
                "thuyet_minh_total": thuyet_minh_total,  # Tổng số tờ thuyết minh
                "special_items_count": len(all_special_items),  # Số tờ khai đặc biệt không tải được
                "results_count": actual_results_count,  # Số items đã tìm thấy (có thể > files nếu download thất bại)
                "total_rows_processed": total_count,  # Số rows đã xử lý (để debug)
                "files_count": actual_files_count,  # Số file trong ZIP (để kiểm tra)
                "total_size": total_size,
                "download_id": download_id,
                "zip_filename": zip_filename,
                "tokhai_type": tokhai_type_label,
                "is_all_types": is_all_types,
                "has_zip": download_id is not None,
                "message": completion_message,
                "special_items": all_special_items if len(all_special_items) > 0 else None,
            }
            
            if download_id and zip_base64:
                chunk_size = 5 * 1024 * 1024
                if len(zip_base64) > chunk_size:
                    logger.info(f"Zip base64 is large ({len(zip_base64)/1024/1024:.2f} MB), sending in chunks")
                    for i in range(0, len(zip_base64), chunk_size):
                        chunk = zip_base64[i:i+chunk_size]
                        yield {
                            "type": "zip_chunk",
                            "download_id": download_id,
                            "chunk_index": i // chunk_size,
                            "chunk_data": chunk,
                            "is_last": (i + chunk_size) >= len(zip_base64)
                        }
                else:
                    yield {
                        "type": "zip_data",
                        "download_id": download_id,
                        "zip_base64": zip_base64,
                        "zip_filename": zip_filename
            }
            
        except Exception as e:
            logger.error(f"Error in crawl_tokhai: {e}")
            error_msg = str(e)
            if "timeout" in error_msg.lower() or "phiên giao dịch" in error_msg.lower():
                yield {"type": "error", "error": "Phiên giao dịch hết hạn. Vui lòng đăng nhập lại.", "error_code": "SESSION_EXPIRED"}
            else:
                yield {"type": "error", "error": f"Lỗi khi tra cứu tờ khai: {error_msg}", "error_code": "CRAWL_ERROR"}
        
        finally:
            debug_files = []
            try:
                if os.path.exists(temp_dir):
                    for file in os.listdir(temp_dir):
                        if file.startswith('DEBUG_'):
                            debug_files.append(os.path.join(temp_dir, file))
                
                if debug_files:
                    logger.warning(f"⚠️ Found {len(debug_files)} debug files in {temp_dir}:")
                    for debug_file in debug_files:
                        file_size = os.path.getsize(debug_file) if os.path.exists(debug_file) else 0
                        logger.warning(f"  - {os.path.basename(debug_file)} ({file_size} bytes)")
                    logger.warning(f"⚠️ Debug files will be kept for inspection. Temp dir: {temp_dir}")
                else:
                    shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception as e:
                logger.warning(f"Error checking debug files: {e}")
            shutil.rmtree(temp_dir, ignore_errors=True)
    
    def _remove_accents(self, text: str) -> str:
        """Remove Vietnamese accents"""
        try:
            import unidecode
            return unidecode.unidecode(text)
        except:
            return text
    
    async def _extract_pagination_info(self, frame) -> Optional[Dict[str, int]]:
        """
        Extract pagination info từ div #currAcc
        Returns: {"current_page": 1, "total_pages": 2, "total_records": 13} hoặc None
        """
        try:
            pagination_div = frame.locator('#currAcc').first
            if await pagination_div.count() == 0:
                logger.warning("⚠️ Cannot find #currAcc div")
                return None
            
            html_content = await pagination_div.inner_html()
            if html_content:
                import re
                page_match = re.search(r'Trang\s+(\d+)\s*/\s*<b>(\d+)</b>', html_content)
                records_match = re.search(r'Có\s+<b>(\d+)</b>\s+bản\s+ghi', html_content)
                
                if page_match:
                    current_page = int(page_match.group(1))
                    total_pages = int(page_match.group(2))
                    total_records = int(records_match.group(1)) if records_match else 0
                    return {
                        "current_page": current_page,
                        "total_pages": total_pages,
                        "total_records": total_records
                    }
            
            text = await pagination_div.text_content()
            if not text:
                logger.warning("⚠️ Pagination div has no content")
                return None
            
            import re
            page_match = re.search(r'Trang\s+(\d+)\s*/\s*(\d+)', text)
            records_match = re.search(r'Có\s+(\d+)\s+bản\s+ghi', text)
            
            if page_match:
                current_page = int(page_match.group(1))
                total_pages = int(page_match.group(2))
                total_records = int(records_match.group(1)) if records_match else 0
                return {
                    "current_page": current_page,
                    "total_pages": total_pages,
                    "total_records": total_records
                }
            
            logger.warning(f"⚠️ Cannot parse pagination info from text: {text[:100]}")
            return None
            
        except Exception as e:
            logger.warning(f"Error extracting pagination info: {e}")
            return None
    
    async def _navigate_to_page(self, frame, target_page: int) -> bool:
        """
        Navigate đến trang cụ thể bằng link có sẵn trong HTML
        Returns: True nếu navigate thành công, False nếu không tìm thấy link
        """
        try:
            try:
                pagination_div = frame.locator('#currAcc').first
                await pagination_div.wait_for(timeout=3000)
            except:
                logger.warning(f"⚠️ Cannot find #currAcc div, trying to navigate anyway")
                pass
            
            link = None
            
            pagination_div = frame.locator('#currAcc').first
            if await pagination_div.count() > 0:
                link = pagination_div.locator(f'a[href*="&pn={target_page}"]').first
                if await link.count() == 0:
                    link = pagination_div.locator(f'a:has-text("{target_page}")').first
            
            if not link or await link.count() == 0:
                link = frame.locator(f'a[href*="&pn={target_page}"]').first
                if await link.count() == 0:
                    link = frame.locator(f'a:has-text("{target_page}")').first
            
            if not link or await link.count() == 0:
                logger.warning(f"⚠️ Cannot find link to page {target_page}")
                return False
            
            # Click link
            await link.click()
            
            try:
                table_body = frame.locator('#allResultTableBody, table.md_list2 tbody, table#data_content_onday tbody').first
                await table_body.wait_for(timeout=5000)
                await asyncio.sleep(1)
            except:
                logger.warning(f"⚠️ Table not found after navigating to page {target_page}")
            
            pagination_info = await self._extract_pagination_info(frame)
            if pagination_info and pagination_info["current_page"] == target_page:
                logger.info(f"✅ Navigated to page {target_page}")
            else:
                logger.info(f"✅ Navigated to page {target_page} (verification: current_page={pagination_info.get('current_page') if pagination_info else 'unknown'})")
            
            return True
                
        except Exception as e:
            logger.error(f"Error navigating to page {target_page}: {e}")
            return False
    
    async def _take_screenshot_on_download_error(
        self,
        session_id: str,
        ma_tkhai: str,
        error_reason: str,
        frame=None
    ) -> Optional[str]:
        """
        Chụp màn hình khi download fail để debug
        
        Args:
            session_id: Session ID để lấy page
            ma_tkhai: Mã giao dịch của tờ khai
            error_reason: Lý do lỗi (để đặt tên file)
            frame: Frame hiện tại (nếu có)
        
        Returns:
            Path đến file screenshot hoặc None nếu fail
        """
        try:
            # Lấy session để có page
            session = self.session_manager.get_session(session_id)
            if not session or not session.page:
                logger.warning(f"[{ma_tkhai}] Cannot take screenshot: No session or page")
                return None
            
            page = session.page
            
            # Tạo thư mục screenshots trong project
            # Lấy đường dẫn project (tool-go-soft)
            # File này ở: tool-gotax/tool-go-soft/services/tax_crawler.py
            # Muốn tới: tool-gotax/tool-go-soft/screenshots/
            current_dir = os.path.dirname(os.path.abspath(__file__))  # .../services/
            services_dir = os.path.dirname(current_dir)  # .../tool-go-soft/
            screenshots_dir = os.path.join(services_dir, "screenshots")  # .../tool-go-soft/screenshots/
            os.makedirs(screenshots_dir, exist_ok=True)
            
            # Tạo tên file với timestamp và ma_tkhai
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_error_reason = error_reason.replace(" ", "_").replace("/", "_").replace("\\", "_")[:50]
            screenshot_filename = f"ERROR_{ma_tkhai}_{safe_error_reason}_{timestamp}.png"
            screenshot_path = os.path.join(screenshots_dir, screenshot_filename)
            
            # Chụp màn hình
            try:
                # ✅ FIX: Frame không có method screenshot() trực tiếp
                # Dùng page.screenshot() để chụp toàn bộ page (bao gồm frame)
                # Hoặc nếu muốn chụp chỉ frame, dùng frame.locator('body').screenshot()
                if frame:
                    # Thử chụp frame content trước (nếu được)
                    try:
                        frame_body = frame.locator('body')
                        await frame_body.screenshot(path=screenshot_path, timeout=5000)
                        if os.path.exists(screenshot_path):
                            file_size = os.path.getsize(screenshot_path)
                            logger.warning(f"📸 [{ma_tkhai}] Screenshot saved (frame): {screenshot_path} ({file_size} bytes) - Error: {error_reason}")
                            return screenshot_path
                    except Exception as frame_e:
                        logger.debug(f"📸 [{ma_tkhai}] Cannot screenshot frame directly: {frame_e}, trying page screenshot")
                
                # Fallback: Chụp toàn bộ page (sẽ bao gồm frame)
                await page.screenshot(path=screenshot_path, full_page=True, timeout=5000)
                
                if os.path.exists(screenshot_path):
                    file_size = os.path.getsize(screenshot_path)
                    logger.warning(f"📸 [{ma_tkhai}] Screenshot saved (page): {screenshot_path} ({file_size} bytes) - Error: {error_reason}")
                    return screenshot_path
                else:
                    logger.warning(f"📸 [{ma_tkhai}] Screenshot file not created: {screenshot_path}")
                    return None
            except Exception as e:
                logger.warning(f"📸 [{ma_tkhai}] Error taking screenshot: {e}")
                return None
                
        except Exception as e:
            logger.warning(f"📸 [{ma_tkhai}] Error in _take_screenshot_on_download_error: {e}")
            return None
    
    async def _navigate_to_page(self, frame, target_page: int) -> bool:
        """
        Navigate đến trang cụ thể bằng cách click vào link pagination
        
        Args:
            frame: Frame hiện tại
            target_page: Số trang cần navigate (1, 2, 3, ...)
        
        Returns:
            True nếu navigate thành công, False nếu fail
        """
        try:
            # ✅ FIX: Check trang hiện tại trước, nếu đang ở target page thì không cần navigate
            pagination_info = await self._extract_pagination_info(frame)
            if pagination_info:
                current_page = pagination_info.get("current_page", 0)
                if current_page == target_page:
                    logger.info(f"✅ Already on page {target_page}")
                    return True
            
            # Nếu là trang 1 và không có pagination info, giả định đang ở trang 1
            if target_page == 1 and not pagination_info:
                logger.info(f"✅ Assuming already on page 1 (no pagination info)")
                return True
            
            # Đợi pagination div xuất hiện
            try:
                pagination_div = frame.locator('#currAcc').first
                await pagination_div.wait_for(timeout=5000)
            except:
                logger.warning(f"⚠️ Cannot find #currAcc div")
                # Nếu không có pagination div và target là trang 1, giả định OK
                if target_page == 1:
                    return True
                return False
            
            # ✅ FIX: Tìm link bằng nhiều cách
            # Cách 1: Tìm tất cả link trong pagination và check href có chứa pn={target_page}
            # (Handle cả &&pn= và &pn=)
            link = None
            all_links = pagination_div.locator('a')
            link_count = await all_links.count()
            
            for i in range(link_count):
                link_elem = all_links.nth(i)
                href = await link_elem.get_attribute('href') or ''
                # Check nhiều pattern: pn=2, &pn=2, &&pn=2
                if f'pn={target_page}' in href:
                    link = link_elem
                    break
            
            # Cách 2: Nếu không tìm thấy qua href, tìm link có text chính xác = target_page
            if not link:
                for i in range(link_count):
                    link_elem = all_links.nth(i)
                    text = await link_elem.text_content()
                    if text and text.strip() == str(target_page):
                        link = link_elem
                        break
            
            # Click vào link nếu tìm thấy
            if link:
                await link.click()
                await asyncio.sleep(1)  # Đợi page load
                
                # Verify navigation thành công bằng cách check pagination info
                try:
                    await asyncio.sleep(0.5)  # Đợi thêm một chút
                    pagination_info = await self._extract_pagination_info(frame)
                    if pagination_info:
                        current_page = pagination_info.get("current_page", 0)
                        if current_page == target_page:
                            logger.info(f"✅ Navigated to page {target_page} (verified)")
                            return True
                        else:
                            logger.warning(f"⚠️ Navigation verification failed: expected page {target_page}, got {current_page}")
                    else:
                        logger.warning(f"⚠️ Cannot verify navigation (no pagination info)")
                        # Vẫn return True vì đã click được
                        return True
                except Exception as verify_e:
                    logger.warning(f"⚠️ Error verifying navigation: {verify_e}")
                    # Vẫn return True vì đã click được
                    return True
            else:
                logger.warning(f"⚠️ Cannot find link to page {target_page}")
                return False
                
        except Exception as e:
            logger.warning(f"⚠️ Error navigating to page {target_page}: {e}")
            return False
    
    async def _extract_pagination_info(self, frame) -> Optional[Dict[str, int]]:
        """
        Extract pagination info từ #currAcc div
        Returns: {"current_page": 1, "total_pages": 2, "total_records": 13} hoặc None
        """
        try:
            pagination_div = frame.locator('#currAcc').first
            if await pagination_div.count() == 0:
                return None
            
            # Lấy innerHTML để có thể parse <b> tags
            html_content = await pagination_div.inner_html()
            
            # Pattern 1: "Trang 1/<b>2</b>. Có <b>13</b> bản ghi."
            import re
            pattern_html = r"Trang\s+(\d+)/<b>(\d+)</b>\.\s+Có\s+<b>(\d+)</b>\s+bản\s+ghi"
            match = re.search(pattern_html, html_content)
            
            if match:
                current_page = int(match.group(1))
                total_pages = int(match.group(2))
                total_records = int(match.group(3))
                return {
                    "current_page": current_page,
                    "total_pages": total_pages,
                    "total_records": total_records
                }
            
            # Pattern 2: "Trang 1/2. Có 13 bản ghi." (plain text)
            text_content = await pagination_div.text_content()
            pattern_text = r"Trang\s+(\d+)/(\d+)\.\s+Có\s+(\d+)\s+bản\s+ghi"
            match = re.search(pattern_text, text_content)
            
            if match:
                current_page = int(match.group(1))
                total_pages = int(match.group(2))
                total_records = int(match.group(3))
                return {
                    "current_page": current_page,
                    "total_pages": total_pages,
                    "total_records": total_records
                }
            
            return None
            
        except Exception as e:
            logger.warning(f"⚠️ Error extracting pagination info: {e}")
            return None
    
    async def _extract_download_params(self, frame) -> Optional[Dict[str, str]]:
        """
        Lấy các tham số dse_* từ form (cần thiết để build URL download)
        ✅ FIX: Loại bỏ dse_pageId và pn (pagination params), dùng dse_pageId cố định khi download
        """
        try:
            params = {}
            
            # Lấy từ hidden inputs trong form
            form = frame.locator('form[name="traCuuKhaiForm"], form#traCuuKhaiForm').first
            if await form.count() > 0:
                inputs = form.locator('input[type="hidden"]')
                input_count = await inputs.count()
                
                for i in range(input_count):
                    input_elem = inputs.nth(i)
                    name = await input_elem.get_attribute('name')
                    value = await input_elem.get_attribute('value')
                    if name and value:
                        # ✅ FIX: Bỏ qua dse_pageId và pn (pagination params)
                        # dse_pageId trong form là số trang pagination (6, 7, ...)
                        # Khi download cần dùng dse_pageId cố định (14 hoặc 8)
                        if name in ['dse_pageId', 'pn']:
                            continue
                        params[name] = value
            
            # Thêm các params cố định
            params.update({
                'dse_operationName': 'traCuuToKhaiProc',
                'dse_processorState': 'viewTraCuuTkhai',
                'dse_nextEventName': 'downTkhai',
                'dse_applicationId': '-1'
            })
            
            logger.info(f"Extracted download params: {list(params.keys())}")
            return params if params.get('dse_sessionId') else None
            
        except Exception as e:
            logger.error(f"Error extracting download params: {e}")
            return None
    
    async def _download_one_via_url(
        self, 
        session_id: str,
        ma_tkhai: str, 
        item: Dict,
        base_params: Dict[str, str],
        temp_dir: str,
        frame=None
    ) -> Optional[Dict]:
        """
        Download 1 file bằng cách gọi URL trực tiếp với httpx
        ✅ FIX: Thêm frame parameter để có thể navigate về đúng trang nếu cần
        """
        try:
            # Build URL với params
            params = base_params.copy()
            params['messageId'] = ma_tkhai
            
            # ✅ FIX: Set dse_pageId cố định (không dùng từ form)
            # Nếu có dse_processorId, dùng pageId=8, ngược lại dùng pageId=14
            if params.get('dse_processorId'):
                params['dse_pageId'] = '8'
            else:
                params['dse_pageId'] = '14'
            
            # ✅ FIX: Thêm pn (page number) dựa trên page_number của item
            # pn cần match với trang của item để download đúng
            page_number = item.get("page_number", 1)
            params['pn'] = str(page_number)
            
            # Xác định download_type để set dse_nextEventName
            download_type = item.get("download_type", "downloadTkhai")
            if download_type == "downloadBke":
                params['dse_nextEventName'] = 'downBke'
            else:
                params['dse_nextEventName'] = 'downTkhai'
            
            download_url = "https://thuedientu.gdt.gov.vn/etaxnnt/Request"
            
            # Lấy httpx client (đã có cookies từ session)
            http_client = await self._get_http_client(session_id)
            if not http_client:
                logger.warning(f"No http client for {ma_tkhai}")
                # ✅ Chụp màn hình khi không có http client
                await self._take_screenshot_on_download_error(
                    session_id, ma_tkhai, "No_http_client", frame
                )
                return None
            
            # ✅ FIX: Retry logic cho tờ thuyết minh và các file có thể fail
            max_retries = 2 if download_type == "downloadBke" else 1
            retry_delay = 1.0
            
            for retry in range(max_retries):
                if retry > 0:
                    await asyncio.sleep(retry_delay)
                
                # ✅ FIX QUAN TRỌNG: Navigate về đúng trang TRƯỚC KHI download (cả lần đầu và retry)
                # Điều này đảm bảo browser đang ở đúng trang của item, tránh server trả về HTML của trang khác
                if frame:
                    try:
                        navigate_success = await self._navigate_to_page(frame, page_number)
                        if navigate_success:
                            # Đợi table load sau khi navigate
                            try:
                                table_body = frame.locator('#allResultTableBody, table.md_list2 tbody, table#data_content_onday tbody').first
                                await table_body.wait_for(timeout=5000)
                                await asyncio.sleep(0.5)  # Đợi thêm một chút để đảm bảo page đã load xong
                            except:
                                pass
                            
                            fresh_params = await self._extract_download_params(frame)
                            if fresh_params:
                                params.update({
                                    'dse_sessionId': fresh_params.get('dse_sessionId', params.get('dse_sessionId')),
                                    'dse_processorId': fresh_params.get('dse_processorId', params.get('dse_processorId')),
                                })
                        else:
                            pass
                    except Exception as nav_e:
                        pass
                
                try:
                    response = await http_client.get(download_url, params=params, timeout=30.0)
                except Exception as e:
                    if retry < max_retries - 1:
                        continue
                    # ✅ Chụp màn hình khi fail sau khi retry hết
                    await self._take_screenshot_on_download_error(
                        session_id, ma_tkhai, f"Request_error_{str(e)[:30]}", frame
                    )
                    return None
            
                if response.status_code != 200:
                    if retry < max_retries - 1:
                        continue
                    await self._take_screenshot_on_download_error(
                        session_id, ma_tkhai, f"HTTP_{response.status_code}", frame
                    )
                    return None
                
                if len(response.content) == 0:
                    location = response.headers.get('location', '')
                    if retry < max_retries - 1:
                        continue
                    # ✅ Chụp màn hình khi fail sau khi retry hết
                    await self._take_screenshot_on_download_error(
                        session_id, ma_tkhai, "Empty_response_0_bytes", frame
                    )
                    return None
                
                content = response.content
                content_type = response.headers.get('content-type', '').lower()
                
                # ✅ FIX: Check content-type header trước
                is_xml_by_type = 'xml' in content_type or 'text/xml' in content_type
                is_xlsx_by_type = 'spreadsheet' in content_type or 'excel' in content_type or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type
                is_xls_by_type = 'application/vnd.ms-excel' in content_type or 'application/msexcel' in content_type
                
                # Validate XML/XLSX/XLS by content
                is_xml = content.startswith(b'<?xml') or b'<HSoTKhai' in content or b'<TKhai' in content or b'<BKe' in content
                
                # ✅ FIX: XLSX files start with PK (ZIP signature)
                # ZIP local file header: PK\x03\x04
                # ZIP end of central directory: PK\x05\x06
                # Check ở đầu file (có thể có BOM nhưng thường không)
                is_xlsx = (len(content) >= 2 and content[:2] == b'PK') or \
                          (len(content) >= 4 and content[0:2] == b'PK' and content[2:4] in [b'\x03\x04', b'\x05\x06'])
                
                # ✅ FIX: XLS files (Excel 97-2003) start with OLE2 signature
                # OLE2 signature: D0 CF 11 E0 A1 B1 1A E1 (8 bytes đầu)
                # ✅ FIX: Mở rộng check cho cả downloadTkhai khi content-type là application/octet-stream
                # (BCTC có type downloadTkhai nhưng server có thể trả về XLS file)
                is_xls = False
                if len(content) >= 8:
                    xls_signature = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'
                    if content[:8] == xls_signature:
                        # Accept XLS cho downloadBke hoặc khi content-type không rõ ràng
                        if download_type == "downloadBke" or 'application/octet-stream' in content_type or content_type == '':
                            is_xls = True
                            logger.info(f"[{ma_tkhai}] Detected XLS file (Excel 97-2003) - OLE2 signature")
                
                # ✅ FIX: Check xem có phải là HTML error page không (ưu tiên check này trước)
                # Check trong toàn bộ content (không chỉ 500 bytes đầu) vì HTML có thể nằm ở giữa
                is_html = (b'<!DOCTYPE html>' in content or 
                          b'<html' in content.lower()[:1000] or 
                          b'<HTML' in content[:1000] or
                          'html' in content_type or
                          (len(content) > 1000 and b'<body' in content.lower()[:2000] and b'<head' in content.lower()[:2000]))
                
                # ✅ FIX: Nếu là HTML, save file tạm để debug và reject
                if is_html:
                    # Save file tạm để inspect
                    debug_file = os.path.join(temp_dir, f"DEBUG_{ma_tkhai}_retry{retry+1}.html")
                    try:
                        with open(debug_file, 'wb') as f:
                            f.write(content)
                        logger.warning(f"[{ma_tkhai}] Server returned HTML page (saved to {debug_file}, content-type: {content_type}, size: {len(content)} bytes, retry: {retry+1}/{max_retries})")
                    except:
                        pass
                    
                    content_preview = content[:500].decode('utf-8', errors='ignore')
                    logger.debug(f"[{ma_tkhai}] Content preview: {content_preview[:200]}")
                    if retry < max_retries - 1:
                        continue
                    # ✅ Chụp màn hình khi fail sau khi retry hết
                    await self._take_screenshot_on_download_error(
                        session_id, ma_tkhai, "HTML_response", frame
                    )
                    return None
                
                # ✅ FIX: Nếu content hợp lệ (XML/XLSX/XLS), break khỏi retry loop
                if (is_xml or is_xlsx or is_xls) or (is_xml_by_type or is_xlsx_by_type or is_xls_by_type):
                    break
                
                # ✅ FIX: Nếu content-type là application/octet-stream và size hợp lý (>1000 bytes)
                # Có thể là XLSX file (server không set đúng content-type)
                # Check xem có phải là binary file không (không phải text/HTML)
                # ✅ FIX: Check với 'in' thay vì '==' vì có thể có thêm charset=utf-8
                if 'application/octet-stream' in content_type or content_type == '':
                    if len(content) > 1000:
                        # ✅ FIX: Check XLS signature trước (OLE2 - 8 bytes đầu)
                        # ✅ FIX: Mở rộng check cho cả downloadTkhai (BCTC có type downloadTkhai nhưng có thể là XLS)
                        if len(content) >= 8:
                            xls_signature = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'
                            if content[:8] == xls_signature:
                                logger.info(f"[{ma_tkhai}] Detected XLS file (Excel 97-2003) despite content-type: {content_type}")
                                is_xls = True
                                break
                        
                        # ✅ FIX: Check PK signature ở nhiều vị trí (có thể bị wrap hoặc offset)
                        # Check ở đầu file (0-100 bytes)
                        pk_positions = []
                        for offset in range(0, min(100, len(content) - 2)):
                            if content[offset:offset+2] == b'PK':
                                pk_positions.append(offset)
                                # Check xem có phải là ZIP/XLSX signature hợp lệ không
                                if offset + 4 <= len(content):
                                    next_bytes = content[offset+2:offset+4]
                                    if next_bytes in [b'\x03\x04', b'\x05\x06', b'\x07\x08']:
                                        logger.info(f"[{ma_tkhai}] Detected XLSX/ZIP file (PK signature at offset {offset}) despite content-type: {content_type}")
                                        is_xlsx = True
                                        break
                        
                        if is_xlsx:
                            break
                            
                        # XML signature: <?xml (check ở đầu file)
                        first_bytes = content[:10] if len(content) >= 10 else content
                        if first_bytes[:5] == b'<?xml':
                            logger.info(f"[{ma_tkhai}] Detected XML file despite content-type: {content_type}")
                            is_xml = True
                            break
                        
                        # ✅ FIX: Nếu tìm thấy PK nhưng không phải signature hợp lệ, log để debug
                        if pk_positions:
                            logger.debug(f"[{ma_tkhai}] Found PK at positions {pk_positions} but not valid ZIP signature")
                        # Nếu không phải HTML và size hợp lý, có thể là binary file hợp lệ
                        # Đặc biệt cho downloadBke (tờ thuyết minh) - thường là XLSX
                        elif len(content) > 5000:
                            # ✅ FIX: Bỏ lưu .bin → Lưu HTML của frame browser để debug
                            logger.warning(f"[{ma_tkhai}] Unknown binary content (content-type: {content_type}, size: {len(content)} bytes). First bytes (hex): {first_bytes.hex()[:40]}")
                            try:
                                # Lưu HTML của frame hiện tại để debug (thay vì lưu .bin)
                                if frame:
                                    frame_html = await frame.content()
                                    debug_html_file = os.path.join(temp_dir, f"DEBUG_{ma_tkhai}_retry{retry+1}_page.html")
                                    with open(debug_html_file, 'w', encoding='utf-8') as f:
                                        f.write(frame_html)
                                    logger.warning(f"[{ma_tkhai}] Saved browser page HTML to {debug_html_file}")
                                # ✅ Chụp màn hình ngay khi lưu file debug
                                await self._take_screenshot_on_download_error(
                                    session_id, ma_tkhai, f"Unknown_binary_{len(content)}bytes", frame
                                )
                            except Exception as e:
                                logger.warning(f"[{ma_tkhai}] Error saving debug HTML: {e}")
                            
                            # Log để debug
                            logger.debug(f"[{ma_tkhai}] First 100 bytes (hex): {content[:100].hex()}")
                            # Check xem có phải là text/HTML không (check một phần content)
                            try:
                                text_sample = content[:1000].decode('utf-8', errors='ignore')
                                if '<html' in text_sample.lower() or '<!doctype' in text_sample.lower():
                                    logger.warning(f"[{ma_tkhai}] Content appears to be HTML despite size")
                                    if retry < max_retries - 1:
                                        continue
                                    return None
                            except:
                                pass
                            
                            # Retry để xem có phải là timing issue không
                            if retry < max_retries - 1:
                                continue
                            # Nếu là downloadBke và size lớn, có thể accept (server có thể trả về file nhưng không đúng format)
                            # Nhưng để an toàn, reject nếu không detect được signature
                            # ✅ Chụp màn hình khi fail sau khi retry hết
                            await self._take_screenshot_on_download_error(
                                session_id, ma_tkhai, f"Unknown_binary_{len(content)}bytes", frame
                            )
                            return None
                
                # ✅ FIX: Nếu không detect được, save file debug và log
                logger.warning(f"[{ma_tkhai}] Not XML/XLSX/XLS content (content-type: {content_type}, size: {len(content)} bytes, retry: {retry+1}/{max_retries})")
                
                # ✅ FIX: Bỏ lưu .bin → Lưu HTML của frame browser để debug
                if len(content) > 1000 and frame:
                    try:
                        frame_html = await frame.content()
                        debug_html_file = os.path.join(temp_dir, f"DEBUG_{ma_tkhai}_retry{retry+1}_page.html")
                        with open(debug_html_file, 'w', encoding='utf-8') as f:
                            f.write(frame_html)
                        logger.warning(f"[{ma_tkhai}] Saved browser page HTML to {debug_html_file}")
                        # ✅ Chụp màn hình ngay khi lưu file debug
                        await self._take_screenshot_on_download_error(
                            session_id, ma_tkhai, f"Invalid_content_{len(content)}bytes", frame
                        )
                    except Exception as e:
                        logger.warning(f"[{ma_tkhai}] Failed to save debug HTML: {e}")
                
                # Log thêm để debug
                if len(content) > 0:
                    first_bytes = content[:100] if len(content) >= 100 else content
                    logger.debug(f"[{ma_tkhai}] First bytes (hex): {first_bytes.hex()[:200]}")
                    # Thử decode để xem có phải text không
                    try:
                        text_preview = content[:200].decode('utf-8', errors='ignore')
                        logger.debug(f"[{ma_tkhai}] First 200 chars (text): {text_preview[:100]}")
                    except:
                        pass
                
                # Retry nếu còn lượt
                if retry < max_retries - 1:
                    continue
                # ✅ Chụp màn hình khi fail sau khi retry hết
                await self._take_screenshot_on_download_error(
                    session_id, ma_tkhai, f"Invalid_content_{content_type}_{len(content)}bytes", frame
                )
                return None
            
            # ✅ FIX: Nếu ra khỏi loop mà vẫn không có content hợp lệ, return None
            if not (is_xml or is_xlsx or is_xls) and not (is_xml_by_type or is_xlsx_by_type or is_xls_by_type):
                logger.warning(f"[{ma_tkhai}] Failed after {max_retries} retries")
                # ✅ Chụp màn hình khi fail sau khi retry hết
                await self._take_screenshot_on_download_error(
                    session_id, ma_tkhai, f"Failed_after_{max_retries}_retries", frame
                )
                return None
            
            # Lưu file với extension đúng
            file_name = item["file_name"]
            # ✅ FIX: Xác định extension dựa trên file type thực tế
            if is_xls:
                file_ext = ".xls"
            elif is_xlsx or (download_type == "downloadBke" and not is_xml):
                file_ext = ".xlsx"
            else:
                file_ext = ".xml"
            
            # ✅ FIX: Validate mã giao dịch trong file XML để đảm bảo không bị lộn file
            if is_xml:
                try:
                    content_str = content.decode('utf-8', errors='ignore')
                    # Check xem mã giao dịch có trong file không
                    if ma_tkhai not in content_str:
                        logger.warning(f"⚠️ [{ma_tkhai}] Mã giao dịch không tìm thấy trong XML content - có thể file bị lộn!")
                        # Vẫn lưu nhưng log warning để debug
                except:
                    pass
            
            # Nếu file_name đã có extension hợp lệ, giữ nguyên, nếu không thì thêm extension
            if not file_name.endswith((".xml", ".xlsx", ".xls")):
                save_path = os.path.join(temp_dir, file_name + file_ext)
            else:
                save_path = os.path.join(temp_dir, file_name)
            
            with open(save_path, 'wb') as f:
                f.write(content)
            
            # Verify
            if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                logger.info(f"✅ [{ma_tkhai}] Downloaded {len(content)} bytes -> {os.path.basename(save_path)}")
                return item
            else:
                logger.warning(f"❌ [{ma_tkhai}] File not saved")
                # ✅ Chụp màn hình khi file không được save
                await self._take_screenshot_on_download_error(
                    session_id, ma_tkhai, "File_not_saved", frame
                )
                return None
                
        except asyncio.TimeoutError:
            logger.warning(f"⏱️ [{ma_tkhai}] Timeout")
            # ✅ Chụp màn hình khi timeout
            await self._take_screenshot_on_download_error(
                session_id, ma_tkhai, "Timeout", frame
            )
            return None
        except Exception as e:
            logger.warning(f"❌ [{ma_tkhai}] Error: {e}")
            # ✅ Chụp màn hình khi có exception
            await self._take_screenshot_on_download_error(
                session_id, ma_tkhai, f"Exception_{str(e)[:30]}", frame
            )
            return None
    
    async def _batch_download_one(self, session: SessionData, item: Dict, temp_dir: str, ssid: str, frame, session_id: str = None) -> Optional[Dict]:
        """
        Download một item bằng cách gọi JS function và intercept response
        ✅ FIXED: Wait for navigation, then extract XML from final page
        """
        try:
            id_tk = item["id"]
            file_name = item["file_name"]
            has_link = item.get("has_link", False)
            download_type = item.get("download_type", "downloadTkhai")
            
            # Bỏ qua trường hợp không có link
            if not has_link:
                logger.info(f"Skipping {id_tk} - no download link (special tokhai)")
                return None
            
            logger.info(f"[DEBUG] Downloading {id_tk} via {download_type}")
            
            page = session.page
            
            # ============================================
            # ✅ STRATEGY: Gọi JS → Đợi navigation → Lấy XML từ page cuối cùng
            # ============================================
            
            try:
                # Bước 1: Gọi JavaScript function (trigger navigation)
                js_function = "downloadBke" if download_type == "downloadBke" else "downloadTkhai"
                logger.info(f"[DEBUG] Calling {js_function}('{id_tk}')")
                
                # Track navigation
                navigation_promise = page.wait_for_load_state('domcontentloaded', timeout=30000)
                
                # Gọi function trong frame
                await frame.evaluate(f"{js_function}('{id_tk}')")
                logger.info(f"[DEBUG] Called {js_function}('{id_tk}')")
                
                # Bước 2: Đợi navigation hoàn tất
                try:
                    await navigation_promise
                    logger.info(f"[DEBUG] Navigation completed for {id_tk}")
                except:
                    # Nếu không có navigation event, đợi một chút
                    await asyncio.sleep(2)
                    logger.info(f"[DEBUG] No navigation event, waited 2s for {id_tk}")
                
                # Bước 3: Đợi content load xong
                await asyncio.sleep(1)
                
                # Bước 4: Lấy content từ page hiện tại
                page_content = await page.content()
                logger.info(f"[DEBUG] Got page content, length: {len(page_content)}")
                
                # Bước 5: Validate content
                is_xml = '<?xml' in page_content or '<HSoTKhai' in page_content or '<TKhai' in page_content
                is_error_page = 'timeout.jsp' in page.url or 'error' in page.url.lower()
                is_html_page = '<!DOCTYPE html>' in page_content or '<html' in page_content
                
                if is_error_page:
                    logger.warning(f"[DEBUG] Error page detected for {id_tk}")
                    return None
                
                if is_xml and not is_html_page:
                    # Content là XML thuần túy
                    response_data = page_content.encode('utf-8')
                    logger.info(f"[DEBUG] Valid XML content for {id_tk}: {len(response_data)} bytes")
                elif '<?xml' in page_content:
                    # XML nằm trong HTML (có thể trong <pre> tag hoặc embedded)
                    # Extract XML từ page
                    try:
                        # Tìm XML trong page content
                        xml_start = page_content.find('<?xml')
                        if xml_start >= 0:
                            # Tìm closing tag cuối cùng
                            # Giả sử XML kết thúc bằng </HSoTKhai> hoặc tag tương tự
                            closing_tags = ['</HSoTKhai>', '</TKhai>', '</BKe>']
                            xml_end = -1
                            for tag in closing_tags:
                                pos = page_content.rfind(tag)
                                if pos > xml_start:
                                    xml_end = pos + len(tag)
                                    break
                            
                            if xml_end > xml_start:
                                xml_content = page_content[xml_start:xml_end]
                                response_data = xml_content.encode('utf-8')
                                logger.info(f"[DEBUG] Extracted XML from HTML for {id_tk}: {len(response_data)} bytes")
                            else:
                                logger.warning(f"[DEBUG] Could not find XML closing tag for {id_tk}")
                                return None
                        else:
                            logger.warning(f"[DEBUG] No XML found in page content for {id_tk}")
                            return None
                    except Exception as e:
                        logger.warning(f"[DEBUG] Error extracting XML for {id_tk}: {e}")
                        return None
                else:
                    # Không phải XML
                    logger.warning(f"[DEBUG] Content is not XML for {id_tk}")
                    logger.debug(f"[DEBUG] First 500 chars: {page_content[:500]}")
                    return None
                
                # Bước 6: Lưu file
                if response_data and len(response_data) > 100:
                    file_ext = ".xlsx" if download_type == "downloadBke" else ".xml"
                    save_path = os.path.join(temp_dir, file_name + file_ext if not file_name.endswith((".xml", ".xlsx")) else file_name)
                    
                    # Lưu file
                    with open(save_path, 'wb') as f:
                        f.write(response_data)
                    
                    # Verify file
                    if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                        # Final validation: check file content
                        with open(save_path, 'rb') as f:
                            first_bytes = f.read(100)
                            if b'<?xml' in first_bytes or b'<HSoTKhai' in first_bytes:
                                logger.info(f"✅ Downloaded {id_tk} ({len(response_data)} bytes) -> {save_path}")
                                return item
                            else:
                                logger.warning(f"❌ Downloaded file is not valid XML for {id_tk}")
                                # Delete invalid file
                                os.remove(save_path)
                                return None
                    else:
                        logger.warning(f"❌ File not saved properly for {id_tk}")
                        return None
                else:
                    logger.warning(f"❌ No valid response data for {id_tk}")
                    return None
                    
            except Exception as inner_e:
                logger.warning(f"❌ Error in download process for {id_tk}: {inner_e}")
                return None
                
        except Exception as e:
            logger.warning(f"❌ Error downloading {item.get('id', 'unknown')}: {e}")
            return None
    
    async def _batch_download_optimized(
        self,
        session_id: str,
        download_queue: List[Dict],
        temp_dir: str,
        frame,
        batch_size: int = 3,
        page_params_map: Dict[int, Dict[str, str]] = None,
        progress_callback=None  # ✅ Callback để yield progress sau mỗi file
    ) -> List[Dict]:
        """
        Download với batching tối ưu:
        - Dùng params tương ứng với từng trang
        - Download theo batch (3-5 files/batch)
        - Dùng httpx thay vì playwright navigation
        
        Returns: List các item đã download thành công
        """
        
        # ✅ FIX: Group items theo page_number để dùng đúng params
        items_by_page = {}
        for item in download_queue:
            page_num = item.get("page_number", 1)
            if page_num not in items_by_page:
                items_by_page[page_num] = []
            items_by_page[page_num].append(item)
        
        logger.info(f"📋 Grouped {len(download_queue)} items into {len(items_by_page)} pages")
        
        if not page_params_map:
            logger.error("❌ No page_params_map provided, cannot download")
            return []
        
        # Bước 2: Download từng page với params tương ứng
        total = len(download_queue)
        successful_downloads = []
        
        logger.info(f"📦 Starting batch download: {total} files, batch_size={batch_size}")
        
        # Download từng page
        for page_num, page_items in sorted(items_by_page.items()):
            # Lấy params cho page này
            base_params = page_params_map.get(page_num)
            if not base_params:
                logger.warning(f"⚠️ No params for page {page_num}, skipping {len(page_items)} items")
                continue
            
            logger.info(f"📄 Downloading page {page_num}: {len(page_items)} items")
            
            # ✅ FIX: Navigate về đúng trang trước khi download (đảm bảo params đúng)
            # Chỉ navigate nếu không phải trang 1 (vì đã ở trang 1 rồi)
            if page_num > 1:
                navigate_success = await self._navigate_to_page(frame, page_num)
                if not navigate_success:
                    logger.warning(f"⚠️ Cannot navigate to page {page_num}, skipping {len(page_items)} items")
                    continue
                
                # Đợi table load và extract params lại (đảm bảo params mới nhất)
                try:
                    table_body = frame.locator('#allResultTableBody, table.md_list2 tbody, table#data_content_onday tbody').first
                    await table_body.wait_for(timeout=5000)
                    await asyncio.sleep(1)
                    
                    # Extract params lại từ trang này (có thể đã thay đổi sau khi navigate)
                    fresh_params = await self._extract_download_params(frame)
                    if fresh_params:
                        base_params = fresh_params
                        logger.info(f"✅ Refreshed params for page {page_num}")
                except:
                    logger.warning(f"⚠️ Cannot refresh params for page {page_num}, using cached params")
            
            # ✅ PHƯƠNG ÁN 3: Sequential Download với Refresh Params
            # Download tuần tự từng file, refresh params trước mỗi file để đảm bảo ổn định
            logger.info(f"📦 Page {page_num}: Downloading {len(page_items)} files sequentially...")
            
            for idx, item in enumerate(page_items):
                file_num = idx + 1
                logger.info(f"📄 Page {page_num}, File {file_num}/{len(page_items)}: {item['id']}")
                
                # ✅ Refresh params trước mỗi file để đảm bảo state đúng
                # (Tránh trường hợp params bị outdated sau khi download file trước)
                try:
                    fresh_params = await self._extract_download_params(frame)
                    if fresh_params:
                        base_params = fresh_params
                        # logger.debug(f"✅ Refreshed params for file {file_num}")
                except Exception as e:
                    logger.warning(f"⚠️ Cannot refresh params for file {file_num}, using cached params: {e}")
                
                # Download file
                try:
                    result = await self._download_one_via_url(
                        session_id,
                        item["id"],
                        item,
                        base_params,
                        temp_dir,
                        frame=frame  # ✅ Truyền frame để có thể navigate khi retry
                    )
                    
                    if result and not isinstance(result, Exception):
                        successful_downloads.append(result)
                        logger.info(f"✅ Page {page_num}, File {file_num}/{len(page_items)}: Success")
                        
                        # ✅ Yield progress sau mỗi file download xong
                        if progress_callback:
                            await progress_callback(result, successful_downloads, page_items)
                    else:
                        logger.warning(f"❌ Page {page_num}, File {file_num}/{len(page_items)}: Failed")
                
                except Exception as e:
                    logger.error(f"❌ Page {page_num}, File {file_num}/{len(page_items)}: Error - {e}")
                
                # ✅ Delay nhỏ giữa mỗi file (0.1s) để:
                # - Tránh spam server
                # - Đảm bảo browser state ổn định
                # - Vẫn tải đủ nhanh (50 files ~ 5 giây chỉ tính delay)
                await asyncio.sleep(0.1)
            
            # Sequential download đã tự động tạo khoảng cách giữa các page
            # Không cần delay thêm
        
        logger.info(f"🎉 Total downloaded: {len(successful_downloads)} / {total}")
        return successful_downloads
    
    async def _batch_download(self, session: SessionData, download_queue: List[Dict], temp_dir: str, ssid: str, frame, session_id: str = None, page_params_map: Dict[int, Dict[str, str]] = None):
        """
        ✅ FIXED VERSION: Download với batching tối ưu dùng httpx
        
        Returns: List các item download thành công
        """
        if not session_id:
            logger.warning("No session_id provided, cannot use optimized download")
            return []
        
        # Filter chỉ lấy items có link (bỏ qua special tokhai)
        download_queue_filtered = [item for item in download_queue if item.get("has_link", False)]
        
        if not download_queue_filtered:
            logger.warning("No items with download links")
            return []
        
        if not page_params_map:
            logger.warning("No page_params_map provided, cannot download")
            return []
        
        # Gọi download optimized
        return await self._batch_download_optimized(
            session_id=session_id,
            download_queue=download_queue_filtered,
            temp_dir=temp_dir,
            frame=frame,
            batch_size=3,  # Tối ưu: 3 files/batch
            page_params_map=page_params_map  # ✅ FIX: Truyền page_params_map
        )
    
    async def _batch_download_old(self, session: SessionData, download_queue: List[Dict], temp_dir: str, ssid: str, frame, session_id: str = None):
        """
        [DEPRECATED] Download nhiều file song song (tối ưu tốc độ)
        Limit concurrent downloads = 5 để không quá tải
        
        Returns: List các item download thành công
        """
        semaphore = asyncio.Semaphore(5)  # Max 5 downloads cùng lúc
        page = session.page
        successful_downloads = []  # Track những item download thành công
        
        async def download_one(item: Dict):
            async with semaphore:
                try:
                    id_tk = item["id"]
                    file_name = item["file_name"]
                    cols = item["cols"]
                    has_link = item.get("has_link", False)
                    download_type = item.get("download_type")  # "downloadTkhai", "downloadBke", hoặc None
                    
                    if has_link:
                        # Có link - click để download (bình thường)
                        download_link = cols.nth(2).locator('a')
                        # Bắt download event từ page (không phải frame)
                        async with page.expect_download(timeout=30000) as download_info:
                            await download_link.first.click()
                        
                        download = await download_info.value
                        # File thuyết minh có thể là .xlsx, còn tờ khai thường là .xml
                        file_ext = ".xlsx" if download_type == "downloadBke" else ".xml"
                        save_path = os.path.join(temp_dir, file_name + file_ext if not file_name.endswith((".xml", ".xlsx")) else file_name)
                        await download.save_as(save_path)
                        
                        # Kiểm tra file đã được lưu thành công
                        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                            logger.info(f"Downloaded {id_tk} ({download_type}) -> {file_name}")
                            successful_downloads.append(item)
                        else:
                            logger.warning(f"Download failed: File not saved or empty for {id_tk}")
                    else:
                        # Tờ khai đặc biệt - không có link <a> download (không có onclick="downloadTkhai" hoặc title="Tải tệp")
                        # Hàm downloadTkhai(msgId) dùng window.location.href để navigate, không trigger download event trên page
                        # Nên cần build URL và dùng new_page.goto() để trigger download event
                        logger.info(f"Special tokhai (no download link) detected: {id_tk}, building download URL")
                        
                        current_ssid = ssid
                        
                        # Lấy session ID từ form
                        if not current_ssid or current_ssid == "NotFound":
                            try:
                                dse_session_input = frame.locator('form[name="traCuuKhaiForm"] input[name="dse_sessionId"], form#traCuuKhaiForm input[name="dse_sessionId"], input[name="dse_sessionId"]').first
                                if await dse_session_input.count() > 0:
                                    current_ssid = await dse_session_input.get_attribute('value') or ""
                                    if current_ssid:
                                        logger.info(f"Retrieved dse_sessionId from form input: {current_ssid[:30]}...")
                            except Exception as e:
                                logger.warning(f"Error getting dse_sessionId from form input: {e}")
                        
                        # Lấy từ frame URL nếu chưa có
                        if not current_ssid or current_ssid == "NotFound":
                            try:
                                frame_url = frame.url
                                match = re.search(r"[&?]dse_sessionId=([^&]+)", frame_url)
                                if match:
                                    current_ssid = match.group(1)
                                    logger.info(f"Retrieved dse_sessionId from frame URL: {current_ssid[:30]}...")
                            except Exception as e:
                                logger.warning(f"Error getting dse_sessionId from frame URL: {e}")
                        
                        if current_ssid and current_ssid != "NotFound":
                            # Lấy processor ID từ form
                            dse_processor_id = ""
                            try:
                                processor_id_input = frame.locator('form[name="traCuuKhaiForm"] input[name="dse_processorId"], form#traCuuKhaiForm input[name="dse_processorId"], input[name="dse_processorId"]').first
                                if await processor_id_input.count() > 0:
                                    dse_processor_id = await processor_id_input.first.get_attribute('value') or ""
                                    if dse_processor_id:
                                        logger.info(f"Retrieved dse_processorId from form: {dse_processor_id[:30]}...")
                            except:
                                pass
                            
                            # Build URL giống như hàm downloadTkhai() làm
                            # downloadTkhai() làm: window.location.href='/etaxnnt/Request?dse_sessionId=...&dse_applicationId=-1&dse_operationName=traCuuToKhaiProc&dse_pageId=8&dse_processorState=viewTraCuuTkhai&dse_processorId=...&dse_nextEventName=downTkhai&messageId='+msgId
                            if dse_processor_id:
                                # Có processor ID: dùng pageId=10 (hoặc có thể là 8 như trong HTML mẫu)
                                download_url = f"{BASE_URL}/etaxnnt/Request?dse_sessionId={current_ssid}&dse_applicationId=-1&dse_operationName=traCuuToKhaiProc&dse_pageId=8&dse_processorState=viewTraCuuTkhai&dse_processorId={dse_processor_id}&dse_nextEventName=downTkhai&messageId={id_tk}"
                            else:
                                # Không có processor ID: dùng pageId=14
                                download_url = f"{BASE_URL}/etaxnnt/Request?dse_sessionId={current_ssid}&dse_applicationId=-1&dse_operationName=traCuuToKhaiProc&dse_pageId=14&dse_processorState=viewTraCuuTkhai&dse_nextEventName=downTkhai&messageId={id_tk}"
                            
                            logger.info(f"Downloading special (no link) {id_tk} via new_page.goto(): {download_url[:100]}...")
                            
                            new_page = None
                            try:
                                new_page = await session.context.new_page()
                                new_page.set_default_timeout(30000)
                                
                                # Intercept response để bắt file download
                                download_occurred = False
                                response_data = None
                                
                                async def handle_response(response):
                                    nonlocal download_occurred, response_data
                                    content_type = response.headers.get('content-type', '').lower()
                                    # Kiểm tra nếu response là XML file
                                    if 'xml' in content_type or response.url.endswith('.xml') or 'application/xml' in content_type or 'text/xml' in content_type:
                                        download_occurred = True
                                        response_data = await response.body()
                                        logger.info(f"Got XML response for {id_tk}, size: {len(response_data)} bytes")
                                
                                new_page.on("response", handle_response)
                                
                                # Navigate đến URL
                                response = await new_page.goto(download_url, wait_until="domcontentloaded", timeout=30000)
                                
                                # Chờ một chút để response được xử lý
                                await asyncio.sleep(1)
                                
                                # Nếu có download event, bắt nó
                                if download_occurred and response_data:
                                    save_path = os.path.join(temp_dir, file_name + ".xml" if not file_name.endswith(".xml") else file_name)
                                    with open(save_path, 'wb') as f:
                                        f.write(response_data)
                                    
                                    if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                                        logger.info(f"Downloaded special (no link) {id_tk} -> {file_name}")
                                        successful_downloads.append(item)
                                    else:
                                        logger.warning(f"Download failed: File not saved or empty for special {id_tk}")
                                else:
                                    # Fallback: thử bắt download event
                                    try:
                                        async with new_page.expect_download(timeout=5000) as download_info:
                                            # Trigger download bằng cách click hoặc navigate lại
                                            await new_page.reload(wait_until="domcontentloaded")
                                        
                                        download = await download_info.value
                                        save_path = os.path.join(temp_dir, file_name + ".xml" if not file_name.endswith(".xml") else file_name)
                                        await download.save_as(save_path)
                                        
                                        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                                            logger.info(f"Downloaded special (no link) {id_tk} via download event -> {file_name}")
                                            successful_downloads.append(item)
                                        else:
                                            logger.warning(f"Download failed: File not saved or empty for special {id_tk}")
                                    except:
                                        # Nếu không có download event, thử lấy content từ response
                                        if response:
                                            content = await response.body()
                                            if content and len(content) > 100:  # Có thể là XML file
                                                save_path = os.path.join(temp_dir, file_name + ".xml" if not file_name.endswith(".xml") else file_name)
                                                with open(save_path, 'wb') as f:
                                                    f.write(content)
                                                
                                                if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                                                    logger.info(f"Downloaded special (no link) {id_tk} from response body -> {file_name}")
                                                    successful_downloads.append(item)
                                                else:
                                                    logger.warning(f"Download failed: File not saved or empty for special {id_tk}")
                                            else:
                                                logger.warning(f"No valid content in response for {id_tk}")
                                        else:
                                            logger.warning(f"No response received for {id_tk}")
                            except Exception as e2:
                                logger.warning(f"Error downloading special {id_tk} via new_page.goto(): {e2}")
                                
                                # Fallback: thử dùng httpx client với cookies
                                if session_id:
                                    try:
                                        logger.info(f"Trying httpx fallback for {id_tk}")
                                        http_client = await self._get_http_client(session_id)
                                        if http_client:
                                            response = await http_client.get(download_url, timeout=30.0)
                                            if response.status_code == 200:
                                                content = response.content
                                                if content and len(content) > 100:
                                                    save_path = os.path.join(temp_dir, file_name + ".xml" if not file_name.endswith(".xml") else file_name)
                                                    with open(save_path, 'wb') as f:
                                                        f.write(content)
                                                    
                                                    if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                                                        logger.info(f"Downloaded special (no link) {id_tk} via httpx -> {file_name}")
                                                        successful_downloads.append(item)
                                                    else:
                                                        logger.warning(f"httpx download failed: File not saved or empty for {id_tk}")
                                                else:
                                                    logger.warning(f"httpx response has no valid content for {id_tk}")
                                            else:
                                                logger.warning(f"httpx response status {response.status_code} for {id_tk}")
                                    except Exception as e3:
                                        logger.warning(f"httpx fallback also failed for {id_tk}: {e3}")
                                    finally:
                                        if new_page:
                                            try:
                                                await new_page.close()
                                            except:
                                                pass
                        else:
                            logger.warning(f"No valid session ID for special download: {id_tk}. ssid={ssid}")
                except Exception as e:
                    logger.warning(f"Error downloading {item.get('id', 'unknown')}: {e}")
        
        # Download tất cả song song (max 5 cùng lúc)
        await asyncio.gather(*[download_one(item) for item in download_queue], return_exceptions=True)
        
        return successful_downloads
    
    async def _download_xml(self, client: httpx.AsyncClient, url: str, temp_dir: str, file_id: str):
        """Download XML file (async)"""
        try:
            response = await client.get(url)
            if response.status_code == 200:
                file_path = os.path.join(temp_dir, f"{file_id}.xml")
                with open(file_path, 'wb') as f:
                    f.write(response.content)
        except Exception as e:
            logger.error(f"Error downloading {file_id}: {e}")
    
    async def _download_single_thongbao(self, session: SessionData, item: Dict, temp_dir: str, max_retries: int = 2) -> bool:
        """
        Download 1 file thông báo với retry logic (dùng Playwright expect_download như cũ)
        
        Returns:
            True nếu download thành công
        """
        page = session.page
        id_tb = item["id"]
        file_name = item.get("file_name", id_tb)
        
        for retry in range(max_retries + 1):
            try:
                # Ưu tiên dùng download_link đã tìm sẵn
                download_link = item.get("download_link")
                
                if not download_link:
                    # Fallback: tìm lại từ cols
                    cols = item.get("cols")
                    col_idx = item.get("col_index", 10)
                    if cols:
                        download_link = cols.nth(col_idx).locator('a:has-text("Tải về")')
                
                if download_link and await download_link.count() > 0:
                    async with page.expect_download(timeout=30000) as download_info:
                        await download_link.first.click()
                    
                    download = await download_info.value
                    save_path = os.path.join(temp_dir, file_name + ".xml" if not file_name.endswith(".xml") else file_name)
                    await download.save_as(save_path)
                    
                    # Verify file exists and has content
                    if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                        logger.info(f"Downloaded thongbao {id_tb} -> {file_name}")
                        return True
                    else:
                        raise Exception("File empty or not saved")
                else:
                    logger.warning(f"No download link for thongbao {id_tb}")
                    return False
                    
            except Exception as e:
                logger.warning(f"Error downloading thongbao {id_tb} (attempt {retry + 1}/{max_retries + 1}): {e}")
                if retry < max_retries:
                    await asyncio.sleep(1)  # Wait before retry
        
        return False
    
    async def _download_xml_with_name(self, client: httpx.AsyncClient, url: str, temp_dir: str, file_id: str, file_name: str):
        """Download XML file với tên file custom (async)"""
        try:
            response = await client.get(url)
            if response.status_code == 200:
                # Đảm bảo tên file hợp lệ
                safe_name = file_name.replace("/", "_").replace("\\", "_").replace(":", "_")
                if not safe_name.endswith(".xml"):
                    safe_name += ".xml"
                file_path = os.path.join(temp_dir, safe_name)
                with open(file_path, 'wb') as f:
                    f.write(response.content)
                return {"id": file_id, "file_name": safe_name, "size": len(response.content)}
        except Exception as e:
            logger.error(f"Error downloading {file_id}: {e}")
            return None
    
    async def crawl_thongbao(
        self,
        session_id: str,
        start_date: str,
        end_date: str,
        job_id: Optional[str] = None
    ) -> AsyncGenerator[Dict[str, Any], None]:
        session = self.session_manager.get_session(session_id)
        if not session:
            yield {"type": "error", "error": "Session không tồn tại hoặc đã hết hạn", "error_code": "SESSION_NOT_FOUND"}
            return
        
        if not session.is_logged_in:
            yield {"type": "error", "error": "Chưa đăng nhập. Vui lòng đăng nhập lại.", "error_code": "NOT_LOGGED_IN"}
            return
        
        page = session.page
        
        # ✅ FIX: Tạo temp directory trong source code thay vì system temp (giống tờ khai)
        # Lấy đường dẫn project (tool-go-soft)
        current_dir = os.path.dirname(os.path.abspath(__file__))  # .../services/
        services_dir = os.path.dirname(current_dir)  # .../tool-go-soft/
        temp_base_dir = os.path.join(services_dir, "temp")  # .../tool-go-soft/temp/
        os.makedirs(temp_base_dir, exist_ok=True)
        
        # Tạo temp directory với timestamp để tránh conflict
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_dir = os.path.join(temp_base_dir, f"thongbao_{timestamp}")
        os.makedirs(temp_dir, exist_ok=True)
        
        logger.info(f"📁 Temp directory for thongbao files: {temp_dir}")  # ✅ Log temp_dir path để dễ tìm file debug
        ssid = session.dse_session_id
        
        try:
            yield {"type": "info", "message": "Đang xử lý ..."}
            
            # Navigate đến trang tra cứu thông báo qua connectSSO (giống tờ khai)
            success = await self._navigate_to_thongbao_page(page, ssid)
            
            if not success:
                yield {"type": "error", "error": "Không thể navigate đến trang tra cứu thông báo. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            # Tìm frame từ iframe SSO (giống tờ khai)
            frame = None
            try:
                frames = page.frames
                for f in frames:
                    if 'thuedientu.gdt.gov.vn' in f.url:
                        frame = f
                        logger.info(f"Found frame for thongbao: {frame.url[:100]}...")
                        break
            except Exception as e:
                logger.warning(f"Error finding frame: {e}")
            
            if not frame:
                yield {"type": "error", "error": "Không tìm thấy iframe sau khi navigate. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            # Đợi frame load và kiểm tra form thông báo
            try:
                await frame.wait_for_load_state('domcontentloaded', timeout=15000)
                await asyncio.sleep(1)
                await frame.wait_for_selector('#qryFromDate', timeout=15000)
                logger.info("Tra cuu thong bao form loaded successfully")
            except Exception as e:
                logger.warning(f"Frame found but form not found: {e}")
                
                # ✅ Screenshot khi có lỗi không tìm thấy form
                try:
                    screenshot_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "screenshots", f"thongbao_{session_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
                    os.makedirs(screenshot_dir, exist_ok=True)
                    
                    # Screenshot page
                    page_screenshot = os.path.join(screenshot_dir, "01_error_page.png")
                    await page.screenshot(path=page_screenshot, full_page=True)
                    logger.info(f"📸 Screenshot page saved: {page_screenshot}")
                    
                    # Screenshot frame
                    try:
                        frame_screenshot = os.path.join(screenshot_dir, "02_error_frame.png")
                        await frame.screenshot(path=frame_screenshot, full_page=True)
                        logger.info(f"📸 Screenshot frame saved: {frame_screenshot}")
                    except Exception as frame_e:
                        logger.warning(f"⚠️ Cannot screenshot frame: {frame_e}")
                    
                    # Lấy HTML của frame để debug
                    try:
                        frame_html = await frame.content()
                        html_file = os.path.join(screenshot_dir, "03_error_frame.html")
                        with open(html_file, 'w', encoding='utf-8') as f:
                            f.write(frame_html)
                        pass
                    except Exception as html_e:
                        pass
                    
                except Exception as screenshot_e:
                    logger.error(f"❌ Error taking screenshot: {screenshot_e}")
                
                yield {"type": "error", "error": "Không tìm thấy form tra cứu thông báo. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            # Check session timeout
            if await self._check_session_timeout(page):
                yield {
                    "type": "error",
                    "error": "Phiên giao dịch hết hạn. Vui lòng đăng nhập lại.",
                    "error_code": "SESSION_EXPIRED"
                }
                return
            
            # Chia khoảng thời gian
            date_ranges = self._get_date_ranges(start_date, end_date)
            
            # ✅ TÍNH % THEO CÔNG THỨC MỚI: Tính số ngày của từng khoảng
            total_days = 0
            range_days = []  # Số ngày của từng khoảng
            for date_range in date_ranges:
                days = self._calculate_days_between(date_range[0], date_range[1])
                range_days.append(days)
                total_days += days
            
            # Tính % cho mỗi khoảng dựa trên số ngày
            range_percentages = []
            for days in range_days:
                if total_days > 0:
                    percent = (days / total_days) * 100
                else:
                    percent = 100.0 if len(date_ranges) == 1 else 0.0
                range_percentages.append(percent)
            
            total_count = 0
            results = []
            files_info = []
            total_size = 0
            accumulated_total_so_far = 0  # Tổng số file đã biết từ các khoảng trước
            accumulated_percent_so_far = 0.0  # % tích lũy từ các khoảng trước
            accumulated_downloaded_so_far = 0  # Số file đã download từ các khoảng trước
            
            yield {"type": "info", "message": f"Bắt đầu crawl {len(date_ranges)} khoảng thời gian..."}
            
            def check_cancelled():
                if not job_id:
                    return False
                try:
                    from shared.redis_client import get_redis_client
                    redis_client = get_redis_client()
                    cancelled = redis_client.get(f"job:{job_id}:cancelled")
                    if cancelled:
                        cancelled = cancelled.decode('utf-8') if isinstance(cancelled, bytes) else str(cancelled).strip()
                        return cancelled == '1'
                    return False
                except Exception as e:
                    return False
            
            for range_idx, date_range in enumerate(date_ranges):
                if check_cancelled():
                    yield {
                        "type": "error",
                        "error": "Job đã bị hủy",
                        "error_code": "JOB_CANCELLED"
                    }
                    return
                accumulated_percent_so_far_at_range_start = accumulated_percent_so_far
                # ✅ Giữ nguyên percent hiện tại khi chuyển khoảng (không reset về 0)
                yield {
                    "type": "progress", 
                    "current": range_idx + 1, 
                    "total": len(date_ranges),
                    "message": f"Đang xử lý khoảng {date_range[0]} - {date_range[1]}...",
                    "percent": int(round(accumulated_percent_so_far)),  # ✅ Giữ nguyên percent, không reset về 0
                    "accumulated_percent": int(round(accumulated_percent_so_far)),
                    "accumulated_total": accumulated_total_so_far,  # ✅ Tổng tích lũy từ các khoảng trước
                    "accumulated_downloaded": accumulated_downloaded_so_far
                }
                
                try:
                    # Nhập ngày bắt đầu - dùng id qryFromDate theo HTML form
                    start_input = frame.locator('#qryFromDate')
                    await start_input.fill('')
                    await start_input.fill(date_range[0])
                    
                    # Nhập ngày kết thúc - dùng id qryToDate theo HTML form
                    end_input = frame.locator('#qryToDate')
                    await end_input.click()
                    await end_input.fill('')
                    await end_input.fill(date_range[1])
                    
                    # Click tìm kiếm - button "Tra cứu"
                    search_btn = frame.locator('input[value="Tra cứu"]')
                    await search_btn.click()
                    
                    await asyncio.sleep(2)
                    
                    logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đã click search cho khoảng: {date_range[0]} - {date_range[1]}")
                    
                    # ✅ Đợi một chút để đảm bảo request đã được gửi
                    await asyncio.sleep(1)
                    
                    # ✅ Tìm lại frame mới sau khi click search (iframe có thể reload khi chuyển khoảng thời gian)
                    try:
                        frames = page.frames
                        for f in frames:
                            if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                frame = f  # Cập nhật frame object mới
                                logger.info(f"🔄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đã tìm lại frame mới sau khi click search: {frame.url[:100]}...")
                                break
                    except Exception as refind_frame_e:
                        logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể tìm lại frame mới sau khi click search: {refind_frame_e}")
                    
                    # ✅ Đợi frame load xong trước khi đợi table
                    try:
                        await frame.wait_for_load_state('networkidle', timeout=5000)
                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Frame đã load xong (networkidle)")
                    except Exception as frame_load_e:
                        logger.debug(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể đợi frame networkidle: {frame_load_e}")
                    
                    # ✅ Đợi table load xong để đảm bảo đã chuyển sang khoảng mới
                    try:
                        logger.info(f"⏳ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đang đợi table load sau khi click search...")
                        table_body_check = frame.locator('#allResultTableBody, table.result_table tbody, table#data_content_onday tbody').first
                        await table_body_check.wait_for(timeout=10000, state='visible')
                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Table đã load xong sau khi click search")
                        
                        # ✅ Đợi thêm một chút để đảm bảo dữ liệu đã được render xong
                        await asyncio.sleep(1.5)
                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đã đợi thêm để đảm bảo dữ liệu đã render xong")
                    except Exception as wait_table_e:
                        logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể đợi table load sau khi click search: {wait_table_e}")
                        # Tiếp tục xử lý, sẽ kiểm tra "Không có dữ liệu" ở bước tiếp theo
                    
                    # Xử lý phân trang
                    check_pages = True
                    page_num = 0
                    range_total_records = None  # Tổng số bản ghi trong khoảng này (parse từ currAcc)
                    range_downloaded_so_far = 0  # Tổng số file đã download trong khoảng này (từ các trang trước)
                    max_pages = 100  # ✅ Giới hạn số trang tối đa để tránh vòng lặp vô hạn
                    previous_row_count = 0  # ✅ Lưu số rows của trang trước để verify table đã chuyển trang
                    previous_first_row_id = None  # ✅ Lưu mã giao dịch của row đầu tiên trang trước để verify table đã chuyển trang
                    while check_pages and page_num < max_pages:
                        # ✅ Check cancelled trước khi xử lý trang tiếp theo
                        if check_cancelled():
                            logger.info(f"[THONGBAO] Job {job_id} đã bị cancel, dừng crawl")
                            yield {
                                "type": "error",
                                "error": "Job đã bị hủy",
                                "error_code": "JOB_CANCELLED"
                            }
                            return
                        
                        page_num += 1
                        logger.info(f"📄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đang xử lý trang {page_num}... (check_pages={check_pages})")
                        # Tìm bảng kết quả - theo HTML: #allResultTableBody hoặc table.result_table tbody
                        try:
                            # ✅ Tăng timeout và thêm retry để đảm bảo table được load
                            table_body = frame.locator('#allResultTableBody, table.result_table tbody, table#data_content_onday tbody').first
                            await table_body.wait_for(timeout=10000, state='visible')
                        except Exception as e:
                            # ✅ LOG chi tiết khi không tìm thấy table
                            logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không tìm thấy bảng kết quả cho khoảng {date_range[0]} - {date_range[1]}: {e}")
                            
                            # ✅ Parse currAcc để kiểm tra xem có dữ liệu không (ngay cả khi không có table)
                            try:
                                curr_acc = frame.locator('#currAcc').first
                                if await curr_acc.count() > 0:
                                    curr_acc_text = await curr_acc.text_content()
                                    import re
                                    match = re.search(r'Có\s*<b>(\d+)</b>\s*bản\s*ghi|Có\s*(\d+)\s*bản\s*ghi', curr_acc_text)
                                    if match:
                                        range_total_records = int(match.group(1) or match.group(2))
                                        logger.info(f"📊 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Parse tổng số bản ghi từ currAcc (không có table): {range_total_records}")
                                        # Nếu có số bản ghi nhưng không có table, có thể là lỗi load trang - RETRY
                                        logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] CÓ {range_total_records} bản ghi nhưng không tìm thấy table! Đang retry...")
                                        # Retry: đợi thêm và thử lại
                                        await asyncio.sleep(3)
                                        try:
                                            table_body = frame.locator('#allResultTableBody, table.result_table tbody, table#data_content_onday tbody').first
                                            await table_body.wait_for(timeout=10000, state='visible')
                                            logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Retry thành công, đã tìm thấy table!")
                                            # Tiếp tục xử lý bình thường (không break)
                                        except Exception as retry_e:
                                            logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Retry vẫn thất bại: {retry_e}")
                                            # Nếu retry vẫn thất bại, break và bỏ qua khoảng này
                                            if total_count == 0:
                                                yield {
                                                    "type": "info", 
                                                    "message": f"Không thể tải dữ liệu trong khoảng {date_range[0]} - {date_range[1]} (có {range_total_records} bản ghi nhưng không load được table)",
                                                    "percent": int(round(accumulated_percent_so_far)),
                                                    "accumulated_percent": int(round(accumulated_percent_so_far)),
                                                    "accumulated_total": accumulated_total_so_far,
                                                    "accumulated_downloaded": accumulated_downloaded_so_far
                                                }
                                            accumulated_percent_so_far += range_percentages[range_idx]
                                            break
                            except Exception as parse_e:
                                logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể parse currAcc: {parse_e}")
                            
                            if total_count == 0:
                                # ✅ Giữ nguyên percent hiện tại khi không có dữ liệu
                                yield {
                                    "type": "info", 
                                    "message": f"Không có thông báo trong khoảng {date_range[0]} - {date_range[1]}",
                                    "percent": int(round(accumulated_percent_so_far)),
                                    "accumulated_percent": int(round(accumulated_percent_so_far)),
                                    "accumulated_total": accumulated_total_so_far,
                                    "accumulated_downloaded": accumulated_downloaded_so_far
                                }
                            # ✅ Vẫn cộng % của khoảng này khi không có dữ liệu
                            accumulated_percent_so_far += range_percentages[range_idx]
                            break
                        
                        # ✅ Parse tổng số bản ghi từ phần currAcc (chỉ parse ở trang đầu tiên)
                        if page_num == 1:
                            try:
                                curr_acc = frame.locator('#currAcc').first
                                if await curr_acc.count() > 0:
                                    curr_acc_text = await curr_acc.text_content()
                                    # Parse pattern: "Có <b>34</b> bản ghi" hoặc "Có 34 bản ghi"
                                    import re
                                    match = re.search(r'Có\s*<b>(\d+)</b>\s*bản\s*ghi|Có\s*(\d+)\s*bản\s*ghi', curr_acc_text)
                                    if match:
                                        range_total_records = int(match.group(1) or match.group(2))
                                        logger.info(f"📊 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Parse tổng số bản ghi từ currAcc: {range_total_records}")
                            except Exception as e:
                                logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể parse tổng số bản ghi từ currAcc: {e}")
                        
                        rows = table_body.locator('tr')
                        row_count = await rows.count()
                        
                        # ✅ Tính % cho khoảng này
                        range_percent = range_percentages[range_idx]  # % của khoảng này
                        
                        logger.info(f"📊 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Tìm thấy {row_count} rows, Range %: {range_percent:.2f}%, Accumulated %: {accumulated_percent_so_far:.2f}%")
                        
                        yield {
                            "type": "progress", 
                            "current": total_count, 
                            "message": f"Đang xử lý {row_count} thông báo (trang hiện tại)...",
                            "percent": int(round(min(accumulated_percent_so_far, 100))),  # ✅ Đảm bảo không vượt quá 100%
                            "accumulated_percent": int(round(min(accumulated_percent_so_far, 100))),  # ✅ Đảm bảo không vượt quá 100%
                            "accumulated_total": accumulated_total_so_far,
                            "accumulated_downloaded": accumulated_downloaded_so_far
                        }
                        
                        download_queue = []
                        page_valid_count = 0
                        range_total_items = 0  # Tổng số items trong khoảng này
                        
                        for i in range(row_count):
                            try:
                                row = rows.nth(i)
                                cols = row.locator('td')
                                col_count = await cols.count()
                                
                                if col_count < 6:
                                    continue
                                
                                # Theo HTML contentthongbao.html:
                                # Cột 0: STT
                                # Cột 1: CQ thông báo (Ngân hàng/Cơ quan thuế)
                                # Cột 2: Mã giao dịch
                                # Cột 3: Loại thông báo
                                # Cột 4: Số thông báo
                                # Cột 5: Ngày thông báo
                                # Cột 6-9: Số GNT, Mã hiệu chứng từ, Số chứng từ, Ngày nộp thuế
                                # Cột 10: Chi tiết | Tải về
                                
                                # Cột 2: Mã giao dịch
                                ma_giao_dich = await cols.nth(2).text_content()
                                ma_giao_dich = ma_giao_dich.strip() if ma_giao_dich else ""
                                
                                if not ma_giao_dich or len(ma_giao_dich) < 5:
                                    continue
                                
                                # Chỉ đếm khi item hợp lệ
                                page_valid_count += 1
                                
                                # Cột 1: CQ thông báo
                                cq_thong_bao = await cols.nth(1).text_content()
                                cq_thong_bao = cq_thong_bao.strip() if cq_thong_bao else ""
                                
                                # Cột 3: Loại thông báo
                                loai_thong_bao = await cols.nth(3).text_content()
                                loai_thong_bao = loai_thong_bao.strip() if loai_thong_bao else ""
                                
                                # Cột 4: Số thông báo
                                so_thong_bao = await cols.nth(4).text_content()
                                so_thong_bao = so_thong_bao.strip() if so_thong_bao else ""
                                
                                # Cột 5: Ngày thông báo
                                ngay_thong_bao = await cols.nth(5).text_content()
                                ngay_thong_bao = ngay_thong_bao.strip() if ngay_thong_bao else ""
                                
                                result = {
                                    "id": ma_giao_dich,
                                    "ma_giao_dich": ma_giao_dich,
                                    "cq_thong_bao": cq_thong_bao,
                                    "loai_thong_bao": loai_thong_bao,
                                    "so_thong_bao": so_thong_bao,
                                    "ngay_thong_bao": ngay_thong_bao,
                                    "type": "thongbao"
                                }
                                results.append(result)
                                
                                # ✅ Thêm accumulated fields vào event item để frontend hiển thị đúng
                                yield {
                                    "type": "item", 
                                    "data": result,
                                    "accumulated_total": accumulated_total_so_far,
                                    "accumulated_downloaded": accumulated_downloaded_so_far,
                                    "accumulated_percent": int(round(min(accumulated_percent_so_far, 100)))
                                }
                                
                                # Tìm link "Tải về" trong các cột
                                # Thử tìm trong cột cuối cùng trước, sau đó tìm trong tất cả các cột
                                download_link_found = None
                                download_col_index = None
                                
                                # Cách 1: Tìm trong cột cuối cùng (thường là cột 10 hoặc cuối cùng)
                                last_col_index = col_count - 1
                                if last_col_index >= 0:
                                    last_col = cols.nth(last_col_index)
                                    download_link = last_col.locator('a:has-text("Tải về"), a[title*="Tải"], a[href*="download"]')
                                    if await download_link.count() > 0:
                                        download_link_found = download_link
                                        download_col_index = last_col_index
                                
                                # Cách 2: Nếu không tìm thấy, tìm trong tất cả các cột
                                if not download_link_found:
                                    for col_idx in range(col_count - 1, -1, -1):  # Tìm từ cuối lên đầu
                                        col = cols.nth(col_idx)
                                        download_link = col.locator('a:has-text("Tải về"), a:has-text("Tải"), a[title*="Tải"], a[href*="download"]')
                                        if await download_link.count() > 0:
                                            download_link_found = download_link
                                            download_col_index = col_idx
                                            break
                                
                                if download_link_found:
                                    # Tạo tên file từ thông tin thông báo
                                    ngay_clean = ngay_thong_bao.replace("/", "-").replace(":", "-").replace(" ", "_")
                                    file_name = f"{ma_giao_dich} - {loai_thong_bao[:40]} - {ngay_clean}"
                                    file_name = self._remove_accents(file_name)
                                    file_name = file_name.replace("/", "_").replace(":", "_").replace("\\", "_")
                                    
                                    download_queue.append({
                                        "id": ma_giao_dich,
                                        "loai_thong_bao": loai_thong_bao,
                                        "ngay_thong_bao": ngay_thong_bao,
                                        "file_name": file_name,
                                        "download_link": download_link_found,
                                        "cols": cols,
                                        "col_index": download_col_index
                                    })
                                else:
                                    logger.debug(f"Không tìm thấy link download cho thông báo {ma_giao_dich}, có {col_count} cột")
                            
                            except Exception as e:
                                logger.error(f"Error processing row: {e}")
                                continue
                        
                        # Cộng số items hợp lệ vào range_total_items
                        range_total_items += page_valid_count
                        
                        logger.info(f"📋 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Có {page_valid_count} items hợp lệ, {len(download_queue)} items có link download")
                        
                        # Download từng file và yield progress
                        if download_queue:
                            queue_total = len(download_queue)
                            
                            # ✅ Tính % cho mỗi file download
                            # Nếu có range_total_records, dùng nó để tính % chính xác (cho tất cả các trang)
                            if range_total_records:
                                # Tính % dựa trên tổng số bản ghi trong khoảng (dùng cho tất cả các trang)
                                percent_per_file = range_percent / range_total_records
                                logger.info(f"📊 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Dùng range_total_records={range_total_records} để tính % per file: {percent_per_file:.4f}%")
                            elif queue_total > 0:
                                # Nếu không có range_total_records, tính % dựa trên số file trên trang hiện tại
                                percent_per_file = range_percent / queue_total
                            else:
                                percent_per_file = 0.0
                            
                            # ✅ Cập nhật accumulated_total khi biết số file cần download
                            # Nếu có range_total_records và đang ở trang đầu, dùng nó để cập nhật accumulated_total
                            if range_total_records and page_num == 1:
                                # Chỉ cập nhật accumulated_total ở trang đầu tiên với tổng số bản ghi
                                accumulated_total_so_far += range_total_records
                                logger.info(f"📊 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Cập nhật accumulated_total với range_total_records={range_total_records}, accumulated_total_so_far={accumulated_total_so_far}")
                            elif not range_total_records:
                                # Nếu không có range_total_records, cộng số file trên trang hiện tại
                                accumulated_total_so_far += queue_total
                            
                            # Hiển thị tổng số file sẽ tải (dùng range_total_records nếu có, nếu không dùng queue_total)
                            display_total = range_total_records if range_total_records else queue_total
                            
                            logger.info(f"⬇️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Bắt đầu download {queue_total} files (tổng khoảng: {display_total}), Range %: {range_percent:.2f}%, Percent per file: {percent_per_file:.4f}%, Accumulated total: {accumulated_total_so_far}, Accumulated %: {accumulated_percent_so_far:.2f}%")
                            
                            # ✅ CHỈ publish download_start khi bắt đầu khoảng mới (trang 1), không publish khi chuyển trang
                            if page_num == 1:
                                yield {
                                    "type": "download_start",
                                    "total_to_download": display_total,  # ✅ Hiển thị tổng số file sẽ tải trong khoảng
                                    "current_page_download": queue_total,  # Số file trên trang hiện tại
                                    "date_range": f"{date_range[0]} - {date_range[1]}",
                                    "range_index": range_idx + 1,
                                    "total_ranges": len(date_ranges),
                                    "accumulated_total": accumulated_total_so_far,
                                    "accumulated_downloaded": accumulated_downloaded_so_far,
                                    "range_percent": range_percent,  # % của khoảng này
                                    "accumulated_percent": int(round(min(accumulated_percent_so_far, 100))),  # ✅ Đảm bảo không vượt quá 100%
                                    "message": f"Bắt đầu tải {display_total} thông báo trong khoảng {date_range[0]} - {date_range[1]}..."
                                }
                            
                            downloaded = 0
                            
                            for item_idx, item in enumerate(download_queue, 1):
                                try:
                                    logger.info(f"📥 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang download file {item_idx}/{queue_total}: {item.get('id', 'N/A')}...")
                                    success = await self._download_single_thongbao(session, item, temp_dir)
                                    if success:
                                        downloaded += 1
                                        accumulated_downloaded_so_far += 1
                                        range_downloaded_so_far += 1  # ✅ Cộng dồn số file đã download trong khoảng này
                                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã download thành công file {item_idx}/{queue_total}: {item.get('id', 'N/A')}")
                                    else:
                                        logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Download thất bại file {item_idx}/{queue_total}: {item.get('id', 'N/A')}")
                                except Exception as download_e:
                                    logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi download file {item_idx}/{queue_total} ({item.get('id', 'N/A')}): {download_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    # Tiếp tục download file tiếp theo
                                    continue
                                
                                # ✅ Tính % tích lũy: % từ các khoảng trước + % của các file đã download trong khoảng này
                                # QUAN TRỌNG: Dùng accumulated_percent_so_far_at_range_start (không phải accumulated_percent_so_far)
                                # để tránh cộng dồn sai khi đã cập nhật accumulated_percent_so_far trong vòng lặp
                                if range_total_records:
                                    # Tính % dựa trên tổng số bản ghi trong khoảng
                                    # % của khoảng này = (số file đã download / tổng số file trong khoảng) * % của khoảng
                                    range_accumulated_percent = (range_downloaded_so_far / range_total_records) * range_percent
                                    # Cộng với % tích lũy từ các khoảng trước (tại thời điểm bắt đầu khoảng này)
                                    current_accumulated_percent = accumulated_percent_so_far_at_range_start + range_accumulated_percent
                                else:
                                    # Tính % dựa trên số file trên trang hiện tại
                                    current_accumulated_percent = accumulated_percent_so_far_at_range_start + (downloaded * percent_per_file)
                                
                                # ✅ Đảm bảo không vượt quá 100%
                                current_accumulated_percent = min(current_accumulated_percent, 100.0)
                                
                                # ✅ CẬP NHẬT accumulated_percent_so_far liên tục trong quá trình download
                                accumulated_percent_so_far = current_accumulated_percent
                                
                                # Hiển thị tổng số file đã download trong khoảng (dùng range_total_records nếu có)
                                display_total = range_total_records if range_total_records else queue_total
                                display_downloaded = range_downloaded_so_far if range_total_records else downloaded
                                
                                if item_idx % 5 == 0 or item_idx == queue_total:  # Log mỗi 5 file hoặc file cuối
                                    logger.info(f"⬇️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã download {display_downloaded}/{display_total} files (trang: {downloaded}/{queue_total}), Current accumulated %: {accumulated_percent_so_far:.2f}%")
                                
                                # ✅ Yield progress event với exception handling
                                try:
                                    yield {
                                        "type": "download_progress",
                                        "downloaded": display_downloaded,  # ✅ Hiển thị tổng số file đã download trong khoảng
                                        "total": display_total,  # ✅ Hiển thị tổng số file sẽ tải trong khoảng
                                        "current_page_downloaded": downloaded,  # Số file đã download trên trang hiện tại
                                        "current_page_total": queue_total,  # Số file trên trang hiện tại
                                        "percent": round(display_downloaded / display_total * 100, 1) if display_total > 0 else 0,
                                        "current_item": item.get("id", ""),
                                        "accumulated_total": accumulated_total_so_far,
                                        "accumulated_downloaded": accumulated_downloaded_so_far,
                                        "accumulated_percent": int(round(accumulated_percent_so_far)),  # ✅ Dùng accumulated_percent_so_far đã được cập nhật
                                        "message": f"Đã tải {display_downloaded}/{display_total} ({round(display_downloaded / display_total * 100, 1) if display_total > 0 else 0}%)"
                                    }
                                except Exception as yield_e:
                                    logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi yield progress event: {yield_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    # Tiếp tục download file tiếp theo, không dừng vì lỗi yield
                                    pass
                            
                            # ✅ Cập nhật previous_row_count sau khi xử lý xong trang này
                            previous_row_count = row_count
                            
                            # ✅ Cập nhật accumulated_percent_so_far sau khi download xong khoảng này (chỉ ở trang cuối cùng)
                            # Chỉ cập nhật khi không còn trang tiếp theo và đã download hết tất cả file trong khoảng
                            if not check_pages:  # Nếu không còn trang tiếp theo
                                # Đảm bảo accumulated_percent_so_far đạt đúng % của khoảng này
                                # Nếu có range_total_records, đã tính % dựa trên số file download, không cần cộng thêm
                                # Nếu không có range_total_records, cộng % của khoảng này
                                if not range_total_records:
                                    accumulated_percent_so_far += range_percent
                                # ✅ Đảm bảo không vượt quá 100%
                                accumulated_percent_so_far = min(accumulated_percent_so_far, 100.0)
                            
                            # Hiển thị tổng số file đã download trong khoảng (dùng range_total_records nếu có)
                            display_total = range_total_records if range_total_records else queue_total
                            display_downloaded = range_downloaded_so_far if range_total_records else downloaded
                            
                            logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Hoàn thành download {display_downloaded}/{display_total} files (trang: {downloaded}/{queue_total}), Accumulated %: {accumulated_percent_so_far:.2f}%")
                            
                            yield {
                                "type": "download_complete",
                                "downloaded": display_downloaded,  # ✅ Hiển thị tổng số file đã download trong khoảng
                                "total": display_total,  # ✅ Hiển thị tổng số file sẽ tải trong khoảng
                                "current_page_downloaded": downloaded,  # Số file đã download trên trang hiện tại
                                "current_page_total": queue_total,  # Số file trên trang hiện tại
                                "accumulated_total": accumulated_total_so_far,
                                "accumulated_downloaded": accumulated_downloaded_so_far,
                                "accumulated_percent": int(round(accumulated_percent_so_far)),  # ✅ Đã đảm bảo không vượt quá 100%
                                # ✅ KHÔNG gửi message để frontend không hiển thị "Hoàn thành tải..."
                                # "message": f"Hoàn thành tải {display_downloaded}/{display_total} thông báo"
                            }
                        
                        # Chỉ cộng số items hợp lệ vào total_count
                        total_count += page_valid_count
                        
                        # Check pagination - next page
                        try:
                            logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang kiểm tra nút next...")
                            next_btn = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                            next_btn_count = await next_btn.count()
                            logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Số lượng nút next: {next_btn_count}")
                            
                            if next_btn_count > 0:
                                logger.info(f"➡️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Có trang tiếp theo, đang chuyển trang...")
                                
                                # ✅ Click với timeout và logging
                                try:
                                    logger.info(f"🖱️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang click nút next...")
                                    await asyncio.wait_for(next_btn.click(), timeout=10.0)
                                    logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã click nút next thành công")
                                except asyncio.TimeoutError:
                                    logger.error(f"⏱️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Timeout khi click nút next (10s)")
                                    check_pages = False
                                    continue
                                except Exception as click_e:
                                    logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi click nút next: {click_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    check_pages = False
                                    continue
                                
                                # ✅ Đợi trang load xong trước khi tiếp tục
                                logger.info(f"⏳ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đợi 2 giây sau khi click...")
                                await asyncio.sleep(2)
                                logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã đợi xong 2 giây, bắt đầu đợi table load...")
                                
                                # ✅ Tìm lại frame mới sau khi click next (iframe có thể reload khi chuyển trang)
                                try:
                                    frames = page.frames
                                    for f in frames:
                                        if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                            frame = f  # Cập nhật frame object mới
                                            logger.info(f"🔄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đã tìm lại frame mới sau khi click next: {frame.url[:100]}...")
                                            break
                                except Exception as refind_frame_e:
                                    logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể tìm lại frame mới sau khi click next: {refind_frame_e}")
                                
                                # ✅ Kiểm tra lại xem có trang tiếp theo không (sau khi click)
                                try:
                                    logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang đợi table load cho trang {page_num + 1}...")
                                    # ✅ Kiểm tra frame còn tồn tại không
                                    try:
                                        frame_url = frame.url
                                        logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Frame URL: {frame_url[:100]}...")
                                    except Exception as frame_check_e:
                                        logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Frame không còn tồn tại sau khi click: {frame_check_e}")
                                        import traceback
                                        logger.error(f"Traceback: {traceback.format_exc()}")
                                        check_pages = False
                                        continue
                                    
                                    # Đợi table load để đảm bảo trang đã chuyển (tăng timeout lên 15 giây)
                                    logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang tìm table locator...")
                                    try:
                                        table_body_check = frame.locator('#allResultTableBody, table.result_table tbody, table#data_content_onday tbody').first
                                        logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã tìm thấy table locator, đang đợi table visible...")
                                    except Exception as locator_e:
                                        logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Lỗi khi tìm table locator: {locator_e}")
                                        import traceback
                                        logger.error(f"Traceback: {traceback.format_exc()}")
                                        check_pages = False
                                        continue
                                    
                                    try:
                                        await asyncio.wait_for(
                                            table_body_check.wait_for(timeout=15000, state='visible'),
                                            timeout=20.0  # Tổng timeout 20 giây
                                        )
                                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Table đã visible, đang verify table đã chuyển trang...")
                                        
                                        # ✅ Đợi frame load xong trước khi verify table
                                        try:
                                            await frame.wait_for_load_state('networkidle', timeout=5000)
                                            logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Frame đã load xong (networkidle) sau khi click next")
                                        except Exception as frame_load_e:
                                            logger.debug(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể đợi frame networkidle: {frame_load_e}")
                                        
                                        # ✅ Đợi thêm một chút để đảm bảo table đã load xong và render đúng
                                        await asyncio.sleep(1.5)
                                        
                                        # ✅ Verify table đã thực sự chuyển trang bằng cách so sánh mã giao dịch của row đầu tiên
                                        try:
                                            rows_check = table_body_check.locator('tr')
                                            row_count_check = await rows_check.count()
                                            logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Verify: Table có {row_count_check} rows sau khi click next (trang trước: {previous_row_count} rows)")
                                            
                                            # ✅ Lấy mã giao dịch của row đầu tiên để verify
                                            first_row_id = None
                                            if row_count_check > 0:
                                                try:
                                                    first_row = rows_check.first
                                                    first_cols = first_row.locator('td')
                                                    col_count = await first_cols.count()
                                                    if col_count > 2:
                                                        # Mã giao dịch ở cột 2 (theo HTML structure)
                                                        first_row_id = await first_cols.nth(2).text_content()
                                                        first_row_id = first_row_id.strip() if first_row_id else None
                                                        logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Mã giao dịch row đầu tiên sau click next: {first_row_id}")
                                                except Exception as get_id_e:
                                                    logger.debug(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể lấy mã giao dịch row đầu tiên: {get_id_e}")
                                            
                                            # Nếu table vẫn có cùng số rows như trang trước, kiểm tra mã giao dịch
                                            if row_count_check == previous_row_count and previous_row_count > 0:
                                                logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Table vẫn có {row_count_check} rows giống trang trước, đợi thêm và kiểm tra mã giao dịch...")
                                                await asyncio.sleep(2)
                                                
                                                # Lấy lại mã giao dịch sau khi đợi
                                                first_row_id_after_wait = None
                                                if row_count_check > 0:
                                                    try:
                                                        first_row_after = rows_check.first
                                                        first_cols_after = first_row_after.locator('td')
                                                        col_count_after = await first_cols_after.count()
                                                        if col_count_after > 2:
                                                            first_row_id_after_wait = await first_cols_after.nth(2).text_content()
                                                            first_row_id_after_wait = first_row_id_after_wait.strip() if first_row_id_after_wait else None
                                                    except Exception as get_id_e2:
                                                        logger.debug(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể lấy mã giao dịch sau khi đợi: {get_id_e2}")
                                                
                                                row_count_check = await rows_check.count()
                                                logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Sau khi đợi thêm: Table có {row_count_check} rows, mã giao dịch: {first_row_id_after_wait}")
                                                
                                                # ✅ Nếu số rows vẫn giống, kiểm tra mã giao dịch
                                                if row_count_check == previous_row_count:
                                                    # So sánh với mã giao dịch của trang trước
                                                    if previous_first_row_id and first_row_id_after_wait:
                                                        if previous_first_row_id == first_row_id_after_wait:
                                                            logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Table vẫn chưa chuyển trang sau khi click next! (Mã giao dịch giống nhau: {previous_first_row_id} == {first_row_id_after_wait})")
                                                            check_pages = False
                                                            continue
                                                        else:
                                                            logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Table đã chuyển trang (mã giao dịch khác: {previous_first_row_id} → {first_row_id_after_wait})")
                                                    elif not previous_first_row_id or not first_row_id_after_wait:
                                                        # Nếu không lấy được mã giao dịch, chỉ dựa vào số rows
                                                        logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể verify bằng mã giao dịch, nhưng số rows vẫn giống, tiếp tục thử...")
                                                        # Tiếp tục xử lý, có thể table đã chuyển nhưng không verify được
                                            else:
                                                # Số rows khác nhau → table đã chuyển trang
                                                logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Table đã chuyển trang (số rows khác: {previous_row_count} → {row_count_check})")
                                                
                                        except Exception as verify_e:
                                            logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể verify table: {verify_e}")
                                        
                                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num + 1} đã load xong, tiếp tục xử lý...")
                                        
                                        # ✅ Cập nhật previous_row_count và previous_first_row_id cho lần verify tiếp theo
                                        previous_row_count = row_count_check
                                        previous_first_row_id = first_row_id_after_wait if first_row_id_after_wait else first_row_id
                                        
                                        # ✅ Tiếp tục vòng lặp (check_pages vẫn True)
                                    except Exception as wait_table_e:
                                        logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Lỗi khi đợi table visible: {wait_table_e}")
                                        import traceback
                                        logger.error(f"Traceback: {traceback.format_exc()}")
                                        # Re-raise để được xử lý bởi except block bên ngoài
                                        raise
                                except asyncio.TimeoutError:
                                    logger.error(f"⏱️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Timeout khi đợi table load cho trang {page_num + 1} (20s)")
                                    # ✅ Retry: Đợi thêm và thử lại
                                    logger.info(f"🔄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Retry đợi table load...")
                                    await asyncio.sleep(3)
                                    try:
                                        table_body_check_retry = frame.locator('#allResultTableBody, table.result_table tbody, table#data_content_onday tbody').first
                                        await asyncio.wait_for(
                                            table_body_check_retry.wait_for(timeout=15000, state='visible'),
                                            timeout=20.0
                                        )
                                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Retry thành công, trang {page_num + 1} đã load xong")
                                        
                                        # ✅ Tìm lại frame mới sau khi retry (iframe có thể reload)
                                        try:
                                            frames = page.frames
                                            for f in frames:
                                                if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                                    frame = f  # Cập nhật frame object mới
                                                    logger.info(f"🔄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đã tìm lại frame mới sau retry: {frame.url[:100]}...")
                                                    break
                                        except Exception as refind_frame_e:
                                            logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể tìm lại frame mới sau retry: {refind_frame_e}")
                                    except Exception as retry_e:
                                        logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Retry vẫn thất bại: {retry_e}")
                                        import traceback
                                        logger.error(f"Traceback: {traceback.format_exc()}")
                                        # Kiểm tra lại nút next sau khi đợi
                                        await asyncio.sleep(2)
                                        try:
                                            next_btn_check = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                                            next_btn_check_count = await next_btn_check.count()
                                            logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Sau retry, số lượng nút next: {next_btn_check_count}")
                                            if next_btn_check_count == 0:
                                                logger.info(f"🏁 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Sau khi click, không còn nút next, kết thúc phân trang")
                                                check_pages = False
                                            else:
                                                logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Vẫn còn nút next nhưng table không load, kết thúc phân trang để tránh hang")
                                                check_pages = False
                                        except Exception as check_e:
                                            logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Lỗi khi kiểm tra nút next sau retry: {check_e}")
                                            check_pages = False
                                except Exception as wait_e:
                                    logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num + 1} chưa load xong sau khi click next: {wait_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    # ✅ Retry: Đợi thêm và thử lại
                                    logger.info(f"🔄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Retry đợi table load...")
                                    await asyncio.sleep(3)
                                    try:
                                        table_body_check_retry = frame.locator('#allResultTableBody, table.result_table tbody, table#data_content_onday tbody').first
                                        await asyncio.wait_for(
                                            table_body_check_retry.wait_for(timeout=15000, state='visible'),
                                            timeout=20.0
                                        )
                                        logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Retry thành công, trang {page_num + 1} đã load xong")
                                        
                                        # ✅ Tìm lại frame mới sau khi retry (iframe có thể reload)
                                        try:
                                            frames = page.frames
                                            for f in frames:
                                                if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                                    frame = f  # Cập nhật frame object mới
                                                    logger.info(f"🔄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Đã tìm lại frame mới sau retry: {frame.url[:100]}...")
                                                    break
                                        except Exception as refind_frame_e:
                                            logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể tìm lại frame mới sau retry: {refind_frame_e}")
                                    except Exception as retry_e:
                                        logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Retry vẫn thất bại: {retry_e}")
                                        import traceback
                                        logger.error(f"Traceback: {traceback.format_exc()}")
                                        
                                        # ✅ Kiểm tra "Không có dữ liệu" khi table không load
                                        try:
                                            no_data_text = frame.locator('div:has-text("Không có dữ liệu"), strong:has-text("Không có dữ liệu"), div.align-center:has-text("Không có dữ liệu")').first
                                            if await no_data_text.count() > 0:
                                                no_data_content = await no_data_text.text_content()
                                                if "Không có dữ liệu" in (no_data_content or ""):
                                                    logger.info(f"📊 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Phát hiện 'Không có dữ liệu' sau retry, dừng pagination")
                                                    check_pages = False
                                                    continue
                                        except Exception as no_data_check_e3:
                                            logger.debug(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể kiểm tra 'Không có dữ liệu' sau retry: {no_data_check_e3}")
                                        
                                        # Kiểm tra lại nút next sau khi đợi
                                        await asyncio.sleep(2)
                                        try:
                                            next_btn_check = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                                            next_btn_check_count = await next_btn_check.count()
                                            logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Sau retry, số lượng nút next: {next_btn_check_count}")
                                            if next_btn_check_count == 0:
                                                logger.info(f"🏁 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Sau khi click, không còn nút next, kết thúc phân trang")
                                                check_pages = False
                                                continue
                                            else:
                                                logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Vẫn còn nút next nhưng table không load, kết thúc phân trang để tránh hang")
                                                check_pages = False
                                                continue
                                        except Exception as check_e:
                                            logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Lỗi khi kiểm tra nút next sau retry: {check_e}")
                                            check_pages = False
                                            continue
                            else:
                                logger.info(f"🏁 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Không còn trang tiếp theo")
                                check_pages = False  # ✅ CHỈ set False khi không còn nút next
                            
                            # ✅ Log trạng thái sau khi xử lý pagination
                            logger.info(f"📊 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Sau khi xử lý pagination: check_pages={check_pages}, page_num={page_num}")
                        except Exception as pagination_e:
                            logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi xử lý phân trang: {pagination_e}")
                            import traceback
                            logger.error(f"Traceback: {traceback.format_exc()}")
                            # ✅ Sau khi có lỗi, kiểm tra lại xem có nút next không
                            try:
                                await asyncio.sleep(2)
                                next_btn_retry = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                                next_btn_retry_count = await next_btn_retry.count()
                                logger.info(f"🔍 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Sau lỗi, số lượng nút next: {next_btn_retry_count}")
                                if next_btn_retry_count > 0:
                                    logger.warning(f"⚠️ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Vẫn còn nút next sau lỗi, nhưng dừng lại để tránh hang")
                                    check_pages = False
                                else:
                                    logger.info(f"🏁 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không còn nút next sau lỗi, kết thúc phân trang")
                                    check_pages = False
                            except Exception as retry_e:
                                # Nếu không kiểm tra được, dừng lại để tránh vòng lặp vô hạn
                                logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Không thể kiểm tra nút next sau lỗi: {retry_e}")
                                import traceback
                                logger.error(f"Traceback: {traceback.format_exc()}")
                                check_pages = False
                        
                        # ✅ Log trước khi tiếp tục vòng lặp
                        if check_pages:
                            logger.info(f"🔄 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Tiếp tục vòng lặp pagination, sẽ xử lý trang tiếp theo...")
                        else:
                            logger.info(f"🛑 [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Dừng vòng lặp pagination, đã xử lý xong {page_num} trang")
                
                except Exception as e:
                    logger.error(f"❌ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Lỗi xử lý khoảng {date_range[0]} - {date_range[1]}: {e}")
                    # ✅ Giữ nguyên percent hiện tại khi có lỗi
                    yield {
                        "type": "warning", 
                        "message": f"Lỗi xử lý khoảng {date_range}: {str(e)}",
                        "percent": int(round(min(accumulated_percent_so_far, 100))),  # ✅ Đảm bảo không vượt quá 100%
                        "accumulated_percent": int(round(min(accumulated_percent_so_far, 100))),  # ✅ Đảm bảo không vượt quá 100%
                        "accumulated_total": accumulated_total_so_far,
                        "accumulated_downloaded": accumulated_downloaded_so_far
                    }
                    # ✅ Vẫn cộng % của khoảng này (đã xử lý một phần hoặc không có dữ liệu)
                    accumulated_percent_so_far += range_percentages[range_idx]
                    accumulated_percent_so_far = min(accumulated_percent_so_far, 100.0)  # ✅ Đảm bảo không vượt quá 100%
                    continue
                
                # ✅ Log khi hoàn thành xử lý khoảng này
                logger.info(f"✅ [THONGBAO] [{range_idx + 1}/{len(date_ranges)}] Hoàn thành xử lý khoảng {date_range[0]} - {date_range[1]}: Tổng {total_count} items, Accumulated %: {accumulated_percent_so_far:.2f}%")
            
            # ✅ Log tổng kết sau khi xử lý tất cả các khoảng
            logger.info(f"🏁 [THONGBAO] Hoàn thành crawl tất cả {len(date_ranges)} khoảng: Tổng {total_count} items, Accumulated %: {accumulated_percent_so_far:.2f}%")
            
            # Parse downloaded files và rename
            parsed_results = []
            files_in_temp_dir = os.listdir(temp_dir) if os.path.exists(temp_dir) else []
            logger.info(f"crawl_thongbao: Found {len(files_in_temp_dir)} files in temp_dir")
            
            if files_in_temp_dir:
                for file_name in files_in_temp_dir:
                    file_path = os.path.join(temp_dir, file_name)
                    if not os.path.isfile(file_path):
                        continue
                    
                    try:
                        # Parse XML để lấy thông tin
                        with open(file_path, 'r', encoding='utf-8') as f:
                            soup = BeautifulSoup(f, 'html.parser')
                        
                        mgd = soup.find('magiaodichdtu')
                        mgd = mgd.text if mgd else ""
                        
                        ttb = soup.find('tentbao')
                        ttb = ttb.text if ttb else ""
                        
                        ma_tbao = soup.find('matbao')
                        ma_tbao = ma_tbao.text if ma_tbao else ""
                        
                        
                        if "Tiếp nhận" in ttb:
                            ttb = "Tiếp nhận"
                        if "Xác nhận" in ttb:
                            ttb = "Xác nhận"
                        if ma_tbao == "844":
                            ttb = "Không chấp nhận"
                        elif ma_tbao == "451":
                            ttb = "Chấp nhận"
                        
                        ttb_2 = "X"
                        try:
                            if ttb == "Tiếp nhận":
                                ngay_tbao_elem = soup.find('ngaytbao')
                                ttb_2 = ngay_tbao_elem.text if ngay_tbao_elem else "X"
                            else:
                                ngay_chap_nhan_elem = soup.find('ngaychapnhan')
                                if ngay_chap_nhan_elem:
                                    ttb_2 = ngay_chap_nhan_elem.text.split("T")[0] if "T" in ngay_chap_nhan_elem.text else ngay_chap_nhan_elem.text
                        except:
                            ttb_2 = "X"
                        
                        ttb_2_clean = ttb_2.replace("/", "-")
                        new_file_name = f"{mgd} - {ttb} - {ttb_2_clean}.xml"
                        new_file_name = self._remove_accents(new_file_name)
                        
                        # Rename file
                        new_file_path = os.path.join(temp_dir, new_file_name)
                        if os.path.exists(file_path):
                            try:
                                os.rename(file_path, new_file_path)
                                file_name = new_file_name
                                file_path = new_file_path
                            except Exception as rename_err:
                                logger.warning(f"Error renaming {file_name}: {rename_err}")
                        
                        parsed_results.append({
                            "ma_giao_dich": mgd,
                            "ten_thong_bao": ttb,
                            "ma_thong_bao": ma_tbao,
                            "ngay_thong_bao": ttb_2.replace("-", "/") if ttb_2 != "X" else ""  # Trả về format gốc
                        })
                        
                        file_size = os.path.getsize(file_path)
                        total_size += file_size
                        files_info.append({"name": file_name, "size": file_size})
                    except Exception as e:
                        logger.warning(f"Error parsing/renaming file {file_name}: {e}")
                        # Nếu parse lỗi, vẫn thêm vào files_info với tên cũ
                        try:
                            file_size = os.path.getsize(file_path)
                            total_size += file_size
                            files_info.append({"name": file_name, "size": file_size})
                        except:
                            pass
                        continue
                
                # Tạo download_id (UUID) để worker có thể download sau (giống tờ khai)
                download_id = str(uuid.uuid4())
                zip_filename = f"thongbao_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                zip_file_path = os.path.join(self.ZIP_STORAGE_DIR, f"{download_id}.zip")
                
                # Lưu zip vào disk thay vì chỉ tạo base64 (giống tờ khai)
                final_files = os.listdir(temp_dir)
                logger.info(f"crawl_thongbao: Found {len(final_files)} files in temp_dir")
                logger.info(f"crawl_thongbao: Creating ZIP from {len(final_files)} files")
                
                if final_files:
                    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for file_name in final_files:
                            file_path = os.path.join(temp_dir, file_name)
                            if os.path.isfile(file_path):
                                zf.write(file_path, file_name)
                                logger.debug(f"Added to ZIP: {file_name}")
                    
                    # Đọc file để tạo base64 (vẫn cần cho Redis)
                    with open(zip_file_path, 'rb') as f:
                        zip_base64 = base64.b64encode(f.read()).decode('utf-8')
                    
                    logger.info(f"✅ Đã tạo file ZIP: {zip_filename} (download_id: {download_id})")
                    
                    # Lưu download_id vào Redis (giống tờ khai)
                    try:
                        from shared.redis_client import get_redis_client
                        redis_client = get_redis_client()
                        redis_key = f"session:{session_id}:download_id"
                        redis_client.setex(redis_key, 3600, download_id.encode('utf-8'))
                    except Exception as redis_err:
                        logger.warning(f"⚠️ Không thể lưu download_id vào Redis: {redis_err}")
                else:
                    zip_base64 = None
                    download_id = None
                    logger.warning("crawl_thongbao: No files to add to ZIP")
            else:
                zip_base64 = None
                download_id = None
                zip_filename = f"thongbao_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                logger.warning("crawl_thongbao: No files in temp_dir")
            
            actual_files_count = len(files_info)
            actual_results_count = len(parsed_results)
            
            # ✅ Log trước khi yield complete
            logger.info(f"📦 [THONGBAO] Chuẩn bị yield complete: total_count={total_count}, actual_results_count={actual_results_count}, actual_files_count={actual_files_count}, zip_base64_length={len(zip_base64) if zip_base64 else 0}, download_id={download_id}")
            
            # Trả về total là số rows đã xử lý (số items tìm thấy) để hiển thị đúng
            # zip_base64 sẽ là None nếu không có files, button sẽ disabled
            yield {
                "type": "complete",
                "total": total_count,  # Số items đã tìm thấy (total_rows_processed)
                "results_count": actual_results_count,  # Số items đã parse
                "total_rows_processed": total_count,  # Số rows đã xử lý (để debug)
                "results": parsed_results,
                "files": files_info,
                "files_count": actual_files_count,  # Số file thực tế trong ZIP
                "total_size": total_size,
                "zip_base64": zip_base64,  # None nếu không có files
                "zip_filename": zip_filename,  # ✅ Dùng zip_filename đã tạo ở trên
                "download_id": download_id  # ✅ Thêm download_id (giống tờ khai)
            }
            
        except Exception as e:
            logger.error(f"❌ [THONGBAO] Error in crawl_thongbao: {e}")
            import traceback
            logger.error(f"❌ [THONGBAO] Traceback: {traceback.format_exc()}")
            error_msg = str(e)
            
            # ✅ Đảm bảo yield complete event ngay cả khi có lỗi (với files đã download)
            try:
                # Parse downloaded files nếu có
                parsed_results = []
                files_in_temp_dir = os.listdir(temp_dir) if os.path.exists(temp_dir) else []
                files_info = []
                total_size = 0
                zip_base64 = None
                
                if files_in_temp_dir:
                    for file_name in files_in_temp_dir:
                        file_path = os.path.join(temp_dir, file_name)
                        if os.path.isfile(file_path):
                            try:
                                file_size = os.path.getsize(file_path)
                                total_size += file_size
                                files_info.append({"name": file_name, "size": file_size})
                            except:
                                pass
                    
                    if files_info:
                        # Tạo download_id và lưu ZIP vào disk (giống tờ khai)
                        try:
                            download_id = str(uuid.uuid4())
                            zip_filename = f"thongbao_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                            zip_file_path = os.path.join(self.ZIP_STORAGE_DIR, f"{download_id}.zip")
                            
                            with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                                for file_name in files_in_temp_dir:
                                    file_path = os.path.join(temp_dir, file_name)
                                    if os.path.isfile(file_path):
                                        zf.write(file_path, file_name)
                            
                            # Đọc file để tạo base64
                            with open(zip_file_path, 'rb') as f:
                                zip_base64 = base64.b64encode(f.read()).decode('utf-8')
                            
                        except Exception as zip_e:
                            logger.error(f"❌ [THONGBAO] Lỗi tạo ZIP: {zip_e}")
                            download_id = None
                            zip_filename = f"thongbao_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                    else:
                        download_id = None
                        zip_filename = f"thongbao_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                
                # Yield complete event với files đã download
                yield {
                    "type": "complete",
                    "total": total_count if 'total_count' in locals() else 0,
                    "results_count": len(parsed_results),
                    "results": parsed_results,
                    "files": files_info,
                    "files_count": len(files_info),
                    "total_size": total_size,
                    "zip_base64": zip_base64,
                    "zip_filename": zip_filename,
                    "download_id": download_id,  # ✅ Thêm download_id
                    "error": error_msg
                }
            except Exception as final_e:
                logger.error(f"❌ [THONGBAO] Lỗi khi yield complete event sau lỗi: {final_e}")
            # Kiểm tra session timeout
            if "timeout" in error_msg.lower() or "phiên giao dịch" in error_msg.lower():
                yield {"type": "error", "error": "Phiên giao dịch hết hạn. Vui lòng đăng nhập lại.", "error_code": "SESSION_EXPIRED"}
            else:
                yield {"type": "error", "error": f"Lỗi khi tra cứu thông báo: {error_msg}", "error_code": "CRAWL_ERROR"}
        
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
    
    
    async def _download_single_giaynoptien(self, session: SessionData, item: Dict, temp_dir: str, max_retries: int = 2) -> bool:
        """
        Download một file giấy nộp tiền
        item: {
            "id": id_gnt,
            "row": row,
            "col_index": col_idx,
            "link_locator": links.first
        }
        """
        id_gnt = item.get("id")
        if not id_gnt:
            return False
        
        page = session.page
        frame = None
        
        # Tìm frame chứa form giấy nộp tiền
        try:
            frames = page.frames
            for f in frames:
                if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                    frame = f
                    break
        except:
            pass
        
        if not frame:
            logger.error(f"Không tìm thấy frame để download giấy nộp tiền {id_gnt}")
            return False
        
        # Lấy tham số từ form reportForm
        form_params = {}
        try:
            form = frame.locator('form[name="reportForm"], form#reportForm').first
            if await form.count() > 0:
                inputs = form.locator('input[type="hidden"]')
                input_count = await inputs.count()
                for i in range(input_count):
                    try:
                        input_elem = inputs.nth(i)
                        name = await input_elem.get_attribute('name')
                        value = await input_elem.get_attribute('value')
                        if name and value:
                            form_params[name] = value
                    except:
                        continue
        except Exception as e:
            logger.warning(f"Lỗi khi lấy tham số từ form cho {id_gnt}: {e}")
        
        # Lấy các tham số cần thiết
        dse_session_id = form_params.get('dse_sessionId', session.dse_session_id)
        dse_application_id = form_params.get('dse_applicationId', '-1')
        dse_operation_name = form_params.get('dse_operationName', 'corpQueryTaxProc')
        dse_page_id = form_params.get('dse_pageId', '35')
        dse_processor_state = form_params.get('dse_processorState', 'viewQueryPage')
        dse_processor_id = form_params.get('dse_processorId', '')
        
        # Xây dựng URL download (giống khi click downloadGNT)
        download_url = (
            f"{BASE_URL}/etaxnnt/Request?"
            f"dse_sessionId={dse_session_id}&"
            f"dse_applicationId={dse_application_id}&"
            f"dse_operationName={dse_operation_name}&"
            f"dse_pageId={dse_page_id}&"
            f"dse_processorState={dse_processor_state}&"
            f"dse_processorId={dse_processor_id}&"
            f"dse_nextEventName=download&"
            f"ctuId={id_gnt}"
        )
        
        # Thử download với retry
        for retry in range(max_retries + 1):
            try:
                # Thử click link trước (nhanh hơn)
                link_locator = item.get("link_locator")
                if link_locator:
                    try:
                        async with page.expect_download(timeout=30000) as download_info:
                            await link_locator.click()
                        
                        download = await download_info.value
                        file_path = os.path.join(temp_dir, f"gnt_{id_gnt}.xml")
                        await download.save_as(file_path)
                        
                        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                            logger.info(f"Downloaded giay nop tien {id_gnt} via click")
                            return True
                    except:
                        pass
                
                # Fallback: dùng URL trực tiếp
                new_page = None
                try:
                    new_page = await session.context.new_page()
                    new_page.set_default_timeout(30000)
                    
                    async with new_page.expect_download(timeout=30000) as download_info:
                        await new_page.goto(download_url, wait_until="domcontentloaded")
                    
                    download = await download_info.value
                    file_path = os.path.join(temp_dir, f"gnt_{id_gnt}.xml")
                    await download.save_as(file_path)
                    
                    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                        logger.info(f"Downloaded giay nop tien {id_gnt} via URL")
                        return True
                except asyncio.TimeoutError:
                    logger.warning(f"Timeout khi download giay nop tien {id_gnt}")
                except Exception as e:
                    logger.warning(f"Lỗi khi download giay nop tien {id_gnt} qua URL: {e}")
                finally:
                    if new_page:
                        try:
                            await new_page.close()
                        except:
                            pass
                
                # Nếu vẫn không được, thử dùng httpx (nếu có session_id từ crawl_giay_nop_tien)
                # Note: session_id không có trong hàm này, bỏ qua httpx fallback
                
            except Exception as e:
                logger.warning(f"Error downloading giaynoptien {id_gnt} (attempt {retry + 1}/{max_retries + 1}): {e}")
                if retry < max_retries:
                    await asyncio.sleep(1)
        
        return False
    
    async def crawl_giay_nop_tien(
        self,
        session_id: str,
        start_date: str,
        end_date: str,
        job_id: Optional[str] = None
    ) -> AsyncGenerator[Dict[str, Any], None]:
        session = self.session_manager.get_session(session_id)
        if not session:
            yield {"type": "error", "error": "Session không tồn tại hoặc đã hết hạn", "error_code": "SESSION_NOT_FOUND"}
            return
        
        if not session.is_logged_in:
            yield {"type": "error", "error": "Chưa đăng nhập. Vui lòng đăng nhập lại.", "error_code": "NOT_LOGGED_IN"}
            return
        
        page = session.page
        
        # ✅ FIX: Tạo temp directory trong source code thay vì system temp (giống tờ khai)
        # Lấy đường dẫn project (tool-go-soft)
        current_dir = os.path.dirname(os.path.abspath(__file__))  # .../services/
        services_dir = os.path.dirname(current_dir)  # .../tool-go-soft/
        temp_base_dir = os.path.join(services_dir, "temp")  # .../tool-go-soft/temp/
        os.makedirs(temp_base_dir, exist_ok=True)
        
        # Tạo temp directory với timestamp để tránh conflict
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_dir = os.path.join(temp_base_dir, f"giaynoptien_{timestamp}")
        os.makedirs(temp_dir, exist_ok=True)
        
        # ✅ Chỉ log temp_dir khi có lỗi (không log khi mới bắt đầu)
        # logger.info(f"📁 Temp directory for debug files: {temp_dir}")
        
        # ✅ FIX: Không tạo folder screenshot ngay từ đầu, chỉ tạo khi có lỗi thực sự
        screenshots_dir = None  # Sẽ được tạo khi cần screenshot
        
        ssid = session.dse_session_id
        
        try:
            yield {"type": "info", "message": "Đang xử lý giấy nộp tiền..."}
            
            # Navigate đến trang giấy nộp tiền qua connectSSO (giống tờ khai)
            success = await self._navigate_to_giaynoptien_page(page, ssid)
            
            if not success:
                # Chụp màn hình khi navigate thất bại
                try:
                    screenshot_path = os.path.join(screenshots_dir, "01_navigate_failed.png")
                    logger.info(f"Attempting to save screenshot to: {screenshot_path}")
                    await page.screenshot(path=screenshot_path, full_page=True)
                    if os.path.exists(screenshot_path):
                        file_size = os.path.getsize(screenshot_path)
                        logger.info(f"✅ Screenshot saved: {screenshot_path} ({file_size} bytes)")
                    else:
                        logger.error(f"❌ Screenshot file not created: {screenshot_path}")
                except Exception as e:
                    logger.error(f"❌ Error saving screenshot 01_navigate_failed: {e}")
                yield {"type": "error", "error": "Không thể navigate đến trang tra cứu giấy nộp thuế. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            # Tìm frame từ iframe SSO (giống tờ khai)
            # Đợi frame xuất hiện trong page.frames (có thể mất thời gian)
            frame = None
            max_wait = 30  # Đợi tối đa 15 giây (30 * 0.5)
            for i in range(max_wait):
                try:
                    frames = page.frames
                    for f in frames:
                        if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                            frame = f
                            logger.info(f"Found frame for giaynoptien: {frame.url[:100]}...")
                            # Kiểm tra xem frame đã load chưa
                            try:
                                await frame.wait_for_load_state('domcontentloaded', timeout=2000)
                                break
                            except:
                                # Frame chưa load xong, tiếp tục đợi
                                frame = None
                                pass
                    
                    if frame:
                        break
                except Exception as e:
                    logger.debug(f"Waiting for frame (attempt {i + 1}/{max_wait}): {e}")
                
                await asyncio.sleep(0.5)
            
            if not frame:
                yield {"type": "error", "error": "Không tìm thấy iframe sau khi navigate. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            # Đợi frame load và kiểm tra form giấy nộp tiền
            try:
                await frame.wait_for_load_state('domcontentloaded', timeout=15000)
                await asyncio.sleep(1)
                await frame.wait_for_selector('input[name="ngay_lap_tu_ngay"], #ngay_lap_tu_ngay', timeout=15000)
                logger.info("Tra cuu giay nop tien form loaded successfully")
                
                # ✅ Không chụp screenshot khi form load thành công (chỉ chụp khi có lỗi)
            except Exception as e:
                logger.warning(f"Frame found but form not found: {e}")
                # Chụp màn hình khi form không load được
                try:
                    # ✅ Tạo folder screenshot khi có lỗi
                    if screenshots_dir is None:
                        screenshots_base_dir = os.path.join(services_dir, "screenshots")
                        os.makedirs(screenshots_base_dir, exist_ok=True)
                        screenshots_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        screenshots_dir = os.path.join(screenshots_base_dir, f"giaynoptien_{session_id[:8]}_{screenshots_timestamp}")
                        os.makedirs(screenshots_dir, exist_ok=True)
                    
                    screenshot_path = os.path.join(screenshots_dir, "02_form_not_found.png")
                    logger.info(f"Attempting to save screenshot to: {screenshot_path}")
                    await page.screenshot(path=screenshot_path, full_page=True)
                    if os.path.exists(screenshot_path):
                        file_size = os.path.getsize(screenshot_path)
                        logger.info(f"✅ Screenshot saved: {screenshot_path} ({file_size} bytes)")
                    else:
                        logger.error(f"❌ Screenshot file not created: {screenshot_path}")
                except Exception as e:
                    logger.error(f"❌ Error saving screenshot 02_form_not_found: {e}")
                yield {"type": "error", "error": "Không tìm thấy form tra cứu giấy nộp tiền. Vui lòng thử lại.", "error_code": "NAVIGATION_ERROR"}
                return
            
            # Check session timeout
            if await self._check_session_timeout(page):
                yield {
                    "type": "error",
                    "error": "Phiên giao dịch hết hạn. Vui lòng đăng nhập lại.",
                    "error_code": "SESSION_EXPIRED"
                }
                return
            
            # Chia khoảng thời gian
            date_ranges = self._get_date_ranges(start_date, end_date, days_interval=360)
            
            # ✅ Tính % cho từng khoảng thời gian (giống tờ khai)
            total_days = (datetime.strptime(end_date, "%d/%m/%Y") - datetime.strptime(start_date, "%d/%m/%Y")).days + 1
            days_per_range = 360
            range_percentages = []
            for i, dr in enumerate(self._get_date_ranges(start_date, end_date, days_interval=days_per_range)):
                start_dt = datetime.strptime(dr[0], "%d/%m/%Y")
                end_dt = datetime.strptime(dr[1], "%d/%m/%Y")
                range_days = (end_dt - start_dt).days + 1
                range_percent = (range_days / total_days) * 100 if total_days > 0 else 0
                range_percentages.append(range_percent)
            
            total_count = 0
            results = []
            files_info = []
            total_size = 0
            
            # ✅ Khởi tạo accumulated variables (giống thông báo)
            accumulated_total_so_far = 0
            accumulated_downloaded_so_far = 0
            accumulated_percent_so_far = 0.0
            
            # ✅ Tạo screenshot_dir một lần duy nhất cho toàn bộ job (không dùng timestamp)
            screenshot_dir = None
            def get_screenshot_dir():
                nonlocal screenshot_dir
                if screenshot_dir is None:
                    # Dùng job_id nếu có, nếu không thì dùng session_id
                    folder_name = f"giaynoptien_{job_id[:8] if job_id else session_id[:8]}"
                    screenshot_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "screenshots", folder_name)
                    os.makedirs(screenshot_dir, exist_ok=True)
                return screenshot_dir
            
            yield {"type": "info", "message": f"Bắt đầu crawl {len(date_ranges)} khoảng thời gian..."}
            
            for range_idx, date_range in enumerate(date_ranges):
                # ✅ Check cancelled trước khi xử lý khoảng tiếp theo
                if job_id and await self._check_cancelled(job_id):
                    logger.info(f"Job {job_id} đã bị cancel, dừng crawl")
                    yield {"type": "error", "error": "Job đã bị hủy", "error_code": "JOB_CANCELLED"}
                    return
                
                yield {
                    "type": "progress", 
                    "current": range_idx + 1, 
                    "total": len(date_ranges),
                    "message": f"Đang xử lý khoảng {date_range[0]} - {date_range[1]}...",
                    "accumulated_total": accumulated_total_so_far,
                    "accumulated_downloaded": accumulated_downloaded_so_far,
                    "accumulated_percent": int(round(accumulated_percent_so_far))
                }
                
                try:
                    # Nhập ngày bắt đầu (dùng name attribute)
                    start_input = frame.locator('input[name="ngay_lap_tu_ngay"], input#ngay_lap_tu_ngay')
                    await start_input.fill('')
                    await start_input.fill(date_range[0])
                    
                    # Nhập ngày kết thúc (dùng name attribute)
                    end_input = frame.locator('input[name="ngay_lap_den_ngay"], input#ngay_lap_den_ngay')
                    await end_input.click()
                    from playwright.async_api import Keyboard
                    await end_input.press('Control+a')
                    await end_input.fill(date_range[1])
                    
                    # Click tìm kiếm (dùng value hoặc onclick)
                    search_btn = frame.locator('input[value="Tra cứu"], input[onclick*="traCuuChungTu"]')
                    await search_btn.click()
                    
                    await asyncio.sleep(2)
                    
                    logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Đã click search cho khoảng: {date_range[0]} - {date_range[1]}")
                    
                    # ✅ Đợi một chút để đảm bảo request đã được gửi
                    await asyncio.sleep(1)
                    
                    # ✅ Tìm lại frame mới sau khi click search (iframe có thể reload khi chuyển khoảng thời gian)
                    try:
                        frames = page.frames
                        for f in frames:
                            if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                frame = f  # Cập nhật frame object mới
                                logger.info(f"🔄 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Đã tìm lại frame mới sau khi click search: {frame.url[:100]}...")
                                break
                    except Exception as refind_frame_e:
                        logger.warning(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể tìm lại frame mới sau khi click search: {refind_frame_e}")
                    
                    # ✅ Đợi frame load xong trước khi đợi table
                    try:
                        await frame.wait_for_load_state('networkidle', timeout=5000)
                        logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Frame đã load xong (networkidle)")
                    except Exception as frame_load_e:
                        logger.debug(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể đợi frame networkidle: {frame_load_e}")
                    
                    # ✅ Đợi table load xong để đảm bảo đã chuyển sang khoảng mới
                    try:
                        logger.info(f"⏳ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Đang đợi table load sau khi click search...")
                        table_body_check = frame.locator('table#data_content_onday tbody#allResultTableBody, #allResultTableBody').first
                        await table_body_check.wait_for(timeout=10000, state='visible')
                        logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Table đã load xong sau khi click search")
                        
                        # ✅ Đợi thêm một chút để đảm bảo dữ liệu đã được render xong
                        await asyncio.sleep(1.5)
                        logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Đã đợi thêm để đảm bảo dữ liệu đã render xong")
                    except Exception as wait_table_e:
                        logger.warning(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể đợi table load sau khi click search: {wait_table_e}")
                        # Tiếp tục xử lý, sẽ kiểm tra "Không có dữ liệu" ở bước tiếp theo
                    
                    # ✅ Kiểm tra "Không có dữ liệu" ngay sau khi search (trước khi vào pagination)
                    try:
                        no_data_text = frame.locator('div:has-text("Không có dữ liệu"), strong:has-text("Không có dữ liệu"), div.align-center:has-text("Không có dữ liệu")').first
                        if await no_data_text.count() > 0:
                            no_data_content = await no_data_text.text_content()
                            if "Không có dữ liệu" in (no_data_content or ""):
                                logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Phát hiện 'Không có dữ liệu' cho khoảng {date_range[0]} - {date_range[1]}")
                                yield {
                                    "type": "info", 
                                    "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}",
                                    "accumulated_percent": int(round(accumulated_percent_so_far)),
                                    "accumulated_total": accumulated_total_so_far,
                                    "accumulated_downloaded": accumulated_downloaded_so_far
                                }
                                accumulated_percent_so_far += range_percentages[range_idx] if range_idx < len(range_percentages) else 0
                                continue  # Bỏ qua khoảng này, chuyển sang khoảng tiếp theo
                    except Exception as no_data_check_e:
                        logger.debug(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể kiểm tra 'Không có dữ liệu': {no_data_check_e}")
                    
                    # Xử lý phân trang (giống thông báo)
                    check_pages = True
                    page_num = 0
                    range_total_records = None  # Tổng số bản ghi trong khoảng này (parse từ currAcc)
                    range_downloaded_so_far = 0  # Tổng số file đã download trong khoảng này (từ các trang trước)
                    max_pages = 100  # ✅ Giới hạn số trang tối đa để tránh vòng lặp vô hạn
                    previous_row_count = 0  # ✅ Lưu số rows của trang trước để verify table đã chuyển trang
                    
                    # ✅ Lưu % tích lũy tại thời điểm bắt đầu khoảng này (để tính % cho khoảng này chính xác) (giống thông báo)
                    accumulated_percent_so_far_at_range_start = accumulated_percent_so_far
                    
                    # ✅ Tính % cho khoảng này
                    range_percent = range_percentages[range_idx] if range_idx < len(range_percentages) else 0
                    
                    while check_pages and page_num < max_pages:
                        # ✅ Check cancelled trước khi xử lý trang tiếp theo
                        if job_id and await self._check_cancelled(job_id):
                            logger.info(f"[GIAYNOPTIEN] Job {job_id} đã bị cancel, dừng crawl")
                            yield {
                                "type": "error",
                                "error": "Job đã bị hủy",
                                "error_code": "JOB_CANCELLED"
                            }
                            return
                        
                        page_num += 1
                        logger.info(f"📄 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Đang xử lý trang {page_num}... (check_pages={check_pages})")
                        
                        # ✅ Kiểm tra "Không có dữ liệu" trước khi tìm table
                        try:
                            no_data_text = frame.locator('div:has-text("Không có dữ liệu"), strong:has-text("Không có dữ liệu")').first
                            if await no_data_text.count() > 0:
                                no_data_content = await no_data_text.text_content()
                                if "Không có dữ liệu" in (no_data_content or ""):
                                    logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Phát hiện 'Không có dữ liệu' cho khoảng {date_range[0]} - {date_range[1]}")
                                    yield {
                                        "type": "info", 
                                        "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}",
                                        "accumulated_percent": int(round(accumulated_percent_so_far)),
                                        "accumulated_total": accumulated_total_so_far,
                                        "accumulated_downloaded": accumulated_downloaded_so_far
                                    }
                                    accumulated_percent_so_far += range_percent
                                    break
                        except Exception as no_data_check_e:
                            logger.debug(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể kiểm tra 'Không có dữ liệu': {no_data_check_e}")
                        
                        # Tìm bảng kết quả
                        try:
                            table_body = frame.locator('table#data_content_onday tbody#allResultTableBody, #allResultTableBody').first
                            await table_body.wait_for(timeout=10000, state='visible')
                        except Exception as e:
                            logger.warning(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không tìm thấy bảng kết quả cho khoảng {date_range[0]} - {date_range[1]}: {e}")
                            
                            # ✅ Kiểm tra "Không có dữ liệu" TRƯỚC KHI screenshot
                            has_no_data = False
                            try:
                                no_data_text = frame.locator('div:has-text("Không có dữ liệu"), strong:has-text("Không có dữ liệu"), div.align-center:has-text("Không có dữ liệu")').first
                                if await no_data_text.count() > 0:
                                    no_data_content = await no_data_text.text_content()
                                    if "Không có dữ liệu" in (no_data_content or ""):
                                        has_no_data = True
                                        logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Phát hiện 'Không có dữ liệu' (không có table)")
                                        yield {
                                            "type": "info", 
                                            "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}",
                                            "accumulated_percent": int(round(accumulated_percent_so_far)),
                                            "accumulated_total": accumulated_total_so_far,
                                            "accumulated_downloaded": accumulated_downloaded_so_far
                                        }
                                        accumulated_percent_so_far += range_percent
                                        break
                            except Exception as no_data_check_e2:
                                logger.debug(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể kiểm tra 'Không có dữ liệu' lần 2: {no_data_check_e2}")
                            
                            # ✅ CHỈ screenshot khi thực sự có lỗi (không phải do không có dữ liệu)
                            if not has_no_data:
                                try:
                                    screenshot_dir = get_screenshot_dir()
                                    
                                    if 'page' in locals() and page:
                                        page_screenshot = os.path.join(screenshot_dir, f"no_table_page_{range_idx + 1}_page_{page_num}.png")
                                        await page.screenshot(path=page_screenshot, full_page=True)
                                        logger.info(f"📸 Screenshot page saved: {page_screenshot}")
                                    
                                    if 'frame' in locals() and frame:
                                        try:
                                            frame_screenshot = os.path.join(screenshot_dir, f"no_table_frame_{range_idx + 1}_page_{page_num}.png")
                                            await frame.screenshot(path=frame_screenshot, full_page=True)
                                            logger.info(f"📸 Screenshot frame saved: {frame_screenshot}")
                                        except Exception as frame_screenshot_e:
                                            logger.warning(f"⚠️ Cannot screenshot frame: {frame_screenshot_e}")
                                        
                                        try:
                                            frame_html = await frame.content()
                                            html_file = os.path.join(screenshot_dir, f"no_table_frame_{range_idx + 1}_page_{page_num}.html")
                                            with open(html_file, 'w', encoding='utf-8') as f:
                                                f.write(frame_html)
                                            logger.info(f"📄 Frame HTML saved: {html_file}")
                                        except Exception as html_e:
                                            logger.warning(f"⚠️ Cannot save frame HTML: {html_e}")
                                    
                                    logger.info(f"📸 Screenshots saved to: {screenshot_dir}")
                                except Exception as screenshot_e:
                                    logger.error(f"❌ Error taking screenshot: {screenshot_e}")
                            else:
                                # Không có dữ liệu, không cần screenshot
                                logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không có dữ liệu, bỏ qua screenshot")
                            
                            # ✅ Nếu không có "Không có dữ liệu" và không có table, bỏ qua khoảng này
                            logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không có table và không có 'Không có dữ liệu', bỏ qua khoảng này")
                            if total_count == 0:
                                yield {
                                    "type": "info", 
                                    "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}",
                                    "accumulated_percent": int(round(accumulated_percent_so_far)),
                                    "accumulated_total": accumulated_total_so_far,
                                    "accumulated_downloaded": accumulated_downloaded_so_far
                                }
                            accumulated_percent_so_far += range_percent
                            break
                        
                        rows = table_body.locator('tr')
                        row_count = await rows.count()
                        
                        # ✅ Lưu row_count của trang hiện tại để verify sau khi click next
                        if page_num == 1:
                            previous_row_count = row_count
                        
                        # ✅ CHỈ parse currAcc nếu có rows (tránh parse sai khi không có dữ liệu)
                        # ✅ Parse tổng số bản ghi từ phần currAcc (chỉ parse ở trang đầu tiên và khi có rows)
                        if page_num == 1 and row_count > 0:
                            try:
                                curr_acc = frame.locator('#currAcc').first
                                if await curr_acc.count() > 0:
                                    curr_acc_text = await curr_acc.text_content()
                                    import re
                                    match = re.search(r'Có\s*<b>(\d+)</b>\s*bản\s*ghi|Có\s*(\d+)\s*bản\s*ghi', curr_acc_text)
                                    if match:
                                        range_total_records = int(match.group(1) or match.group(2))
                                        
                                        # ✅ Parse số trang từ pagination info
                                        pagination_info = await self._extract_pagination_info(frame)
                                        total_pages = pagination_info.get("total_pages", 1) if pagination_info else 1
                                        
                                        logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Parse tổng số bản ghi từ currAcc: {range_total_records} trong {total_pages} trang")
                                        
                                        # ✅ KHÔNG cộng range_total_records vào accumulated_total ngay lập tức
                                        # Sẽ cộng sau khi biết số file thực sự cần download (sau khi filter duplicate)
                                        
                                        yield {
                                            "type": "info",
                                            "message": f"Tìm thấy {range_total_records} bản ghi trong {total_pages} trang. Bắt đầu tải...",
                                            "accumulated_total": accumulated_total_so_far,
                                            "accumulated_downloaded": accumulated_downloaded_so_far,
                                            "accumulated_percent": int(round(accumulated_percent_so_far))
                                        }
                            except Exception as e:
                                logger.warning(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể parse tổng số bản ghi từ currAcc: {e}")
                        
                        # ✅ Kiểm tra nếu không có rows (table rỗng)
                        if row_count == 0:
                            logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Table rỗng (0 rows)")
                            if page_num == 1 and total_count == 0:
                                yield {
                                    "type": "info", 
                                    "message": f"Không có dữ liệu trong khoảng {date_range[0]} - {date_range[1]}",
                                    "accumulated_percent": int(round(accumulated_percent_so_far)),
                                    "accumulated_total": accumulated_total_so_far,
                                    "accumulated_downloaded": accumulated_downloaded_so_far
                                }
                                accumulated_percent_so_far += range_percent
                                break
                            else:
                                # Không có rows trên trang này, dừng pagination
                                check_pages = False
                                break
                        
                        logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Tìm thấy {row_count} rows, Range %: {range_percent:.2f}%, Accumulated %: {accumulated_percent_so_far:.2f}%")
                        
                        yield {
                            "type": "progress", 
                            "current": total_count, 
                            "message": f"Đang xử lý {row_count} giấy nộp tiền (trang hiện tại)...",
                            "percent": int(round(min(accumulated_percent_so_far, 100))),
                            "accumulated_percent": int(round(min(accumulated_percent_so_far, 100))),
                            "accumulated_total": accumulated_total_so_far,
                            "accumulated_downloaded": accumulated_downloaded_so_far
                        }
                        
                        download_queue = []
                        page_valid_count = 0
                        
                        for i in range(row_count):
                            try:
                                row = rows.nth(i)
                                cols = row.locator('td')
                                col_count = await cols.count()
                                
                                if col_count < 5:
                                    continue
                                
                                # Lấy id_gnt từ link chiTietCT(id) trong cột 5 (index 4)
                                # Hoặc từ link downloadGNT(id) trong cột 19 (index 18)
                                id_gnt = None
                                
                                # Thử lấy từ cột 5 (chiTietCT)
                                try:
                                    if col_count > 4:
                                        col5_links = cols.nth(4).locator('a[href*="chiTietCT"]')
                                        if await col5_links.count() > 0:
                                            href = await col5_links.first.get_attribute('href')
                                            if href and 'chiTietCT(' in href:
                                                match = re.search(r'chiTietCT\((\d+)\)', href)
                                                if match:
                                                    id_gnt = match.group(1)
                                except:
                                    pass
                                
                                # Nếu không lấy được từ cột 5, thử lấy từ cột 19 (downloadGNT)
                                if not id_gnt:
                                    try:
                                        if col_count > 18:
                                            col19_links = cols.nth(18).locator('a[href*="downloadGNT"]')
                                            if await col19_links.count() > 0:
                                                href = await col19_links.first.get_attribute('href')
                                                if href and 'downloadGNT(' in href:
                                                    match = re.search(r'downloadGNT\((\d+)\)', href)
                                                    if match:
                                                        id_gnt = match.group(1)
                                    except:
                                        pass
                                
                                # Fallback: Lấy từ cột 2
                                if not id_gnt:
                                    try:
                                        id_gnt = await cols.nth(2).text_content()
                                        id_gnt = id_gnt.strip() if id_gnt else ""
                                        if not id_gnt or len(id_gnt) < 4:
                                            id_gnt = None
                                    except:
                                        pass
                                
                                if not id_gnt:
                                    continue
                                
                                # Chỉ đếm khi item hợp lệ (giống thông báo)
                                page_valid_count += 1
                                total_count += 1
                                
                                # Tìm link download từ các cột 17-20 (cột 19 là cột # có link downloadGNT)
                                download_link_found = None
                                download_col_index = None
                                
                                for col_idx in [17, 18, 19, 20]:
                                    if col_count > col_idx and not download_link_found:
                                        try:
                                            links = cols.nth(col_idx).locator('a[href*="downloadGNT"], a[onclick*="downloadGNT"]')
                                            link_count = await links.count()
                                            if link_count > 0:
                                                download_link_found = links.first
                                                download_col_index = col_idx
                                                logger.info(f"Found download link for giaynoptien {id_gnt} in column {col_idx}")
                                                break
                                        except Exception as e:
                                            logger.debug(f"Error checking column {col_idx} for download link: {e}")
                                            pass
                                
                                if download_link_found:
                                                download_queue.append({
                                                    "id": id_gnt,
                                        "download_link": download_link_found,
                                        "cols": cols,
                                        "col_index": download_col_index
                                    })
                            
                            except Exception as e:
                                logger.error(f"Error processing row: {e}")
                                continue
                        
                        logger.info(f"📋 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Có {page_valid_count} items hợp lệ, {len(download_queue)} items có link download")
                        
                        # Download từng file và yield progress (giống thông báo)
                        if download_queue:
                            queue_total = len(download_queue)
                            
                            # ✅ Tính % cho mỗi file download (giống thông báo)
                            if range_total_records:
                                # Tính % dựa trên tổng số bản ghi trong khoảng (dùng cho tất cả các trang)
                                percent_per_file = range_percent / range_total_records
                                logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Dùng range_total_records={range_total_records} để tính % per file: {percent_per_file:.4f}%")
                            elif queue_total > 0:
                                # Nếu không có range_total_records, tính % dựa trên số file trên trang hiện tại
                                percent_per_file = range_percent / queue_total
                            else:
                                percent_per_file = 0.0
                            
                            # ✅ Cập nhật accumulated_total khi biết số file cần download (giống thông báo)
                            if range_total_records and page_num == 1:
                                # Chỉ cập nhật accumulated_total ở trang đầu tiên với tổng số bản ghi
                                accumulated_total_so_far += range_total_records
                                logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Cập nhật accumulated_total với range_total_records={range_total_records}, accumulated_total_so_far={accumulated_total_so_far}")
                            elif not range_total_records:
                                # Nếu không có range_total_records, cộng số file trên trang hiện tại
                                accumulated_total_so_far += queue_total
                            
                            # Hiển thị tổng số file sẽ tải (dùng range_total_records nếu có, nếu không dùng queue_total)
                            display_total = range_total_records if range_total_records else queue_total
                            
                            logger.info(f"⬇️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Bắt đầu download {queue_total} files (tổng khoảng: {display_total}), Range %: {range_percent:.2f}%, Percent per file: {percent_per_file:.4f}%, Accumulated total: {accumulated_total_so_far}, Accumulated %: {accumulated_percent_so_far:.2f}%")
                            
                            # ✅ CHỈ publish download_start khi bắt đầu khoảng mới (trang 1), không publish khi chuyển trang
                            if page_num == 1:
                                yield {
                                    "type": "download_start",
                                    "total_to_download": display_total,  # ✅ Hiển thị tổng số file sẽ tải trong khoảng
                                    "current_page_download": queue_total,  # Số file trên trang hiện tại
                                    "date_range": f"{date_range[0]} - {date_range[1]}",
                                    "range_index": range_idx + 1,
                                    "total_ranges": len(date_ranges),
                                    "accumulated_total": accumulated_total_so_far,
                                    "accumulated_downloaded": accumulated_downloaded_so_far,
                                    "range_percent": range_percent,  # % của khoảng này
                                    "accumulated_percent": int(round(min(accumulated_percent_so_far, 100))),  # ✅ Đảm bảo không vượt quá 100%
                                    "message": f"Bắt đầu tải {display_total} giấy nộp tiền trong khoảng {date_range[0]} - {date_range[1]}..."
                                }
                            
                            downloaded = 0
                            
                            for item_idx, item in enumerate(download_queue, 1):
                                try:
                                    logger.info(f"📥 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang download file {item_idx}/{queue_total}: {item.get('id', 'N/A')}...")
                                    success = await self._download_single_giaynoptien(session, item, temp_dir)
                                    if success:
                                        downloaded += 1
                                        accumulated_downloaded_so_far += 1
                                        range_downloaded_so_far += 1  # ✅ Cộng dồn số file đã download trong khoảng này
                                        logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã download thành công file {item_idx}/{queue_total}: {item.get('id', 'N/A')}")
                                    else:
                                        logger.warning(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Download thất bại file {item_idx}/{queue_total}: {item.get('id', 'N/A')}")
                                except Exception as download_e:
                                    logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi download file {item_idx}/{queue_total} ({item.get('id', 'N/A')}): {download_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    # Tiếp tục download file tiếp theo
                                    continue
                                
                                # ✅ Tính % tích lũy: % từ các khoảng trước + % của các file đã download trong khoảng này
                                # QUAN TRỌNG: Dùng accumulated_percent_so_far_at_range_start (không phải accumulated_percent_so_far)
                                # để tránh cộng dồn sai khi đã cập nhật accumulated_percent_so_far trong vòng lặp
                                if range_total_records:
                                    # Tính % dựa trên tổng số bản ghi trong khoảng
                                    # % của khoảng này = (số file đã download / tổng số file trong khoảng) * % của khoảng
                                    range_accumulated_percent = (range_downloaded_so_far / range_total_records) * range_percent
                                    # Cộng với % tích lũy từ các khoảng trước (tại thời điểm bắt đầu khoảng này)
                                    current_accumulated_percent = accumulated_percent_so_far_at_range_start + range_accumulated_percent
                                else:
                                    # Tính % dựa trên số file trên trang hiện tại
                                    current_accumulated_percent = accumulated_percent_so_far_at_range_start + (downloaded * percent_per_file)
                                
                                # ✅ Đảm bảo không vượt quá 100%
                                current_accumulated_percent = min(current_accumulated_percent, 100.0)
                                
                                # ✅ CẬP NHẬT accumulated_percent_so_far liên tục trong quá trình download
                                accumulated_percent_so_far = current_accumulated_percent
                                
                                current_accumulated_percent = min(current_accumulated_percent, 100.0)
                                accumulated_percent_so_far = current_accumulated_percent
                                
                                display_downloaded = range_downloaded_so_far if range_total_records else downloaded
                                # Hiển thị tổng số file đã download trong khoảng (dùng range_total_records nếu có)
                                display_total = range_total_records if range_total_records else queue_total
                                display_downloaded = range_downloaded_so_far if range_total_records else downloaded
                                
                                if item_idx % 5 == 0 or item_idx == queue_total:  # Log mỗi 5 file hoặc file cuối
                                    logger.info(f"⬇️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã download {display_downloaded}/{display_total} files (trang: {downloaded}/{queue_total}), Current accumulated %: {accumulated_percent_so_far:.2f}%")
                                
                                # ✅ Yield progress event với exception handling (giống thông báo)
                                try:
                                    yield {
                                        "type": "download_progress",
                                        "downloaded": display_downloaded,  # ✅ Hiển thị tổng số file đã download trong khoảng
                                        "total": display_total,  # ✅ Hiển thị tổng số file sẽ tải trong khoảng
                                        "current_page_downloaded": downloaded,  # Số file đã download trên trang hiện tại
                                        "current_page_total": queue_total,  # Số file trên trang hiện tại
                                        "percent": round(display_downloaded / display_total * 100, 1) if display_total > 0 else 0,
                                        "current_item": item.get("id", ""),
                                        "accumulated_total": accumulated_total_so_far,
                                        "accumulated_downloaded": accumulated_downloaded_so_far,
                                        "accumulated_percent": int(round(accumulated_percent_so_far)),  # ✅ Dùng accumulated_percent_so_far đã được cập nhật
                                        "message": f"Đã tải {display_downloaded}/{display_total} ({round(display_downloaded / display_total * 100, 1) if display_total > 0 else 0}%)"
                                    }
                                except Exception as yield_e:
                                    logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi yield progress event: {yield_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    # Tiếp tục download file tiếp theo, không dừng vì lỗi yield
                                    pass
                            
                            # ✅ Cập nhật previous_row_count sau khi xử lý xong trang này
                            previous_row_count = row_count
                            
                            # ✅ Cập nhật accumulated_percent_so_far sau khi download xong khoảng này (chỉ ở trang cuối cùng) (giống thông báo)
                            # Chỉ cập nhật khi không còn trang tiếp theo và đã download hết tất cả file trong khoảng
                            if not check_pages:  # Nếu không còn trang tiếp theo
                                # Đảm bảo accumulated_percent_so_far đạt đúng % của khoảng này
                                # Nếu có range_total_records, đã tính % dựa trên số file download, không cần cộng thêm
                                # Nếu không có range_total_records, cộng % của khoảng này
                                if not range_total_records:
                                    accumulated_percent_so_far += range_percent
                                # ✅ Đảm bảo không vượt quá 100%
                                accumulated_percent_so_far = min(accumulated_percent_so_far, 100.0)
                            
                            # Hiển thị tổng số file đã download trong khoảng (dùng range_total_records nếu có)
                            display_total = range_total_records if range_total_records else queue_total
                            display_downloaded = range_downloaded_so_far if range_total_records else downloaded
                            
                            logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Hoàn thành download {display_downloaded}/{display_total} files (trang: {downloaded}/{queue_total}), Accumulated %: {accumulated_percent_so_far:.2f}%")
                            
                            yield {
                                "type": "download_complete",
                                "downloaded": display_downloaded,
                                "total": display_total,
                                "current_page_downloaded": downloaded,
                                "current_page_total": queue_total,
                                "accumulated_total": accumulated_total_so_far,
                                "accumulated_downloaded": accumulated_downloaded_so_far,
                                "accumulated_percent": int(round(accumulated_percent_so_far))
                            }
                        
                        # Chỉ cộng số items hợp lệ vào total_count
                        total_count += page_valid_count
                        
                        # Check pagination - next page (giống thông báo)
                        try:
                            logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang kiểm tra nút next...")
                            next_btn = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                            next_btn_count = await next_btn.count()
                            logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Số lượng nút next: {next_btn_count}")
                            
                            if next_btn_count > 0:
                                logger.info(f"➡️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Có trang tiếp theo, đang chuyển trang...")
                                
                                try:
                                    logger.info(f"🖱️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang click nút next...")
                                    await asyncio.wait_for(next_btn.click(), timeout=10.0)
                                    logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã click nút next thành công")
                                except asyncio.TimeoutError:
                                    logger.error(f"⏱️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Timeout khi click nút next (10s)")
                                    
                                    # ✅ Screenshot khi timeout click next
                                    try:
                                        screenshot_dir = get_screenshot_dir()
                                        
                                        if 'page' in locals() and page:
                                            page_screenshot = os.path.join(screenshot_dir, f"timeout_click_next_page_{range_idx + 1}_page_{page_num}.png")
                                            await page.screenshot(path=page_screenshot, full_page=True)
                                            logger.info(f"📸 Screenshot page saved: {page_screenshot}")
                                        
                                        if 'frame' in locals() and frame:
                                            try:
                                                frame_screenshot = os.path.join(screenshot_dir, f"timeout_click_next_frame_{range_idx + 1}_page_{page_num}.png")
                                                await frame.screenshot(path=frame_screenshot, full_page=True)
                                                logger.info(f"📸 Screenshot frame saved: {frame_screenshot}")
                                            except Exception as frame_screenshot_e:
                                                logger.warning(f"⚠️ Cannot screenshot frame: {frame_screenshot_e}")
                                        
                                        logger.info(f"📸 Screenshots saved to: {screenshot_dir}")
                                    except Exception as screenshot_e:
                                        logger.error(f"❌ Error taking screenshot: {screenshot_e}")
                                    
                                    check_pages = False
                                    continue
                                except Exception as click_e:
                                    logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi click nút next: {click_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    
                                    # ✅ Screenshot khi lỗi click next
                                    try:
                                        screenshot_dir = get_screenshot_dir()
                                        
                                        if 'page' in locals() and page:
                                            page_screenshot = os.path.join(screenshot_dir, f"error_click_next_page_{range_idx + 1}_page_{page_num}.png")
                                            await page.screenshot(path=page_screenshot, full_page=True)
                                            logger.info(f"📸 Screenshot page saved: {page_screenshot}")
                                        
                                        if 'frame' in locals() and frame:
                                            try:
                                                frame_screenshot = os.path.join(screenshot_dir, f"error_click_next_frame_{range_idx + 1}_page_{page_num}.png")
                                                await frame.screenshot(path=frame_screenshot, full_page=True)
                                                logger.info(f"📸 Screenshot frame saved: {frame_screenshot}")
                                            except Exception as frame_screenshot_e:
                                                logger.warning(f"⚠️ Cannot screenshot frame: {frame_screenshot_e}")
                                        
                                        logger.info(f"📸 Screenshots saved to: {screenshot_dir}")
                                    except Exception as screenshot_e:
                                        logger.error(f"❌ Error taking screenshot: {screenshot_e}")
                                    
                                    check_pages = False
                                    continue
                                
                                logger.info(f"⏳ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đợi 2 giây sau khi click...")
                                await asyncio.sleep(2)
                                logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã đợi xong 2 giây, bắt đầu đợi table load...")
                                
                                # ✅ Kiểm tra lại xem có trang tiếp theo không (sau khi click)
                                try:
                                    logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang đợi table load cho trang {page_num + 1}...")
                                    # ✅ Kiểm tra frame còn tồn tại không
                                    try:
                                        frame_url = frame.url
                                        logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Frame URL: {frame_url[:100]}...")
                                    except Exception as frame_check_e:
                                        logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Frame không còn tồn tại sau khi click: {frame_check_e}")
                                        import traceback
                                        logger.error(f"Traceback: {traceback.format_exc()}")
                                        check_pages = False
                                        continue
                                    
                                    # Đợi table load để đảm bảo trang đã chuyển (tăng timeout lên 15 giây)
                                    logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đang tìm table locator...")
                                    try:
                                        table_body_check = frame.locator('table#data_content_onday tbody#allResultTableBody, #allResultTableBody').first
                                        logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Đã tìm thấy table locator, đang đợi table visible...")
                                    except Exception as locator_e:
                                        logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Lỗi khi tìm table locator: {locator_e}")
                                        import traceback
                                        logger.error(f"Traceback: {traceback.format_exc()}")
                                        check_pages = False
                                        continue
                                    
                                    try:
                                        await asyncio.wait_for(
                                            table_body_check.wait_for(timeout=15000, state='visible'),
                                            timeout=20.0  # Tổng timeout 20 giây
                                        )
                                        await asyncio.sleep(1)
                                        
                                        try:
                                            rows_check = table_body_check.locator('tr')
                                            row_count_check = await rows_check.count()
                                            
                                            if row_count_check == previous_row_count and previous_row_count > 0:
                                                await asyncio.sleep(2)
                                                row_count_check = await rows_check.count()
                                                
                                                if row_count_check == previous_row_count:
                                                    check_pages = False
                                                    continue
                                        except Exception as verify_e:
                                            pass
                                        
                                        # ✅ Tiếp tục vòng lặp (check_pages vẫn True) - giống thông báo
                                    except Exception as wait_table_e:
                                        raise
                                except asyncio.TimeoutError:
                                    await asyncio.sleep(3)
                                    try:
                                        table_body_check_retry = frame.locator('table#data_content_onday tbody#allResultTableBody, #allResultTableBody').first
                                        await asyncio.wait_for(
                                            table_body_check_retry.wait_for(timeout=15000, state='visible'),
                                            timeout=20.0
                                        )
                                        
                                        try:
                                            frames = page.frames
                                            for f in frames:
                                                if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                                    frame = f
                                                    break
                                        except Exception as refind_frame_e:
                                            pass
                                    except Exception as retry_e:
                                        try:
                                            no_data_text = frame.locator('div:has-text("Không có dữ liệu"), strong:has-text("Không có dữ liệu"), div.align-center:has-text("Không có dữ liệu")').first
                                            if await no_data_text.count() > 0:
                                                no_data_content = await no_data_text.text_content()
                                                if "Không có dữ liệu" in (no_data_content or ""):
                                                    check_pages = False
                                                    continue
                                        except Exception as no_data_check_e3:
                                            pass
                                        
                                        await asyncio.sleep(2)
                                        try:
                                            next_btn_check = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                                            next_btn_check_count = await next_btn_check.count()
                                            if next_btn_check_count == 0:
                                                check_pages = False
                                                continue
                                        except Exception as check_next_e:
                                            check_pages = False
                                            continue
                                        
                                        # Screenshot khi retry thất bại
                                        try:
                                            screenshot_dir = get_screenshot_dir()
                                            
                                            if 'page' in locals() and page:
                                                page_screenshot = os.path.join(screenshot_dir, f"table_not_load_page_{range_idx + 1}_page_{page_num}.png")
                                                await page.screenshot(path=page_screenshot, full_page=True)
                                                logger.info(f"📸 Screenshot page saved: {page_screenshot}")
                                            
                                            if 'frame' in locals() and frame:
                                                try:
                                                    frame_screenshot = os.path.join(screenshot_dir, f"table_not_load_frame_{range_idx + 1}_page_{page_num}.png")
                                                    await frame.screenshot(path=frame_screenshot, full_page=True)
                                                    logger.info(f"📸 Screenshot frame saved: {frame_screenshot}")
                                                except Exception as frame_screenshot_e:
                                                    logger.warning(f"⚠️ Cannot screenshot frame: {frame_screenshot_e}")
                                                
                                                try:
                                                    frame_html = await frame.content()
                                                    html_file = os.path.join(screenshot_dir, f"table_not_load_frame_{range_idx + 1}_page_{page_num}.html")
                                                    with open(html_file, 'w', encoding='utf-8') as f:
                                                        f.write(frame_html)
                                                    logger.info(f"📄 Frame HTML saved: {html_file}")
                                                except Exception as html_e:
                                                    logger.warning(f"⚠️ Cannot save frame HTML: {html_e}")
                                            
                                            logger.info(f"📸 Screenshots saved to: {screenshot_dir}")
                                        except Exception as screenshot_e:
                                            logger.error(f"❌ Error taking screenshot: {screenshot_e}")
                                        
                                        check_pages = False
                                        continue
                                    
                                except Exception as wait_e:
                                    logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num + 1} chưa load xong sau khi click next: {wait_e}")
                                    import traceback
                                    logger.error(f"Traceback: {traceback.format_exc()}")
                                    
                                    # ✅ Screenshot khi table không load sau khi click next
                                    try:
                                        screenshot_dir = get_screenshot_dir()
                                        
                                        if 'page' in locals() and page:
                                            page_screenshot = os.path.join(screenshot_dir, f"table_not_load_page_{range_idx + 1}_page_{page_num}.png")
                                            await page.screenshot(path=page_screenshot, full_page=True)
                                            logger.info(f"📸 Screenshot page saved: {page_screenshot}")
                                        
                                        if 'frame' in locals() and frame:
                                            try:
                                                frame_screenshot = os.path.join(screenshot_dir, f"table_not_load_frame_{range_idx + 1}_page_{page_num}.png")
                                                await frame.screenshot(path=frame_screenshot, full_page=True)
                                                logger.info(f"📸 Screenshot frame saved: {frame_screenshot}")
                                            except Exception as frame_screenshot_e:
                                                logger.warning(f"⚠️ Cannot screenshot frame: {frame_screenshot_e}")
                                            
                                            try:
                                                frame_html = await frame.content()
                                                html_file = os.path.join(screenshot_dir, f"table_not_load_frame_{range_idx + 1}_page_{page_num}.html")
                                                with open(html_file, 'w', encoding='utf-8') as f:
                                                    f.write(frame_html)
                                                logger.info(f"📄 Frame HTML saved: {html_file}")
                                            except Exception as html_e:
                                                logger.warning(f"⚠️ Cannot save frame HTML: {html_e}")
                                        
                                        logger.info(f"📸 Screenshots saved to: {screenshot_dir}")
                                    except Exception as screenshot_e:
                                        logger.error(f"❌ Error taking screenshot: {screenshot_e}")
                                    await asyncio.sleep(3)
                                    try:
                                        table_body_check_retry = frame.locator('table#data_content_onday tbody#allResultTableBody, #allResultTableBody').first
                                        await asyncio.wait_for(
                                            table_body_check_retry.wait_for(timeout=15000, state='visible'),
                                            timeout=20.0
                                        )
                                        logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Retry thành công, trang {page_num + 1} đã load xong")
                                        
                                        # ✅ Tìm lại frame mới sau khi retry (iframe có thể reload)
                                        try:
                                            frames = page.frames
                                            for f in frames:
                                                if 'thuedientu.gdt.gov.vn' in f.url and 'etaxnnt' in f.url:
                                                    frame = f  # Cập nhật frame object mới
                                                    logger.info(f"🔄 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Đã tìm lại frame mới sau retry: {frame.url[:100]}...")
                                                    break
                                        except Exception as refind_frame_e:
                                            logger.warning(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Không thể tìm lại frame mới sau retry: {refind_frame_e}")
                                    except Exception as retry_e:
                                        logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Retry vẫn thất bại: {retry_e}")
                                        await asyncio.sleep(2)
                                        try:
                                            next_btn_check = frame.locator('img[src="/etaxnnt/static/images/pagination_right.gif"]')
                                            next_btn_check_count = await next_btn_check.count()
                                            logger.info(f"🔍 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Sau retry, số lượng nút next: {next_btn_check_count}")
                                            if next_btn_check_count == 0:
                                                logger.info(f"🏁 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Sau khi click, không còn nút next, kết thúc phân trang")
                                                check_pages = False
                                            else:
                                                logger.warning(f"⚠️ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Vẫn còn nút next nhưng table không load, kết thúc phân trang để tránh hang")
                                                check_pages = False
                                        except Exception as check_e:
                                            logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Lỗi khi kiểm tra nút next sau retry: {check_e}")
                                            check_pages = False
                            else:
                                logger.info(f"🏁 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Không còn trang tiếp theo")
                                check_pages = False  # ✅ CHỈ set False khi không còn nút next
                            
                            # ✅ Log trạng thái sau khi xử lý pagination
                            logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Sau khi xử lý pagination: check_pages={check_pages}, page_num={page_num}")
                        except Exception as pagination_e:
                            logger.error(f"❌ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Trang {page_num}: Lỗi khi xử lý phân trang: {pagination_e}")
                            import traceback
                            logger.error(f"Traceback: {traceback.format_exc()}")
                            check_pages = False
                        
                        # ✅ Log trước khi tiếp tục vòng lặp (giống thông báo)
                        if check_pages:
                            logger.info(f"🔄 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Tiếp tục vòng lặp pagination, sẽ xử lý trang tiếp theo...")
                        else:
                            logger.info(f"🛑 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Dừng vòng lặp pagination, đã xử lý xong {page_num} trang")
                    
                    # ✅ Điều chỉnh accumulated_total_so_far sau khi download xong khoảng này
                    # Nếu đã cộng range_total_records ở trang đầu, nhưng số file thực sự download ít hơn (do duplicate),
                    # thì điều chỉnh lại accumulated_total_so_far
                    if range_total_records:
                        # Đã cộng range_total_records vào accumulated_total_so_far ở trang đầu
                        # Nhưng số file thực sự download là range_downloaded_so_far
                        # Điều chỉnh: accumulated_total_so_far = accumulated_total_so_far - range_total_records + range_downloaded_so_far
                        actual_files_downloaded = range_downloaded_so_far
                        if actual_files_downloaded < range_total_records:
                            # Có duplicate files, điều chỉnh accumulated_total_so_far
                            adjustment = range_total_records - actual_files_downloaded
                            accumulated_total_so_far -= adjustment
                            logger.info(f"📊 [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Điều chỉnh accumulated_total: -{adjustment} (duplicate files), actual={actual_files_downloaded}, expected={range_total_records}, accumulated_total_so_far={accumulated_total_so_far}")
                    
                    logger.info(f"✅ [GIAYNOPTIEN] [{range_idx + 1}/{len(date_ranges)}] Hoàn thành xử lý khoảng {date_range[0]} - {date_range[1]}: Tổng {range_downloaded_so_far if range_total_records else total_count} items, Accumulated %: {accumulated_percent_so_far:.2f}%")
                
                except Exception as e:
                    logger.error(f"Error processing date range {date_range}: {e}")
                    import traceback
                    logger.error(f"Traceback: {traceback.format_exc()}")
                    
                    # ✅ Screenshot khi có lỗi (lưu vào D:\tool-gotax\tool-gotax\tool-go-soft\screenshots)
                    try:
                        # Đảm bảo đường dẫn đúng: tool-go-soft/screenshots/giaynoptien_...
                        screenshot_dir = get_screenshot_dir()
                        logger.info(f"📸 Screenshot directory: {screenshot_dir}")
                        
                        # Screenshot page (nếu có)
                        if 'page' in locals() and page:
                            try:
                                page_screenshot = os.path.join(screenshot_dir, f"01_error_page_range_{range_idx + 1}.png")
                                await page.screenshot(path=page_screenshot, full_page=True)
                                logger.info(f"📸 Screenshot page saved: {page_screenshot}")
                            except Exception as page_e:
                                logger.warning(f"⚠️ Cannot screenshot page: {page_e}")
                        
                        # Screenshot frame (nếu có)
                        if 'frame' in locals() and frame:
                            try:
                                frame_screenshot = os.path.join(screenshot_dir, f"02_error_frame_range_{range_idx + 1}.png")
                                await frame.screenshot(path=frame_screenshot, full_page=True)
                                logger.info(f"📸 Screenshot frame saved: {frame_screenshot}")
                            except Exception as frame_e:
                                logger.warning(f"⚠️ Cannot screenshot frame: {frame_e}")
                            
                            # Lấy HTML của frame để debug
                            try:
                                frame_html = await frame.content()
                                html_file = os.path.join(screenshot_dir, f"03_error_frame_range_{range_idx + 1}.html")
                                with open(html_file, 'w', encoding='utf-8') as f:
                                    f.write(frame_html)
                                logger.info(f"📄 Frame HTML saved: {html_file}")
                            except Exception as html_e:
                                logger.warning(f"⚠️ Cannot save frame HTML: {html_e}")
                        
                        # Lấy HTML của page để debug (nếu có)
                        if 'page' in locals() and page:
                            try:
                                page_html = await page.content()
                                html_file = os.path.join(screenshot_dir, f"04_error_page_range_{range_idx + 1}.html")
                                with open(html_file, 'w', encoding='utf-8') as f:
                                    f.write(page_html)
                                logger.info(f"📄 Page HTML saved: {html_file}")
                            except Exception as html_e:
                                logger.warning(f"⚠️ Cannot save page HTML: {html_e}")
                        
                        logger.info(f"📸 Screenshots saved to: {screenshot_dir}")
                    except Exception as screenshot_e:
                        logger.error(f"❌ Error taking screenshot: {screenshot_e}")
                        import traceback
                        logger.error(f"Screenshot error traceback: {traceback.format_exc()}")
                    
                    yield {
                        "type": "warning", 
                        "message": f"Lỗi xử lý khoảng {date_range}: {str(e)}",
                        "accumulated_total": accumulated_total_so_far,
                        "accumulated_downloaded": accumulated_downloaded_so_far,
                        "accumulated_percent": int(round(accumulated_percent_so_far))
                    }
                    continue
            
            # Parse downloaded files và rename
            parsed_results = []
            files_in_temp_dir = os.listdir(temp_dir) if os.path.exists(temp_dir) else []
            logger.info(f"crawl_giay_nop_tien: Found {len(files_in_temp_dir)} files in temp_dir")
            
            # ✅ Không log screenshots khi không có lỗi (chỉ log khi có lỗi thực sự)
            
            if files_in_temp_dir:
                nnn = 0
                
                for file_name in files_in_temp_dir:
                    file_path = os.path.join(temp_dir, file_name)
                    if not os.path.isfile(file_path):
                        continue
                    
                    try:
                        # Parse XML để lấy thông tin
                        with open(file_path, 'r', encoding='utf-8') as f:
                            soup = BeautifulSoup(f, 'html.parser')
                        
                        ma_ndkt = soup.find('ma_ndkt')
                        ma_ndkt = ma_ndkt.text if ma_ndkt else ""
                        
                        ngay_lap = soup.find('ngay_lap')
                        ngay_lap = ngay_lap.text if ngay_lap else ""
                        ngay_lap = ngay_lap.replace("/", "-")
                        
                        ma_chuong = soup.find('ma_chuong')
                        ma_chuong = ma_chuong.text if ma_chuong else ""
                        
                        ky_thue = soup.find('ky_thue')
                        ky_thue = ky_thue.text if ky_thue else ""
                        ky_thue = ky_thue.replace("/", "-")
                        
                        # Rename file theo format
                        nnn += 1
                        new_file_name = f"{ma_ndkt} - {ma_chuong} - Kynopthue - {ky_thue} - Ngaynopthue - {ngay_lap} [{nnn}].xml"
                        new_file_name = self._remove_accents(new_file_name)
                        
                        # Rename file
                        new_file_path = os.path.join(temp_dir, new_file_name)
                        if os.path.exists(file_path):
                            try:
                                os.rename(file_path, new_file_path)
                                file_name = new_file_name
                                file_path = new_file_path
                            except Exception as rename_err:
                                logger.warning(f"Error renaming {file_name}: {rename_err}")
                        
                        parsed_results.append({
                            "ma_noi_dung_kinh_te": ma_ndkt,
                            "ngay_lap": ngay_lap.replace("-", "/") if ngay_lap else "",  # Trả về format gốc
                            "ma_chuong": ma_chuong,
                            "ky_thue": ky_thue.replace("-", "/") if ky_thue else ""  # Trả về format gốc
                        })
                        
                        file_size = os.path.getsize(file_path)
                        total_size += file_size
                        files_info.append({"name": file_name, "size": file_size})
                    except Exception as e:
                        logger.warning(f"Error parsing/renaming file {file_name}: {e}")
                        # Nếu parse lỗi, vẫn thêm vào files_info với tên cũ
                        try:
                            file_size = os.path.getsize(file_path)
                            total_size += file_size
                            files_info.append({"name": file_name, "size": file_size})
                        except:
                            pass
                        continue
                
                download_id = str(uuid.uuid4())
                zip_filename = f"giaynoptien_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                zip_file_path = os.path.join(self.ZIP_STORAGE_DIR, f"{download_id}.zip")
                
                final_files = os.listdir(temp_dir)
                logger.info(f"crawl_giay_nop_tien: Found {len(final_files)} files in temp_dir")
                logger.info(f"crawl_giay_nop_tien: Creating ZIP from {len(final_files)} files")
                
                if final_files:
                    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for file_name in final_files:
                            file_path = os.path.join(temp_dir, file_name)
                            if os.path.isfile(file_path):
                                zf.write(file_path, file_name)
                                logger.debug(f"Added to ZIP: {file_name}")
                    
                    with open(zip_file_path, 'rb') as f:
                        zip_base64 = base64.b64encode(f.read()).decode('utf-8')
                    
                    logger.info(f"✅ Đã tạo file ZIP: {zip_filename} (download_id: {download_id})")
                    
                    try:
                        from shared.redis_client import get_redis_client
                        redis_client = get_redis_client()
                        redis_key = f"session:{session_id}:download_id"
                        redis_client.setex(redis_key, 3600, download_id.encode('utf-8'))
                    except Exception as redis_err:
                        logger.warning(f"⚠️ Không thể lưu download_id vào Redis: {redis_err}")
                else:
                    zip_base64 = None
                    download_id = None
                    logger.warning("crawl_giay_nop_tien: No files to add to ZIP")
            else:
                zip_base64 = None
                download_id = None
                zip_filename = f"giaynoptien_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
                logger.warning("crawl_giay_nop_tien: No files in temp_dir")
            
            actual_files_count = len(files_info)
            actual_results_count = len(parsed_results)
            
            yield {
                "type": "complete",
                "total": actual_files_count,  # Số file thực tế trong ZIP
                "results_count": actual_results_count,  # Số items đã parse
                "total_rows_processed": total_count,  # Số rows đã xử lý (để debug)
                "results": parsed_results,
                "files": files_info,
                "files_count": actual_files_count,
                "total_size": total_size,
                "zip_base64": zip_base64,
                "zip_filename": zip_filename,
                "download_id": download_id
            }
            
        except Exception as e:
            logger.error(f"Error in crawl_giay_nop_tien: {e}")
            error_msg = str(e)
            # Kiểm tra session timeout
            if "timeout" in error_msg.lower() or "phiên giao dịch" in error_msg.lower():
                yield {"type": "error", "error": "Phiên giao dịch hết hạn. Vui lòng đăng nhập lại.", "error_code": "SESSION_EXPIRED"}
            else:
                yield {"type": "error", "error": f"Lỗi khi tra cứu giấy nộp tiền: {error_msg}", "error_code": "CRAWL_ERROR"}
        
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
    
    _gnt_download_counter = 0
    
    async def _download_single_giaynoptien(self, session: SessionData, item: Dict, temp_dir: str, max_retries: int = 2) -> bool:
        page = session.page
        id_gnt = item["id"]
        
        for retry in range(max_retries + 1):
            try:
                # Nếu đã có link_locator, dùng trực tiếp
                if "link_locator" in item:
                    download_link = item["link_locator"]
                else:
                    # Fallback: tìm lại link từ row và col_index
                    row = item.get("row")
                    col_idx = item.get("col_index")
                    if row and col_idx is not None:
                        cols = row.locator('td')
                        links = cols.nth(col_idx).locator('a[href*="downloadGNT"]')
                        link_count = await links.count()
                        
                        # Nếu có 2 links thì click link thứ 2, nếu không thì click link đầu
                        if link_count >= 2:
                            download_link = links.nth(1)
                        elif link_count >= 1:
                            download_link = links.first
                        else:
                            logger.warning(f"No download link found for {id_gnt}")
                            return False
                    else:
                        logger.warning(f"Missing link_locator or row/col_index for {id_gnt}")
                        return False
                
                # Download file
                async with page.expect_download(timeout=30000) as download_info:
                    await download_link.click()
                
                download = await download_info.value
                
                # Lưu file với tên tạm unique
                TaxCrawlerService._gnt_download_counter += 1
                temp_name = f"chungtu_{id_gnt}_{TaxCrawlerService._gnt_download_counter}.xml"
                save_path = os.path.join(temp_dir, temp_name)
                await download.save_as(save_path)
                
                # Verify file exists and has content
                await asyncio.sleep(0.3)
                if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                    logger.info(f"Downloaded giaynoptien {id_gnt} -> {temp_name}")
                    return True
                else:
                    raise Exception("File empty or not saved")
                    
            except Exception as e:
                logger.warning(f"Error downloading giaynoptien {id_gnt} (attempt {retry + 1}/{max_retries + 1}): {e}")
                if retry < max_retries:
                    await asyncio.sleep(1)  # Wait before retry
        
        return False
    
    async def convert_xml_to_xlsx(self, xml_files_base64: str) -> Dict[str, Any]:
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Giải nén ZIP
            zip_bytes = base64.b64decode(xml_files_base64)
            zip_buffer = BytesIO(zip_bytes)
            
            with zipfile.ZipFile(zip_buffer, 'r') as zf:
                zf.extractall(temp_dir)
            
            # Tạo workbook
            workbook = Workbook()
            worksheet = workbook.active
            
            # Headers
            headers = [
                'Tên', 'Kỳ tính thuế Tháng/Quý', 'Lần', 'Năm',
                'VAT đầu kỳ', 'Giá trị HH mua vào', 'VAT mua vào',
                'VAT được khấu trừ kỳ này', 'Giá trị HH bán ra', 'VAT bán ra',
                'Điều chỉnh tăng', 'Điều chỉnh giảm', 'Thuế vãng lai ngoại tỉnh',
                'VAT còn phải nộp', 'VAT còn được khấu trừ chuyển kỳ sau'
            ]
            worksheet.append(headers)
            
            # Parse each XML file
            for filename in os.listdir(temp_dir):
                if not filename.endswith('.xml'):
                    continue
                
                file_path = os.path.join(temp_dir, filename)
                
                try:
                    tree = ET.parse(file_path)
                    root = tree.getroot()
                    
                    # Get namespace
                    namespace = {'ns0': root.tag.split('}')[0][1:]} if '}' in root.tag else {}
                    
                    # Extract data
                    def get_element_text(tag):
                        if namespace:
                            elem = root.find(f'.//ns0:{tag}', namespace)
                        else:
                            elem = root.find(f'.//{tag}')
                        return elem.text if elem is not None else ''
                    
                    ky_kkhai = get_element_text('kyKKhai')
                    ky = ky_kkhai.split("/")[0] if "/" in ky_kkhai else ''
                    nam = ky_kkhai.split("/")[1] if "/" in ky_kkhai else ''
                    
                    try:
                        so_lan = filename.split("-")[2] + " " + filename.split("-")[3]
                    except:
                        so_lan = ""
                    
                    row = [
                        filename,
                        ky,
                        so_lan,
                        nam,
                        get_element_text('ct22'),
                        get_element_text('ct23'),
                        get_element_text('ct24'),
                        get_element_text('ct25'),
                        get_element_text('ct34'),
                        get_element_text('ct35'),
                        get_element_text('ct38'),
                        get_element_text('ct37'),
                        get_element_text('ct39'),
                        get_element_text('ct40'),
                        get_element_text('ct43'),
                    ]
                    
                    worksheet.append(row)
                    
                except Exception as e:
                    logger.error(f"Error parsing XML {filename}: {e}")
                    continue
            
            # Format worksheet
            header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.border = thin_border
            
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    if col >= 5:
                        if cell.value:
                            try:
                                cell.value = float(cell.value)
                                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                            except ValueError:
                                pass
            
            for col in range(1, worksheet.max_column + 1):
                header_text = worksheet.cell(row=1, column=col).value
                worksheet.column_dimensions[get_column_letter(col)].width = len(str(header_text)) + 5
            
            xlsx_buffer = BytesIO()
            workbook.save(xlsx_buffer)
            xlsx_base64 = base64.b64encode(xlsx_buffer.getvalue()).decode('utf-8')
            
            return {
                "success": True,
                "xlsx_base64": xlsx_base64,
                "row_count": worksheet.max_row - 1
            }
            
        except Exception as e:
            logger.error(f"Error in convert_xml_to_xlsx: {e}")
            return {"success": False, "error": str(e)}
        
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
    
    async def get_tokhai_types(self, session_id: str) -> Dict[str, Any]:
        session = self.session_manager.get_session(session_id)
        if not session:
            return {"success": False, "error": "Session not found"}
        
        if not session.is_logged_in:
            return {"success": False, "error": "Not logged in"}
        
        page = session.page
        
        try:
            # Navigate đến trang tra cứu tờ khai bằng JavaScript (nhanh hơn)
            success = await self._navigate_to_tokhai_page(page, session.dse_session_id)
            
            if not success:
                return {"success": False, "error": "Không thể navigate đến trang tra cứu. Vui lòng thử lại."}
            
            frame = page.frame('mainframe')
            if not frame:
                return {"success": False, "error": "Không tìm thấy mainframe"}
            
            # Tìm dropdown id="maTKhai"
            select = frame.locator('#maTKhai')
            await select.wait_for(timeout=10000)
            
            options = await select.locator('option').all()
            tokhai_types = []
            
            # Thêm option "Tất cả" vào đầu danh sách
            tokhai_types.append({
                "value": "00",
                "label": "--Tất cả--"
            })
            
            for option in options:
                value = await option.get_attribute('value')
                text = await option.text_content()
                # Bỏ qua header groups (value="--") và "Tất cả" (value="00") vì đã thêm ở trên
                if value and value not in ['--', '00'] and text:
                    tokhai_types.append({
                        "value": value,
                        "label": text.strip()
                    })
            
            return {
                "success": True,
                "tokhai_types": tokhai_types
            }
            
        except Exception as e:
            logger.error(f"Error getting tokhai types: {e}")
            return {"success": False, "error": str(e)}
    
    async def crawl_batch(
        self,
        session_id: str,
        start_date: str,
        end_date: str,
        crawl_types: List[str],
        tokhai_type: str = "00"
    ) -> AsyncGenerator[Dict[str, Any], None]:
        """
        Crawl nhiều loại dữ liệu đồng thời (tờ khai, thông báo, giấy nộp tiền)
        
        Args:
            session_id: Session ID đã đăng nhập
            start_date: Ngày bắt đầu (dd/mm/yyyy)
            end_date: Ngày kết thúc (dd/mm/yyyy)
            crawl_types: Danh sách loại cần crawl ["tokhai", "thongbao", "giaynoptien"]
            tokhai_type: Loại tờ khai (chỉ áp dụng nếu crawl tokhai)
        
        Yields:
            Dict với progress và kết quả từng loại
        """
        session = self.session_manager.get_session(session_id)
        if not session:
            yield {"type": "error", "error": "Session không tồn tại hoặc đã hết hạn", "error_code": "SESSION_NOT_FOUND"}
            return
        
        if not session.is_logged_in:
            yield {"type": "error", "error": "Chưa đăng nhập. Vui lòng đăng nhập lại.", "error_code": "NOT_LOGGED_IN"}
            return
        
        # Validate crawl_types
        valid_types = ["tokhai", "thongbao", "giaynoptien"]
        crawl_types = [t for t in crawl_types if t in valid_types]
        
        if not crawl_types:
            yield {"type": "error", "error": "Không có loại crawl hợp lệ. Chọn từ: tokhai, thongbao, giaynoptien", "error_code": "INVALID_CRAWL_TYPES"}
            return
        
        total_types = len(crawl_types)
        yield {
            "type": "batch_start",
            "message": f"Bắt đầu crawl {total_types} loại dữ liệu: {', '.join(crawl_types)}",
            "crawl_types": crawl_types,
            "total_types": total_types
        }
        
        # Kết quả tổng hợp
        batch_results = {
            "tokhai": None,
            "thongbao": None,
            "giaynoptien": None
        }
        
        # Xử lý từng loại tuần tự (vì cùng dùng 1 session/page)
        for idx, crawl_type in enumerate(crawl_types):
            yield {
                "type": "batch_progress",
                "current_type": crawl_type,
                "type_index": idx + 1,
                "total_types": total_types,
                "message": f"Đang crawl {crawl_type} ({idx + 1}/{total_types})..."
            }
            
            try:
                if crawl_type == "tokhai":
                    # Crawl tờ khai - thứ tự: session_id, tokhai_type, start_date, end_date
                    async for result in self.crawl_tokhai(session_id, tokhai_type, start_date, end_date):
                        # Forward progress events với prefix
                        if result.get("type") == "complete":
                            batch_results["tokhai"] = result
                            yield {
                                "type": "type_complete",
                                "crawl_type": "tokhai",
                                "result": result
                            }
                        elif result.get("type") == "zip_data":
                            # Lưu zip_data vào batch_results
                            if batch_results.get("tokhai"):
                                batch_results["tokhai"]["zip_base64"] = result.get("zip_base64")
                            # Forward event
                            yield {
                                **result,
                                "crawl_type": "tokhai"
                            }
                        elif result.get("type") == "error":
                            yield {
                                "type": "type_error",
                                "crawl_type": "tokhai",
                                "error": result.get("error")
                            }
                        else:
                            # Forward info/progress events
                            yield {
                                **result,
                                "crawl_type": "tokhai"
                            }
                
                elif crawl_type == "thongbao":
                    # Crawl thông báo
                    async for result in self.crawl_thongbao(session_id, start_date, end_date):
                        if result.get("type") == "complete":
                            batch_results["thongbao"] = result
                            # Nếu có zip_base64 trong complete event, giữ lại
                            yield {
                                "type": "type_complete",
                                "crawl_type": "thongbao",
                                "result": result
                            }
                        elif result.get("type") == "zip_data":
                            if batch_results.get("thongbao"):
                                batch_results["thongbao"]["zip_base64"] = result.get("zip_base64")
                            yield {
                                **result,
                                "crawl_type": "thongbao"
                            }
                        elif result.get("type") == "error":
                            yield {
                                "type": "type_error",
                                "crawl_type": "thongbao",
                                "error": result.get("error")
                            }
                        else:
                            yield {
                                **result,
                                "crawl_type": "thongbao"
                            }
                
                elif crawl_type == "giaynoptien":
                    # Crawl giấy nộp tiền
                    async for result in self.crawl_giay_nop_tien(session_id, start_date, end_date):
                        if result.get("type") == "complete":
                            batch_results["giaynoptien"] = result
                            yield {
                                "type": "type_complete",
                                "crawl_type": "giaynoptien",
                                "result": result
                            }
                        elif result.get("type") == "zip_data":
                            if batch_results.get("giaynoptien"):
                                batch_results["giaynoptien"]["zip_base64"] = result.get("zip_base64")
                            yield {
                                **result,
                                "crawl_type": "giaynoptien"
                            }
                        elif result.get("type") == "error":
                            yield {
                                "type": "type_error",
                                "crawl_type": "giaynoptien",
                                "error": result.get("error")
                            }
                        else:
                            yield {
                                **result,
                                "crawl_type": "giaynoptien"
                            }
                
            except Exception as e:
                logger.error(f"Error crawling {crawl_type}: {e}")
                yield {
                    "type": "type_error",
                    "crawl_type": crawl_type,
                    "error": str(e)
                }
        
        # Tổng hợp kết quả cuối cùng
        # Merge tất cả ZIP files thành 1 ZIP duy nhất
        merged_zip_buffer = BytesIO()
        total_files = 0
        total_size = 0
        all_results = []
        
        with zipfile.ZipFile(merged_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as merged_zip:
            for crawl_type, result in batch_results.items():
                if result and result.get("zip_base64"):
                    try:
                        # Decode ZIP của từng loại
                        type_zip_bytes = base64.b64decode(result["zip_base64"])
                        type_zip_buffer = BytesIO(type_zip_bytes)
                        
                        with zipfile.ZipFile(type_zip_buffer, 'r') as type_zip:
                            for file_info in type_zip.filelist:
                                # Thêm prefix folder theo loại
                                new_name = f"{crawl_type}/{file_info.filename}"
                                file_data = type_zip.read(file_info.filename)
                                merged_zip.writestr(new_name, file_data)
                                total_files += 1
                                total_size += len(file_data)
                        
                        # Collect results
                        if result.get("results"):
                            for r in result["results"]:
                                r["crawl_type"] = crawl_type
                                all_results.append(r)
                    except Exception as e:
                        logger.warning(f"Error merging ZIP for {crawl_type}: {e}")
        
        # Encode merged ZIP
        merged_zip_base64 = base64.b64encode(merged_zip_buffer.getvalue()).decode('utf-8') if total_files > 0 else None
        zip_filename = f"batch_crawl_{start_date.replace('/', '')}_{end_date.replace('/', '')}.zip"
        
        yield {
            "type": "batch_complete",
            "message": f"Hoàn thành crawl {total_types} loại dữ liệu",
            "total_files": total_files,
            "total_size": total_size,
            "results": all_results,
            "batch_results": {
                crawl_type: {
                    "total": result.get("total", 0) if result else 0,
                    "files_count": result.get("files_count", 0) if result else 0,
                    "total_size": result.get("total_size", 0) if result else 0,
                    "zip_base64": result.get("zip_base64") if result else None,
                    "zip_filename": result.get("zip_filename") if result else None,
                    "results": result.get("results", []) if result else []
                }
                for crawl_type, result in batch_results.items()
                if crawl_type in crawl_types
            },
            "zip_base64": merged_zip_base64,
            "zip_filename": zip_filename
        }
    
    async def _extract_pagination_info(self, frame) -> Optional[Dict[str, int]]:
        """
        Extract pagination info từ giấy nộp tiền page.
        Format: "Trang 1/<b>2</b>. Có <b>11</b> bản ghi."
        Returns: {"current_page": 1, "total_pages": 2, "total_records": 11} hoặc None
        """
        try:
            # Tìm pagination div: id="currAcc" với class "table_headerto"
            pagination_div = frame.locator('#currAcc.table_headerto, #currAcc, .table_headerto')
            if await pagination_div.count() == 0:
                return None
            
            pagination_text = await pagination_div.text_content()
            if not pagination_text:
                return None
            
            # Parse: "Trang 1/<b>2</b>. Có <b>11</b> bản ghi."
            # Hoặc: "Trang 1/2. Có 11 bản ghi."
            # Tìm "Trang X/Y" hoặc "Trang X/<b>Y</b>"
            page_match = re.search(r'Trang\s+(\d+)\s*/\s*(?:<b>)?(\d+)(?:</b>)?', pagination_text)
            if not page_match:
                return None
            
            current_page = int(page_match.group(1))
            total_pages = int(page_match.group(2))
            
            # Tìm "Có X bản ghi" hoặc "Có <b>X</b> bản ghi"
            records_match = re.search(r'Có\s+(?:<b>)?(\d+)(?:</b>)?\s+bản ghi', pagination_text)
            total_records = int(records_match.group(1)) if records_match else 0
            
            return {
                "current_page": current_page,
                "total_pages": total_pages,
                "total_records": total_records
            }
        except Exception as e:
            logger.warning(f"Error extracting pagination info: {e}")
            return None
    
    async def _navigate_to_page(self, frame, page_num: int) -> bool:
        """
        Navigate đến trang page_num của giấy nộp tiền.
        Có thể dùng link hoặc JavaScript gotoPage().
        """
        try:
            # Thử click vào link số trang trước (nếu có)
            # Link format: <a href="...&pn=2">2</a>
            page_link = frame.locator(f'a[href*="pn={page_num}"]:has-text("{page_num}")')
            if await page_link.count() > 0:
                await page_link.first.click()
                await asyncio.sleep(1)
                
                # Verify navigation: check xem có đúng trang không
                pagination_info = await self._extract_pagination_info(frame)
                if pagination_info and pagination_info["current_page"] == page_num:
                    logger.info(f"✅ Navigated to page {page_num} via link")
                    return True
                else:
                    logger.warning(f"⚠️ Navigation verification failed: expected page {page_num}, got {pagination_info.get('current_page') if pagination_info else 'unknown'}")
            
            # Nếu link không work, thử dùng JavaScript gotoPage()
            try:
                # Tìm input field: id="gotoPageNO_objectList"
                goto_input = frame.locator('#gotoPageNO_objectList')
                if await goto_input.count() > 0:
                    # Fill page number
                    await goto_input.fill(str(page_num))
                    await asyncio.sleep(0.3)
                    
                    # Click nút "go" (img với src="/etaxnnt/static/images/pagination_go.gif")
                    go_btn = frame.locator('a[href*="gotoPage"] img[src*="pagination_go"], a:has(img[src*="pagination_go"])')
                    if await go_btn.count() > 0:
                        await go_btn.first.click()
                        await asyncio.sleep(1)
                        
                        # Verify navigation
                        pagination_info = await self._extract_pagination_info(frame)
                        if pagination_info and pagination_info["current_page"] == page_num:
                            logger.info(f"✅ Navigated to page {page_num} via JavaScript gotoPage")
                            return True
            except Exception as js_e:
                logger.debug(f"JavaScript gotoPage failed: {js_e}")
            
            # Nếu cả 2 cách đều không work, thử click vào nút "next" (pagination_right.gif) nhiều lần
            # Nhưng cách này không chính xác, chỉ dùng khi không có cách nào khác
            current_page = 1
            pagination_info = await self._extract_pagination_info(frame)
            if pagination_info:
                current_page = pagination_info["current_page"]
            
            if current_page < page_num:
                # Click nút "next" (pagination_right.gif) cho đến khi đến đúng trang
                next_btn = frame.locator('a[href*="pn="] img[src*="pagination_right"], a:has(img[src*="pagination_right"])')
                clicks_needed = page_num - current_page
                for _ in range(min(clicks_needed, 10)):  # Giới hạn tối đa 10 lần click
                    if await next_btn.count() > 0:
                        await next_btn.first.click()
                        await asyncio.sleep(1)
                        
                        # Check xem đã đến đúng trang chưa
                        pagination_info = await self._extract_pagination_info(frame)
                        if pagination_info and pagination_info["current_page"] == page_num:
                            logger.info(f"✅ Navigated to page {page_num} via next button")
                            return True
                        elif pagination_info and pagination_info["current_page"] > page_num:
                            # Đã vượt quá trang cần đến
                            break
                    else:
                        break
                
                # Verify sau khi click
                pagination_info = await self._extract_pagination_info(frame)
                if pagination_info and pagination_info["current_page"] == page_num:
                    return True
            
            logger.warning(f"⚠️ Cannot navigate to page {page_num}")
            return False
            
        except Exception as e:
            logger.error(f"Error navigating to page {page_num}: {e}")
            return False
    
    async def _download_single_giaynoptien(self, session: SessionData, item: Dict, temp_dir: str, max_retries: int = 2) -> bool:
        """
        Download 1 file giấy nộp tiền với retry logic (giống thông báo)
        
        Args:
            session: SessionData object
            item: Dict chứa thông tin file cần download (id, download_link, cols, col_index)
            temp_dir: Thư mục tạm để lưu file
            max_retries: Số lần retry tối đa
        
        Returns:
            True nếu download thành công
        """
        page = session.page
        id_gnt = item["id"]
        file_name = item.get("file_name", f"chungtu_{id_gnt}")
        
        for retry in range(max_retries + 1):
            try:
                # Ưu tiên dùng download_link đã tìm sẵn
                download_link = item.get("download_link")
                
                if not download_link:
                    # Fallback: tìm lại từ cols
                    cols = item.get("cols")
                    col_idx = item.get("col_index", 18)
                    if cols:
                        download_link = cols.nth(col_idx).locator('a[href*="downloadGNT"], a[onclick*="downloadGNT"]')
                
                if download_link and await download_link.count() > 0:
                    async with page.expect_download(timeout=30000) as download_info:
                        await download_link.first.click()
                    
                    download = await download_info.value
                    save_path = os.path.join(temp_dir, file_name + ".xml" if not file_name.endswith(".xml") else file_name)
                    await download.save_as(save_path)
                    
                    # Verify file exists and has content
                    if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
                        logger.info(f"Downloaded giaynoptien {id_gnt} -> {file_name}")
                        return True
                    else:
                        raise Exception("File empty or not saved")
                else:
                    logger.warning(f"No download link for giaynoptien {id_gnt}")
                    return False
                    
            except Exception as e:
                logger.warning(f"Error downloading giaynoptien {id_gnt} (attempt {retry + 1}/{max_retries + 1}): {e}")
                if retry < max_retries:
                    await asyncio.sleep(1)  # Wait before retry
        
        return False


# Singleton instance - sẽ được khởi tạo với session_manager
_tax_crawler_instance = None

def get_tax_crawler() -> TaxCrawlerService:
    global _tax_crawler_instance
    if _tax_crawler_instance is None:
        from .session_manager import session_manager
        _tax_crawler_instance = TaxCrawlerService(session_manager)
    return _tax_crawler_instance

# Backwards compatibility
tax_crawler = None  # Will be lazy-initialized
