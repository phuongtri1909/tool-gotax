"""
Routes cho tool-gobot
Được gọi từ api_server.py chung

Luồng xử lý:
1. POST /api/go-bot/lookup: Tra cứu đồng bộ (list taxcodes + type_taxcode)
2. POST /api/go-bot/lookup/queue: Nhận job_id + params, chạy lookup trong background, ghi progress/result vào Redis
"""

import os
import sys
import json
import time
import traceback
import threading
import logging
import uuid as _uuid
from flask import request, jsonify, Response

# Quart (api_server): dùng Quart request + trả response bằng Response(json.dumps(...)) để không cần app context.
try:
    from quart import request as quart_request, Response as QuartResponse
    QUART_AVAILABLE = True
except ImportError:
    quart_request = None
    QuartResponse = None
    QUART_AVAILABLE = False

# Thêm parent directory vào path để import backend
_api_dir = os.path.dirname(os.path.abspath(__file__))
_tool_root = os.path.dirname(_api_dir)
sys.path.insert(0, _tool_root)

# Gốc dự án = tool-gotax BÊN TRONG (chứa shared/redis_client.py, api_server, run_all). Folder ngoài cùng chỉ là nơi gom.
_gotax_root = os.path.dirname(_tool_root) if os.path.basename(_tool_root) == 'toolgobot' else _tool_root
if _gotax_root not in sys.path:
    sys.path.insert(0, _gotax_root)

# MOVED to register_routes(): from backend_.backend_service import BackendService
# Reason: Lazy-load onnxruntime DLL only when route is registered (Windows DLL safety)

try:
    from shared.redis_client import get_redis_client, publish_progress
except ImportError:
    get_redis_client = None
    publish_progress = None

logger = logging.getLogger(__name__)

# Model cache
_gobot_cache = {
    'backend': None,
    'lock': None
}

def get_gobot_backend(proxy_url=None):
    """Lazy load BackendService instance - tạo mới nếu có proxy"""
    global _gobot_cache
    
    # Nếu có proxy, luôn tạo instance mới để sử dụng proxy đó
    if proxy_url:
        print(f"🔄 Đang khởi tạo BackendService với proxy...")
        try:
            return BackendService(proxy_url=proxy_url)
        except Exception as e:
            print(f"⚠️  Cảnh báo: BackendService khởi tạo có vấn đề: {e}")
            return BackendService(proxy_url=proxy_url)
    
    # Nếu không có proxy, sử dụng cache singleton
    if _gobot_cache['backend'] is not None:
        return _gobot_cache['backend']
    
    print("🔄 Đang khởi tạo BackendService...")
    try:
        _gobot_cache['backend'] = BackendService()
        print("✅ BackendService đã được khởi tạo")
    except Exception as e:
        print(f"⚠️  Cảnh báo: BackendService khởi tạo có vấn đề: {e}")
        _gobot_cache['backend'] = BackendService()
    
    return _gobot_cache['backend']

def detect_id_type(value: str, provided_type: str = None) -> str:
    """
    Nhận dạng loại ID từ độ dài của giá trị
    
    Nếu provided_type được cung cấp, sử dụng giá trị đó.
    Nếu không, tự động detect dựa vào độ dài:
    - 9 ký tự: CMT (Chứng minh thư)
    - 10 ký tự: MST (Mã số thuế)
    - 12 ký tự: CCCD (Căn cước công dân)
    
    Args:
        value: Chuỗi ID
        provided_type: Loại ID được cung cấp từ request ('cmt', 'mst', 'cccd')
                      Nếu None, tự động detect từ độ dài
    
    Returns:
        'CMT', 'MST', hoặc 'CCCD'
    """
    # Nếu được cung cấp, sử dụng giá trị đó (chuyển thành uppercase)
    if provided_type:
        type_upper = provided_type.upper()
        if type_upper in ['CMT', 'MST', 'CCCD']:
            return type_upper
    
    # Nếu không cung cấp, detect dựa vào độ dài
    length = len(value.strip())
    if length == 9:
        return "CMT"
    elif length == 12:
        return "CCCD"
    elif length == 10:
        return "MST"
    else:
        # Default to MST nếu không match
        return "MST"


# Go-Bot dùng subprocess (run_lookup_standalone.py) vì API server chạy nhiều tool: go-invoice và go-bot
# đều có package "backend_" → nếu chạy lookup trong process như go-invoice/go-soft thì Python cache
# backend_ từ go-invoice, Go-Bot thiếu BaseServiceCMT. Các tool kia progress trong routes vì không bị trùng package.
_RUN_LOOKUP_SCRIPT = os.path.join(_tool_root, "api", "run_lookup_standalone.py")
_LOOKUP_TIMEOUT = int(os.getenv("GOBOT_LOOKUP_TIMEOUT", "900"))  # seconds (default 15 min; tang neu tra cuu cham/TensorFlow load)


def _run_lookup_job(job_id, taxcodes, type_taxcode, id_type, proxy):
    """
    Chạy lookup trong subprocess (run_lookup_standalone.py). Progress do script đó gửi Redis.
    """
    import subprocess
    import tempfile
    redis_client = None
    if get_redis_client:
        try:
            redis_client = get_redis_client()
        except Exception as e:
            logger.error(f"[Job {job_id}] Redis connect error: {e}")
    if not redis_client:
        logger.error(f"[Job {job_id}] Redis client not available")
        return
    try:
        redis_client.set(f"job:{job_id}:status", "processing".encode("utf-8"))
        if publish_progress:
            publish_progress(job_id, 0, "Bắt đầu tra cứu...", data={"total": len(taxcodes), "processed": 0})
    except Exception as e:
        logger.error(f"[Job {job_id}] Redis set status error: {e}")
        return

    if not os.path.isfile(_RUN_LOOKUP_SCRIPT):
        err = f"Go-Bot standalone script not found: {_RUN_LOOKUP_SCRIPT}"
        logger.error(f"[Job {job_id}] {err}")
        try:
            redis_client.set(f"job:{job_id}:status", b"failed")
            redis_client.set(f"job:{job_id}:error", err.encode("utf-8"))
            if publish_progress:
                publish_progress(job_id, 0, f"Lỗi: {err}", data={"type": "error", "error": err})
        except Exception:
            pass
        return

    params = {
        "job_id": job_id,
        "taxcodes": taxcodes,
        "type_taxcode": type_taxcode,
        "id_type": id_type,
        "proxy": proxy,
    }
    gotax_root = os.path.normpath(os.path.abspath(_gotax_root))
    gobot_root = os.path.normpath(os.path.abspath(_tool_root))
    env = os.environ.copy()
    env["GOTAX_ROOT"] = gotax_root
    env["GOBOT_ROOT"] = gobot_root
    env["PYTHONIOENCODING"] = "utf-8"  # Tránh UnicodeEncodeError khi subprocess print/log emoji trên Windows
    env["PYTHONUNBUFFERED"] = "1"  # Subprocess print ra ngay de API server doc real-time
    fd, temp_path = None, None
    try:
        fd, temp_path = tempfile.mkstemp(suffix=".json", prefix="gobot_")
        os.write(fd, json.dumps(params, ensure_ascii=False).encode("utf-8"))
        os.close(fd)
        fd = None
        proc = subprocess.Popen(
            [sys.executable, _RUN_LOOKUP_SCRIPT, temp_path, gotax_root, gobot_root],
            cwd=gotax_root,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        # Doc stdout/stderr real-time de biet subprocess dang o dau (tranh treo 600s khong ro ly do)
        out_lines, err_lines = [], []

        def read_stream(stream, lines_list, prefix):
            try:
                for line in iter(stream.readline, b""):
                    if line:
                        decoded = line.decode("utf-8", errors="replace").strip()
                        lines_list.append(decoded)
                        logger.info(f"[Job {job_id}] [{prefix}] {decoded}")
            except Exception as e:
                logger.debug(f"[Job {job_id}] stream read error: {e}")
            finally:
                try:
                    stream.close()
                except Exception:
                    pass

        t_out = threading.Thread(target=read_stream, args=(proc.stdout, out_lines, "out"), daemon=True)
        t_err = threading.Thread(target=read_stream, args=(proc.stderr, err_lines, "err"), daemon=True)
        t_out.start()
        t_err.start()

        waited = 0
        while proc.poll() is None and waited < _LOOKUP_TIMEOUT:
            time.sleep(10)
            waited += 10
            logger.info(f"[Job {job_id}] Dang xu ly... (da doi {waited}s)")
        if proc.poll() is None:
            proc.kill()
            proc.wait(timeout=5)
            raise subprocess.TimeoutExpired(proc.args, _LOOKUP_TIMEOUT)
        t_out.join(timeout=2)
        t_err.join(timeout=2)
        proc_stdout = "\n".join(out_lines)
        proc_stderr = "\n".join(err_lines)
        if proc.returncode != 0:
            err = (proc_stderr or "").strip() or f"Subprocess exit code {proc.returncode}"
            logger.error(f"[Job {job_id}] Lookup subprocess failed: {err}")
            try:
                redis_client.set(f"job:{job_id}:status", b"failed")
                redis_client.set(f"job:{job_id}:error", err.encode("utf-8"))
                if publish_progress:
                    publish_progress(job_id, 0, f"Lỗi: {err}", data={"type": "error", "error": err})
            except Exception:
                pass
    except subprocess.TimeoutExpired:
        err = f"Lookup timeout after {_LOOKUP_TIMEOUT}s"
        logger.error(f"[Job {job_id}] {err}")
        try:
            redis_client.set(f"job:{job_id}:status", b"failed")
            redis_client.set(f"job:{job_id}:error", err.encode("utf-8"))
            if publish_progress:
                publish_progress(job_id, 0, f"Lỗi: {err}", data={"type": "error", "error": err})
        except Exception:
            pass
    except Exception as e:
        err_msg = str(e)
        logger.exception(f"[Job {job_id}] Lookup failed: {err_msg}")
        try:
            redis_client.set(f"job:{job_id}:status", b"failed")
            redis_client.set(f"job:{job_id}:error", err_msg.encode("utf-8"))
            if publish_progress:
                publish_progress(job_id, 0, f"Lỗi: {err_msg}", data={"type": "error", "error": err_msg})
        except Exception:
            pass
    finally:
        if fd is not None:
            try:
                os.close(fd)
            except Exception:
                pass
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


def register_routes(app, prefix):
    """
    Đăng ký routes cho tool-gobot
    
    Args:
        app: Flask app instance
        prefix: URL prefix (ví dụ: '/api/go-bot')
    """
    # Lazy-load BackendService only when routes are registered (Windows DLL safety)
    from toolgobot.backend_.backend_service import BackendService
    
    @app.route(f'{prefix}/health', methods=['GET'])
    def go_bot_health_check():
        """Health check cho tool này"""
        return jsonify({
            "status": "success",
            "message": "Go Bot API is running",
            "version": "1.0"
        })
    
    @app.route(f'{prefix}/lookup', methods=['POST'])
    def go_bot_lookup():
        """
        Tra cứu thông tin từ danh sách mã số thuế
        
        Request JSON:
        {
            "taxcodes": ["0311111111", "123456789", "123456789012"],
            "type_taxcode": "cn" or "dn",
            "id_type": "cmt" or "mst" or "cccd" (optional, default "mst"),
            "proxy": "http://proxy:port" (optional)
        }
        
        Returns:
        {
            "status": "success",
            "data": {...}
        }
        """
        try:
            if not request.is_json:
                return jsonify({
                    "status": "error",
                    "message": "Request must be JSON"
                }), 400
            
            data = request.get_json()
            
            # Validate required fields
            taxcodes = data.get("taxcodes")
            type_taxcode = data.get("type_taxcode")
            id_type = data.get("id_type", "mst")  # Default to "mst" nếu không cung cấp
            proxy = data.get("proxy")
            
            if not taxcodes:
                return jsonify({
                    "status": "error",
                    "message": "Missing 'taxcodes' in request"
                }), 400
            
            if not isinstance(taxcodes, list):
                return jsonify({
                    "status": "error",
                    "message": "'taxcodes' must be a list"
                }), 400
            
            if len(taxcodes) == 0:
                return jsonify({
                    "status": "error",
                    "message": "'taxcodes' list cannot be empty"
                }), 400
            
            if not type_taxcode:
                return jsonify({
                    "status": "error",
                    "message": "Missing 'type_taxcode' in request"
                }), 400
            
            if type_taxcode not in ['cn', 'dn']:
                return jsonify({
                    "status": "error",
                    "message": "'type_taxcode' must be 'cn' (cá nhân) or 'dn' (doanh nghiệp)"
                }), 400
            
            # Validate id_type nếu được cung cấp
            if id_type and id_type.lower() not in ['cmt', 'mst', 'cccd']:
                return jsonify({
                    "status": "error",
                    "message": "'id_type' must be 'cmt', 'mst', or 'cccd' (case-insensitive)"
                }), 400
            
            # Nhận dạng ID type từ taxcode đầu tiên hoặc từ request
            first_taxcode = taxcodes[0].strip()
            detected_id_type = detect_id_type(first_taxcode, id_type)
            
            if type_taxcode == 'dn' and detected_id_type != "MST":
                return jsonify({
                    "status": "error",
                    "message": f"For business (dn), ID type must be MST, but got {detected_id_type}"
                }), 400
            if type_taxcode == 'cn' and detected_id_type not in ["CMT", "CCCD", "MST"]:
                return jsonify({
                    "status": "error",
                    "message": f"For individual (cn), ID type must be CMT, CCCD or MST, but got {detected_id_type}"
                }), 400
            
            # Sync lookup: chạy subprocess (cùng script queue) rồi đọc result từ Redis
            import subprocess
            import tempfile
            import uuid
            sync_job_id = str(uuid.uuid4())
            params = {"job_id": sync_job_id, "taxcodes": taxcodes, "type_taxcode": type_taxcode, "id_type": id_type, "proxy": proxy}
            gotax_root = os.path.normpath(os.path.abspath(_gotax_root))
            gobot_root = os.path.normpath(os.path.abspath(_tool_root))
            sync_env = os.environ.copy()
            sync_env["GOTAX_ROOT"] = gotax_root
            sync_env["GOBOT_ROOT"] = gobot_root
            sync_env["PYTHONIOENCODING"] = "utf-8"
            if not os.path.isfile(_RUN_LOOKUP_SCRIPT):
                return jsonify({"status": "error", "message": "Go-Bot standalone script not found"}), 500
            fd, temp_path = None, None
            try:
                fd, temp_path = tempfile.mkstemp(suffix=".json", prefix="gobot_sync_")
                os.write(fd, json.dumps(params, ensure_ascii=False).encode("utf-8"))
                os.close(fd)
                fd = None
                proc = subprocess.run(
                    [sys.executable, _RUN_LOOKUP_SCRIPT, temp_path, gotax_root, gobot_root],
                    cwd=gotax_root,
                    timeout=_LOOKUP_TIMEOUT,
                    env=sync_env,
                    capture_output=True,
                )
                if get_redis_client:
                    rc = get_redis_client()
                    status = (rc.get(f"job:{sync_job_id}:status") or b"").decode("utf-8")
                    if status == "completed":
                        raw = rc.get(f"job:{sync_job_id}:result")
                        result = json.loads((raw or b"{}").decode("utf-8")) if raw else {}
                        return jsonify({"status": "success", "data": result}), 200
                    err = (rc.get(f"job:{sync_job_id}:error") or b"").decode("utf-8") or (proc.stderr or b"").decode("utf-8", errors="replace")
                    return jsonify({"status": "error", "message": err or f"Subprocess exit {proc.returncode}"}), 500
                return jsonify({"status": "error", "message": "Redis not available"}), 500
            except subprocess.TimeoutExpired:
                return jsonify({"status": "error", "message": f"Lookup timeout after {_LOOKUP_TIMEOUT}s"}), 504
            finally:
                if fd is not None:
                    try:
                        os.close(fd)
                    except Exception:
                        pass
                if temp_path and os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except Exception:
                        pass
            
        except Exception as e:
            return jsonify({
                "status": "error",
                "message": str(e),
                "detail": traceback.format_exc() if app.config.get('DEBUG') else None
            }), 500

    @app.route(f'{prefix}/lookup/upload', methods=['POST'])
    async def go_bot_lookup_upload():
        """
        Tra hàng loạt: Upload file txt hoặc xlsx, mỗi dòng 1 mã số thuế.
        Form: file, type_taxcode (cn|dn), id_type (cmt|mst|cccd, optional cho cn)
        Trả về job_id để poll progress/result.
        """
        def _parse_txt(content):
            lines = content.decode("utf-8", errors="replace").strip().splitlines()
            return [ln.strip() for ln in lines if ln.strip()]

        def _parse_xlsx(content):
            try:
                from openpyxl import load_workbook
                from io import BytesIO
                wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                codes = []
                for row in ws.iter_rows(min_row=1, values_only=True):
                    val = row[0] if row else None
                    if val is not None:
                        s = str(val).strip()
                        if s:
                            codes.append(s)
                wb.close()
                return codes
            except Exception as e:
                logger.warning("parse xlsx error: %s", e)
                return []

        def _json_response(obj, status=200):
            body = json.dumps(obj, ensure_ascii=False)
            if QUART_AVAILABLE and QuartResponse is not None:
                return QuartResponse(body, status=status, mimetype="application/json")
            return body, status, {"Content-Type": "application/json; charset=utf-8"}

        req = quart_request if QUART_AVAILABLE else request
        try:
            files = await req.files
            form = await req.form
            file = files.get("file") if files else None
            if not file or not (getattr(file, 'filename', None) or getattr(file, 'name', None)):
                return _json_response({"status": "error", "message": "Missing 'file' in form"}, 400)
            type_taxcode = (form.get("type_taxcode") or "").strip().lower()
            id_type = (form.get("id_type") or "").strip().lower() or None
            proxy = form.get("proxy") or None
            if type_taxcode not in ("cn", "dn"):
                return _json_response({"status": "error", "message": "'type_taxcode' must be 'cn' or 'dn'"}, 400)
            _r = file.read()
            content = await _r if hasattr(_r, '__await__') else _r
            if isinstance(content, str):
                content = content.encode("utf-8")
            fn = (getattr(file, 'filename', '') or "").lower()
            if fn.endswith(".xlsx") or fn.endswith(".xls"):
                taxcodes = _parse_xlsx(content)
            else:
                taxcodes = _parse_txt(content)
            taxcodes = list(dict.fromkeys(t for t in taxcodes if t))
            if not taxcodes:
                return _json_response({"status": "error", "message": "File không có mã số thuế hợp lệ"}, 400)
            if type_taxcode == "dn":
                id_type = id_type or "mst"
            else:
                id_type = id_type or detect_id_type(taxcodes[0]).lower() if taxcodes else "mst"
            if not get_redis_client:
                return _json_response({"status": "error", "message": "Redis not available"}, 503)
            job_id = str(_uuid.uuid4())
            job_data = {
                "job_id": job_id,
                "params": {"taxcodes": taxcodes, "type_taxcode": type_taxcode, "id_type": id_type, "proxy": proxy},
            }
            redis_client = get_redis_client()
            redis_client.lpush("go-bot:jobs", json.dumps(job_data, ensure_ascii=False))
            return _json_response({"status": "accepted", "job_id": job_id, "total": len(taxcodes)}, 202)
        except Exception as e:
            logger.exception("go_bot_lookup_upload error")
            return _json_response({"status": "error", "message": str(e)}, 500)

    @app.route(f'{prefix}/lookup/queue', methods=['POST'])
    async def go_bot_lookup_queue():
        """
        Nhận job_id + params, chạy lookup trong background, ghi progress/result vào Redis.
        Trả về 202 Accepted ngay. Handler async + Quart request để có request context (sync view chạy trong thread → lỗi context).
        """
        def _json_response(obj, status=200):
            body = json.dumps(obj, ensure_ascii=False)
            if QUART_AVAILABLE and QuartResponse is not None:
                return QuartResponse(body, status=status, mimetype="application/json")
            return body, status, {"Content-Type": "application/json; charset=utf-8"}

        req = quart_request if QUART_AVAILABLE else request
        try:
            if not (req.content_type and "application/json" in req.content_type):
                return _json_response({"status": "error", "message": "Request must be JSON"}, 400)
            if QUART_AVAILABLE:
                data = await req.get_json(silent=True)
            else:
                data = req.get_json(silent=True)
            if data is None:
                return _json_response({"status": "error", "message": "Invalid or empty JSON"}, 400)
            job_id = data.get("job_id")
            taxcodes = data.get("taxcodes")
            type_taxcode = data.get("type_taxcode")
            id_type = data.get("id_type")
            proxy = data.get("proxy")
            if not job_id:
                return _json_response({"status": "error", "message": "Missing 'job_id'"}, 400)
            if not taxcodes:
                return _json_response({"status": "error", "message": "Missing 'taxcodes'"}, 400)
            if not isinstance(taxcodes, list) or len(taxcodes) == 0:
                return _json_response({"status": "error", "message": "'taxcodes' must be a non-empty list"}, 400)
            if not type_taxcode or type_taxcode not in ["cn", "dn"]:
                return _json_response({"status": "error", "message": "'type_taxcode' must be 'cn' or 'dn'"}, 400)
            if id_type and str(id_type).lower() not in ["cmt", "mst", "cccd"]:
                return _json_response({"status": "error", "message": "'id_type' must be 'cmt', 'mst', or 'cccd'"}, 400)
            if not get_redis_client:
                return _json_response({"status": "error", "message": "Redis not available for queue"}, 503)
            thread = threading.Thread(
                target=_run_lookup_job,
                args=(job_id, taxcodes, type_taxcode, id_type, proxy),
                daemon=True,
            )
            thread.start()
            return _json_response({"status": "accepted", "job_id": job_id}, 202)
        except Exception as e:
            logger.exception("go_bot_lookup_queue error")
            return _json_response({"status": "error", "message": str(e)}, 500)

    # ==================== DOWNLOAD ENDPOINT (Stream chunk 8KB – giai đoạn download client) ====================
    # ==================== DOWNLOAD ENDPOINT (Stream chunk 8KB – giai đoạn download client) ====================
    # ==================== DOWNLOAD ENDPOINT (Stream chunk 8KB – giai đoạn download client) ====================
    @app.route(f'{prefix}/download/<download_id>', methods=['GET'])
    async def go_bot_download(download_id: str):
        """
        Download file từ disk. Stream theo chunk 8KB để tránh load toàn bộ vào memory.
        """
        debug_info = {}
        try:
            # Manual import to avoid sys.path issues
            import importlib.util
            shared_dir = os.path.join(_tool_root, 'shared')
            ds_path = os.path.join(shared_dir, 'download_service.py')
            
            debug_info['ds_path_candidates'] = [ds_path]
            
            if os.path.exists(ds_path):
                spec = importlib.util.spec_from_file_location("download_service_manual", ds_path)
                ds_mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(ds_mod)
                get_file_path = ds_mod.get_file_path
                SERVICE_STORAGE_DIR = ds_mod.STORAGE_DIR
                debug_info['imported_via'] = 'manual_path'
            else:
                # Fallback to sys.path import
                if _tool_root not in sys.path:
                    sys.path.insert(0, _tool_root)
                from shared.download_service import get_file_path, STORAGE_DIR as SERVICE_STORAGE_DIR
                debug_info['imported_via'] = 'sys_path'
            
            # Use quart_request for async context
            args = {}
            if QUART_AVAILABLE and quart_request:
                args = quart_request.args
                ext = args.get('ext', 'xlsx')
                filename = args.get('filename', f'{download_id}.{ext}')
                debug_info['request_type'] = 'quart'
            else:
                args = request.args
                ext = args.get('ext', 'xlsx')
                filename = args.get('filename', f'{download_id}.{ext}')
                debug_info['request_type'] = 'flask'
            
            debug_info['args'] = dict(args)
            
            file_path = get_file_path(download_id, ext)
            debug_info['file_path'] = file_path
            debug_info['storage_dir'] = SERVICE_STORAGE_DIR
            debug_info['exists'] = os.path.exists(file_path) if file_path else False
            
            if not file_path or not os.path.exists(file_path):
                return QuartResponse(
                    json.dumps({
                        "status": "error", 
                        "message": f"File not found: {file_path}",
                        "debug": debug_info
                    }),
                    status=404,
                    content_type='application/json'
                )
            
            file_size = os.path.getsize(file_path)
            mime_types = {'zip': 'application/zip', 'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'pdf': 'application/pdf'}
            mime_type = mime_types.get(ext, 'application/octet-stream')
            chunk_size = 8192  # 8KB (giống Go Invoice)

            async def generate():
                with open(file_path, 'rb') as f:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk

            return QuartResponse(
                generate(),
                mimetype=mime_type,
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"',
                    'Content-Length': str(file_size),
                },
            )
        except Exception as e:
            logger.exception(f"go_bot_download error: {e}")
            return QuartResponse(
                json.dumps({
                    "status": "error", 
                    "message": str(e),
                    "debug": debug_info
                }),
                status=500,
                content_type='application/json'
            )



