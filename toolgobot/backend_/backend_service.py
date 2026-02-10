import requests,random,csv
import os,sys,time,shutil,io,base64,json
import zipfile
import logging
from bs4 import BeautifulSoup, Comment
from datetime import datetime, timedelta
from toolgobot.backend_.base_service import BaseService
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd
from playwright.sync_api import sync_playwright
from random_user_agent.user_agent import UserAgent
from random_user_agent.params import SoftwareName, OperatingSystem
from toolgobot.backend_.base_service import BaseServiceCMT
from toolgobot.backend_.getmst_info2 import process_tax_codes
# Logging trong base: file + stderr (de biet luong chay khi subprocess/queue)
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
_log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "logs")
if not os.path.exists(_log_dir):
    os.makedirs(_log_dir)
_log_file = os.path.join(_log_dir, f"gobot_{datetime.now().strftime('%Y%m%d')}.log")
_fh = logging.FileHandler(_log_file, encoding="utf-8")
_fh.setLevel(logging.DEBUG)
_sh = logging.StreamHandler(sys.stderr)
_sh.setLevel(logging.INFO)
_fmt = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
_fh.setFormatter(_fmt)
_sh.setFormatter(_fmt)
logger.addHandler(_fh)
logger.addHandler(_sh)

# Đọc file ds canboqlt (nếu có); path theo thư mục tool-gobot để chạy đúng khi subprocess cwd khác
_canboqlt_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "__pycache__", "canboqlt.txt")
try:
    with open(os.path.normpath(_canboqlt_path), "r", encoding="utf-8") as f:
        ds_canbo = [line.strip() for line in f.readlines() if line.strip()]
except Exception:
    ds_canbo = []

class JobCancelledException(Exception):
    """Exception raised when a job is cancelled via Redis."""
    pass


class BackendService(BaseService):
    def __init__(self, proxy_url=None):
        super().__init__(proxy_url=proxy_url)
        self.proxy_url = proxy_url  # ✅ Lưu proxy URL để recreate session
        self.svCMT = BaseServiceCMT(proxy_url=proxy_url)
        self.size_res = 0  # ✅ Track bandwidth khi cào 1 MST
        self.current_mst = None  # ✅ MST hiện tại đang cào
        self._job_id = None  # ✅ Set từ run_lookup_standalone để check cancelled
        self._redis_client = None  # ✅ Redis client từ run_lookup_standalone

    def _check_cancelled(self):
        """Check if job is cancelled from Redis. Raise JobCancelledException if cancelled."""
        if not self._job_id or not self._redis_client:
            return False
        try:
            cancelled = self._redis_client.get(f"job:{self._job_id}:cancelled")
            if cancelled:
                c = cancelled.decode('utf-8') if isinstance(cancelled, bytes) else str(cancelled).strip()
                if c == '1':
                    logger.info("[Job %s] Job đã bị cancel (detected in BackendService)", self._job_id)
                    raise JobCancelledException(f"Job {self._job_id} đã bị hủy")
            status = self._redis_client.get(f"job:{self._job_id}:status")
            s = ""
            if status:
                s = status.decode('utf-8') if isinstance(status, bytes) else str(status).strip()
                if s == 'cancelled':
                    logger.info("[Job %s] Job status=cancelled (detected in BackendService)", self._job_id)
                    raise JobCancelledException(f"Job {self._job_id} đã bị hủy")
            if s == 'processing':
                import time as _time
                current = int(_time.time())
                last_poll_time = self._redis_client.get(f"job:{self._job_id}:last_poll_time")
                start_time_str = self._redis_client.get(f"job:{self._job_id}:start_time")
                start_time = None
                if start_time_str is not None:
                    start_time = int(start_time_str) if isinstance(start_time_str, bytes) else int(start_time_str)
                
                if last_poll_time is not None:
                    last_poll = int(last_poll_time) if isinstance(last_poll_time, bytes) else int(last_poll_time)
                    if start_time and abs(last_poll - start_time) < 2:
                        elapsed = current - start_time
                        if elapsed > 15:
                            logger.info("[Job %s] Client disconnect detected (no poll after job start, elapsed %ds)", self._job_id, elapsed)
                            self._redis_client.set(f"job:{self._job_id}:cancelled", "1")
                            self._redis_client.set(f"job:{self._job_id}:status", "cancelled")
                            raise JobCancelledException(f"Job {self._job_id} đã bị hủy (client disconnect - no poll)")
                    if current - last_poll > 12:
                        logger.info("[Job %s] Client disconnect detected (no poll in %ds)", self._job_id, current - last_poll)
                        self._redis_client.set(f"job:{self._job_id}:cancelled", "1")
                        self._redis_client.set(f"job:{self._job_id}:status", "cancelled")
                        raise JobCancelledException(f"Job {self._job_id} đã bị hủy (client disconnect)")
                else:
                    if start_time:
                        elapsed = current - start_time
                        if elapsed > 15:
                            logger.info("[Job %s] Client disconnect detected (no poll after %ds, job started %ds ago)", self._job_id, elapsed, elapsed)
                            self._redis_client.set(f"job:{self._job_id}:cancelled", "1")
                            self._redis_client.set(f"job:{self._job_id}:status", "cancelled")
                            raise JobCancelledException(f"Job {self._job_id} đã bị hủy (client disconnect - no poll)")
        except JobCancelledException:
            raise
        except Exception:
            pass
        return False
    
    def _recreate_session_with_new_proxy(self):
        """
        ✅ Tạo session mới + add proxy lại (IP tự đổi)
        Thay vì delay/backoff khi 429
        """
        logger.info("Recreating session + rotating proxy IP...")
        # Tạo session mới
        self.session = requests.Session()
        
        # Add proxy lại (Luna Proxy tự đổi IP)
        if self.proxy_url:
            self.session.proxies = {
                'http': self.proxy_url
            }
        
        return self.session
    
    def _format_bytes(self, bytes_val: int) -> str:
        """Format bytes thành KB/MB"""
        if bytes_val < 1024:
            return f"{bytes_val} B"
        elif bytes_val < 1024 * 1024:
            return f"{bytes_val / 1024:.2f} KB"
        else:
            return f"{bytes_val / (1024 * 1024):.2f} MB"
    def handle_request(self, request_data=None):
        request_data = request_data or {}
        self.lookup_data = request_data.get("type_data", "")  # 1:DN, 2:CN
        self.raw_data = request_data.get("raw_data", []) or []
        id_type_param = (request_data.get("id_type") or "").strip().upper()
        if id_type_param:
            self.type_id = {"CMT": "CMT", "CCCD": "CCCD", "MST": "MST", "CMND": "CMT"}.get(id_type_param) or id_type_param
        else:
            raw = self.check_id_type(self.raw_data[0] if self.raw_data else "")
            self.type_id = "CMT" if raw == "CMND" else raw
        logger.info("handle_request: type_data=%s, raw_data len=%s, type_id=%s", self.lookup_data, len(self.raw_data), self.type_id)
        try:
            if self.lookup_data == "1":
                return self.lookup_business(self.type_id,self.raw_data)
            elif self.lookup_data == "2":
                return self.lookup_individual(self.type_id,self.raw_data)
            else:
                return {"status":"error","message":"Loại tra cứu không hợp lệ."}
        except JobCancelledException as e:
            logger.info("Job cancelled: %s", e)
            return {"status": "error", "message": str(e)}
#================================================
    def fix_csv(self,input_file, output_file):
        import csv
        rows = []
        with open(input_file, 'r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            for row in reader:
                rows.append(row)
        max_length = max(len(row) for row in rows)
        for i, row in enumerate(rows):
            if len(row) < max_length:
                rows[i] = row + [''] * (max_length - len(row))
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file)
            writer.writerows(rows)
    def get_infoCMT(self,payload,cmt):
        session = payload["session"]
        headers = payload["headers"]
        dse_sessionId = payload["dse_sessionId"]
        dse_ProcessorId = payload["dse_processorId"]
        dse_PageId = payload["dse_pageId"]
        captcha = payload["captcha"]
        url = "https://canhantmdt.gdt.gov.vn/ICanhan/Request"
        payload_post = {
            "dse_sessionId": dse_sessionId,
            "dse_applicationId": "-1",
            "dse_operationName": "retailTraCuuMSTCNTMDTProc",
            "dse_pageId": dse_PageId ,
            "dse_processorState": "traCuuMSTViewPage",
            "dse_processorId": dse_ProcessorId,
            "dse_errorPage": "error_page.jsp",
            "dse_nextEventName": "ok",
            "loaiGiayTo": "1010" if len(cmt) == 9 else "2080",
            "_tin":"",
            "_cmnd": cmt,
            "capcha": captcha
        }
        r1 = session.post(url, data=payload_post, headers=headers, timeout=15)
        while 1:
            soup1 = BeautifulSoup(r1.text, "html.parser")
            error_td = soup1.find("td", {"colspan": "2", "align": "center"})
            pid = soup1.find("input", {"name": "dse_pageId"})["value"]
            if error_td and "Mã xác thực không đúng" in str(error_td):
                logger.warning("CMT captcha sai, thu lai 1")
                captcha_text = self.service_cmt.get_captcha(headers)
                payload_post["dse_pageId"] = pid    
                payload_post["capcha"] = captcha_text
                r1 = session.post(url, data=payload_post, headers=headers, timeout=15)
                continue
            break
        payload_post2 = payload_post
        payload_post2["dse_nextEventName"] = "view"
        payload_post2["dse_pageId"] = pid
        captcha_text = self.service_cmt.get_captcha(headers)
        payload_post2["capcha"] = captcha_text

        r2 = session.post(url, data=payload_post2, headers=headers, timeout=15)
        while 1:
            soup2 = BeautifulSoup(r2.text, "html.parser")
            # with open("t2_cmt.html", "w", encoding="utf-8") as f:
            #     f.write(r2.text)
            error_td = soup2.find("td", {"colspan": "2", "align": "center"})
            pid = soup2.find("input", {"name": "dse_pageId"})["value"]
            if error_td and "Mã xác thực không đúng" in str(error_td):
                logger.warning("CMT captcha sai, thu lai 2")
                captcha_text = self.service_cmt.get_captcha(headers)
                payload_post2["dse_pageId"] = pid    
                payload_post2["capcha"] = captcha_text
                r2 = session.post(url, data=payload_post2, headers=headers, timeout=15)
                continue
            break
        
        # Trích xuất dữ liệu từ bảng soup2
        tables = soup2.find_all('table', class_='confirm_tableTMDT')
        if len(tables) >= 0:
            rows = tables[0].find_all('tr')
            if len(rows) > 1:
                # Lấy tr thứ 2 (index 1)
                tr = rows[1]
                tds = tr.find_all('td')
                
                # Lấy từ td thứ 2 đến td thứ 7 (index 1 đến 6)
                if len(tds) >= 7:
                    info = {
                        "Mã số thuế": tds[1].text.strip(),
                        "Tên người nộp thuế": tds[2].text.strip(),
                        "Cơ quan thuế": tds[3].text.strip(),
                        "CCCD/CMT": tds[4].text.strip(),
                        "Ngày cấp": tds[5].text.strip(),
                        "Trạng thái": tds[6].text.strip()
                    }
                    return {
                        "status": "success",
                        "data": info
                    }
        
        return {
            "status": "error",
            "message": "Không tìm thấy dữ liệu trong bảng"
        }
    
    def check_nnt(self, cmt, url, type_id="", type_lookup="CN", row_index=None):
        # ✅ Check cancelled trước mỗi MST
        self._check_cancelled()

        try:
            r = self.session.get("https://api.ipify.org?format=json", timeout=10)
            logger.info("IP check: %s", (r.text[:80] if r.text else "?")[:80])
        except Exception as e:
            logger.debug("IP check skip: %s", e)
        # Reset size_res khi bắt đầu cào 1 MST mới
        
        self.current_mst = cmt
        logger.info("Dang tra cuu: mst=%s, loai=%s, id_type=%s", cmt, type_lookup, type_id)
        if type_id in ("CMT", "CCCD", "CMND"):
            # ✅ Tạo instance CMT (proxy được handle trong base_service)
            self.service_cmt = BaseServiceCMT()
            payload_data = self.service_cmt.get_dse()
            rs = self.get_infoCMT(payload_data,cmt)
            
            return rs
        # software_names = [SoftwareName.CHROME.value]
        # operating_systems = [OperatingSystem.WINDOWS.value, OperatingSystem.LINUX.value]   
        # user_agent_rotator = UserAgent(software_names=software_names, operating_systems=operating_systems, limit=100)
        #user_agent = user_agent_rotator.get_random_user_agent()    
        # Không delay khi get_captcha, cần gọi nhanh
        self._check_cancelled()
        logger.info("Lay captcha MST...")
        text, cookies = self.get_captcha()
        if text is None or not text.strip():
            err_msg = (
                f"Khong the giai captcha (TensorFlow loi tren may nay - DLL load failed). {text}"

            )
            logger.error("%s", err_msg)
            return {"status": "error", "message": err_msg}
        logger.info("Captcha xong, gui POST tra cuu...")
        data = {
                'cm': "cm",
                'mst': cmt,
                'fullname': '',
                'address': '',
                'cmt': '',
                'captcha': text
        }

        retry_count = 0
        max_retries = 5  # Giảm retry vì IP sẽ đổi

        while retry_count < max_retries:
                self._check_cancelled()
                logger.info("Attempt %s/%s POST tra cuu...", retry_count + 1, max_retries)
                try:
                    # ✅ Proxy đã được setup trong session
                    response = self.session.post(url, data=data, cookies=cookies, timeout=30)
                    
                    # ✅ Xử lý 429: Tạo session mới + rotate IP
                    if response.status_code == 429:
                        logger.warning("429 Too Many Requests - rotating proxy...")
                        self._recreate_session_with_new_proxy()
                        retry_count += 1
                        continue
                    
                    if response.status_code == 200:
                        logger.info("POST 200 OK")
                        response_size = len(response.content)
                        self.size_res += response_size
                        logger.info("Response size: %s (Total: %s)", self._format_bytes(response_size), self._format_bytes(self.size_res))
                        break
                    else:
                        logger.warning("Status %s - rotating proxy...", response.status_code)
                        self._recreate_session_with_new_proxy()
                        retry_count += 1
                        continue
                
                except Exception as e:
                    logger.warning("Exception: %s - rotating proxy...", e)
                    self._recreate_session_with_new_proxy()
                    retry_count += 1
        soup = BeautifulSoup(response.text, 'html.parser')
        self.html_buffer.write(response.text)  # ✅ Lưu vào memory thay vì file
        tables = soup.find_all('table', class_='ta_border')
        #==========================
        if len(tables) == 0:
            logger.warning("Sai captcha, retry tra cuu...")
            proxy_info = self.session.proxies
            self.session = requests.Session()  # Reset session để tránh cookie cũ
            self.session.proxies = proxy_info  # Giữ nguyên proxy nếu có
            
            return self.check_nnt(cmt=cmt, url=url, type_id=type_id, type_lookup=type_lookup, row_index=row_index)
        tds = tables[0].find_all('td')

        cmt = str(cmt)
        with open(self.risklist_file, 'r', encoding='utf-8') as file:
            self.lines = file.readlines()
        rr = " "
        for i in self.lines:
            if cmt in i:
                rr = "THUỘC NHÓM DN RỦI RO CAO VỀ THUẾ"
        self.cmt__ = cmt
        add_tax_info = []
        if type_lookup == "DN":
            try:    
                # Get td elements from index 2 to 5 (inclusive)
                csv_row = [td.text.strip("\n") for td in tds[1:6]]
                if len(tds) > 11:
                    if tds[7] == tds[1]:
                        add_tax_info.append([td.text.strip("\n") for td in tds[7:12]])
                    if len(tds) > 15:
                        if tds[13] == tds[1]:
                            add_tax_info.append([td.text.strip("\n") for td in tds[13:18]])
                if len(csv_row) < 5:
                    csv_row = [cmt] + ['CHƯA ĐĂNG KÝ']
                else:
                    csv_row.append(rr)
                    tax_codes = [cmt]
                    try:
                        result = process_tax_codes(tax_codes, self.proxy_url)[0]
                        if result.get('status_code') == 200 and result.get('company_data'):
                            company_data2 = result['company_data']
                            loai_hinh_dn = company_data2.get('company_type', '') or ''
                            nguoi_dai_dien_pl = company_data2.get('representative_name', '') or ''
                            nganh_nghe_chinh = company_data2.get('main_industry', '') or ''
                            ds_nganh_nghe_list = company_data2.get('industries_list', []) or []
                            
                            # Format industry list for Excel
                            ds_nganh_nghe_str = ""
                            if ds_nganh_nghe_list:
                                items = []
                                for item in ds_nganh_nghe_list:
                                    if isinstance(item, dict):
                                        code = str(item.get('code', '') or item.get('ma', ''))
                                        job = str(item.get('job', '') or item.get('ten', ''))
                                        if code and job:
                                            items.append(f"{code} - {job}")
                                        elif job:
                                            items.append(job)
                                    elif isinstance(item, str):
                                        items.append(item)
                                ds_nganh_nghe_str = "; ".join(items)

                            can_bo_qlt = ""
                            phone_canbo = ""
                            email_canbo = ""
                            for data in ds_canbo:
                                parts = (data or "").split('+')
                                if len(parts) >= 4 and cmt == parts[3].strip():
                                    can_bo_qlt = parts[0].strip() if len(parts) > 0 else ""
                                    phone_canbo = parts[1].strip() if len(parts) > 1 else ""
                                    email_canbo = parts[2].strip() if len(parts) > 2 else ""
                                    break
                            csv_row.extend([loai_hinh_dn, nguoi_dai_dien_pl, nganh_nghe_chinh, can_bo_qlt, phone_canbo, email_canbo, ds_nganh_nghe_str])
                            if row_index is not None:
                                self._industries_by_row[row_index] = ds_nganh_nghe_list
                        else:
                            logger.warning("process_tax_codes failed or empty for MST=%s (status=%s)", cmt, result.get('status_code'))
                            csv_row.extend([""] * 7)
                    except Exception as ext:
                        logger.warning("process_tax_codes error for MST=%s: %s", cmt, ext)
                        csv_row.extend([""] * 7)
            except Exception as e:
                logger.warning("Error processing td elements (DN): %s", e)
                if len(csv_row) == 6:
                    csv_row.extend([""] * 7)
            
        else:
            try:    
                csv_row = [td.text.strip("\n") for td in tds[1:5]]
                logger.debug("tds count=%s", len(tds))
                if len(tds) > 6:
                    if tds[6] == tds[1]:
                        add_tax_info.append([td.text.strip("\n") for td in tds[6:10]])
                    if len(tds) > 13:
                        if tds[11] == tds[1]:
                            add_tax_info.append([td.text.strip("\n") for td in tds[11:15]])
                if len(csv_row) < 3:
                    csv_row = [cmt] + ['CHƯA ĐĂNG KÝ']
            except Exception as e:
                logger.warning("Error processing td elements (CN): %s", e)
        logger.info("csv_row: %s", csv_row[:3] if len(csv_row) > 3 else csv_row)
        #trùng ?????? ++++++++++++++++++++++++++++++++++++++++++++++
        if 1:
            csv_content = self.csv_buffer.getvalue()
            line_count = len(csv_content.split('\n')) if csv_content else 1
            
            # Ghi dòng mới vào buffer (sử dụng @ thay vì ,)
            csv_row_str = '@'.join(str(x) for x in csv_row) + '\n'
            self.csv_buffer.write(csv_row_str)
            
            if len(add_tax_info) > 0:
                for add_info in add_tax_info:
                    line_count += 1
                    add_info.insert(0, line_count)
                    add_info[1] = "'" + str(add_info[1])
                    add_info_str = '@'.join(str(x) for x in add_info) + '\n'
                    self.csv_buffer.write(add_info_str)
        
        company_name = csv_row[1] if len(csv_row) > 1 else "Unknown"
        logger.info("Hoan thanh tra cuu MST: mst=%s, cong_ty=%s, dung_luong=%s", self.current_mst, company_name, self._format_bytes(self.size_res))
        self.size_res = 0  # Reset size_res sau mỗi MST 
        return None
    def convert(self, type_id="", type_lookup="CN"):
        # Template theo thu muc tool-gobot (cung cap base_temp_dir tu BaseService, da la path tuyet doi)
        if type_lookup == "DN":
            template_path = os.path.join(self.base_temp_dir, "maursdn.xlsx")
            output_path = 'Kết quả tra cứu MST DN.xlsx'
        elif type_lookup == "CN":
            template_path = os.path.join(self.base_temp_dir, "maurscn.xlsx")
            output_path = 'Kết quả tra cứu MST CN.xlsx'

        if not os.path.isfile(template_path):
            err = "Thieu file template: %s. Vui long them maursdn.xlsx hoac maurscn.xlsx vao thu muc tool-gobot/__pycache__/" % template_path
            logger.error("%s", err)
            return {"status": "error", "message": err}

        # ✅ Lấy dữ liệu từ in-memory buffer
        csv_content = self.csv_buffer.getvalue()
        from io import StringIO, BytesIO
        import base64
        import csv as csv_module
        
        csv_lines = csv_content.strip().split('\n')
        headers = []
        data_rows = []
        
        if csv_lines:
            # Dòng đầu = headers (delimiter: @)
            headers = csv_lines[0].split('@')
            
            # Dòng 2+ = values (delimiter: @)
            for line in csv_lines[1:]:
                if line.strip():
                    row = line.split('@')
                    data_rows.append(row)
        
        # ✅ Tạo JSON theo format: STT -> {field: value}
        looked_info = {}
        industries_by_row = getattr(self, '_industries_by_row', {})
        for row_idx, row in enumerate(data_rows, start=1):
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_dict[header] = row[col_idx]
            if type_lookup == "DN" and row_idx in industries_by_row:
                row_dict["industries_list"] = industries_by_row[row_idx]
            looked_info[row_idx] = row_dict
        
        # ✅ Tạo JSON response (ưu tiên download_id, fallback bytes_excel)
        json_response = {
            "status": "success",
            "looked_info": looked_info,
            "bytes_excel": None,
            "download_id": None,
            "excel_filename": output_path,
        }
        
        from openpyxl.styles import Border, Side, Alignment
        from datetime import datetime
        from io import BytesIO

        try:
            excel_buffer = BytesIO()
            if type_lookup == "DN":
                workbook = load_workbook(template_path)
                sheet = workbook.active
                # Cột "Danh sách ngành nghề" là cột thứ 13 (column 14 = N)
                ds_nganh_nghe_col = 14
                header_row_dn = 10
                for row in range(1, min(11, sheet.max_row + 1)):
                    found = False
                    for col in range(1, 14):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value and "EMAIL" in str(cell.value).upper():
                            header_row_dn = row
                            found = True
                            break
                    if found:
                        break
                for row_idx, row_data in enumerate(data_rows, start=11):
                    sheet.cell(row=row_idx, column=1, value=row_idx - 10)
                    for col_idx, value in enumerate(row_data, start=2):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                # Ghi header "Danh sách ngành nghề" vào cột N (template thường chỉ có 12 cột header)
                sheet.cell(row=header_row_dn, column=ds_nganh_nghe_col, value="Danh sách ngành nghề")
                now = datetime.now().strftime("%H:%M:%S %d/%m/%Y")
                sheet['D6'] = now
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
                alignment_wrap = Alignment(wrap_text=True, vertical='top')
                column_widths = {}
                # Xác định cột "Ngành nghề chính" (index 8 trong headers, column 10 trong Excel)
                nganh_nghe_col_idx = None
                # Xác định cột "Danh sách ngành nghề" (cột cuối cùng, index 12 trong headers, column 14 = N trong Excel)
                ds_nganh_nghe_col_idx = None
                if len(headers) > 8:
                    for idx, header in enumerate(headers):
                        if 'Ngành nghề chính' in header or 'ngành nghề chính' in header:
                            nganh_nghe_col_idx = idx + 2  # +2 vì column 1 là STT, column 2 là MSTDN
                        if 'Danh sách ngành nghề' in header or 'danh sách ngành nghề' in header:
                            ds_nganh_nghe_col_idx = idx + 2
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_len = len(str(cell.value))
                            column_letter = cell.column_letter
                            if column_letter not in column_widths or cell_len > column_widths[column_letter]:
                                column_widths[column_letter] = cell_len
                            cell.border = border
                            # Set wrap text cho cột "Ngành nghề chính" và "Danh sách ngành nghề"
                            if (nganh_nghe_col_idx and cell.column == nganh_nghe_col_idx) or \
                               (ds_nganh_nghe_col_idx and cell.column == ds_nganh_nghe_col_idx):
                                cell.alignment = alignment_wrap
                
                for column_letter, max_len in column_widths.items():
                    col_num = ord(column_letter) - ord('A') + 1
                    # Giới hạn độ rộng cột "Ngành nghề chính" (max 60 ký tự)
                    if nganh_nghe_col_idx and col_num == nganh_nghe_col_idx:
                        sheet.column_dimensions[column_letter].width = min(max_len + 2, 60)
                    # Giới hạn độ rộng cột "Danh sách ngành nghề" (max 80 ký tự vì text rất dài)
                    elif ds_nganh_nghe_col_idx and col_num == ds_nganh_nghe_col_idx:
                        sheet.column_dimensions[column_letter].width = min(max_len + 2, 80)
                    else:
                        sheet.column_dimensions[column_letter].width = max_len + 2
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
            else:
                workbook = load_workbook(template_path)
                sheet = workbook.active
                for row_idx, row_data in enumerate(data_rows, start=11):
                    sheet.cell(row=row_idx, column=1, value=row_idx - 10)
                    for col_idx, value in enumerate(row_data, start=2):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                now = datetime.now().strftime("%H:%M:%S %d/%m/%Y")
                sheet['D6'] = now
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
                alignment_wrap = Alignment(wrap_text=True, vertical='top')
                column_widths = {}
                # Xác định cột có text dài (nếu có) để set wrap text
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_len = len(str(cell.value))
                            column_letter = cell.column_letter
                            if column_letter not in column_widths or cell_len > column_widths[column_letter]:
                                column_widths[column_letter] = cell_len
                            cell.border = border
                            # Set wrap text cho các cell có text dài (> 50 ký tự)
                            if cell_len > 50:
                                cell.alignment = alignment_wrap
                
                for column_letter, max_len in column_widths.items():
                    # Giới hạn độ rộng tối đa 60 ký tự cho các cột có text dài
                    if max_len > 50:
                        sheet.column_dimensions[column_letter].width = min(max_len + 2, 60)
                    else:
                        sheet.column_dimensions[column_letter].width = max_len + 2
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
            excel_bytes = excel_buffer.getvalue()
            # ✅ Giai đoạn ghi: lưu file xuống disk theo chunk 8KB, trả download_id
            try:
                import sys
                _tool_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                if _tool_root not in sys.path:
                    sys.path.insert(0, _tool_root)
                from shared.download_service import save_file_to_disk
                download_id, _ = save_file_to_disk(excel_bytes, 'xlsx')
                json_response["download_id"] = download_id
                json_response["excel_filename"] = output_path
                # bytes_excel để None khi có download_id (client tải qua /download/{id}) — tránh lỗi JSON serializable
            except Exception as save_err:
                logger.warning("save_file_to_disk error (fallback bytes_excel): %s", save_err)
                json_response["bytes_excel"] = base64.b64encode(excel_bytes).decode('utf-8')
        except Exception as e:
            logger.warning("convert Excel error (still return looked_info): %s", e)
        return json_response  

    def lookup_individual(self,type_id = "", raw_data = []):
        """
        Tra cứu cá nhân (CN):
        - Luôn build CSV + Excel (maurscn.xlsx) để trả về file tải về (download_id) cho cả MST và CMT/CCCD.
        - looked_info được sinh từ CSV (giống DN) để frontend dùng chung.
        """
        from io import StringIO
        logger.info("lookup_individual: type_id=%s, so_luong=%s", type_id, len(raw_data or []))
        self.csv_buffer = StringIO()
        url = 'https://tracuunnt.gdt.gov.vn/tcnnt/mstcn.jsp'
        header_row = '@'.join(['MSTCN','Tên người nộp thuế', 'Cơ quan thuế', 'Trạng thái']) + '\n'
        self.csv_buffer.write(header_row)
        ds_cmt = {}
        step = 1
        for taxcode in raw_data:
            rs = self.check_nnt(cmt=taxcode, url=url, type_id=type_id, type_lookup="CN")
            if isinstance(rs, dict) and rs.get("status") == "error":
                return rs
            if rs is not None:  # tra CMT/CCCD trực tiếp
                ds_cmt[step] = rs["data"]
                step += 1

        # Ghi thêm dữ liệu ds_cmt vào csv_buffer để Excel có nội dung (CMT/CCCD)
        if ds_cmt:
            for _, data in ds_cmt.items():
                try:
                    mst = (data.get("Mã số thuế") or "").strip()
                    ten = (data.get("Tên người nộp thuế") or "").strip()
                    cqthue = (data.get("Cơ quan thuế") or "").strip()
                    trang_thai = (data.get("Trạng thái") or "").strip()
                    row = '@'.join([mst, ten, cqthue, trang_thai]) + '\n'
                    self.csv_buffer.write(row)
                except Exception:
                    continue

        # Trước đây: nếu type_id là CMT/CCCD/CMND thì chỉ trả JSON, không tạo Excel → không có download_id
        # Yêu cầu mới: tra hàng loạt cá nhân (dù nhập CMT/CCCD hay MST) đều phải trả file Excel để tải.
        resp = self.convert(type_id=type_id, type_lookup="CN")

        # Nếu cần giữ lại cấu trúc ds_cmt cho client cũ, có thể merge vào looked_info khi convert không có dữ liệu
        try:
            if isinstance(resp, dict) and resp.get("status") == "success":
                looked = resp.get("looked_info") or {}
                if not looked and ds_cmt:
                    resp["looked_info"] = ds_cmt
        except Exception:
            pass

        return resp

    def lookup_business(self, type_id="", raw_data=[]):
        from io import StringIO
        logger.info("lookup_business: type_id=%s, so_luong=%s", type_id, len(raw_data or []))
        self.csv_buffer = StringIO()
        self._industries_by_row = {}
        url = 'https://tracuunnt.gdt.gov.vn/tcnnt/mstdn.jsp'
        if type_id == "MST":
            header_row = '@'.join(['MSTDN','Tên người nộp thuế', 'Địa chỉ trụ sở', 'Cơ quan thuế', 'Trạng thái','Rủi ro','Loại hình doanh nghiệp','Người đại diện pháp luật','Ngành nghề chính','Cán bộ QLT','Điện thoại','Email','Danh sách ngành nghề']) + '\n'
            self.csv_buffer.write(header_row)
        for row_index, taxcode in enumerate(raw_data, start=1):
            self.size_res = 0
            self.check_nnt(cmt=taxcode, url=url, type_id=type_id, type_lookup="DN", row_index=row_index)
        return self.convert(type_id=type_id, type_lookup="DN")